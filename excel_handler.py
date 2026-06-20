import openpyxl
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime, date

logger = logging.getLogger(__name__)

STATUS_FILL_MAP = {
    "Pending":                              "FFF9C4",
    "Documents Verification In Progress":   "FFE0B2",
    "Document Required":                    "FFCC80",
    "Verified":                             "C8E6C9",
    "Approved":                             "B2DFDB",
    "Admission Confirmed":                  "69F0AE",
    "Admitted":                             "BBDEFB",
    "Rejected":                             "FFCDD2",
    "Fetch Error":                          "E0E0E0",
    "Not Found":                            "F8BBD0",
    "Unknown":                              "F5F5F5",
}

def get_fill(label):
    colour = STATUS_FILL_MAP.get(label, "F5F5F5")
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _clean(val):
    if not val:
        return ""
    return str(val).strip().lower().replace(".", "").replace("_", " ").replace("-", " ")

def _find_col(headers_clean, keywords):
    for kw in keywords:
        for i, h in enumerate(headers_clean):
            if kw in h:
                return i + 1
    return None

def detect_source(raw_headers):
    """Decide where a sheet came from by its header style.
    The MVS student portal exports camelCase headers (referenceNo, examSession,
    enrollmentNo, niosAdmissionStatus); manually-prepared sheets use spaced/UPPER
    headers (REFERENCE NUMBER, ADMISSION SESSION...). Returns 'mvs_portal' | 'mvs_tracker'."""
    blob = " ".join(str(h or "").strip().lower() for h in raw_headers)
    portal_sigs = ("referenceno", "examsession", "niosadmissionstatus",
                   "enrollmentno", "admissionstatus")
    return "mvs_portal" if any(sig in blob for sig in portal_sigs) else "mvs_tracker"

def read_students_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hc = [_clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    source = detect_source(raw_headers)

    ref_col     = _find_col(hc, ["reference number", "reference no", "ref no", "refno", "reference"])
    name_col    = _find_col(hc, ["student name", "name"])
    mobile_col  = _find_col(hc, ["mobile no", "mobile", "phone no", "phone"])
    alt_col     = _find_col(hc, ["alternate number", "alternate mobile", "alternate no",
                                 "alternate phone", "alt number", "alt mobile", "alt no",
                                 "second number", "2nd number", "other number", "whatsapp number"])
    class_col   = _find_col(hc, ["class"])
    email_col   = _find_col(hc, ["email"])
    dob_col     = _find_col(hc, ["date of birth", "dob", "d o b", "birth"])
    session_col = _find_col(hc, ["admission session", "session"])
    enroll_col  = _find_col(hc, ["enrollment number", "enrolment number", "enrollment no",
                                 "enrolment no", "enrol no", "enroll no", "enrol", "enrolment", "enrollment"])

    logger.info(f"Cols ref:{ref_col} name:{name_col} email:{email_col} dob:{dob_col} session:{session_col} enroll:{enroll_col}")

    if ref_col is None and email_col is None and enroll_col is None:
        wb.close()
        raise ValueError(f"Need 'Reference Number', 'Email' or 'Enrollment Number' column. Found: {hc}")

    def cell(row, col):
        if not col:
            return ""
        v = ws.cell(row, col).value
        if v is None:
            return ""
        # Excel often stores enrollment / mobile as numbers -> "220035253029.0".
        # Convert whole-number floats to plain integers so NIOS login doesn't break.
        if isinstance(v, float):
            return str(int(v)) if v.is_integer() else str(v)
        if isinstance(v, int):
            return str(v)
        if isinstance(v, (datetime, date)):
            return v.strftime("%d-%m-%Y")
        return str(v).strip()

    def session_cell(row, col):
        """Format session: dates become 'Month YYYY', text stays as-is."""
        if not col:
            return ""
        v = ws.cell(row, col).value
        if isinstance(v, (datetime, date)):
            return v.strftime("%B %Y")   # e.g. April 2027
        return str(v or "").strip()

    students = []
    for row in range(2, ws.max_row + 1):
        ref    = cell(row, ref_col)
        email  = cell(row, email_col)
        enroll = cell(row, enroll_col)
        if not ref and not email and not enroll:
            continue
        # Stable unique key. Reference number is NIOS's TRUE unique id, so it decides
        # identity FIRST — siblings who share an email/phone but have DIFFERENT reference
        # numbers are kept separate (never merged). Email is only a fallback when there is
        # no reference, and placeholder emails ("temp"/"na"/...) never become the key.
        _em = (email or "").strip().lower()
        _real_email = ("@" in _em and "." in _em.split("@")[-1]
                       and _em not in ("temp", "na", "none", "nil", "-", "null", "n/a"))
        if ref:
            rk = f"ref:{ref}"
        elif _real_email:
            rk = f"email:{_em}"
        elif enroll:
            rk = f"enr:{enroll}"
        else:
            rk = f"row:{source}:{row}"   # last resort: unique per row, never merged
        students.append({
            "row_index":    row,
            "reference_no": ref,
            "enrollment_no": enroll,
            "email":        email,
            "dob":          cell(row, dob_col),
            "student_name": cell(row, name_col),
            "mobile":       cell(row, mobile_col),
            "alt_mobile":   cell(row, alt_col),
            "class_level":  cell(row, class_col),
            "session":      session_cell(row, session_col),
            "source":       source,
            # Stable key: prefer email, then reference, then enrollment (SYC students).
            "row_key":      rk,
        })
    wb.close()
    # Secondary source signal: the MVS student portal exports DOB as a JavaScript
    # Date string ("Wed Aug 08 2007 12:30:00 GMT+0530 (India Standard Time)").
    # If the header style didn't flag the sheet but the DOBs look like that, it's
    # portal data — reclassify the whole sheet.
    if source != "mvs_portal":
        _jsd = re.compile(r"^[A-Za-z]{3}\s+[A-Za-z]{3}\s+\d{1,2}\s+\d{4}\b")
        if any(("gmt" in str(st.get("dob", "")).lower()) or _jsd.match(str(st.get("dob", "")))
               for st in students):
            source = "mvs_portal"
            for st in students:
                st["source"] = "mvs_portal"
    logger.info(f"Read {len(students)} students from Excel (source={source})")
    return students

def dedupe_students(students):
    """Remove duplicate rows (same row_key). Keeps the first occurrence.
    Returns (unique_list, duplicate_count)."""
    seen = set()
    unique = []
    dups = 0
    for s in students:
        k = s.get("row_key")
        if k in seen:
            dups += 1
            continue
        seen.add(k)
        unique.append(s)
    return unique, dups

def write_status_to_excel(filepath, updates):
    """updates: list of dicts with row_key/reference_no/email matching + status fields."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    hr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hc = [_clean(v) for v in hr]

    ref_col   = _find_col(hc, ["reference number", "reference no", "ref no", "refno", "reference"])
    email_col = _find_col(hc, ["email"])

    # Output columns — reuse the existing column whether the sheet uses the old
    # Excel names (Admission Status, Reference Number, ...) or the MVS student-portal
    # camelCase names (niosAdmissionStatus, referenceNo, examSession, ...).
    status_col = _ensure(ws, "Admission Status", ["nios admission status", "admission status", "status"])
    remark_col = _ensure(ws, "Remarks", ["remark"])
    idc_col    = _ensure(ws, "Download ID Card", ["download id card", "id card", "idcard", "icard"])
    app_col    = _ensure(ws, "Download Application Form", ["download application form", "application form",
                                                           "registration summary", "appform"])
    hall_col   = _ensure(ws, "Hall Ticket", ["hall ticket", "halltkt", "hallticket"])
    chk_col    = _ensure(ws, "Last Checked", ["last checked"])
    ref_col    = _find_col(hc, ["reference number", "reference no", "ref no", "refno", "reference"])
    email_col  = _find_col(hc, ["email"])

    # Build lookup by ref and by email
    by_ref   = {u["reference_no"]: u for u in updates if u.get("reference_no")}
    by_email = {u["email"]: u for u in updates if u.get("email")}

    for row in range(2, ws.max_row + 1):
        ref   = str(ws.cell(row, ref_col).value or "").strip() if ref_col else ""
        email = str(ws.cell(row, email_col).value or "").strip() if email_col else ""
        upd = by_ref.get(ref) or by_email.get(email)
        if not upd:
            continue

        now = upd.get("last_checked", datetime.now().strftime("%Y-%m-%d %H:%M"))
        label = upd.get("status_label", "Unknown")

        # If reference was discovered via email, write it back
        if upd.get("reference_no") and ref_col and not ref:
            rc = ws.cell(row, ref_col)
            rc.value = upd["reference_no"]
            rc.font = Font(bold=True, color="1565C0")

        sc = ws.cell(row, status_col)
        sc.value = label
        sc.fill = get_fill(label)
        sc.alignment = Alignment(horizontal="center")
        sc.border = thin_border()
        if upd.get("changed"):
            sc.font = Font(bold=True)

        if upd.get("remark"):
            rmc = ws.cell(row, remark_col)
            rmc.value = upd["remark"]
            rmc.alignment = Alignment(wrap_text=True)

        if upd.get("id_card_link"):
            ws.cell(row, idc_col).value = upd["id_card_link"]
        if upd.get("app_form_link"):
            ws.cell(row, app_col).value = upd["app_form_link"]
        if upd.get("hall_ticket_link"):
            ws.cell(row, hall_col).value = upd["hall_ticket_link"]

        cc = ws.cell(row, chk_col)
        cc.value = now
        cc.alignment = Alignment(horizontal="center")

    # Auto width
    for c in range(1, ws.max_column + 1):
        ml = 0
        cl = get_column_letter(c)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, c).value
            if v:
                ml = max(ml, len(str(v)))
        ws.column_dimensions[cl].width = min(ml + 3, 45)

    wb.save(filepath)
    wb.close()
    logger.info(f"Excel updated: {filepath}")

def _ensure(ws, name, aliases=None):
    """Return the column for `name`. First tries to find an existing column that
    matches the name OR any alias (ignoring case/spaces, so 'Admission Status' and
    'niosAdmissionStatus' are treated the same). Only creates a new column if none
    of the aliases are found — so MVS portal sheets aren't given duplicate columns."""
    cands = [name] + list(aliases or [])
    headers = [_clean(ws.cell(1, i).value).replace(" ", "") for i in range(1, ws.max_column + 1)]
    for cand in cands:
        cc = _clean(cand).replace(" ", "")
        if not cc:
            continue
        for i, h in enumerate(headers):
            if h and (cc in h or h in cc):
                return i + 1
    nc = ws.max_column + 1
    cell = ws.cell(1, nc)
    cell.value = name
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    return nc
