import os
import logging
import json
import threading as _threading
import time as _time
os.environ["TZ"] = "Asia/Kolkata"
try:
    _time.tzset()
except Exception:
    pass
from datetime import datetime, timedelta, timezone

from fastapi import FastAPI, HTTPException, Depends, BackgroundTasks, UploadFile, File, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from jose import JWTError, jwt
from passlib.context import CryptContext
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import aiofiles

from database import init_db, get_db, get_setting, set_setting
from job_runner import run_status_check

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger(__name__)

import secrets as _secrets
# SECRET_KEY signs the login tokens. NEVER fall back to a public hardcoded value (that
# would let anyone forge a valid token). If the env var is missing we generate a strong
# random key for this process — tokens then simply reset on restart (re-login needed).
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = _secrets.token_hex(32)
    logger.warning("SECRET_KEY not set in environment — using a random key for this run. "
                   "Set SECRET_KEY so logins survive restarts.")
PORTAL_USER = os.environ.get("PORTAL_USER", "admin")
PORTAL_PASS = os.environ.get("PORTAL_PASS", "MVS2025")
EXCEL_PATH  = os.environ.get("EXCEL_PATH",  os.path.join(os.environ.get("DATA_DIR", "."), "students.xlsx"))
ALGORITHM   = "HS256"
TOKEN_EXPIRE_HOURS = 12

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
bearer  = HTTPBearer()

app = FastAPI(title="NIOS Status Tracker", version="2.0.0")
# CORS: the portal UI is served from the SAME origin as the API, so cross-origin access
# isn't needed. Restrict to the known production domains (override with ALLOWED_ORIGINS)
# so other websites cannot call the data API from a browser.
_allowed = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if not _allowed:
    _allowed = ["https://status.mvsfoundation.in",
                "https://web-production-09671.up.railway.app"]
app.add_middleware(CORSMiddleware, allow_origins=_allowed,
                   allow_methods=["GET", "POST"], allow_headers=["*"],
                   allow_credentials=False)

PORTAL_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NIOS Status Tracker — MVS Foundation</title>
<style>
  :root{
    --primary:#4F46E5; --primary-dark:#4338CA; --primary-light:#EEF2FF;
    --sidebar:#1E293B; --sidebar-hover:#334155; --sidebar-text:#CBD5E1; --sidebar-muted:#64748B;
    --bg:#F1F5F9; --card:#FFFFFF; --border:#E2E8F0; --text:#0F172A;
    --muted:#64748B; --success:#16A34A; --danger:#DC2626; --warn:#EA580C;
    --th-bg:#F8FAFC; --row-hover:#FAFBFF; --chip:#F1F5F9; --soft:#F8FAFC; --dup-bg:#FFF7ED;
    --shadow:0 1px 3px rgba(0,0,0,.08),0 1px 2px rgba(0,0,0,.04);
    --shadow-lg:0 10px 25px rgba(0,0,0,.08);
  }
  html[data-theme="dark"]{
    --primary:#6366F1; --primary-dark:#4F46E5; --primary-light:#1E1B4B;
    --sidebar:#0B1220; --sidebar-hover:#1E293B; --sidebar-text:#CBD5E1; --sidebar-muted:#64748B;
    --bg:#0F172A; --card:#1E293B; --border:#334155; --text:#E2E8F0;
    --muted:#94A3B8; --success:#22C55E; --danger:#F87171; --warn:#FB923C;
    --th-bg:#172033; --row-hover:#243043; --chip:#334155; --soft:#172033; --dup-bg:#3a2a18;
    --shadow:0 1px 3px rgba(0,0,0,.4); --shadow-lg:0 10px 25px rgba(0,0,0,.5);
  }
  *{margin:0;padding:0;box-sizing:border-box}
  html,body{overflow-x:hidden;max-width:100%}
  body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}
  img{max-width:100%}
  a{text-decoration:none}
  ::-webkit-scrollbar{width:8px;height:8px}
  ::-webkit-scrollbar-thumb{background:#CBD5E1;border-radius:4px}

  #login-screen{position:fixed;inset:0;display:flex;align-items:center;justify-content:center;
    background:linear-gradient(135deg,#4F46E5 0%,#7C3AED 100%);z-index:1000;padding:16px}
  .login-card{background:var(--card);border-radius:18px;padding:42px;width:380px;max-width:100%;
    box-shadow:0 20px 60px rgba(0,0,0,.3)}
  @media(max-width:480px){.login-card{padding:28px 22px}}
  .login-card .logo{width:60px;height:60px;background:linear-gradient(135deg,#4F46E5,#7C3AED);
    border-radius:16px;display:flex;align-items:center;justify-content:center;font-size:28px;margin:0 auto 18px}
  .login-card h2{text-align:center;font-size:21px;margin-bottom:4px}
  .login-card p.sub{text-align:center;color:var(--muted);font-size:13px;margin-bottom:26px}
  .login-card label{display:block;font-size:13px;font-weight:600;margin-bottom:6px;color:#334155}
  .login-card input{width:100%;padding:12px 14px;border:2px solid var(--border);
    border-radius:10px;font-size:14px;margin-bottom:16px;transition:.2s}
  .login-card input:focus{outline:none;border-color:var(--primary)}
  .login-card button{width:100%;padding:13px;background:var(--primary);color:#fff;border:none;
    border-radius:10px;font-size:15px;font-weight:600;cursor:pointer;transition:.2s}
  .login-card button:hover{background:var(--primary-dark)}
  #login-error{color:var(--danger);font-size:13px;text-align:center;margin-top:12px;min-height:18px}

  #app{display:none;min-height:100vh}
  .sidebar{position:fixed;left:0;top:0;bottom:0;width:240px;background:var(--sidebar);
    padding:20px 0 12px;display:flex;flex-direction:column;z-index:100}
  .sidebar .brand{padding:0 22px 20px;display:flex;align-items:center;gap:11px;
    border-bottom:1px solid rgba(255,255,255,.08);margin-bottom:12px}
  .sidebar .brand .ic{width:38px;height:38px;background:linear-gradient(135deg,#4F46E5,#7C3AED);
    border-radius:10px;display:flex;align-items:center;justify-content:center}
  .sidebar .brand .ic svg{width:21px;height:21px}
  .sidebar .brand .tx b{color:#fff;font-size:15px;display:block}
  .sidebar .brand .tx span{color:#94A3B8;font-size:11px}
  .nav-scroll{flex:1;overflow-y:auto}
  .side-foot{border-top:1px solid rgba(255,255,255,.08);padding-top:8px;margin-top:6px}
  .nav-item{display:flex;align-items:center;gap:12px;padding:11px 22px;color:var(--sidebar-text);
    cursor:pointer;font-size:14px;font-weight:500;transition:.15s;border-left:3px solid transparent}
  .nav-item:hover{background:var(--sidebar-hover);color:#fff}
  .nav-item.active{background:var(--sidebar-hover);color:#fff;border-left-color:#818CF8;font-weight:600}
  .nav-item.active .ic{color:#A5B4FC}
  .nav-item.logout:hover{background:rgba(248,113,113,.14);color:#FCA5A5}
  .nav-item .ic{width:20px;height:20px;display:flex;align-items:center;justify-content:center;flex-shrink:0}
  .nav-item .ic svg{width:19px;height:19px}
  .nav-item .badge-count{margin-left:auto;background:var(--danger);color:#fff;
    font-size:11px;padding:1px 7px;border-radius:10px;font-weight:700;display:none}
  .nav-item .nav-num{margin-left:auto;background:rgba(255,255,255,.14);color:#fff;
    font-size:11px;padding:1px 8px;border-radius:10px;font-weight:700}
  .nav-item.active .nav-num{background:rgba(255,255,255,.22)}
  .nav-sep{padding:14px 22px 6px;color:var(--sidebar-muted);font-size:11px;font-weight:700;
    text-transform:uppercase;letter-spacing:.5px}

  .main{margin-left:240px;min-height:100vh}
  .topbar{background:var(--card);border-bottom:1px solid var(--border);padding:0 28px;height:64px;
    display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50}
  .topbar h1{font-size:19px;font-weight:700}
  .topbar .right{display:flex;align-items:center;gap:16px}

  .bell-wrap{position:relative;cursor:pointer}
  .bell-btn{width:42px;height:42px;border-radius:10px;background:var(--bg);display:flex;
    align-items:center;justify-content:center;font-size:19px;transition:.15s;border:1px solid var(--border)}
  .bell-btn:hover{background:var(--primary-light)}
  .bell-badge{position:absolute;top:-4px;right:-4px;background:var(--danger);color:#fff;
    font-size:10px;font-weight:700;min-width:18px;height:18px;border-radius:9px;
    display:none;align-items:center;justify-content:center;padding:0 4px;border:2px solid #fff}
  .bell-dropdown{position:absolute;right:0;top:52px;width:360px;max-height:440px;overflow-y:auto;
    background:var(--card);border:1px solid var(--border);border-radius:14px;box-shadow:var(--shadow-lg);
    display:none;z-index:200}
  .bell-dropdown.open{display:block}
  .bell-head{padding:14px 18px;border-bottom:1px solid var(--border);font-weight:700;
    font-size:14px;display:flex;align-items:center;justify-content:space-between}
  .bell-head .cnt{background:var(--warn);color:#fff;font-size:11px;padding:2px 9px;border-radius:10px}
  .notif-item{padding:13px 18px;border-bottom:1px solid var(--border);display:flex;gap:11px}
  .notif-item:last-child{border-bottom:none}
  .notif-item .nm{color:var(--text)}
  .notif-item .dot{width:9px;height:9px;border-radius:50%;background:var(--warn);margin-top:5px;flex-shrink:0}
  .notif-item .nm{font-weight:600;font-size:13px}
  .notif-item .rf{font-size:12px;color:var(--primary);font-family:monospace}
  .notif-item .rk{font-size:11px;color:var(--muted);margin-top:3px;line-height:1.4}
  .notif-empty{padding:34px 18px;text-align:center;color:var(--muted);font-size:13px}

  .content{padding:28px}
  .page-section{display:none}
  .page-section.active{display:block}

  .card{background:var(--card);border:1px solid var(--border);border-radius:14px;
    padding:22px;box-shadow:var(--shadow);margin-bottom:20px}
  .card h3{font-size:15px;margin-bottom:16px;font-weight:700}
  /* Settings — professional card headers with icon chips */
  .set-head{display:flex;align-items:flex-start;gap:13px;margin-bottom:18px}
  .set-ico{width:40px;height:40px;border-radius:11px;display:flex;align-items:center;
    justify-content:center;color:#fff;flex-shrink:0;box-shadow:0 3px 9px rgba(15,23,42,.12)}
  .set-ico svg{width:20px;height:20px}
  .set-head .set-tt{flex:1;min-width:0}
  .set-head h3{font-size:15.5px;font-weight:700;margin:0;line-height:1.3}
  .set-head .set-sub{font-size:12.5px;color:var(--muted);margin-top:4px;line-height:1.5}
  .set-head .set-act{margin-left:auto;flex-shrink:0}
  .iv-row{display:flex;gap:12px;align-items:center;margin-bottom:11px;flex-wrap:wrap}
  .iv-row .iv-lbl{width:250px;display:flex;align-items:center;gap:11px;font-weight:600;font-size:14px}
  .iv-row .iv-dot{width:11px;height:11px;border-radius:4px;flex-shrink:0}
  .iv-row input,.iv-row select{padding:10px 12px;border:2px solid var(--border);
    border-radius:10px;font-size:14px;background:var(--card);color:var(--text)}
  .iv-row input{width:84px}
  .iv-row input:focus,.iv-row select:focus{outline:none;border-color:var(--primary)}
  .set-note{font-size:11.5px;color:var(--muted);margin-top:10px;line-height:1.5}
  .set-foot{margin-top:18px;padding-top:16px;border-top:1px solid var(--border)}

  .stat-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:16px;margin-bottom:22px}
  .stat{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:20px;
    box-shadow:var(--shadow);position:relative;overflow:hidden}
  .stat .ic{position:absolute;right:16px;top:16px;width:38px;height:38px;border-radius:10px;
    display:flex;align-items:center;justify-content:center}
  .stat .ic svg{width:20px;height:20px}
  .stat .lbl{font-size:13px;color:var(--muted);font-weight:600;margin-bottom:8px}
  .stat .val{font-size:30px;font-weight:800;line-height:1}
  .stat .bar{position:absolute;left:0;bottom:0;height:4px;width:100%}

  .btn{display:inline-flex;align-items:center;gap:8px;padding:11px 20px;border-radius:10px;
    font-size:14px;font-weight:600;cursor:pointer;border:none;transition:.15s}
  .btn-primary{background:var(--primary);color:#fff}
  .btn-primary:hover{background:var(--primary-dark);transform:translateY(-1px);box-shadow:0 4px 12px rgba(79,70,229,.3)}
  .btn-outline{background:var(--card);color:var(--primary);border:2px solid var(--primary)}
  .btn-outline:hover{background:var(--primary-light)}
  .btn-success{background:var(--success);color:#fff}
  .btn-success:hover{filter:brightness(1.08)}
  .btn-sm{padding:8px 15px;font-size:13px;border-radius:8px}
  .btn-dl{padding:6px 11px;font-size:12px;font-weight:600;border-radius:7px;cursor:pointer;
    border:1.5px solid var(--primary);background:var(--primary-light);color:var(--primary);transition:.15s}
  .btn-dl:hover{background:var(--primary);color:#fff}
  .btn-dl.loading{background:var(--primary);color:#fff;pointer-events:none;opacity:.92}
  .btn-dl svg{vertical-align:-2px}
  .dl-spin{display:inline-block;width:11px;height:11px;border:2px solid rgba(255,255,255,.45);
    border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;vertical-align:-1px}

  .filter-bar{display:flex;gap:12px;margin-bottom:18px;flex-wrap:wrap}
  .filter-bar input,.filter-bar select{padding:11px 14px;border:2px solid var(--border);
    border-radius:10px;font-size:14px;transition:.15s;background:var(--card);color:var(--text)}
  .filter-bar input{flex:1;min-width:200px}
  .filter-bar input:focus,.filter-bar select:focus{outline:none;border-color:var(--primary)}
  table{width:100%;border-collapse:collapse}
  thead th{text-align:left;padding:13px 14px;font-size:12px;font-weight:700;color:var(--muted);
    text-transform:uppercase;letter-spacing:.4px;background:var(--th-bg);border-bottom:2px solid var(--border)}
  tbody td{padding:14px;border-bottom:1px solid var(--border);font-size:14px;vertical-align:top}
  tbody tr:hover{background:var(--row-hover)}
  .ref-tag{background:var(--chip);padding:3px 8px;border-radius:6px;font-family:monospace;font-size:13px}

  .run-live{color:#2563EB;font-weight:700;font-size:13px}
  .run-done{color:#16A34A;font-weight:700;font-size:13px}
  .run-cancel{color:#94A3B8;font-weight:700;font-size:13px}
  .run-err{color:var(--danger);font-weight:700;font-size:12px}
  .btn-cancel{margin-left:8px;background:#FEE2E2;color:#B91C1C;border:1px solid #FCA5A5;
    padding:3px 11px;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer}
  .btn-cancel:hover{background:#FCA5A5;color:#7F1D1D}

  .card-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}
  .card-head h3{margin-bottom:0}
  .btn-refresh{display:inline-flex;align-items:center;gap:6px;padding:7px 13px;border-radius:8px;
    font-size:13px;font-weight:600;cursor:pointer;border:1px solid var(--border);
    background:var(--soft);color:var(--text);transition:.15s}
  .btn-refresh:hover{background:var(--primary);color:#fff;border-color:var(--primary)}
  .btn-refresh:hover svg{color:#fff}

  .progress-banner{background:var(--card);border:1px solid var(--primary);border-left:4px solid var(--primary);
    border-radius:12px;padding:16px 18px;margin-bottom:18px;box-shadow:var(--shadow)}
  .pb-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:9px}
  .pb-title{display:flex;align-items:center;gap:9px;font-weight:700;font-size:14px;color:var(--text)}
  .pb-pct{font-weight:800;font-size:18px;color:var(--primary)}
  .pb-track{height:9px;background:var(--chip);border-radius:6px;overflow:hidden}
  .pb-fill{height:100%;background:linear-gradient(90deg,#6366F1,#8B5CF6);border-radius:6px;
    transition:width .5s ease;width:0%}
  .pb-sub{font-size:12px;color:var(--muted);margin-top:7px}
  .pb-spin{width:15px;height:15px;border:2.5px solid var(--chip);border-top-color:var(--primary);
    border-radius:50%;animation:spin .8s linear infinite;display:inline-block}
  @keyframes spin{to{transform:rotate(360deg)}}

  .timer-row{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:18px}
  .timer-chip{flex:1;min-width:230px;display:flex;align-items:center;gap:12px;background:var(--card);
    border:1px solid var(--border);border-radius:12px;padding:14px 16px;box-shadow:var(--shadow)}
  .timer-chip .tc-ic{width:38px;height:38px;border-radius:10px;display:flex;align-items:center;
    justify-content:center;background:var(--primary-light);color:var(--primary);flex-shrink:0}
  .timer-chip.soon{border-color:var(--warn);border-left:4px solid var(--warn)}
  .timer-chip.soon .tc-ic{background:rgba(234,88,12,.12);color:var(--warn)}
  .timer-chip.paused{border-left:4px solid #94a3b8;opacity:.9}
  .timer-chip.paused .tc-ic{background:rgba(148,163,184,.16);color:#64748b}
  .timer-chip.paused .tc-time{color:#64748b}
  .tc-btn{flex-shrink:0;display:inline-flex;align-items:center;gap:5px;border-radius:9px;padding:7px 12px;
    font-size:12.5px;font-weight:700;cursor:pointer;border:1px solid var(--border);transition:all .15s;white-space:nowrap}
  .tc-btn.pause{background:var(--soft);color:var(--text)}
  .tc-btn.pause:hover{background:#fff7ed;border-color:var(--warn);color:var(--warn)}
  .tc-btn.resume{background:#dcfce7;color:#166534;border-color:#bbf7d0}
  .tc-btn.resume:hover{background:#bbf7d0}
  .tc-lbl{font-size:12px;color:var(--muted);font-weight:600}
  .tc-time{font-size:18px;font-weight:800;color:var(--text);font-variant-numeric:tabular-nums}
  .tc-grp{font-size:11px;color:var(--muted)}
  .run-prog{font-size:11px;color:var(--muted);font-weight:600;margin-left:8px}

  .badge{display:inline-block;padding:5px 12px;border-radius:20px;font-size:12px;font-weight:700;white-space:nowrap}
  .b-pending{background:#FEF9C3;color:#854D0E}
  .b-docs{background:#FFEDD5;color:#9A3412}
  .b-required{background:#FED7AA;color:#9A3412}
  .b-verified{background:#DCFCE7;color:#166534}
  .b-approved{background:#CCFBF1;color:#115E59}
  .b-confirmed{background:#86EFAC;color:#14532D}
  .b-rejected{background:#FECACA;color:#991B1B}
  .b-error{background:#E2E8F0;color:#475569}
  .b-syc{background:#EDE9FE;color:#5B21B6}

  .pg-bar{display:flex;align-items:center;justify-content:space-between;margin-top:18px;flex-wrap:wrap;gap:12px}
  .pg-controls{display:flex;gap:6px;align-items:center}
  .pg-controls button{padding:8px 13px;border:1px solid var(--border);background:var(--card);color:var(--text);
    border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;transition:.15s}
  .pg-controls button:hover:not(:disabled){background:var(--primary-light);border-color:var(--primary)}
  .pg-controls button.active{background:var(--primary);color:#fff;border-color:var(--primary)}
  .pg-controls button:disabled{opacity:.4;cursor:not-allowed}
  .perpage{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--muted)}
  .perpage select{padding:7px 10px;border:1px solid var(--border);border-radius:8px;font-size:13px;background:var(--card);color:var(--text)}

  .drop{border:2.5px dashed var(--border);border-radius:14px;padding:48px;text-align:center;
    transition:.2s;cursor:pointer;background:#FAFBFF}
  .drop:hover,.drop.drag{border-color:var(--primary);background:var(--primary-light)}
  .drop .ic{margin-bottom:12px;color:var(--primary)}
  .drop .ic svg{width:42px;height:42px}

  .up-stats{display:flex;gap:14px;flex-wrap:wrap}
  .up-stat{flex:1;min-width:150px;background:var(--soft);border:1px solid var(--border);
    border-radius:12px;padding:14px 16px}
  .up-stat.dup{border-left:4px solid var(--warn)}
  .up-stat.new{border-left:4px solid var(--success)}
  .us-lbl{font-size:12px;color:var(--muted);font-weight:600;margin-bottom:6px}
  .us-val{font-size:26px;font-weight:800;line-height:1}
  .up-stat.dup .us-val{color:var(--warn)}
  .up-stat.new .us-val{color:var(--success)}
  .prev-head{font-size:13px;font-weight:700;margin-bottom:8px;color:var(--text)}
  .prev-box{max-height:380px;overflow:auto;border:1px solid var(--border);border-radius:12px}
  .prev-box table{width:100%}
  .prev-box thead th{position:sticky;top:0;z-index:1}
  .dup-tag{display:inline-block;background:#FEF3C7;color:#B45309;font-size:11px;font-weight:700;
    padding:2px 9px;border-radius:10px}
  .new-tag{display:inline-block;background:#DCFCE7;color:#15803D;font-size:11px;font-weight:700;
    padding:2px 9px;border-radius:10px}
  .sim-tag{display:inline-block;background:#FEF9C3;color:#854D0E;font-size:11px;font-weight:700;
    padding:2px 9px;border-radius:10px}
  .up-stat.sim{border-left:4px solid #CA8A04}
  .up-stat.sim .us-val{color:#CA8A04}
  .up-prog{background:var(--soft);border:1px solid var(--border);border-radius:12px;padding:14px 16px}
  .up-prog-top{display:flex;justify-content:space-between;font-weight:700;font-size:13px;margin-bottom:8px}
  .up-track{height:8px;background:var(--chip);border-radius:5px;overflow:hidden}
  .up-fill{height:100%;width:0%;background:linear-gradient(90deg,#6366F1,#8B5CF6);border-radius:5px;transition:width .2s}
  .up-eta{font-size:12px;color:var(--muted);margin-top:7px}

  .legend-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:12px}
  .legend-item{padding:14px 16px;border-radius:11px}
  .legend-item .nm{font-weight:700;font-size:13px}
  .legend-item .ds{font-size:11px;opacity:.75;font-weight:500}

  .dist-row{display:flex;align-items:center;gap:12px;margin-bottom:11px}
  .dist-row .nm{width:230px;font-size:13px;font-weight:600;flex-shrink:0}
  .dist-row .track{flex:1;height:22px;background:#F1F5F9;border-radius:11px;overflow:hidden}
  .dist-row .fill{height:100%;border-radius:11px;transition:width .5s}
  .dist-row .ct{width:40px;text-align:right;font-weight:700;font-size:13px}

  .empty{text-align:center;color:var(--muted);padding:40px;font-size:14px}

  #toast{position:fixed;bottom:28px;right:28px;background:var(--text);color:#fff;padding:14px 22px;
    border-radius:11px;font-size:14px;box-shadow:var(--shadow-lg);opacity:0;transform:translateY(20px);
    transition:.3s;z-index:2000;max-width:360px}
  #toast.show{opacity:1;transform:translateY(0)}

  .nav-backdrop{display:none}

  /* ===== Tablet / iPad: slim icon-only sidebar, content uses the space ===== */
  @media(max-width:1024px) and (min-width:768px){
    .sidebar{width:64px}
    .sidebar .brand .tx,.nav-item span.lbl,.nav-sep{display:none}
    .nav-item{justify-content:center;padding:14px}
    .main{margin-left:64px}
    .content{padding:20px}
    .topbar{padding:0 18px}
    .bell-dropdown{width:340px}
    .stat-grid{grid-template-columns:repeat(auto-fill,minmax(170px,1fr))}
  }

  /* ===== Phone: off-canvas slide-in drawer, full-width content ===== */
  @media(max-width:767px){
    .sidebar{width:250px;transform:translateX(-100%);
      transition:transform .26s cubic-bezier(.4,0,.2,1);box-shadow:0 0 50px rgba(0,0,0,.35)}
    #app.nav-open .sidebar{transform:translateX(0)}
    .main{margin-left:0}
    /* drawer always shows full labels even if 'minimize' was toggled earlier */
    #app.sidebar-min .sidebar{width:250px}
    #app.sidebar-min .sidebar .brand .tx,
    #app.sidebar-min .nav-item span.lbl,
    #app.sidebar-min .nav-sep,
    #app.sidebar-min .nav-item .nav-num,
    #app.sidebar-min .nav-item .badge-count{display:revert!important}
    #app.sidebar-min .nav-item{justify-content:flex-start;padding:11px 22px}
    #app.sidebar-min .main{margin-left:0}
    .nav-backdrop{display:block;position:fixed;inset:0;background:rgba(15,23,42,.5);
      z-index:90;opacity:0;visibility:hidden;transition:opacity .26s,visibility .26s}
    #app.nav-open .nav-backdrop{opacity:1;visibility:visible}

    .topbar{padding:0 12px;height:58px}
    .topbar h1{font-size:16px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:40vw}
    .topbar .right{gap:8px}
    .content{padding:14px}
    .card{padding:16px;border-radius:12px;margin-bottom:16px}
    .card h3{font-size:14px}
    .stat-grid{grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px}
    .stat{padding:15px}
    .filter-bar{gap:8px;margin-bottom:14px}
    .filter-bar input,.filter-bar select{padding:10px 12px;font-size:14px}
    .filter-bar input{min-width:130px;flex:1 1 100%}
    .bell-dropdown{width:calc(100vw - 20px);right:-8px;max-height:70vh}
    #run-menu{min-width:0;width:calc(100vw - 24px);max-width:310px}
    .btn-sm{padding:9px 12px;font-size:13px}
    .legend-grid{grid-template-columns:1fr}
    table{font-size:13px}
    table th,table td{padding:10px 12px}
    .rn-txt{display:none}
    #run-now-btn{padding:9px 11px}
    /* hard guards so nothing forces a horizontal scroll / zoom-out */
    #app,.main,.content,.page-section,.card{max-width:100%;min-width:0}
    .content{overflow-x:hidden}
    #toast{left:12px;right:12px;bottom:18px;max-width:none;text-align:center}
    .iv-row .iv-lbl{width:auto;flex:1 1 100%}
    .topbar{max-width:100vw}
  }

  /* ===== Big screens / TV: cap content width so it never stretches awkwardly ===== */
  @media(min-width:1700px){
    .content{max-width:1640px;margin-left:auto;margin-right:auto}
  }
  /* Counsellor-toggled collapse (click only) — icon-only sidebar, content expands */
  #app.sidebar-min .sidebar{width:64px}
  #app.sidebar-min .sidebar .brand .tx,
  #app.sidebar-min .nav-item span.lbl,
  #app.sidebar-min .nav-sep,
  #app.sidebar-min .nav-item .nav-num,
  #app.sidebar-min .nav-item .badge-count{display:none!important}
  #app.sidebar-min .nav-item{justify-content:center;padding:14px;position:relative}
  /* small dot so a non-zero count is still visible when collapsed */
  #app.sidebar-min .nav-item .nav-num.has,
  #app.sidebar-min .nav-item .badge-count.has{display:block!important;position:absolute;top:8px;right:10px;
    min-width:0;width:8px;height:8px;padding:0;border-radius:50%;font-size:0;overflow:hidden}
  #app.sidebar-min .main{margin-left:64px}
  .sb-toggle{background:none;border:none;cursor:pointer;color:var(--text);padding:7px;border-radius:9px;
    display:flex;align-items:center;margin-right:6px}
  .sb-toggle:hover{background:var(--soft)}
</style>
<script>
function toggleSidebar(){
  var app=document.getElementById("app");if(!app)return;
  // On phones the hamburger opens an off-canvas drawer; on larger screens it
  // collapses the sidebar to icons (persisted).
  if(window.matchMedia("(max-width:767px)").matches){
    app.classList.toggle("nav-open");
    return;
  }
  var min=app.classList.toggle("sidebar-min");
  try{localStorage.setItem("mvs_sidebar_min",min?"1":"0");}catch(e){}
}
function closeNav(){var a=document.getElementById("app");if(a)a.classList.remove("nav-open");}
function applySidebarPref(){
  try{ if(localStorage.getItem("mvs_sidebar_min")==="1"){var a=document.getElementById("app");if(a)a.classList.add("sidebar-min");} }catch(e){}
}
</script>
</head>
<body>

<div id="login-screen">
  <div class="login-card">
    <img src="/logo.png" alt="MVS Foundation" style="width:68px;height:68px;border-radius:50%;display:block;margin:0 auto 18px;object-fit:cover;box-shadow:0 4px 14px rgba(0,0,0,.18)">
    <h2>NIOS Status Tracker</h2>
    <p class="sub">MVS Foundation — Admin Portal</p>
    <label>Username</label>
    <input type="text" id="lg-user" placeholder="admin" autocomplete="username">
    <label>Password</label>
    <div style="position:relative;margin-bottom:16px">
      <input type="password" id="lg-pass" placeholder="enter password" autocomplete="current-password"
        style="padding-right:44px;margin-bottom:0" onkeydown="if(event.key==='Enter')doLogin()">
      <button type="button" id="lg-pass-toggle" onclick="toggleLgPass()" aria-label="Show password"
        style="position:absolute;right:6px;top:50%;transform:translateY(-50%);width:auto!important;min-width:0;margin:0!important;background:none!important;border:none;cursor:pointer;padding:6px;color:#94a3b8;display:flex;align-items:center;justify-content:center">
        <svg id="lg-eye" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="19" height="19"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"/><circle cx="12" cy="12" r="3"/></svg>
      </button>
    </div>
    <button onclick="doLogin()">Sign In</button>
    <div id="login-error"></div>
  </div>
</div>

<div id="app">
  <aside class="sidebar">
    <div class="brand">
      <img src="/logo.png" alt="MVS" style="width:38px;height:38px;border-radius:50%;object-fit:cover;flex-shrink:0">
      <div class="tx"><b>NIOS Tracker</b><span>MVS Foundation</span></div>
    </div>
    <nav class="nav-scroll">
    <div class="nav-item active" data-page="dashboard" onclick="nav('dashboard')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="7" height="9"/><rect x="14" y="3" width="7" height="5"/><rect x="14" y="12" width="7" height="9"/><rect x="3" y="16" width="7" height="5"/></svg></span><span class="lbl">Dashboard</span></div>
    <div class="nav-item" data-page="students" onclick="nav('students')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87M16 3.13a4 4 0 0 1 0 7.75"/></svg></span><span class="lbl">Active Students</span>
      <span class="nav-num" id="nav-students-badge" style="display:none">0</span></div>
    <div class="nav-item" data-page="confirmed" onclick="nav('confirmed')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg></span><span class="lbl">Confirmed</span>
      <span class="nav-num" id="nav-confirmed-badge" style="display:none">0</span></div>
    <div class="nav-item" data-page="required" onclick="nav('required')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="12" y1="11" x2="12" y2="15"/><line x1="12" y1="18" x2="12" y2="18"/></svg></span><span class="lbl">Required</span>
      <span class="badge-count" id="nav-required-badge">0</span></div>
    <div class="nav-item" data-page="docreq" onclick="nav('docreq')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15a2 2 0 0 1-2 2H7l-4 4V5a2 2 0 0 1 2-2h14a2 2 0 0 1 2 2z"/></svg></span><span class="lbl">Doc Requests</span>
      <span class="badge-count" id="nav-docreq-badge">0</span></div>
    <div class="nav-item" data-page="syc" onclick="nav('syc')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/></svg></span><span class="lbl">SYC Students</span>
      <span class="nav-num" id="nav-syc-badge" style="display:none">0</span></div>
    <div class="nav-item" data-page="failed" onclick="nav('failed')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/></svg></span><span class="lbl">Failed to Run</span>
      <span class="badge-count" id="nav-failed-badge" style="display:none">0</span></div>
    <div class="nav-item" data-page="unknown" onclick="nav('unknown')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><path d="M9.09 9a3 3 0 0 1 5.83 1c0 2-3 3-3 3"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg></span><span class="lbl">Unknown</span>
      <span class="badge-count" id="nav-unknown-badge" style="display:none">0</span></div>
    <div class="nav-item" data-page="notoc" onclick="nav('notoc')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M20.59 13.41l-7.17 7.17a2 2 0 0 1-2.83 0L2 12V2h10l8.59 8.59a2 2 0 0 1 0 2.82z"/><line x1="7" y1="7" x2="7.01" y2="7"/></svg></span><span class="lbl">No-TOC Students</span>
      <span class="badge-count" id="nav-notoc-badge" style="display:none;background:#92400e">0</span></div>
    <div class="nav-item" data-page="tocerror" onclick="nav('tocerror')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg></span><span class="lbl">TOC Status Error</span>
      <span class="badge-count" id="nav-tocerror-badge" style="display:none;background:#b91c1c">0</span></div>
    <div class="nav-item" data-page="pending" onclick="nav('pending')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg></span><span class="lbl">Pending Students</span>
      <span class="badge-count" id="nav-pending-badge" style="display:none;background:#B45309">0</span></div>
    <div class="nav-sep">Activity</div>
    <div class="nav-item" data-page="history" onclick="nav('history')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 3v5h5"/><path d="M3.05 13A9 9 0 1 0 6 5.3L3 8"/><path d="M12 7v5l4 2"/></svg></span><span class="lbl">Change History</span></div>
    <div class="nav-item" data-page="runlogs" onclick="nav('runlogs')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><line x1="3" y1="6" x2="3.01" y2="6"/><line x1="3" y1="12" x2="3.01" y2="12"/><line x1="3" y1="18" x2="3.01" y2="18"/></svg></span><span class="lbl">Run Logs</span></div>
    <div class="nav-item" data-page="transfers" onclick="nav('transfers')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="17 1 21 5 17 9"/><path d="M3 11V9a4 4 0 0 1 4-4h14"/><polyline points="7 23 3 19 7 15"/><path d="M21 13v2a4 4 0 0 1-4 4H3"/></svg></span><span class="lbl">Transfer Data</span></div>
    <div class="nav-sep">Manage</div>
    <div class="nav-item" data-page="upload" onclick="nav('upload')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg></span><span class="lbl">Upload Excel</span></div>
    <div class="nav-item" data-page="settings" onclick="nav('settings')">
      <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 1 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 1 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 1 1 2.83-2.83l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 1 1 2.83 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z"/></svg></span><span class="lbl">Settings</span></div>
    </nav>
    <div class="side-foot">
      <div class="nav-item" onclick="toggleTheme()">
        <span class="ic" id="theme-ic"></span><span class="lbl" id="theme-lbl">Dark Mode</span></div>
      <div class="nav-item logout" onclick="doLogout()">
        <span class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg></span><span class="lbl">Logout</span></div>
    </div>
  </aside>
  <div class="nav-backdrop" onclick="closeNav()"></div>

  <div class="main">
    <div class="topbar">
      <div style="display:flex;align-items:center;gap:4px">
        <button class="sb-toggle" onclick="toggleSidebar()" title="Minimize / maximize menu" aria-label="Toggle sidebar">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="20" height="20"><line x1="3" y1="12" x2="21" y2="12"/><line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="18" x2="21" y2="18"/></svg>
        </button>
        <h1 id="page-title">Dashboard</h1>
      </div>
      <div class="right">
        <div class="run-menu-wrap" style="position:relative">
          <button class="btn btn-success btn-sm" id="run-now-btn" onclick="toggleRunMenu(event)">
            <svg viewBox="0 0 24 24" fill="currentColor" width="14" height="14"><polygon points="5 3 19 12 5 21 5 3"/></svg> <span class="rn-txt">Run Now</span>
            <svg id="run-caret" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round" width="12" height="12" style="margin-left:5px;transition:transform .15s"><polyline points="6 9 12 15 18 9"/></svg></button>
          <div id="run-menu" style="display:none">
            <div class="rm-head"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round" width="13" height="13"><polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"/></svg> Run a status check</div>
            <button class="run-menu-item rmi-all" onclick="runChoice('all')">
              <span class="rmi-ic"><svg viewBox="0 0 24 24" fill="currentColor" width="15" height="15"><polygon points="5 3 19 12 5 21 5 3"/></svg></span>
              <span class="rmi-txt"><span class="rmi-main">Run all data</span>
              <span class="rmi-sub">MVS Tracker + MVS Portal together</span></span>
              <span class="rmi-go">&rsaquo;</span></button>

            <div class="rm-group trk">
              <div class="rm-ghead">
                <span class="rm-gic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><rect x="3" y="4" width="18" height="16" rx="2"/><line x1="3" y1="10" x2="21" y2="10"/><line x1="9" y1="10" x2="9" y2="20"/></svg></span>
                <span class="rm-gname">MVS Tracker</span>
                <span class="rm-ghint">by group</span>
              </div>
              <button class="run-menu-item trk" onclick="runChoice('tracker','ondemand')"><span class="rmi-dot"></span><span class="rmi-main">On Demand</span></button>
              <button class="run-menu-item trk" onclick="runChoice('tracker','stream2')"><span class="rmi-dot"></span><span class="rmi-main">Stream 2</span></button>
              <button class="run-menu-item trk" onclick="runChoice('tracker','public')"><span class="rmi-dot"></span><span class="rmi-main">Public</span></button>
              <button class="run-menu-item trk" onclick="runChoice('tracker','all')"><span class="rmi-dot"></span><span class="rmi-main">All Tracker</span><span class="rmi-tag">all</span></button>
            </div>

            <div class="rm-group por">
              <div class="rm-ghead">
                <span class="rm-gic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><circle cx="12" cy="12" r="9"/><line x1="3" y1="12" x2="21" y2="12"/><path d="M12 3a14 14 0 0 1 0 18 14 14 0 0 1 0-18"/></svg></span>
                <span class="rm-gname">MVS Portal</span>
                <span class="rm-ghint">by group</span>
              </div>
              <button class="run-menu-item por" onclick="runChoice('portal','ondemand')"><span class="rmi-dot"></span><span class="rmi-main">On Demand</span></button>
              <button class="run-menu-item por" onclick="runChoice('portal','stream2')"><span class="rmi-dot"></span><span class="rmi-main">Stream 2</span></button>
              <button class="run-menu-item por" onclick="runChoice('portal','public')"><span class="rmi-dot"></span><span class="rmi-main">Public</span></button>
              <button class="run-menu-item por" onclick="runChoice('portal','all')"><span class="rmi-dot"></span><span class="rmi-main">All Portal</span><span class="rmi-tag">all</span></button>
              <button class="run-menu-item por" onclick="runChoice('portalnew')" style="background:#f0fdf4" title="Check only students newly arrived on the portal — skips already-active data, saves CapSolver credits"><span class="rmi-dot" style="background:#16a34a"></span><span class="rmi-main">New data only</span><span class="rmi-tag" style="background:#dcfce7;color:#15803d">saves credits</span></button>
            </div>
          </div>
        </div>
        <div class="bell-btn" onclick="refreshPage(this)" title="Refresh this page" style="cursor:pointer">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="18" height="18"><polyline points="23 4 23 10 17 10"/><polyline points="1 20 1 14 7 14"/><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"/></svg>
        </div>
        <div class="bell-wrap">
          <div class="bell-btn" onclick="toggleBell(event)">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="19" height="19"><path d="M18 8A6 6 0 0 0 6 8c0 7-3 9-3 9h18s-3-2-3-9"/><path d="M13.73 21a2 2 0 0 1-3.46 0"/></svg>
            <span class="bell-badge" id="bell-badge">0</span></div>
          <div class="bell-dropdown" id="bell-dropdown" onclick="event.stopPropagation()">
            <div class="bell-head">Document Required
              <span class="cnt" id="bell-head-cnt">0</span></div>
            <div id="bell-list"></div>
          </div>
        </div>
      </div>
    </div>

    <div class="content">
      <section id="sec-dashboard" class="page-section active">
        <div id="run-progress" class="progress-banner" style="display:none">
          <div class="pb-top">
            <span class="pb-title"><span class="pb-spin"></span><span id="pb-label">Checking students…</span></span>
            <span class="pb-pct" id="pb-pct">0%</span>
          </div>
          <div class="pb-track"><div class="pb-fill" id="pb-fill" style="width:0%"></div></div>
          <div class="pb-sub" id="pb-sub">0 / 0 done</div>
          <div id="run-queue-line" style="display:none;margin:9px 0 2px;font-size:12.5px;padding:8px 12px;background:#EEF2FF;border:1px solid #C7D2FE;border-radius:9px;color:#3730A3"></div>
          <div id="pb-srcwrap" style="margin-top:10px;display:none">
            <div id="pb-mvs-row">
              <div style="display:flex;justify-content:space-between;font-size:13px;font-weight:600;margin-bottom:3px">
                <span style="display:inline-flex;align-items:center;gap:7px"><span style="width:9px;height:9px;border-radius:50%;background:#7C3AED"></span>MVS Portal</span><span id="pb-mvs-txt">0 / 0</span></div>
              <div class="pb-track" style="height:8px"><div class="pb-fill" id="pb-mvs-fill" style="width:0%;background:#7C3AED"></div></div>
            </div>
            <div id="pb-trk-row">
              <div style="display:flex;justify-content:space-between;font-size:13px;font-weight:600;margin:8px 0 3px">
                <span style="display:inline-flex;align-items:center;gap:7px"><span style="width:9px;height:9px;border-radius:50%;background:#0EA5E9"></span>MVS Tracker</span><span id="pb-trk-txt">0 / 0</span></div>
              <div class="pb-track" style="height:8px"><div class="pb-fill" id="pb-trk-fill" style="width:0%;background:#0EA5E9"></div></div>
            </div>
          </div>
        </div>

        <div id="next-runs" class="timer-row"></div>

        <div class="stat-grid" id="stat-grid"></div>
        <div class="card">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:10px">
            <h3 style="margin:0">Session-wise Students</h3>
            <button onclick="loadReconciliation()" style="background:var(--soft);color:var(--text);border:1px solid var(--border);border-radius:8px;padding:6px 13px;font-size:12.5px;font-weight:700;cursor:pointer">Reconcile with Portal</button>
          </div>
          <div id="reconcile-panel" style="display:none;margin:14px 0 4px;padding:16px;background:var(--soft);border:1px solid var(--border);border-radius:12px"></div>
          <div id="source-counts" style="margin-top:14px"></div>
          <div id="session-counts"><div class="empty">Loading...</div></div>
        </div>
        <div class="card">
          <h3>Status Distribution</h3>
          <div id="distribution"><div class="empty">Loading...</div></div>
        </div>
        <div class="card">
          <div class="card-head">
            <h3>Recent Runs</h3>
            <button class="btn-refresh" onclick="loadDashboard()" title="Refresh">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><polyline points="23 4 23 10 17 10"/><polyline points="1 20 1 14 7 14"/><path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"/></svg>
                Refresh</button>
          </div>
          <style>.recent-runs-scroll{max-height:360px;overflow:auto}.recent-runs-scroll thead th{position:sticky;top:0;background:var(--card);z-index:2}</style>
          <div class="recent-runs-scroll"><table>
            <thead><tr><th>Run At</th><th>Type</th><th>Checked</th><th>Changed</th><th>Failed</th><th>Status</th><th>Action</th></tr></thead>
            <tbody id="recent-runs"></tbody>
          </table></div>
        </div>
      </section>

      <section id="sec-students" class="page-section">
        <div class="card">
          <div class="filter-bar">
            <input type="text" id="s-search" placeholder="Search name / reference / email..." oninput="debounceStudents()">
            <select id="s-status" onchange="loadStudents(1)">
              <option value="">All Statuses</option>
              <option>Pending</option><option>Documents Verification In Progress</option>
              <option>Document Required</option><option>Verified</option>
              <option>Approved</option><option>Rejected</option><option>Fetch Error</option><option>Unknown</option>
            </select>
            <select id="s-session" onchange="loadStudents(1)"><option value="">All Sessions</option></select>
            <select id="s-toc" onchange="loadStudents(1)" title="Filter by tocStatus"><option value="">All TOC</option><option value="yes">TOC: yes</option><option value="no">TOC: no</option><option value="blank">TOC: not set</option><option value="mismatch">TOC mismatch (error)</option></select>
            <select id="s-class" onchange="loadStudents(1)"><option value="">All Classes</option><option value="10">Class 10</option><option value="12">Class 12</option></select>
            <select id="s-source" onchange="loadStudents(1)"><option value="">All Data Types</option><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <select id="s-datepreset" onchange="onDatePreset('s',()=>loadStudents(1))">
              <option value="">All dates</option>
              <option value="today">Today</option>
              <option value="yesterday">Yesterday</option>
              <option value="7d">Last 7 days</option>
              <option value="custom">Custom range…</option>
            </select>
            <select id="s-checkstate" onchange="loadStudents(1)" title="Filter by whether the latest run actually read this student's status. 'New check' = read this run. 'Not checked' = the run fell back (captcha/proxy) — last status kept.">
              <option value="">All checks</option>
              <option value="new">New check</option>
              <option value="stale">Not checked</option>
            </select>
            <button class="btn btn-outline btn-sm" onclick="exportStudents('normal')">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Export Excel</button>
            <button class="btn btn-primary btn-sm" id="s-run-nc-btn" onclick="runNotChecked()" title="Re-check only the students the latest run couldn't read, for the selected session (or all). Confirmed students are skipped.">Run not-checked</button>
          </div>
          <div class="filter-bar" id="s-daterow" style="display:none">
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">From
              <input type="datetime-local" id="s-from"></label>
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">To
              <input type="datetime-local" id="s-to"></label>
            <button class="btn btn-primary btn-sm" onclick="loadStudents(1)">Apply</button>
          </div>
          <div id="s-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div id="s-checksummary" style="font-size:13px;margin-bottom:14px;display:none"></div>
          <div id="sel-bar" style="display:none;align-items:center;gap:12px;flex-wrap:wrap;background:var(--soft);border:1px solid var(--border);border-radius:10px;padding:10px 14px;margin-bottom:14px">
            <span style="font-weight:600;font-size:13.5px"><span id="sel-count">0</span> selected <span style="color:var(--muted);font-weight:400">(max 20)</span></span>
            <button class="btn btn-success btn-sm" id="sel-run-btn" onclick="runSelected()">
              <svg viewBox="0 0 24 24" fill="currentColor" width="13" height="13"><polygon points="5 3 19 12 5 21 5 3"/></svg> Run Selected</button>
            <button class="btn btn-sm" style="background:#7C3AED;color:#fff" onclick="tocCheckSelected('active',this)" title="Read the real NIOS TOC (Previous Subject Details) for the selected students and flag any mismatch with the Portal into 'TOC Status Error'.">Check TOC (selected)</button>
            <button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="deleteSelectedStudents()">&#128465; Delete selected</button>
            <select onchange="changeSelectedSource(this)" style="max-width:170px" title="Move selected students to a data type">
              <option value="">Change source&hellip;</option>
              <option value="mvs_tracker">&rarr; MVS Tracker</option>
              <option value="mvs_portal">&rarr; MVS Portal</option>
            </select>
            <button class="btn btn-sm" style="background:transparent;color:var(--muted);border:1px solid var(--border)" onclick="clearSelection()">Clear</button>
            <span style="font-size:12px;color:var(--muted)">Pick the exact students whose status you need right now — saves credits.</span>
          </div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th style="width:34px"><input type="checkbox" id="sel-all" onclick="toggleSelectAll(this)" title="Select all on this page"></th>
              <th>#</th><th>Reference No</th><th>Student Name</th><th>Session</th>
              <th>Status</th><th>tocStatus</th><th>Last Checked</th><th>Action</th>
            </tr></thead><tbody id="s-body"></tbody></table>
          </div>
          <div class="pg-bar" id="s-pg"></div>
        </div>
      </section>

      <section id="sec-notoc" class="page-section">
        <div class="card">
          <h3>No-TOC Students</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            Students whose documents <b>differ from the normal set</b> for their session:
            <b>On Demand / Stream 2 with tocStatus = &quot;no&quot;</b> (shorter set), and
            <b>Public with tocStatus = &quot;yes&quot;</b> (gets the Application Form too).
            Public &quot;no&quot; is the normal Public case, so it is not listed here.
            If empty even though the Portal has such students, click <b>Sync tocStatus from Portal</b> first.</p>
          <div class="filter-bar">
            <input type="text" id="nt-search" placeholder="Search name / reference / enrollment / mobile..." oninput="debounceNoToc()">
            <select id="nt-session" onchange="loadNoToc(1)"><option value="">All Sessions</option></select>
            <button class="btn btn-primary btn-sm" onclick="syncToc(this)" title="Pull the latest tocStatus from the Portal and update it on the Tracker — NO CapSolver, no credits used. Run this after updating tocStatus on the Portal.">Sync tocStatus from Portal</button>
            <button class="btn btn-success btn-sm" onclick="runNotoc(this)" title="Run a NIOS status check now for the pending (not-yet-confirmed) no-TOC students">
              <svg viewBox="0 0 24 24" fill="currentColor" width="13" height="13"><polygon points="5 3 19 12 5 21 5 3"/></svg> Run no-TOC now</button>
            <button class="btn btn-outline btn-sm" onclick="exportNoToc()">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Export Excel</button>
          </div>
          <div id="nt-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>#</th><th>Reference No</th><th>Student Name</th><th>Session</th>
              <th>Status</th><th>tocStatus</th><th>Last Checked</th>
            </tr></thead><tbody id="nt-body"></tbody></table>
          </div>
          <div class="pg-bar" id="nt-pg"></div>
        </div>
      </section>

      <section id="sec-tocerror" class="page-section">
        <div class="card">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap">
            <h3 style="margin:0">TOC Status Error</h3>
            <div style="display:flex;gap:8px;flex-wrap:wrap">
              <button class="btn btn-sm" style="background:#0F766E;color:#fff" onclick="forceRecheckAll(this)" title="Re-read the real NIOS TOC for EVERY confirmed student — even those already checked or marked verified. Catches students locked on a wrong TOC by an old bad read.">Re-check ALL confirmed (force)</button>
              <button class="btn btn-sm" style="background:#FEE2E2;color:#B91C1C;border:1px solid #FECACA" onclick="recheckFlagged(this)" title="Fresh NIOS read for every still-flagged mismatch — genuine ones are auto-corrected and pushed to the Portal; wrong flags clear.">Re-check flagged</button>
              <button class="btn btn-outline btn-sm" onclick="runTocCheckConfirmed(this)" title="One-time: read the real NIOS TOC for every confirmed student that hasn't been checked yet. Confirmed students are never re-run, so this covers them. Safe to re-click — it only picks up whatever is left.">Check all confirmed students' TOC (one-time)</button>
              <button class="btn btn-primary btn-sm" onclick="runTocCheck(this)" title="Read NIOS TOC for unverified no-TOC students.">Check TOC from NIOS</button>
            </div>
          </div>
          <div id="toc-progress" style="display:none;margin:14px 0 4px;padding:14px 16px;background:var(--soft);border:1px solid var(--border);border-radius:12px"></div>
          <p style="color:var(--muted);font-size:13px;margin:10px 0 12px">
            When a check finds the MVS Portal's <b>tocStatus</b> disagreeing with the NIOS official site
            (Previous Subject Details), the tracker now <b>corrects it automatically</b> — the NIOS value is
            applied, pushed to the Portal, and the WhatsApp goes with the right campaign. No manual Verify
            needed. Every auto-corrected student is listed below (permanent audit list), so you can always
            check who was changed, from what, to what.
          </p>
          <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px">
            <input id="te-search" placeholder="Search name / reference..." oninput="teSearch()"
                   style="flex:1;min-width:220px;max-width:420px;padding:9px 13px;border:1px solid var(--border);border-radius:9px;font-size:13.5px">
          </div>
          <div id="te-count" style="font-size:12.5px;color:var(--muted);margin-bottom:8px"></div>
          <div style="overflow-x:auto">
            <table class="tbl"><thead><tr>
              <th>Student</th><th>Reference</th><th>Session</th><th>Status</th>
              <th>Portal &rarr; NIOS</th><th>TOC subjects (NIOS)</th><th>Fixed at</th><th>Portal push</th><th>WhatsApp</th><th>Action</th>
            </tr></thead><tbody id="te-body"></tbody></table>
          </div>
          <div id="te-pg" style="margin-top:12px"></div>
        </div>
      </section>

      <section id="sec-pending" class="page-section">
        <div class="card">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap">
            <h3 style="margin:0">Pending Students <span style="font-weight:600;color:var(--muted);font-size:13px">(no reference yet)</span></h3>
            <button class="btn btn-outline btn-sm" onclick="loadPending(true)">Refresh from Portal</button>
          </div>
          <p style="color:var(--muted);font-size:13px;margin:10px 0 12px">
            These students exist on the MVS Portal but have <b>no Reference/Enrollment number yet</b>, so the
            tracker cannot check them. This list is read <b>live from the Portal</b> every time (nothing is
            stored) — the moment a reference is filled on the Portal, the next run imports the student
            automatically and they drop off this list. Being here never blocks fetching.
          </p>
          <div id="pend-count" style="font-size:12.5px;color:var(--muted);margin-bottom:8px"></div>
          <div style="overflow-x:auto">
            <table class="tbl"><thead><tr>
              <th>#</th><th>Student Name</th><th>Mobile</th><th>Session</th><th>Class</th><th>Reference on Portal</th><th>Why pending</th><th>Portal ID</th>
            </tr></thead><tbody id="pend-body"><tr><td colspan="8" class="empty">Open this page to load…</td></tr></tbody></table>
          </div>
        </div>
      </section>

      <section id="sec-confirmed" class="page-section">
        <div class="card">
          <h3>Admission Confirmed Students</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            Their admission is confirmed. Documents are sent to WhatsApp <b>automatically in the background</b> —
            you do not need to click Resend. Use the <b>WhatsApp</b> filter below to see who is still pending.</p>
          <div class="filter-bar">
            <input type="text" id="c-search" placeholder="Search name / reference / email..." oninput="debounceConfirmed()">
            <select id="c-status" onchange="loadConfirmed(1)">
              <option value="">All Statuses</option>
              <option>Admission Confirmed</option><option>Admitted</option>
            </select>
            <select id="c-wa" onchange="loadConfirmed(1)">
              <option value="">All WhatsApp</option>
              <option value="delivered">Delivered</option>
              <option value="sent">Sent (pending delivery)</option>
              <option value="pending24">Sent, pending &gt; 24h</option>
              <option value="notsent">Not sent yet</option>
              <option value="failed">Failed</option>
            </select>
            <select id="c-session" onchange="loadConfirmed(1)"><option value="">All Sessions</option></select>
            <select id="c-toc" onchange="loadConfirmed(1)" title="Filter by tocStatus"><option value="">All TOC</option><option value="yes">TOC: yes</option><option value="no">TOC: no</option><option value="blank">TOC: not set</option><option value="mismatch">TOC mismatch (error)</option></select>
            <select id="c-class" onchange="loadConfirmed(1)"><option value="">All Classes</option><option value="10">Class 10</option><option value="12">Class 12</option></select>
            <select id="c-source" onchange="loadConfirmed(1)"><option value="">All Data Types</option><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <select id="c-saved" onchange="loadConfirmed(1)" title="Filter by whether ALL of the student's documents are saved in our database"><option value="">All (saved + not)</option><option value="saved">All documents saved</option><option value="notsaved">Not fully saved (missing docs)</option></select>
            <select id="c-datepreset" onchange="onDatePreset('c',()=>loadConfirmed(1))">
              <option value="">All dates</option>
              <option value="today">Today</option>
              <option value="yesterday">Yesterday</option>
              <option value="7d">Last 7 days</option>
              <option value="custom">Custom range…</option>
            </select>
            <button class="btn btn-primary btn-sm" onclick="autoSendNow(this)" title="Send WhatsApp now to all pending confirmed students">Send all pending now</button>
            <button class="btn btn-outline btn-sm" onclick="resendNoTocConfirmed(this)" title="Re-send the CORRECT documents to confirmed students whose tocStatus is NO (they previously received the wrong documents). Their old links stop working.">Resend no-TOC docs</button>
            <button class="btn btn-outline btn-sm" onclick="exportStudents('confirmed')">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Export Excel</button>
            <button class="btn btn-success btn-sm" id="cache-docs-btn" onclick="cacheDocsNow(this)" title="Fetch every confirmed student document from NIOS once and SAVE it in our database. After this, WhatsApp links open straight from our copy — no NIOS/CapSolver needed, and they keep working even if NIOS is down.">Save all documents to DB</button>
            <button class="btn btn-outline btn-sm" onclick="cacheDocsNow(this,true)" title="Re-fetch and OVERWRITE all saved documents (use if documents changed on NIOS).">Refresh all saved</button>
            <button class="btn btn-sm" style="background:#FEE2E2;color:#B91C1C" onclick="stopCacheDocs(this)" title="Stop the auto-resume of the document save.">&#9632; Stop saving</button>
            <span id="cache-docs-status" style="font-size:12px;color:var(--muted);align-self:center"></span>
          </div>
          <div id="c-bulkbar" style="display:none;align-items:center;gap:12px;flex-wrap:wrap;background:var(--soft);border:1px solid var(--border);border-radius:10px;padding:10px 14px;margin-bottom:12px">
            <span style="font-weight:700;font-size:13px"><span id="c-selcount">0</span> selected</span>
            <button class="btn btn-primary btn-sm" onclick="resendSelected(this)">Resend WhatsApp to selected</button>
            <button class="btn btn-sm" style="background:#7C3AED;color:#fff" onclick="tocCheckSelected('confirmed',this)" title="Read the real NIOS TOC (Previous Subject Details) for the selected students and flag any mismatch with the Portal into 'TOC Status Error'.">Check TOC (selected)</button>
            <button class="btn btn-sm" style="background:#F97316;color:#fff" onclick="refreshSelectedDocs(this)" title="Re-fetch and overwrite the saved documents for the selected students (use if their documents changed on NIOS).">Refresh docs for selected</button>
            <button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="bulkDeleteConf()">&#128465; Delete selected</button>
            <button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="clearConfSel()">Clear</button>
          </div>
          <div id="wa-progress" style="display:none;background:linear-gradient(135deg,#ECFDF5,#F0FDF4);border:1px solid #BBF7D0;border-radius:12px;padding:14px 16px;margin-bottom:14px">
            <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:9px">
              <span style="font-weight:700;font-size:13.5px;color:#065F46"><span id="wap-title">Sending WhatsApp to pending students</span></span>
              <span style="font-weight:800;font-size:15px;color:#047857"><span id="wap-pct">0</span>%</span>
            </div>
            <div style="height:10px;background:#D1FAE5;border-radius:6px;overflow:hidden">
              <div id="wap-bar" style="height:100%;width:0%;background:linear-gradient(90deg,#10B981,#059669);transition:width .4s"></div>
            </div>
            <div style="display:flex;gap:16px;flex-wrap:wrap;margin-top:10px;font-size:12.5px;font-weight:600">
              <span style="color:#047857">Sending: <span id="wap-total">0</span></span>
              <span style="color:#059669">Sent this run: <span id="wap-sent">0</span></span>
              <span style="color:#B45309">Remaining: <span id="wap-remaining">0</span></span>
              <span style="color:#B91C1C">Failed (auto-retry): <span id="wap-failed">0</span></span>
            </div>
          </div>
          <div id="cache-progress" style="display:none;background:linear-gradient(135deg,#EEF2FF,#F5F3FF);border:1px solid #C7D2FE;border-radius:12px;padding:14px 16px;margin-bottom:14px">
            <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:9px">
              <span style="font-weight:700;font-size:13.5px;color:#3730A3"><span id="cap-title">Saving documents to database</span></span>
              <span style="font-weight:800;font-size:15px;color:#4338CA"><span id="cap-pct">0</span>%</span>
            </div>
            <div style="height:10px;background:#E0E7FF;border-radius:6px;overflow:hidden">
              <div id="cap-bar" style="height:100%;width:0%;background:linear-gradient(90deg,#6366F1,#4F46E5);transition:width .4s"></div>
            </div>
            <div style="display:flex;gap:16px;flex-wrap:wrap;margin-top:10px;font-size:12.5px;font-weight:600">
              <span style="color:#4338CA">Progress: <span id="cap-done">0</span>/<span id="cap-total">0</span></span>
              <span style="color:#059669">Saved: <span id="cap-saved">0</span></span>
              <span style="color:#B45309">Retrying: <span id="cap-retry">0</span></span>
              <span style="color:#B91C1C">Needs data fix: <span id="cap-genfail">0</span></span>
              <span style="color:#6D28D9" id="cap-resume"></span>
            </div>
            <div id="cap-err" style="display:none;margin-top:8px;font-size:11.5px;color:#B45309;background:#FEF3C7;border-radius:8px;padding:6px 10px"></div>
          </div>
          <div class="filter-bar" id="c-daterow" style="display:none">
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">From
              <input type="datetime-local" id="c-from"></label>
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">To
              <input type="datetime-local" id="c-to"></label>
            <button class="btn btn-primary btn-sm" onclick="loadConfirmed(1)">Apply</button>
          </div>
          <div id="c-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th style="width:34px"><input type="checkbox" id="c-selall" onclick="toggleConfAll(this)" title="Select all on this page"></th>
              <th>#</th><th>Reference No</th><th>Student Name</th><th>Session</th>
              <th>Status</th><th>tocStatus</th><th>Downloads</th><th>Confirmed On</th><th>Action</th>
            </tr></thead><tbody id="c-body"></tbody></table>
          </div>
          <div class="pg-bar" id="c-pg"></div>
        </div>
      </section>

      <section id="sec-syc" class="page-section">
        <div class="card">
          <h3>SYC Students</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            SYC students are <b>not status-checked</b>. Their Hall Ticket is fetched directly using
            <b>Enrollment Number + Date of Birth</b>. Click "Download Hall Ticket" — it shows progress and opens the hall ticket in the correct format (then use Save as PDF).</p>
          <div class="filter-bar">
            <input type="text" id="syc-search" placeholder="Search name / enrollment / mobile..." oninput="debounceSyc()">
            <button class="btn btn-outline btn-sm" onclick="loadSyc(1)">Refresh</button>
          </div>
          <div id="syc-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>#</th><th>Enrollment No</th><th>Student Name</th><th>Mobile</th><th>DOB</th><th>Status</th><th>Hall Ticket</th><th>Remove</th>
            </tr></thead><tbody id="syc-body"></tbody></table>
          </div>
          <div class="pg-bar" id="syc-pg"></div>
        </div>
      </section>

      <section id="sec-required" class="page-section">
        <div class="card">
          <h3>Document Required — Action Needed</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            These need to be resolved by the counsellor. The exported / WhatsApp Excel includes
            <b>DOB</b> so you can log in directly. Once resolved, hit <b>Re-check Required</b> to update only these students.</p>
          <div class="filter-bar">
            <input type="text" id="r-search" placeholder="Search name / reference / email..." oninput="debounceRequired()">
            <select id="r-status" onchange="loadRequired(1)">
              <option value="">All Statuses</option>
              <option>Document Required</option>
            </select>
            <select id="r-session" onchange="loadRequired(1)"><option value="">All Sessions</option></select>
            <select id="r-source" onchange="loadRequired(1)"><option value="">All Data Types</option><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <button class="btn btn-primary btn-sm" onclick="runRequired(this)">
              <svg viewBox="0 0 24 24" fill="currentColor" width="14" height="14"><polygon points="5 3 19 12 5 21 5 3"/></svg>
              Re-check Required</button>
            <button class="btn btn-outline btn-sm" onclick="exportStudents('required')">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Export Excel</button>
          </div>
          <div id="r-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div id="r-bulkbar" style="display:none;align-items:center;gap:12px;flex-wrap:wrap;background:#FEF2F2;border:1px solid #FECACA;border-radius:10px;padding:10px 14px;margin-bottom:12px"><span style="font-weight:700;font-size:13px"><span id="r-selcount">0</span> selected</span><button class="btn btn-sm btn-success" onclick="runSelectedFrom('r')"><svg viewBox="0 0 24 24" fill="currentColor" width="12" height="12"><polygon points="5 3 19 12 5 21 5 3"/></svg> Run selected</button><button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="bulkDelete('r')">&#128465; Delete selected</button><button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="selClear('r')">Clear</button></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th style="width:34px"><input type="checkbox" id="r-selall" onclick="selAll('r',this)" title="Select all on this page"></th><th>#</th><th>Reference No</th><th>Student Name</th><th>Session</th><th>RC Comment / Remark</th><th>Action</th>
            </tr></thead><tbody id="r-body"></tbody></table>
          </div>
          <div class="pg-bar" id="r-pg"></div>
        </div>
      </section>

      <section id="sec-docreq" class="page-section">
        <div class="card">
          <h3>&#128172; Document Requests &mdash; Review &amp; Send on WhatsApp</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:14px;line-height:1.6">
            For every <b>Document Required</b> student, we read the NIOS remark and prepare a <b>simple, friendly message</b>
            (in your words, not NIOS's technical language). <b>Review each one, edit if needed, attach a demo screenshot if helpful, then send.</b> Nothing goes
            out automatically. Routing is automatic: <b>Public &#8594; public number</b>, <b>On Demand &amp; Stream 2 &#8594; main number</b>.</p>
          <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:14px">
            <button class="btn btn-outline btn-sm" onclick="loadDocReq(this)">Refresh</button>
            <button class="btn btn-sm" style="background:#16a34a;color:#fff" onclick="sendDocReq('selected',this)">Send selected</button>
            <button class="btn btn-sm" style="background:#2563eb;color:#fff" onclick="sendDocReq('all',this)">Send all pending</button>
            <label style="display:flex;align-items:center;gap:7px;font-size:13px;cursor:pointer;margin-left:4px"><input type="checkbox" id="dr-selall" onclick="toggleDrAll(this)"> Select all</label>
            <span id="dr-count" style="font-size:13px;color:var(--muted);margin-left:auto"></span>
          </div>
          <div id="dr-warn" style="display:none;background:#fff7ed;border:1px solid #fed7aa;border-radius:10px;padding:10px 13px;font-size:12.5px;color:#9a3412;margin-bottom:12px"></div>
          <div id="dr-body"></div>
          <div class="pg-bar" id="dr-pg"></div>
        </div>
      </section>

      <section id="sec-failed" class="page-section">
        <div class="card">
          <h3>Failed to Run — Wrong Data (NIOS login failed)</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            These students' NIOS portal could not be opened with the uploaded details (wrong DOB / Reference / Enrollment No),
            so <b>no WhatsApp/document was sent</b> (the link would open to an error and panic the student).
            Click <b>Edit &amp; fix</b>, correct the detail, then <b>Save &amp; Run again</b> — only that one student re-runs.
            If the data is correct it leaves this list and moves to its normal category.<br>
            <span style="color:var(--success);font-weight:600">Auto-fix:</span> every run now auto-retries transient errors and auto-tries a DOB date/month swap.
            Use <b>Re-check all (auto-fix)</b> below to apply this to the current list in one go.</p>
          <div class="filter-bar">
            <input type="text" id="f-search" placeholder="Search name / reference / email..." oninput="debounceFailed()">
            <select id="f-source" onchange="loadFailed(1)"><option value="">All Data Types</option><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <button class="btn btn-outline btn-sm" onclick="loadFailed(1)">Refresh</button>
            <button class="btn btn-success btn-sm" id="f-runall-btn" onclick="runAllFailed(this)" title="Re-run every failed student with auto-retry + DOB date/month auto-swap">Re-check all (auto-fix)</button>
            <button class="btn btn-outline btn-sm" onclick="diagnoseLogin()" title="Check why NIOS login is failing (site-key / bounce / captcha)">&#128295; Diagnose NIOS login</button>
          </div>
          <div id="f-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div id="f-bulkbar" style="display:none;align-items:center;gap:12px;flex-wrap:wrap;background:#FEF2F2;border:1px solid #FECACA;border-radius:10px;padding:10px 14px;margin-bottom:12px"><span style="font-weight:700;font-size:13px"><span id="f-selcount">0</span> selected</span><button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="bulkDelete('f')">&#128465; Delete selected</button><button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="selClear('f')">Clear</button></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th style="width:34px"><input type="checkbox" id="f-selall" onclick="selAll('f',this)" title="Select all on this page"></th><th>Student</th><th>Reference / Enroll</th><th>Session</th><th>Problem</th><th>Action</th>
            </tr></thead><tbody id="f-body"></tbody></table>
          </div>
          <div class="pg-bar" id="f-pg"></div>
        </div>
      </section>

      <section id="sec-unknown" class="page-section">
        <div class="card">
          <div class="card-head">
            <h3>Unknown Status</h3>
            <button class="btn btn-success btn-sm" id="u-runall-btn" onclick="runAllUnknown(this)" title="Re-check only the Unknown students (with auto-retry)">Run Unknown only</button>
          </div>
          <p style="color:var(--muted);font-size:13px;margin-bottom:14px">
            NIOS returned <b>no recognisable status</b> for these students — usually a wrong or not-yet-active Reference No,
            or a status NIOS shows that isn't tracked yet. The <b>reason</b> is shown for each. Fix the reference if needed, or
            click <b>Run Unknown only</b> to re-check just these (errors often clear on a retry).</p>
          <div class="filter-bar">
            <input type="text" id="u-search" placeholder="Search name / reference / email..." oninput="debounceUnknown()">
            <select id="u-session" onchange="loadUnknown(1)"><option value="">All Sessions</option></select>
            <button class="btn btn-outline btn-sm" onclick="loadUnknown(1)">Refresh</button>
          </div>
          <div id="u-count" style="font-size:13px;color:var(--muted);margin-bottom:14px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>Student</th><th>Reference / Enroll</th><th>Session</th><th>Why Unknown</th><th>Last Checked</th><th>Action</th>
            </tr></thead><tbody id="u-body"></tbody></table>
          </div>
          <div class="pg-bar" id="u-pg"></div>
        </div>
      </section>

      <section id="sec-history" class="page-section">
        <div class="card">
          <h3>Status Change History</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:14px">
            Choose a date &amp; time range — e.g. use "Custom range" to view confirmations from yesterday 6:30 PM until now.</p>
          <div class="filter-bar">
            <select id="h-preset" onchange="onHistPreset()">
              <option value="all">All time</option>
              <option value="today">Today</option>
              <option value="yesterday">Yesterday</option>
              <option value="24h">Last 24 hours</option>
              <option value="custom">Custom range…</option>
            </select>
            <select id="h-status" onchange="loadHistory(1)">
              <option value="">All status changes</option>
              <option>Admission Confirmed</option><option>Verified</option>
              <option>Documents Verification In Progress</option><option>Document Required</option>
              <option>Approved</option><option>Rejected</option>
              <option>Unknown</option><option>Fetch Error</option>
            </select>
            <select id="h-source" onchange="loadHistory(1)"><option value="">All Data Types</option><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <input id="h-search" type="text" placeholder="Search name / reference / email…" oninput="histSearchDebounced()"
              style="padding:9px 12px;border:2px solid var(--border);border-radius:9px;font-size:13.5px;min-width:230px">
            <button class="btn btn-outline btn-sm" onclick="exportHistory()">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="15" height="15"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Export Excel</button>
            <button class="btn btn-sm" style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca" onclick="clearHistory()">&#128465; Clear All</button>
          </div>
          <div class="filter-bar" id="h-custom" style="display:none">
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">From
              <input type="datetime-local" id="h-from"></label>
            <label style="font-size:13px;color:var(--muted);display:flex;align-items:center;gap:8px">To
              <input type="datetime-local" id="h-to"></label>
            <button class="btn btn-primary btn-sm" onclick="loadHistory(1)">Apply</button>
          </div>
          <div id="h-count" style="font-size:13px;color:var(--muted);margin-bottom:12px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>Reference No</th><th>Student</th><th>Old Status</th><th>New Status</th><th>Changed At</th><th>Action</th>
            </tr></thead><tbody id="h-body"></tbody></table>
          </div>
          <div class="pg-bar" id="h-pg"></div>
        </div>
      </section>

      <section id="sec-runlogs" class="page-section">
        <div class="card">
          <div class="card-head">
            <h3>Run Logs</h3>
            <div style="display:flex;gap:8px;align-items:center">
              <select id="rl-limit" onchange="loadRunLogs()" style="padding:6px 10px;border:2px solid var(--border);border-radius:8px;font-size:13px">
                <option value="10">10</option><option value="20">20</option><option value="30">30</option>
                <option value="50" selected>50</option><option value="100">100</option></select>
              <button class="btn btn-sm" style="background:#fee2e2;color:#b91c1c" onclick="clearRunLogs()">Clear All</button>
            </div>
          </div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>Run At</th><th>Type</th><th>Checked</th><th>Changed</th><th>Failed</th><th>Status</th><th>Action</th>
            </tr></thead><tbody id="rl-body"></tbody></table>
          </div>
        </div>
      </section>

      <section id="sec-transfers" class="page-section">
        <div class="card" style="border:1px solid #fde68a;background:#fffdf5">
          <div class="card-head">
            <h3>&#128465; Old / Removed Portal Students</h3>
            <button class="btn btn-outline btn-sm" id="stale-find-btn" onclick="findStalePortal(this)">Find old portal data</button>
          </div>
          <p style="color:var(--muted);font-size:13px;margin-bottom:6px">
            Finds students we still track as <b>MVS Portal</b> that are <b>no longer on the live portal</b> — e.g. an old batch
            you removed from the portal. It only <b>previews</b> them; nothing is deleted until you click <b>Move to Trash</b>
            (recoverable from Settings &#8594; Trash any time).</p>
          <div id="stale-box" style="margin-top:10px"></div>
        </div>
        <div class="card">
          <div class="card-head">
            <h3>Transfer Data — Tracker &#8594; Portal</h3>
            <div style="display:flex;gap:8px;flex-wrap:wrap">
              <button class="btn btn-primary btn-sm" id="tr-match-btn" onclick="matchTransfers(this)" title="Match Portal data to your already-checked Tracker data by Reference No and push it to the Portal WITHOUT using CapSolver">&#9889; Match &amp; Transfer (no CapSolver)</button>
              <button class="btn btn-success btn-sm" id="tr-sync-btn" onclick="syncTransfers(this)" title="Push the current status of every transferred Portal student to MVS Portal now">Sync matched to Portal</button>
            </div>
          </div>
          <p style="color:var(--muted);font-size:13px;margin-bottom:10px">
            <b>Match &amp; Transfer</b> matches your already-checked Tracker data to the Portal <b>by Reference No</b> and pushes the
            status across <b>without using CapSolver</b> — so data you already checked never costs credits again. Unmatched students
            are listed with a reason and are left for the normal <b>New Fetch</b>. Every transfer is logged below (Old &#8594; New status).</p>
          <div id="tr-match-result" style="margin-bottom:10px"></div>
          <div class="filter-bar">
            <input id="tr-search" type="text" placeholder="Master search — name / reference / enrollment / mobile…" oninput="trSearchDebounced()"
              style="padding:9px 12px;border:2px solid var(--border);border-radius:9px;font-size:13.5px;min-width:280px">
            <select id="tr-mode" onchange="loadTransfers(1)">
              <option value="">All transfers</option>
              <option value="auto">Automatic (during runs)</option>
              <option value="manual">Manual (Sync button)</option>
            </select>
            <button class="btn btn-outline btn-sm" onclick="loadTransfers(1)">Refresh</button>
            <button class="btn btn-outline btn-sm" onclick="downloadTransfers()" title="Download the Transfer Data sheet (.xlsx)">&#11015; Download sheet</button>
            <button class="btn btn-outline btn-sm" onclick="clearTransferLog()" title="Reset this transfer history list. Students already transferred to the Portal stay transferred — only the log is cleared.">&#128465; Clear log</button>
          </div>
          <div id="tr-count" style="font-size:13px;color:var(--muted);margin-bottom:12px"></div>
          <div style="overflow-x:auto">
            <table><thead><tr>
              <th>Student</th><th>Reference / Enroll</th><th>Session</th><th>Old Status</th><th>New Status</th><th>Transferred At</th><th>Mode</th>
            </tr></thead><tbody id="tr-body"></tbody></table>
          </div>
          <div class="pg-bar" id="tr-pg"></div>
        </div>
      </section>

      <section id="sec-upload" class="page-section">
        <div class="card">
          <h3>Upload Student Excel</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:14px">
            Upload an Excel file (.xlsx). Columns: Student Name, Mobile No, Class, Reference Number, Enrol No, Email, Date of Birth, Admission Session.</p>
          <div style="background:var(--soft);border:1px solid var(--border);border-radius:12px;padding:14px 16px;margin-bottom:18px">
            <div style="font-size:13px;font-weight:600;margin-bottom:4px">&#128196; Not sure about the format? Download a sample sheet:</div>
            <div style="font-size:12px;color:var(--muted);margin-bottom:10px">Fill your data in the same column order, then upload.</div>
            <div style="display:flex;gap:10px;flex-wrap:wrap">
              <button class="btn btn-outline btn-sm" onclick="downloadSample('regular')">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="14" height="14" style="vertical-align:-2px"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
                Sample — Regular (On Demand / Stream 2 / April-October)</button>
              <button class="btn btn-outline btn-sm" onclick="downloadSample('syc')">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="14" height="14" style="vertical-align:-2px"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
                Sample — SYC Students</button>
            </div>
          </div>
          <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:14px">
            <span style="font-size:13px;font-weight:600">Data source for this sheet:</span>
            <select id="upload-source" onchange="saveUploadSource()" style="max-width:260px">
              <option value="">Auto-detect (recommended)</option>
              <option value="mvs_portal">MVS Portal</option>
              <option value="mvs_tracker">MVS Tracker</option>
            </select>
            <span style="font-size:11px;color:var(--muted)">Applied on the next Run Now</span>
          </div>
          <div class="drop" id="drop" onclick="document.getElementById('file-input').click()">
            <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg></div>
            <div style="font-weight:600;font-size:15px">Click or drag Excel file here</div>
            <div style="color:var(--muted);font-size:13px;margin-top:5px">.xlsx files only</div>
          </div>
          <input type="file" id="file-input" accept=".xlsx,.xls" style="display:none" onchange="handleFile(this.files[0])">
          <div id="upload-status" style="margin-top:16px"></div>
          <div id="upload-summary" style="margin-top:16px"></div>
          <div id="upload-preview" style="margin-top:16px"></div>
          <div style="margin-top:18px;display:flex;gap:12px">
            <button class="btn btn-outline btn-sm" onclick="downloadExcel()">Download Updated Excel</button>
          </div>
        </div>

        <div class="card">
          <h3>Bulk Alternate Numbers</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:14px">
            Forgot to add alternate WhatsApp numbers in your sheet? Upload a small sheet here with just two
            columns — <b>Reference No</b> and <b>Alternate Number</b> — to add/update them in bulk. Students are
            matched by reference number. For anyone already <b>confirmed</b>, the documents are sent to the new
            alternate number automatically (the primary number is not messaged again).</p>
          <div style="background:var(--soft);border:1px solid var(--border);border-radius:12px;padding:14px 16px;margin-bottom:18px">
            <div style="font-size:13px;font-weight:600;margin-bottom:4px">&#128196; Need the format? Download a sample:</div>
            <div style="font-size:12px;color:var(--muted);margin-bottom:10px">Two columns: Reference No, Alternate Number. Fill and upload.</div>
            <button class="btn btn-outline btn-sm" onclick="downloadAltSample()">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="14" height="14" style="vertical-align:-2px"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
              Sample — Alternate Numbers</button>
          </div>
          <div class="drop" id="alt-drop" onclick="document.getElementById('alt-file-input').click()">
            <div class="ic"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg></div>
            <div style="font-weight:600;font-size:15px">Click or drag the alternate-numbers sheet here</div>
            <div style="color:var(--muted);font-size:13px;margin-top:5px">.xlsx — Reference No + Alternate Number</div>
          </div>
          <input type="file" id="alt-file-input" accept=".xlsx,.xls" style="display:none" onchange="handleAltFile(this.files[0])">
          <div id="alt-status" style="margin-top:16px"></div>
        </div>
      </section>

      <section id="sec-settings" class="page-section">
        <div class="card" style="border:2px solid #FECACA;background:#FEF2F2">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#DC2626,#B91C1C)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 9v4M12 17h.01M10.29 3.86 1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/></svg></span>
            <div class="set-tt"><h3>NIOS Login Diagnostic</h3>
            <div class="set-sub">Logins failing? Enter a <b>known-confirmed</b> student's Reference No + DOB and click Run — it tells you the REAL cause (site-key change, captcha token, or NIOS rejection).</div></div>
          </div>
          <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-top:6px">
            <input id="dg-ref" placeholder="Reference No" style="padding:9px 11px;border:1px solid var(--border);border-radius:8px;font-size:13px">
            <input id="dg-dob" placeholder="DOB (DD-MM-YYYY)" style="padding:9px 11px;border:1px solid var(--border);border-radius:8px;font-size:13px">
            <button class="btn-primary" onclick="diagnoseLogin()" style="padding:9px 16px">Run diagnostic</button>
          </div>
          <pre id="dg-out" style="display:none;margin-top:10px;background:#0f172a;color:#e2e8f0;padding:12px;border-radius:8px;font-size:12px;white-space:pre-wrap;word-break:break-word;max-height:320px;overflow:auto"></pre>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#4F46E5,#7C3AED)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"/><polyline points="12 7 12 12 15 14"/></svg></span>
            <div class="set-tt"><h3>Recheck Intervals</h3>
            <div class="set-sub">Set a separate auto-run interval for each group. Confirmed students are skipped automatically.</div></div>
          </div>
          <div class="iv-row">
            <span class="iv-lbl"><span class="iv-dot" style="background:#0EA5E9"></span>On Demand</span>
            <input type="number" id="iv-ondemand" min="1" max="43200" value="6">
            <select id="iv-ondemand-unit"><option value="hours">hours</option><option value="minutes">minutes</option><option value="days">days</option></select>
            <select id="iv-ondemand-src" title="Which data this group auto-runs"><option value="both">Both sources</option><option value="mvs_tracker">Tracker only</option><option value="mvs_portal">Portal only</option></select>
          </div>
          <div class="iv-row">
            <span class="iv-lbl"><span class="iv-dot" style="background:#F59E0B"></span>Stream 2</span>
            <input type="number" id="iv-stream2" min="1" max="43200" value="6">
            <select id="iv-stream2-unit"><option value="hours">hours</option><option value="minutes">minutes</option><option value="days">days</option></select>
            <select id="iv-stream2-src" title="Which data this group auto-runs"><option value="both">Both sources</option><option value="mvs_tracker">Tracker only</option><option value="mvs_portal">Portal only</option></select>
          </div>
          <div class="iv-row">
            <span class="iv-lbl"><span class="iv-dot" style="background:#7C3AED"></span>Public Exam <span style="font-weight:400;color:var(--muted);font-size:12px">(April / Oct)</span></span>
            <input type="number" id="iv-public" min="1" max="43200" value="12">
            <select id="iv-public-unit"><option value="hours">hours</option><option value="minutes">minutes</option><option value="days">days</option></select>
            <select id="iv-public-src" title="Which data this group auto-runs"><option value="both">Both sources</option><option value="mvs_tracker">Tracker only</option><option value="mvs_portal">Portal only</option></select>
          </div>
          <div class="iv-row" style="background:#f0fdf4;border-radius:9px;padding:8px 10px;margin-top:4px">
            <span class="iv-lbl"><span class="iv-dot" style="background:#16A34A"></span>MVS Portal — New data <span style="font-weight:400;color:var(--muted);font-size:12px">(saves credits)</span></span>
            <input type="number" id="iv-portalnew" min="1" max="43200" value="3">
            <select id="iv-portalnew-unit"><option value="hours">hours</option><option value="minutes">minutes</option><option value="days">days</option></select>
            <span style="font-size:11.5px;color:var(--muted);align-self:center">Checks only newly-arrived portal students</span>
          </div>
          <div class="set-foot">
            <button class="btn btn-primary btn-sm" onclick="saveIntervals()">Save Intervals</button>
            <div class="set-note">Range: 15 minutes to 30 days (use the <b>days</b> unit to run every few days). Shorter intervals use more CapSolver credits.</div>
            <div id="iv-status" style="margin-top:10px;font-size:13px"></div>
          </div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#10B981,#059669)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="22" y1="2" x2="11" y2="13"/><polygon points="22 2 15 22 11 13 2 9 22 2"/></svg></span>
            <div class="set-tt"><h3>Run Report on WhatsApp</h3>
            <div class="set-sub">After every run, get a WhatsApp summary (Confirmed, Document Required, Error and unchanged counts) plus a link to download the full Excel. Sent to up to 10 numbers.</div></div>
          </div>
          <label style="display:flex;align-items:center;gap:10px;font-size:14px;font-weight:600;margin-bottom:16px;cursor:pointer">
            <input type="checkbox" id="rep-enabled" style="width:18px;height:18px"> Send report after each run
          </label>
          <label style="display:block;font-size:13px;font-weight:600;margin-bottom:6px">WhatsApp numbers <span style="font-weight:400;color:var(--muted)">(one per line or comma-separated)</span></label>
          <textarea id="rep-numbers" rows="3" placeholder="9876543210&#10;9123456789&#10;9988776655"
            style="width:100%;padding:11px 13px;border:2px solid var(--border);border-radius:10px;font-size:14px;font-family:inherit;resize:vertical;background:var(--card);color:var(--text)"></textarea>
          <div class="set-note" style="margin-bottom:16px">10-digit Indian numbers (country code 91 added automatically). Report template must be set up in AiSensy (campaign <b>AISENSY_CAMPAIGN_REPORT</b>).</div>
          <div style="display:flex;gap:10px;flex-wrap:wrap">
            <button class="btn btn-primary btn-sm" onclick="saveReportSettings()">Save</button>
            <button class="btn btn-primary btn-sm" style="background:#16A34A;border-color:#16A34A" onclick="sendLatestReport(this)">Send a run report to all</button>
            <button class="btn btn-outline btn-sm" onclick="testReport(this)">Send test report now</button>
          </div>
          <div id="rep-last" style="margin-top:12px;font-size:12.5px;color:var(--muted)"></div>
          <div id="rep-status" style="margin-top:8px;font-size:13px"></div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#0EA5E9,#0369A1)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="17 1 21 5 17 9"/><path d="M3 11V9a4 4 0 0 1 4-4h14"/><polyline points="7 23 3 19 7 15"/><path d="M21 13v2a4 4 0 0 1-4 4H3"/></svg></span>
            <div class="set-tt"><h3>Change Data Source (MVS Portal &harr; MVS Tracker)</h3>
            <div class="set-sub">Wrongly tagged on upload (e.g. auto-detect put a Tracker sheet under MVS Portal)? Move students from one data type to the other. <b>No data is lost</b> — only the tag changes.</div></div>
            <div class="set-act"><button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="loadSourceCounts()">Refresh counts</button></div>
          </div>
          <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:10px">
            <span style="font-size:13px">Move&nbsp;all from</span>
            <select id="src-from" style="max-width:170px"><option value="mvs_portal">MVS Portal</option><option value="mvs_tracker">MVS Tracker</option></select>
            <span style="font-size:13px">&rarr; to</span>
            <select id="src-to" style="max-width:170px"><option value="mvs_tracker">MVS Tracker</option><option value="mvs_portal">MVS Portal</option></select>
            <button class="btn btn-primary btn-sm" onclick="moveAllSource(this)">Move all</button>
          </div>
          <div id="src-counts" style="font-size:12.5px;color:var(--muted)">Current: MVS Portal — … &middot; MVS Tracker — …</div>
          <div id="src-status" style="margin-top:8px;font-size:13px"></div>
          <div style="font-size:11.5px;color:var(--muted);margin-top:8px;line-height:1.5">Tip: to move only <b>some</b> students, go to the <b>Students</b> tab, filter by Data Type, tick the ones you want, then use <b>Change source</b> in the select bar.</div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#EF4444,#DC2626)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"/><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/></svg></span>
            <div class="set-tt"><h3>Deleted Students (Trash)</h3>
            <div class="set-sub">Removed students land here. <b>Restore</b> brings them back; <b>Delete permanently</b> cannot be undone.</div></div>
            <div class="set-act"><button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="loadTrash()">Refresh</button></div>
          </div>
          <div id="trash-bulkbar" style="display:none;align-items:center;gap:10px;flex-wrap:wrap;background:#FEF2F2;border:1px solid #FECACA;border-radius:10px;padding:10px 14px;margin-bottom:12px">
            <span style="font-weight:700;font-size:13px"><span id="trash-selcount">0</span> selected</span>
            <button class="btn btn-sm" style="background:#16A34A;color:#fff" onclick="restoreSelected()">&#8617; Restore selected</button>
            <button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="purgeSelected()">&#128465; Permanently delete selected</button>
            <button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="selClear('trash')">Clear</button>
          </div>
          <div style="overflow-x:auto"><table>
            <thead><tr><th style="width:34px"><input type="checkbox" id="trash-selall" onclick="selAll('trash',this)" title="Select all"></th><th>Student</th><th>Reference / Enroll</th><th>Session</th><th>Deleted At</th><th>Action</th></tr></thead>
            <tbody id="trash-body"><tr><td colspan="6" class="empty">Loading…</td></tr></tbody>
          </table></div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#22C55E,#16A34A)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 11.5a8.38 8.38 0 0 1-.9 3.8 8.5 8.5 0 0 1-7.6 4.7 8.38 8.38 0 0 1-3.8-.9L3 21l1.9-5.7a8.38 8.38 0 0 1-.9-3.8 8.5 8.5 0 0 1 4.7-7.6 8.38 8.38 0 0 1 3.8-.9h.5a8.48 8.48 0 0 1 8 8v.5z"/></svg></span>
            <div class="set-tt"><h3>WhatsApp Auto-Send</h3>
            <div class="set-sub">When a student becomes <b>Admission Confirmed</b>, their documents are auto-sent to their WhatsApp as secure links (each student receives this <b>only once</b>).</div></div>
          </div>
          <div id="wa-config" style="font-size:13px;margin-bottom:14px;color:var(--muted)">Loading…</div>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;margin-bottom:18px">
            <input type="checkbox" id="wa-enabled" onchange="saveWa()" style="width:18px;height:18px;cursor:pointer">
            <span style="font-weight:600">Enable auto-send</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;margin-bottom:10px">
            <input type="checkbox" id="wa-required-enabled" onchange="saveWa()" style="width:18px;height:18px;cursor:pointer">
            <span style="font-weight:600">Enable <b>Document Requests</b> (counsellor-reviewed reminders)</span>
          </label>
          <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:11px 13px;font-size:12.5px;color:#166534;line-height:1.6;margin-bottom:18px">
            When ON, every <b>Document Required</b> student appears on the <b>Doc Requests</b> page with a simple, friendly
            message prepared from the NIOS remark. <b>The counsellor reviews/edits and sends manually</b> &mdash; nothing goes
            out automatically. Routing: <b>Public &#8594; public number</b>; <b>On Demand &amp; Stream 2 &#8594; main number</b>.<br>
            <span style="color:#15803d">Needs just <b>2 image-header templates</b> ({{1}} = name, {{2}} = document request) &mdash; one per account &mdash; with campaigns in Railway:
            <b>AISENSY_CAMPAIGN_REQUIRED</b> (main account) and <b>AISENSY_CAMPAIGN_REQUIRED_PUBLIC</b> (public account). Every message carries an image: the attached screenshot, or a default MVS banner when none.</span>
          </div>
          <div style="border-top:1px solid var(--border);padding-top:16px">
            <p style="color:var(--muted);font-size:13px;margin-bottom:10px">
              <b>Send a test message</b> — choose a template, enter your number, and a sample message will be sent:</p>
            <div style="display:flex;gap:10px;flex-wrap:wrap">
              <select id="wa-group" style="padding:11px;border:2px solid var(--border);border-radius:10px;font-size:14px">
                <option value="ondemand">On Demand (3inone)</option>
                <option value="stream2">Stream 2 (str2toc1)</option>
                <option value="public">Public April/October</option>
                <option value="syc">SYC (Hall Ticket)</option>
              </select>
              <input type="text" id="wa-num" placeholder="WhatsApp number (e.g. 7065187637)"
                style="flex:1;min-width:180px;padding:11px;border:2px solid var(--border);border-radius:10px;font-size:14px">
              <button class="btn btn-primary btn-sm" onclick="testWa()">Send Test</button>
            </div>
          </div>
          <div id="wa-status" style="margin-top:12px;font-size:13px"></div>
          <div style="border-top:1px solid var(--border);margin-top:16px;padding-top:16px">
            <div style="font-weight:700;font-size:13.5px;margin-bottom:6px">Delivery confirmation (recommended)</div>
            <p style="color:var(--muted);font-size:12.5px;line-height:1.6;margin-bottom:10px">
              AiSensy accepting a message does <b>not</b> guarantee the student received it. Paste this
              <b>Webhook URL</b> into AiSensy (Manage → API/Webhooks → Delivery/Status webhook). Then the
              Confirmed list shows a real <b>Delivered &#10003;</b> or <b>Delivery FAILED</b> per student.</p>
            <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
              <input type="text" id="wh-url" readonly style="flex:1;min-width:220px;padding:10px 12px;border:2px solid var(--border);border-radius:10px;font-size:12.5px;background:var(--soft);color:var(--text)">
              <button class="btn btn-outline btn-sm" onclick="copyWebhook(this)">Copy</button>
            </div>
          </div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#F59E0B,#D97706)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="13.5" cy="6.5" r="2.5"/><circle cx="17.5" cy="10.5" r="2.5"/><circle cx="8.5" cy="7.5" r="2.5"/><circle cx="6.5" cy="12.5" r="2.5"/><path d="M12 2a10 10 0 1 0 0 20 1.5 1.5 0 0 0 1-2.6 1.5 1.5 0 0 1 1-2.6h2A5.6 5.6 0 0 0 22 11 10 10 0 0 0 12 2z"/></svg></span>
            <div class="set-tt"><h3>Status Colour Legend</h3>
            <div class="set-sub">What each status colour means across the portal.</div></div>
          </div>
          <div class="legend-grid">
            <div class="legend-item b-pending"><div class="nm">Pending</div><div class="ds">Awaiting review</div></div>
            <div class="legend-item b-docs"><div class="nm">Documents Verification In Progress</div><div class="ds">Under review</div></div>
            <div class="legend-item b-required"><div class="nm">Document Required</div><div class="ds">Action needed by counsellor</div></div>
            <div class="legend-item b-verified"><div class="nm">Verified</div><div class="ds">Documents verified</div></div>
            <div class="legend-item b-approved"><div class="nm">Approved</div><div class="ds">Application approved</div></div>
            <div class="legend-item b-confirmed"><div class="nm">Admission Confirmed</div><div class="ds">Admission confirmed</div></div>
            <div class="legend-item b-rejected"><div class="nm">Rejected</div><div class="ds">Application rejected</div></div>
            <div class="legend-item b-syc"><div class="nm">SYC</div><div class="ds">SYC student — hall ticket ready</div></div>
          </div>
        </div>
        <div class="card">
          <div class="set-head">
            <span class="set-ico" style="background:linear-gradient(135deg,#64748B,#475569)"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14.7 6.3a1 1 0 0 0 0 1.4l1.6 1.6a1 1 0 0 0 1.4 0l3.77-3.77a6 6 0 0 1-7.94 7.94l-6.91 6.91a2.12 2.12 0 0 1-3-3l6.91-6.91a6 6 0 0 1 7.94-7.94l-3.76 3.76z"/></svg></span>
            <div class="set-tt"><h3>Test Login &amp; Find Download Links</h3>
            <div class="set-sub">Enter a <b>confirmed student's</b> Reference No + DOB to log in and locate their document download links (for verification).</div></div>
          </div>
          <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px">
            <input type="text" id="dbg-ref" placeholder="Reference No (e.g. D0026300046)"
              style="flex:1;min-width:180px;padding:11px;border:2px solid var(--border);border-radius:10px;font-size:14px">
            <input type="text" id="dbg-dob" placeholder="DOB DD-MM-YYYY (e.g. 08-08-2007)"
              style="flex:1;min-width:180px;padding:11px;border:2px solid var(--border);border-radius:10px;font-size:14px">
          </div>
          <button class="btn btn-primary btn-sm" onclick="testLogin()">Test Login & Find Links</button>
          <div style="margin-top:16px;padding-top:16px;border-top:1px solid var(--border)">
            <p style="color:var(--muted);font-size:13px;margin-bottom:10px">
              <b>Inspect document page</b> — run this if a download fails (to see how the PDF is embedded):</p>
            <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center">
              <select id="dbg-kind" style="padding:10px;border:2px solid var(--border);border-radius:10px;font-size:14px">
                <option value="id_card">ID Card</option>
                <option value="app_form">Application Form</option>
                <option value="hall_ticket">Hall Ticket</option>
              </select>
              <button class="btn btn-outline btn-sm" onclick="inspectDoc()">Inspect Doc Page</button>
              <button class="btn btn-outline btn-sm" onclick="findAddr()">Find RC Address (Stream 2)</button>
            </div>
          </div>
          <div id="dbg-status" style="margin-top:12px;font-size:13px"></div>
          <pre id="dbg-result" style="margin-top:12px;background:#0F172A;color:#A5F3FC;padding:16px;
            border-radius:10px;font-size:12px;overflow-x:auto;max-height:400px;display:none;white-space:pre-wrap"></pre>
        </div>
        <div class="card" style="border:1px solid var(--danger)">
          <h3 style="color:var(--danger)">Danger Zone — Clear All Data</h3>
          <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
            This will <b>permanently delete</b> all students, statuses, change history and the uploaded sheet.
            Your settings (intervals, WhatsApp) will be kept. You will need to upload a new sheet afterwards.</p>
          <button class="btn btn-sm" style="background:var(--danger);color:#fff" onclick="resetData()">Clear All Data</button>
          <div id="reset-status" style="margin-top:12px;font-size:13px"></div>
        </div>
      </section>
    </div>
  </div>
</div>

<div id="toast"></div>

<div id="edit-overlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:1000;align-items:center;justify-content:center;padding:16px">
  <div style="background:var(--card,#fff);border-radius:16px;max-width:480px;width:100%;max-height:90vh;overflow:auto;box-shadow:0 20px 60px rgba(0,0,0,.3)">
    <div style="padding:18px 22px;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center">
      <h3 style="margin:0">&#9998; Edit Student</h3>
      <button onclick="closeEdit()" style="background:none;border:none;font-size:22px;cursor:pointer;color:var(--muted)">&times;</button>
    </div>
    <div style="padding:20px 22px">
      <div id="edit-warn" style="display:none;font-size:12.5px;font-weight:600;color:#b91c1c;background:#fee2e2;border:1px solid #fecaca;border-radius:8px;padding:8px 10px;margin-bottom:14px"></div>
      <input type="hidden" id="edit-rowkey">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <label style="font-size:12px;font-weight:600;grid-column:1/3">Student Name<input id="edit-student_name" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Reference No<input id="edit-reference_no" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Enrollment No<input id="edit-enrollment_no" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Date of Birth<input id="edit-dob" class="edit-inp" placeholder="DD-MM-YYYY"></label>
        <label style="font-size:12px;font-weight:600">Mobile No<input id="edit-mobile" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Alternate No <span style="color:var(--muted);font-weight:400">(optional — WhatsApp goes to both)</span><input id="edit-alt_mobile" class="edit-inp" placeholder="2nd number (optional)"></label>
        <label style="font-size:12px;font-weight:600;grid-column:1/3">Email<input id="edit-email" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Class<input id="edit-class_level" class="edit-inp"></label>
        <label style="font-size:12px;font-weight:600">Session<input id="edit-session" class="edit-inp"></label>
      </div>
      <div style="font-size:11.5px;color:var(--muted);margin-top:10px">Tip: if NIOS login failed, fix the <b>Date of Birth</b> or <b>Reference/Enrollment No</b>, then "Save &amp; Run again" to re-check and send.</div>
      <div id="edit-syncline" style="font-size:12px;font-weight:600;margin-top:8px"></div>
      <div id="edit-status" style="font-size:12.5px;font-weight:600;margin-top:10px"></div>
    </div>
    <div style="padding:14px 22px;border-top:1px solid var(--border);display:flex;gap:10px;justify-content:flex-end">
      <button class="btn" style="background:var(--soft);color:var(--text)" onclick="closeEdit()">Cancel</button>
      <button class="btn" style="background:#64748b;color:#fff" onclick="saveEdit(false)">Save</button>
      <button class="btn btn-primary" onclick="saveEdit(true)">&#10227; Save &amp; Run again</button>
    </div>
  </div>
</div>
<style>.edit-inp{width:100%;margin-top:4px;padding:9px 11px;border:2px solid var(--border);border-radius:9px;font-size:14px;font-weight:400}
#run-menu{display:none;position:absolute;right:0;top:calc(100% + 10px);background:var(--card,#fff);border:1px solid var(--border);border-radius:16px;box-shadow:0 20px 50px rgba(15,23,42,.20),0 3px 10px rgba(15,23,42,.07);min-width:280px;z-index:900;overflow:hidden;padding:8px}
.rm-head{display:flex;align-items:center;gap:7px;padding:7px 10px 9px;font-size:10.5px;font-weight:800;letter-spacing:.7px;color:var(--muted);text-transform:uppercase}
.rm-head svg{color:#F59E0B}
.run-menu-item{display:flex;align-items:center;width:100%;text-align:left;padding:9px 11px;border:none;background:none;cursor:pointer;border-radius:10px;transition:background .14s,transform .12s,box-shadow .14s}
.run-menu-item .rmi-main{font-size:13.5px;font-weight:600;color:var(--text)}
.run-menu-item .rmi-dot{width:7px;height:7px;border-radius:50%;margin:0 11px 0 5px;flex-shrink:0;background:var(--muted);transition:transform .14s}
.run-menu-item .rmi-tag{margin-left:auto;font-size:9.5px;font-weight:800;color:#fff;padding:2px 8px;border-radius:20px;text-transform:uppercase;letter-spacing:.4px}

/* Run-all hero */
.rmi-all{margin-bottom:8px;padding:11px;background:linear-gradient(135deg,rgba(16,185,129,.12),rgba(16,185,129,.04));border:1px solid rgba(16,185,129,.25)}
.rmi-all:hover{background:linear-gradient(135deg,rgba(16,185,129,.20),rgba(16,185,129,.09));box-shadow:0 4px 14px rgba(16,185,129,.18)}
.rmi-all .rmi-ic{display:inline-flex;align-items:center;justify-content:center;width:32px;height:32px;border-radius:9px;background:linear-gradient(135deg,#10B981,#059669);color:#fff;margin-right:12px;flex-shrink:0;box-shadow:0 3px 8px rgba(16,185,129,.35)}
.rmi-all .rmi-txt{display:flex;flex-direction:column;line-height:1.3}
.rmi-all .rmi-main{font-size:14px;font-weight:800}
.rmi-all .rmi-sub{font-size:11px;font-weight:500;color:var(--muted)}
.rmi-all .rmi-go{margin-left:auto;font-size:22px;font-weight:700;color:#10B981;line-height:1}

/* Source group cards */
.rm-group{border-radius:13px;padding:5px;margin-bottom:7px;border:1px solid var(--border)}
.rm-group:last-child{margin-bottom:0}
.rm-group.trk{background:linear-gradient(180deg,rgba(14,165,233,.07),rgba(14,165,233,.02));border-color:rgba(14,165,233,.20)}
.rm-group.por{background:linear-gradient(180deg,rgba(124,58,237,.07),rgba(124,58,237,.02));border-color:rgba(124,58,237,.20)}
.rm-ghead{display:flex;align-items:center;gap:9px;padding:7px 9px 8px}
.rm-gic{display:inline-flex;align-items:center;justify-content:center;width:28px;height:28px;border-radius:8px;color:#fff;flex-shrink:0}
.rm-group.trk .rm-gic{background:linear-gradient(135deg,#0EA5E9,#0284C7);box-shadow:0 2px 7px rgba(14,165,233,.32)}
.rm-group.por .rm-gic{background:linear-gradient(135deg,#7C3AED,#6D28D9);box-shadow:0 2px 7px rgba(124,58,237,.32)}
.rm-gname{font-size:13px;font-weight:800;letter-spacing:.2px}
.rm-group.trk .rm-gname{color:#0369A1}
.rm-group.por .rm-gname{color:#6D28D9}
.rm-ghint{margin-left:auto;font-size:9.5px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.4px;opacity:.75}
.rm-group .run-menu-item{padding:8px 10px}
.run-menu-item.trk .rmi-dot{background:#0EA5E9}
.run-menu-item.por .rmi-dot{background:#7C3AED}
.run-menu-item.trk .rmi-tag{background:#0EA5E9}
.run-menu-item.por .rmi-tag{background:#7C3AED}
.run-menu-item.trk:hover{background:#fff;box-shadow:0 3px 10px rgba(14,165,233,.16);transform:translateX(2px)}
.run-menu-item.por:hover{background:#fff;box-shadow:0 3px 10px rgba(124,58,237,.16);transform:translateX(2px)}
.run-menu-item.trk:hover .rmi-dot,.run-menu-item.por:hover .rmi-dot{transform:scale(1.45)}
#run-now-btn.open #run-caret{transform:rotate(180deg)}</style>

<script>
const API = window.location.origin;
let TOKEN = "";
let perPage = 20;
const SELECTED = new Set();   // hand-picked active students for a manual "Run Selected"
const SEL_MAX = 20;
function toggleSel(rk,cb){
  if(cb.checked){
    if(SELECTED.size>=SEL_MAX){ cb.checked=false; showToast("Maximum "+SEL_MAX+" students can be selected at once"); return; }
    SELECTED.add(rk);
  }else{ SELECTED.delete(rk); }
  updateSelBar();
}
function toggleSelectAll(cb){
  const boxes=document.querySelectorAll(".sel-cb");
  if(cb.checked){
    boxes.forEach(b=>{
      if(!b.checked){
        if(SELECTED.size>=SEL_MAX){ b.checked=false; return; }
        b.checked=true; SELECTED.add(b.dataset.rk);
      }
    });
    if(SELECTED.size>=SEL_MAX) showToast("Selected the first "+SEL_MAX+" (maximum) on this page");
  }else{
    boxes.forEach(b=>{ b.checked=false; SELECTED.delete(b.dataset.rk); });
  }
  updateSelBar();
}
function clearSelection(){
  SELECTED.clear();
  document.querySelectorAll(".sel-cb").forEach(b=>b.checked=false);
  const sa=document.getElementById("sel-all");if(sa)sa.checked=false;
  updateSelBar();
}
function updateSelBar(){
  const bar=document.getElementById("sel-bar");if(!bar)return;
  const n=SELECTED.size;
  bar.style.display=n>0?"flex":"none";
  const cnt=document.getElementById("sel-count");if(cnt)cnt.textContent=n;
}
async function runSelected(){
  const keys=[...SELECTED];
  if(keys.length<1){showToast("Select at least 1 student");return;}
  if(keys.length>SEL_MAX){showToast("Maximum "+SEL_MAX+" students");return;}
  const btn=document.getElementById("sel-run-btn");
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/run-selected","POST",{row_keys:keys});
    showToast(r.message+" — running in the background");
    clearSelection();
    const box=document.getElementById("run-progress");
    if(box){box.style.display="block";document.getElementById("pb-pct").textContent="0%";
      document.getElementById("pb-fill").style.width="0%";
      document.getElementById("pb-sub").textContent="Starting…";
      document.getElementById("pb-label").textContent="Checking selected students…";}
    startProgressPoll();
  }catch(e){showToast("Error: "+e.message);}
  finally{ setTimeout(()=>{if(btn){btn.disabled=false;btn.style.opacity="";}},3000); }
}
async function deleteSelectedStudents(){
  const keys=[...SELECTED];
  if(!keys.length){showToast("Select at least 1 student");return;}
  if(!confirm("Remove "+keys.length+" selected student(s)?\\n\\nThey move to Trash — you can restore from Settings → Deleted Students."))return;
  try{
    const r=await api("/api/students-delete-bulk","POST",{row_keys:keys});
    showToast("Moved "+(r.deleted||keys.length)+" student(s) to Trash");
    clearSelection();
    refreshAllTables();
  }catch(e){showToast("Error: "+e.message);}
}
let sessionsLoaded = false;

function toggleLgPass(){
  const inp=document.getElementById("lg-pass");
  const eye=document.getElementById("lg-eye");
  if(!inp)return;
  if(inp.type==="password"){
    inp.type="text";
    if(eye)eye.innerHTML='<path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19m-6.72-1.07a3 3 0 1 1-4.24-4.24"/><line x1="1" y1="1" x2="23" y2="23"/>';
  }else{
    inp.type="password";
    if(eye)eye.innerHTML='<path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"/><circle cx="12" cy="12" r="3"/>';
  }
}
async function doLogin(){
  const u=document.getElementById("lg-user").value;
  const p=document.getElementById("lg-pass").value;
  const err=document.getElementById("login-error");
  err.textContent="";
  try{
    const r=await fetch(API+"/api/login",{method:"POST",headers:{"Content-Type":"application/json"},
      body:JSON.stringify({username:u,password:p})});
    if(!r.ok){err.textContent="Invalid username or password";return;}
    const d=await r.json();
    TOKEN=d.token;
    document.getElementById("login-screen").style.display="none";
    document.getElementById("app").style.display="block";
    applySidebarPref();
    loadDashboard();
    loadNextRuns();
    startProgressPoll();
    setInterval(refreshBell,60000);
    setInterval(loadNextRuns,60000);
  }catch(e){err.textContent="Connection error: "+e.message;}
}

async function api(path,method="GET",body=null){
  const opt={method,headers:{"Authorization":"Bearer "+TOKEN}};
  if(body){opt.headers["Content-Type"]="application/json";opt.body=JSON.stringify(body);}
  const r=await fetch(API+path,opt);
  if(r.status===401){location.reload();throw new Error("Session expired");}
  if(!r.ok){const e=await r.json().catch(()=>({detail:"HTTP "+r.status}));throw new Error(e.detail||("HTTP "+r.status));}
  return r.json();
}

/* ---------- icons ---------- */
const DL_ICON='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="13" height="13" style="vertical-align:-2px"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>';
const SUN='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="5"/><line x1="12" y1="1" x2="12" y2="3"/><line x1="12" y1="21" x2="12" y2="23"/><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"/><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"/><line x1="1" y1="12" x2="3" y2="12"/><line x1="21" y1="12" x2="23" y2="12"/><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"/><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"/></svg>';
const MOON='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"/></svg>';
const CLOCK='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="18" height="18"><circle cx="12" cy="12" r="9"/><polyline points="12 7 12 12 15 14"/></svg>';

/* ---------- theme ---------- */
function applyTheme(t){
  document.documentElement.setAttribute("data-theme",t);
  const ic=document.getElementById("theme-ic"),lb=document.getElementById("theme-lbl");
  if(ic)ic.innerHTML=(t==="dark")?SUN:MOON;
  if(lb)lb.textContent=(t==="dark")?"Light Mode":"Dark Mode";
}
function toggleTheme(){
  const cur=(document.documentElement.getAttribute("data-theme")==="dark")?"light":"dark";
  try{localStorage.setItem("nios_theme",cur);}catch(e){}
  applyTheme(cur);
}
(function(){var t="light";try{t=localStorage.getItem("nios_theme")||"light";}catch(e){}applyTheme(t);})();

/* ---------- logout ---------- */
function doLogout(){
  if(!confirm("Log out of the portal?")) return;
  TOKEN="";stopProgressPoll();if(timerInt){clearInterval(timerInt);timerInt=null;}
  document.getElementById("app").style.display="none";
  document.getElementById("login-screen").style.display="flex";
  const p=document.getElementById("lg-pass");if(p)p.value="";
}

/* ---------- helpers ---------- */
function secActive(id){const e=document.getElementById("sec-"+id);return e&&e.classList.contains("active");}

/* ---------- live progress ---------- */
let progInt=null,wasRunning=false;
function startProgressPoll(){if(progInt)return;pollProgress();progInt=setInterval(pollProgress,2500);}
function stopProgressPoll(){if(progInt){clearInterval(progInt);progInt=null;}}
function setSrcRow(which,o){
  const row=document.getElementById("pb-"+which+"-row");
  if(row)row.style.display=(o.total>0)?"block":"none";
  const txt=document.getElementById("pb-"+which+"-txt");
  if(txt)txt.textContent=o.done+" / "+o.total+" ("+o.percent+"%)";
  const fill=document.getElementById("pb-"+which+"-fill");
  if(fill)fill.style.width=o.percent+"%";
}
async function pollProgress(){
  try{
    const d=await api("/api/progress");
    const box=document.getElementById("run-progress");if(!box)return;
    if(d.running){
      wasRunning=true;box.style.display="block";
      const pct=d.percent||0;
      document.getElementById("pb-pct").textContent=pct+"%";
      document.getElementById("pb-fill").style.width=pct+"%";
      document.getElementById("pb-sub").innerHTML=
        (d.current||0)+" / "+(d.total||0)+" checked &nbsp;·&nbsp; "+
        "<b style='color:var(--success)'>"+(d.changed||0)+"</b> changed &nbsp;·&nbsp; "+
        (d.same||0)+" same &nbsp;·&nbsp; "+
        "<b style='color:#B45309'>"+(d.not_checked||0)+"</b> not checked &nbsp;·&nbsp; "+
        "<b>"+(d.remaining||0)+"</b> remaining";
      // Separate progress for MVS Portal vs MVS Tracker (only show a bar if that
      // source actually has students in THIS run — so a Tracker-only run shows just Tracker)
      const mv=d.mvs||{done:0,total:0,percent:0}, tk=d.trk||{done:0,total:0,percent:0};
      const sw=document.getElementById("pb-srcwrap");
      if(sw){
        sw.style.display=(mv.total>0||tk.total>0)?"block":"none";
        setSrcRow("mvs",mv);setSrcRow("trk",tk);
      }
      const g=d.group_type||"all";
      if((d.phase||"main")==="retry" && (d.retry_total||0)>0){
        // Main pass done — now auto-retrying the fixable (captcha/proxy) failures.
        const rp=d.retry_percent||0;
        document.getElementById("pb-label").textContent="Auto-retrying failed students…";
        document.getElementById("pb-pct").textContent=rp+"%";
        document.getElementById("pb-fill").style.width=rp+"%";
        document.getElementById("pb-sub").innerHTML=
          "Re-checking students that failed (captcha/proxy) &nbsp;·&nbsp; "+
          "<b style='color:var(--success)'>"+(d.retry_done||0)+"</b> / "+(d.retry_total||0)+" resolved";
      } else {
        document.getElementById("pb-label").textContent="Checking "+g+" students…";
      }
      updateNavCounts();   // live badge updates as students move between tabs
      // live filtering: refresh whatever view the counsellor is on, as it happens
      if(secActive("dashboard"))loadDashboard();
      else if(secActive("confirmed"))loadConfirmed(1);
      else if(secActive("required"))loadRequired(1);
      else if(secActive("students"))loadStudents(1);
      else if(secActive("failed"))loadFailed(1);
    }else{
      box.style.display="none";
      if(wasRunning){wasRunning=false;
        updateFailedBadge();
        if(secActive("dashboard"))loadDashboard();
        if(secActive("runlogs"))loadRunLogs();
        if(secActive("failed"))loadFailed(1);}
    }
  }catch(e){}
}

/* ---------- next-run timers ---------- */
let timers=[],timerInt=null,pnTimer=null;
async function loadNextRuns(){
  try{
    const d=await api("/api/next-runs");
    const all=(d.runs||[]);
    timers=all.filter(r=>r.group!=="portalnew").map(r=>({label:r.label,group:r.group,paused:!!r.paused,
      remain:(r.seconds==null?null:r.seconds),at:Date.now()}));
    const pn=all.find(r=>r.group==="portalnew");
    pnTimer=pn?{paused:!!pn.paused,remain:(pn.seconds==null?null:pn.seconds),at:Date.now()}:null;
    renderTimers();
    if(!timerInt)timerInt=setInterval(renderTimers,1000);
  }catch(e){}
}
function fmtDur(s){
  if(s==null)return "Not scheduled";if(s<0)s=0;
  const h=Math.floor(s/3600),m=Math.floor((s%3600)/60),ss=s%60;
  let o="";if(h>0)o+=h+"h ";o+=(m<10&&h>0?"0":"")+m+"m ";o+=(ss<10?"0":"")+ss+"s";return o;
}
function renderTimers(){
  const el=document.getElementById("next-runs");
  if(el){if(!timers.length){el.innerHTML="";}else{
  el.innerHTML=timers.map(t=>{
    const paused=t.paused;
    // When paused, freeze the remaining time (don't subtract elapsed).
    let rem=(t.remain==null)?null:(paused?t.remain:Math.max(0,t.remain-Math.floor((Date.now()-t.at)/1000)));
    const soon=(!paused&&rem!=null&&rem<=1800);
    const btn=paused
      ? '<button class="tc-btn resume" onclick="resumeIntervals(&quot;'+t.group+'&quot;,this)" title="Resume this timer"><svg viewBox="0 0 24 24" fill="currentColor" width="12" height="12"><polygon points="5 3 19 12 5 21 5 3"/></svg> Resume</button>'
      : '<button class="tc-btn pause" onclick="pauseIntervals(&quot;'+t.group+'&quot;,this)" title="Pause this timer"><svg viewBox="0 0 24 24" fill="currentColor" width="11" height="11"><rect x="6" y="4" width="4" height="16"/><rect x="14" y="4" width="4" height="16"/></svg> Pause</button>';
    return '<div class="timer-chip'+(soon?' soon':'')+(paused?' paused':'')+'">'+
      '<div class="tc-ic">'+CLOCK+'</div>'+
      '<div style="flex:1;min-width:0">'+
        '<div class="tc-lbl">'+(paused?'Paused':('Next auto-run'+(soon?' — running soon':'')))+'</div>'+
        '<div class="tc-time">'+fmtDur(rem)+(paused?' <span style="font-size:11px;font-weight:600">(frozen)</span>':'')+'</div>'+
        '<div class="tc-grp">'+t.label+'</div>'+
      '</div>'+btn+'</div>';
  }).join("");
  }}
  updatePnTimer();
}
function updatePnTimer(){
  const t=document.getElementById("pn-timer"),b=document.getElementById("pn-btn");
  if(!t||!pnTimer)return;
  let rem=(pnTimer.remain==null)?null:(pnTimer.paused?pnTimer.remain:Math.max(0,pnTimer.remain-Math.floor((Date.now()-pnTimer.at)/1000)));
  t.textContent=(pnTimer.paused?"Paused · ":"")+fmtDur(rem);
  if(b){b.textContent=pnTimer.paused?"Resume":"Pause";b.disabled=false;b.style.opacity="";
    b.style.background=pnTimer.paused?"#dbeafe":"#dcfce7";b.style.color=pnTimer.paused?"#1d4ed8":"#15803d";
    b.style.borderColor=pnTimer.paused?"#bfdbfe":"#bbf7d0";}
}
function togglePn(btn){
  if(!pnTimer)return;
  if(pnTimer.paused)resumeIntervals("portalnew",btn);
  else pauseIntervals("portalnew",btn);
}
async function pauseIntervals(group,btn){
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{const r=await api("/api/intervals-pause","POST",{group:group});showToast(r.message);await loadNextRuns();}
  catch(e){showToast("Error: "+e.message);if(btn){btn.disabled=false;btn.style.opacity="";}}
}
async function resumeIntervals(group,btn){
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{const r=await api("/api/intervals-resume","POST",{group:group});showToast(r.message);await loadNextRuns();}
  catch(e){showToast("Error: "+e.message);if(btn){btn.disabled=false;btn.style.opacity="";}}
}

const titles={dashboard:"Dashboard",students:"Active Students",confirmed:"Confirmed Students",
  syc:"SYC Students",
  required:"Document Required",failed:"Failed to Run",unknown:"Unknown Status",docreq:"Document Requests",history:"Change History",runlogs:"Run Logs",transfers:"Transfer Data",upload:"Upload Excel",settings:"Settings"};
function refreshPage(btn){
  // reload the data of whichever page is currently open (no full reload, stays logged in)
  if(btn){var ic=btn.querySelector("svg");if(ic){ic.style.transition="transform .6s";ic.style.transform="rotate(360deg)";
    setTimeout(function(){ic.style.transition="";ic.style.transform="";},650);}}
  const sec=document.querySelector(".page-section.active");
  const page=sec?sec.id.replace("sec-",""):"dashboard";
  nav(page);
  refreshBell();
  showToast("Refreshed");
}

function nav(page){
  closeNav();   // close the mobile drawer after picking a page
  document.querySelectorAll(".nav-item").forEach(n=>n.classList.toggle("active",n.dataset.page===page));
  document.querySelectorAll(".page-section").forEach(s=>s.classList.remove("active"));
  document.getElementById("sec-"+page).classList.add("active");
  document.getElementById("page-title").textContent=titles[page];
  if(page==="dashboard")loadDashboard();
  if(page==="students")loadStudents(1);
  if(page==="confirmed")loadConfirmed(1);
  if(page==="syc")loadSyc(1);
  if(page==="required")loadRequired(1);
  if(page==="failed")loadFailed(1);
  if(page==="unknown")loadUnknown(1);
  if(page==="notoc")loadNoToc(1);
  if(page==="tocerror")loadTocErrors();
  if(page==="pending")loadPending();
  if(page==="docreq")loadDocReq();
  if(page==="history")loadHistory();
  if(page==="runlogs")loadRunLogs();
  if(page==="transfers")loadTransfers(1);
  if(page==="settings"){loadIntervals();loadWa();loadTrash();loadReportSettings();loadSourceCounts();}
  if(page==="upload")loadUploadSource();
  updateNavCounts();
}

let drData=[], drPrefix="", drSuffix="", drPage=1, drPerPage=10;
function drEsc(t){return (t||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");}
async function loadDocReq(btn){
  if(btn)btn.disabled=true;
  const body=document.getElementById("dr-body");
  if(body)body.innerHTML='<div style="color:var(--muted);font-size:13px;padding:10px 0">Loading…</div>';
  try{
    const r=await api("/api/doc-requests");
    drData=r.requests||[]; drPrefix=r.prefix||""; drSuffix=r.suffix||"";
    drData.forEach(s=>{s._checked=!s.sent;});
    drPage=1;
    renderDocReq();
  }catch(e){if(body)body.innerHTML='<div style="color:var(--danger);font-size:13px">'+e.message+'</div>';}
  finally{if(btn)btn.disabled=false;}
}
function drItem(key){return drData.find(x=>x.row_key===key);}
function renderDrPg(pages){
  const el=document.getElementById("dr-pg");if(!el)return;
  if(drData.length<=drPerPage){
    el.innerHTML=drData.length?('<div class="perpage" style="margin-left:auto">Per page: <select onchange="drSetPerPage(this.value)">'+
      [10,20,50,100].map(n=>'<option value="'+n+'" '+(n===drPerPage?"selected":"")+'>'+n+'</option>').join("")+'</select></div>'):"";
    return;
  }
  let ctrl='<div class="pg-controls"><button onclick="drGoPage('+(drPage-1)+')" '+(drPage<=1?"disabled":"")+'>&#8249; Prev</button>';
  const start=Math.max(1,drPage-2),end=Math.min(pages,drPage+2);
  for(let i=start;i<=end;i++)ctrl+='<button class="'+(i===drPage?'active':'')+'" onclick="drGoPage('+i+')">'+i+'</button>';
  ctrl+='<button onclick="drGoPage('+(drPage+1)+')" '+(drPage>=pages?"disabled":"")+'>Next &#8250;</button></div>';
  const sel='<div class="perpage">Per page: <select onchange="drSetPerPage(this.value)">'+
    [10,20,50,100].map(n=>'<option value="'+n+'" '+(n===drPerPage?"selected":"")+'>'+n+'</option>').join("")+'</select></div>';
  el.innerHTML=ctrl+sel;
}
function drGoPage(p){const pages=Math.max(1,Math.ceil(drData.length/drPerPage));drPage=Math.min(Math.max(1,p),pages);renderDocReq();var s=document.getElementById("sec-docreq");if(s)s.scrollIntoView({behavior:"smooth",block:"start"});}
function drSetPerPage(v){drPerPage=parseInt(v)||10;drPage=1;renderDocReq();}
function renderDocReq(){
  const body=document.getElementById("dr-body");if(!body)return;
  const badge=document.getElementById("nav-docreq-badge");
  const pending=drData.filter(s=>!s.sent).length;
  if(badge){badge.textContent=pending;badge.style.display=pending?"inline-flex":"none";}
  if(!drData.length){body.innerHTML='<div style="color:var(--success);font-size:14px;padding:10px 0">No Document-Required students right now.</div>';renderDrPg(0);updateDrCount();return;}
  const pages=Math.max(1,Math.ceil(drData.length/drPerPage));
  if(drPage>pages)drPage=pages;
  const slice=drData.slice((drPage-1)*drPerPage, drPage*drPerPage);
  body.innerHTML=slice.map((s)=>{
    const i=drData.indexOf(s);
    const sentBadge=s.sent?'<span style="background:#dcfce7;color:#15803d;font-size:11px;font-weight:700;border-radius:6px;padding:3px 9px">Sent'+(s.sent_at?' &middot; '+s.sent_at:'')+'</span>':'<span style="background:#fef3c7;color:#92400e;font-size:11px;font-weight:700;border-radius:6px;padding:3px 9px">Pending</span>';
    const blank=!s.message;
    return '<div class="dr-card" style="border:1px solid var(--border);border-radius:12px;padding:14px;margin-bottom:12px;background:'+(s.sent?'#f6fef9':'#fff')+'">'+
      '<div style="display:flex;align-items:center;gap:10px;margin-bottom:9px;flex-wrap:wrap">'+
        '<input type="checkbox" class="dr-cb" value="'+s.row_key+'" '+(s._checked?'checked':'')+' onchange="drToggle(&quot;'+s.row_key+'&quot;,this.checked)" style="width:16px;height:16px">'+
        '<b style="font-size:14px">'+(i+1)+'. '+drEsc(s.student_name)+'</b>'+
        '<span style="font-size:12px;color:var(--muted)">'+drEsc(s.session)+' &middot; '+drEsc(s.mobile)+'</span>'+
        '<span style="margin-left:auto">'+sentBadge+'</span>'+
      '</div>'+
      '<div style="font-size:11.5px;color:var(--muted);background:var(--soft);border-radius:8px;padding:8px 10px;margin-bottom:9px"><b>NIOS remark:</b> '+(s.remark?drEsc(s.remark):'<i>none</i>')+'</div>'+
      '<label style="font-size:12px;font-weight:600">Message to send (edit if needed)'+(blank?' <span style="color:#b91c1c">&mdash; please write what to ask for</span>':'')+'</label>'+
      '<textarea class="dr-msg" data-key="'+s.row_key+'" rows="2" oninput="drDirty(this)" style="width:100%;margin-top:5px;padding:10px;border:2px solid var(--border);border-radius:9px;font-size:13.5px;font-family:inherit;resize:vertical">'+drEsc(s.message||"")+'</textarea>'+
      '<div style="display:flex;align-items:center;gap:10px;margin-top:8px;flex-wrap:wrap">'+
        '<button class="btn btn-outline btn-sm" onclick="saveDocReqMsg(this,&quot;'+s.row_key+'&quot;)">Save</button>'+
        '<button class="btn btn-sm" style="background:#16a34a;color:#fff" onclick="sendDocReqOne(this,&quot;'+s.row_key+'&quot;)">Send now</button>'+
        '<span class="dr-saved" data-key="'+s.row_key+'" style="font-size:12px;color:var(--success)"></span>'+
        '<button type="button" onclick="toggleDrPrev(this)" style="margin-left:auto;background:none;border:none;color:#2563eb;font-size:12px;cursor:pointer;text-decoration:underline">Preview full message</button>'+
      '</div>'+
      '<div class="dr-img" data-key="'+s.row_key+'" style="margin-top:9px">'+drImgHtml(s)+'</div>'+
      '<div class="dr-prev" style="display:none;white-space:pre-wrap;background:#ecfdf5;border:1px solid #bbf7d0;border-radius:9px;padding:11px;margin-top:9px;font-size:13px;color:#065f46"></div>'+
    '</div>';
  }).join("");
  renderDrPg(pages);
  updateDrCount();
}
function drToggle(key,checked){const it=drItem(key);if(it)it._checked=checked;updateDrCount();}
function drDirty(ta){const it=drItem(ta.dataset.key);if(it)it.message=ta.value;const tag=document.querySelector('.dr-saved[data-key="'+ta.dataset.key+'"]');if(tag){tag.textContent="unsaved…";tag.style.color="var(--muted)";}}
function drImgHtml(s){
  if(s.image){
    return '<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">'+
      '<img src="'+s.image+'" style="height:52px;border-radius:8px;border:1px solid var(--border);object-fit:cover">'+
      '<span style="font-size:12px;color:var(--success);font-weight:600">&#128247; Screenshot attached</span>'+
      '<button class="btn btn-outline btn-sm" onclick="pickDocImg(&quot;'+s.row_key+'&quot;)">Replace</button>'+
      '<button class="btn btn-outline btn-sm" onclick="pasteDocImg(&quot;'+s.row_key+'&quot;)" title="Paste an image copied to clipboard (e.g. after Win+Shift+S)">&#128203; Paste</button>'+
      '<button class="btn btn-sm" style="background:#fee2e2;color:#b91c1c" onclick="removeDocImg(&quot;'+s.row_key+'&quot;)">Remove</button>'+
      '<input type="file" accept="image/png,image/jpeg,image/webp" style="display:none" id="drf-'+s.row_key+'" onchange="uploadDocImg(this,&quot;'+s.row_key+'&quot;)">'+
    '</div>';
  }
  return '<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">'+
    '<button class="btn btn-outline btn-sm" onclick="pickDocImg(&quot;'+s.row_key+'&quot;)" title="Attach a demo screenshot to send with this WhatsApp message">&#128206; Attach screenshot (optional)</button>'+
    '<button class="btn btn-outline btn-sm" onclick="pasteDocImg(&quot;'+s.row_key+'&quot;)" title="Paste an image copied to clipboard (e.g. after taking a screenshot with Win+Shift+S)">&#128203; Paste from clipboard</button>'+
    '<input type="file" accept="image/png,image/jpeg,image/webp" style="display:none" id="drf-'+s.row_key+'" onchange="uploadDocImg(this,&quot;'+s.row_key+'&quot;)">'+
  '</div>';
}
function drImgBox(key){return Array.from(document.querySelectorAll('.dr-img')).find(b=>b.dataset.key===key);}
function pickDocImg(key){const f=document.getElementById("drf-"+key);if(f)f.click();}
async function uploadDocImgFile(file,key){
  if(!file)return;
  const fd=new FormData();fd.append("row_key",key);fd.append("file",file);
  try{
    const r=await fetch(API+"/api/doc-request-image",{method:"POST",headers:{"Authorization":"Bearer "+TOKEN},body:fd});
    if(!r.ok){const e=await r.json().catch(()=>({}));showToast("Upload failed: "+(e.detail||r.status));return;}
    const d=await r.json();
    const it=drData.find(x=>x.row_key===key);if(it){it.image=d.url;const box=drImgBox(key);if(box)box.innerHTML=drImgHtml(it);}
    showToast("Screenshot attached \u2713");
  }catch(e){showToast("Error: "+e.message);}
}
async function uploadDocImg(input,key){const file=input.files&&input.files[0];await uploadDocImgFile(file,key);}
async function pasteDocImg(key){
  try{
    if(!navigator.clipboard||!navigator.clipboard.read){showToast("Clipboard paste isn't supported here — use Attach instead");return;}
    const items=await navigator.clipboard.read();
    for(const it of items){
      const type=(it.types||[]).find(t=>t.startsWith("image/"));
      if(type){
        const blob=await it.getType(type);
        const ext=(type.split("/")[1]||"png").replace("jpeg","jpg");
        if(["png","jpg","webp"].indexOf(ext)<0){showToast("Pasted image type not supported");return;}
        const file=new File([blob],"pasted."+ext,{type:type});
        showToast("Uploading pasted image\u2026");
        await uploadDocImgFile(file,key);
        return;
      }
    }
    showToast("No image found in clipboard — copy a screenshot first");
  }catch(e){showToast("Couldn't read clipboard: "+(e.message||e)+" (allow clipboard permission, or use Attach)");}
}
async function removeDocImg(key){
  try{await api("/api/doc-request-image-remove","POST",{row_key:key});
    const it=drData.find(x=>x.row_key===key);if(it){it.image="";const box=drImgBox(key);if(box)box.innerHTML=drImgHtml(it);}
    showToast("Screenshot removed");
  }catch(e){showToast("Error: "+e.message);}
}
function toggleDrPrev(btn){
  const card=btn.closest(".dr-card");if(!card)return;
  const prev=card.querySelector(".dr-prev"),ta=card.querySelector(".dr-msg");
  const hasImg=!!card.querySelector(".dr-img img");
  if(prev.style.display==="none"){prev.textContent=(hasImg?"[ screenshot will be attached at the top ]\\n\\n":"")+drPrefix.replace("{name}","<student name>")+(ta.value||"…")+drSuffix;prev.style.display="block";btn.textContent="Hide preview";}
  else{prev.style.display="none";btn.textContent="Preview full message";}
}
function getDrSelected(){return drData.filter(s=>s._checked&&!s.sent).map(s=>s.row_key);}
function updateDrCount(){const el=document.getElementById("dr-count");if(el)el.textContent=getDrSelected().length+" selected · "+drData.filter(s=>!s.sent).length+" pending · "+drData.length+" total";}
function toggleDrAll(cb){drData.forEach(s=>{if(!s.sent)s._checked=cb.checked;});renderDocReq();}
async function saveDocReqMsg(btn,key){
  const ta=document.querySelector('.dr-msg[data-key="'+key+'"]');if(!ta)return;
  try{await api("/api/doc-request-save","POST",{row_key:key,message:ta.value});
    const tag=document.querySelector('.dr-saved[data-key="'+key+'"]');if(tag){tag.textContent="\u2713 saved";tag.style.color="var(--success)";}
    const it=drData.find(x=>x.row_key===key);if(it)it.message=ta.value;
  }catch(e){showToast("Error: "+e.message);}
}
async function sendDocReqOne(btn,key){
  const ta=document.querySelector('.dr-msg[data-key="'+key+'"]');
  if(ta&&!ta.value.trim()){showToast("Please write the document message first");return;}
  if(!confirm("Send this document request on WhatsApp now?"))return;
  if(ta){try{await api("/api/doc-request-save","POST",{row_key:key,message:ta.value});}catch(e){}}
  if(btn){btn.disabled=true;btn.textContent="Sending…";}
  try{const r=await api("/api/doc-request-send","POST",{row_keys:[key]});
    showToast(r.sent?"Sent \u2713":("Failed: "+((r.results&&r.results[0]&&r.results[0].info)||"")));
    loadDocReq();
  }catch(e){showToast("Error: "+e.message);if(btn){btn.disabled=false;btn.textContent="Send now";}}
}
async function sendDocReq(mode,btn){
  let keys=[];
  if(mode==="selected"){keys=getDrSelected();if(!keys.length){showToast("Select at least one student");return;}}
  else{keys=drData.filter(s=>!s.sent).map(s=>s.row_key);}
  if(!keys.length){showToast("Nothing to send");return;}
  if(!confirm("Send document request on WhatsApp to "+keys.length+" student(s)?"))return;
  if(btn)btn.disabled=true;
  try{
    // persist each message from drData (covers edits on any page)
    for(const k of keys){const it=drItem(k);if(it){try{await api("/api/doc-request-save","POST",{row_key:k,message:it.message||""});}catch(_){}}}
    const payload=mode==="all"?{all:true}:{row_keys:keys};
    const r=await api("/api/doc-request-send","POST",payload);
    showToast("Sent: "+r.sent+(r.failed?(" · Failed: "+r.failed):""));
    const warn=document.getElementById("dr-warn");
    if(warn){
      const fails=(r.results||[]).filter(x=>!x.ok);
      if(fails.length){
        warn.style.display="block";
        warn.innerHTML='<b>'+fails.length+' message(s) failed.</b> Reason from WhatsApp gateway:<br>'+
          fails.slice(0,5).map(f=>'&bull; <b>'+(f.student_name||"")+'</b>: '+(f.info||"unknown")).join("<br>")+
          (fails.length>5?'<br>&hellip; and '+(fails.length-5)+' more':'')+
          '<br><span style="color:#7c2d12">Tip: check the campaign name is exactly correct (case-sensitive) and the template is Approved &amp; Live on AiSensy.</span>';
      }else{warn.style.display="none";}
    }
    loadDocReq();
  }catch(e){showToast("Error: "+e.message);}
  finally{if(btn)btn.disabled=false;}
}
async function loadDashboard(){
  try{
    const d=await api("/api/dashboard");
    const c=d.counts||{};
    document.getElementById("stat-grid").innerHTML=
      statCard("Total Students",d.total_students,SI.users,"#4F46E5")+
      statCard("Changes Today",c.changes_today||0,SI.activity,"#0891B2")+
      statCard("Admission Confirmed",c.confirmed||0,SI.check,"#16A34A")+
      statCard("SYC Students",c.syc||0,SI.users,"#7C3AED")+
      statCard("Verified",c.verified||0,SI.shield,"#65A30D")+
      statCard("Document Required",c.document_required||0,SI.file,"#EA580C")+
      statCard("In Verification",c.doc_verification||0,SI.loader,"#D97706");
    renderDistribution(d.status_distribution,d.total_students);
    renderSessionCounts(d.session_counts||[]);
    renderSourceCounts(d.source_counts||[], d.transfer_count||0);
    // Recent Runs is now a scrollable list (no page/limit dropdown) — load a healthy
    // window and let the user scroll; ~7-8 rows are visible at a time.
    try{const rr=await api("/api/run-logs?limit=30");renderRuns(rr,"recent-runs");}
    catch(e){renderRuns(d.recent_runs,"recent-runs");}
    renderBell(d.notifications||[],d.dup_notifications||[],d.unknown_notifications||[]);
  updateFailedBadge();
  }catch(e){showToast("Error: "+e.message);}
}
function renderSessionCounts(arr){
  const el=document.getElementById("session-counts");
  if(!arr.length){el.innerHTML='<div class="empty">No data yet</div>';return;}
  const cell=(label,val,color)=>'<div style="text-align:center;flex:1;min-width:0">'+
    '<div style="font-size:17px;font-weight:800;color:'+color+';line-height:1.1">'+(val||0)+'</div>'+
    '<div style="font-size:10.5px;color:var(--muted);font-weight:600;margin-top:2px">'+label+'</div></div>';
  el.innerHTML='<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(290px,1fr));gap:14px">'+
    arr.map(s=>'<div style="padding:15px 16px;background:var(--soft);border:1px solid var(--border);border-radius:13px">'+
      '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">'+
        '<span style="font-size:14.5px;font-weight:700">'+(s.session||"—")+'</span>'+
        '<span style="font-size:12.5px;color:var(--muted);font-weight:600">Total&nbsp;<b style="color:var(--text);font-size:17px">'+s.cnt+'</b></span></div>'+
      '<div style="display:flex;gap:4px;padding-top:11px;border-top:1px solid var(--border)">'+
        cell("Confirmed",s.confirmed,"#16A34A")+
        cell("Verified",s.verified,"#2563EB")+
        cell("Active",s.active,"#7C3AED")+
        cell("Required",s.required,"#EA580C")+
      '</div></div>').join("")+'</div>';
}
function renderSourceCounts(arr,tcount){
  const el=document.getElementById("source-counts");
  if(!el)return;
  if(!arr.length){el.innerHTML="";return;}
  window._srcData=arr; if(window._srcCombined===undefined)window._srcCombined=false;
  drawSourceCards();
}
function drawSourceCards(){
  const el=document.getElementById("source-counts");
  if(!el)return;
  var arr=window._srcData||[];
  const cell=(label,val,color)=>'<div style="text-align:center;flex:1;min-width:0">'+
    '<div style="font-size:17px;font-weight:800;color:'+color+';line-height:1.1">'+(val||0)+'</div>'+
    '<div style="font-size:10.5px;color:var(--muted);font-weight:600;margin-top:2px">'+label+'</div></div>';
  var enrol=arr.filter(function(s){return s.key==="enrol_portal";})[0];
  var portal=arr.filter(function(s){return s.key==="mvs_portal";})[0];
  var tracker=arr.filter(function(s){return s.key==="mvs_tracker";})[0];
  var combined=window._srcCombined;
  var display=[];
  if(combined && (enrol||portal)){
    display.push({source:"MVS Portal (all)",key:"mvs_portal_all",combinable:true,
      cnt:((enrol&&enrol.cnt)||0)+((portal&&portal.cnt)||0),
      confirmed:((enrol&&enrol.confirmed)||0)+((portal&&portal.confirmed)||0),
      verified:((enrol&&enrol.verified)||0)+((portal&&portal.verified)||0),
      active:((enrol&&enrol.active)||0)+((portal&&portal.active)||0),
      required:((enrol&&enrol.required)||0)+((portal&&portal.required)||0)});
  }else{
    if(enrol)display.push(enrol);
    if(portal){portal.combinable=true;display.push(portal);}
  }
  if(tracker)display.push(tracker);
  var dotFor=function(k){return k==="enrol_portal"?"#7C3AED":(k==="mvs_tracker"?"#0EA5E9":"#16A34A");};
  el.innerHTML='<div style="font-size:11px;font-weight:700;letter-spacing:.4px;color:var(--muted);text-transform:uppercase;margin-bottom:10px">By Data Source</div>'+
    '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(290px,1fr));gap:14px;margin-bottom:18px">'+
    display.map(function(s){
      var isEnrol=(s.key==="enrol_portal");
      var footer;
      if(isEnrol){
        footer='<div style="display:flex;align-items:center;gap:8px;margin-top:11px;padding-top:10px;border-top:1px dashed var(--border);font-size:11.5px;color:var(--muted)">'+
          '<svg viewBox="0 0 24 24" fill="none" stroke="#16A34A" stroke-width="2" width="13" height="13"><circle cx="12" cy="12" r="9"/><polyline points="12 7 12 12 15 14"/></svg>'+
          '<span style="flex:1">New-data run in <b id="pn-timer" style="color:var(--text)">…</b></span>'+
          '<button id="pn-btn" onclick="togglePn(this)" style="background:#dcfce7;color:#15803d;border:1px solid #bbf7d0;border-radius:7px;padding:3px 11px;font-size:11px;font-weight:700;cursor:pointer">Pause</button>'+
        '</div>';
      }else if(s.combinable){
        footer='<div style="display:flex;align-items:center;gap:8px;margin-top:11px;padding-top:10px;border-top:1px dashed var(--border);font-size:11.5px;color:var(--muted)">'+
          '<span style="flex:1">'+(combined?"Enrol + MVS Portal combined.":"Sheet uploads on the Portal + transfers.")+'</span>'+
          '<button onclick="toggleSrcCombine()" style="background:'+(combined?"#EDE9FE":"#dcfce7")+';color:'+(combined?"#6D28D9":"#15803d")+';border:1px solid '+(combined?"#DDD6FE":"#bbf7d0")+';border-radius:7px;padding:3px 11px;font-size:11px;font-weight:700;cursor:pointer">'+(combined?"Split":"Combine")+'</button>'+
        '</div>';
      }else{
        footer='<div style="display:flex;align-items:center;gap:8px;margin-top:11px;padding-top:10px;border-top:1px dashed var(--border);font-size:11.5px;color:var(--muted)">'+
          '<span style="flex:1">Uploaded straight into the tracker.</span></div>';
      }
      return '<div style="padding:15px 16px;background:var(--soft);border:1px solid var(--border);border-radius:13px">'+
        '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">'+
          '<span style="font-size:14.5px;font-weight:700;display:flex;align-items:center;gap:7px">'+
            '<span style="width:9px;height:9px;border-radius:50%;background:'+dotFor(s.key)+';display:inline-block;flex:none"></span>'+(s.source||"—")+'</span>'+
          '<span style="font-size:12.5px;color:var(--muted);font-weight:600">Total&nbsp;<b style="color:var(--text);font-size:17px">'+s.cnt+'</b></span></div>'+
        '<div style="display:flex;gap:4px;padding-top:11px;border-top:1px solid var(--border)">'+
          cell("Confirmed",s.confirmed,"#16A34A")+cell("Verified",s.verified,"#2563EB")+
          cell("Active",s.active,"#7C3AED")+cell("Required",s.required,"#EA580C")+
        '</div>'+footer+'</div>';
    }).join("")+'</div>'+
    '<div style="font-size:11px;font-weight:700;letter-spacing:.4px;color:var(--muted);text-transform:uppercase;margin-bottom:10px">By Session</div>';
  updatePnTimer();
}
function toggleSrcCombine(){ window._srcCombined=!window._srcCombined; drawSourceCards(); }
async function loadReconciliation(refresh){
  var p=document.getElementById("reconcile-panel");
  if(!p)return;
  // Toggle: clicking the button again (or the ×) hides the panel. refresh=true re-renders in place.
  if(!refresh && p.style.display==="block"){p.style.display="none";return;}
  p.style.display="block";
  p.innerHTML='<div style="color:var(--muted);font-size:13px">Loading…</div>';
  try{
    const d=await api("/api/reconciliation");
    var at=d.attention||{}, org=d.origin||{}, bs=d.by_source||{}, c=d.confirmed||{};
    var row=function(lbl,val,color,note){
      return '<div style="display:flex;justify-content:space-between;gap:10px;padding:8px 0;border-bottom:1px solid var(--border)">'+
        '<span style="font-size:13px">'+lbl+(note?(' <span style="color:#94a3b8;font-size:11.5px">'+note+'</span>'):'')+'</span>'+
        '<b style="font-size:14px;color:'+(color||"var(--text)")+'">'+val+'</b></div>';
    };
    p.innerHTML=
      '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px">'+
        '<span style="font-size:14px;font-weight:800">Tracker vs Portal — quick check</span>'+
        '<span style="display:flex;gap:8px;flex-wrap:wrap">'+
          '<button onclick="compareConfirmed(this)" style="background:#FEF3C7;color:#92400E;border:1px solid #FDE68A;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer">Compare confirmed</button>'+
          '<button onclick="auditConfirmed(this)" style="background:#FEF3C7;color:#92400E;border:1px solid #FDE68A;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer">Find confirmed gap</button>'+
          '<button onclick="probeOrigin(this)" style="background:#EDE9FE;color:#6D28D9;border:1px solid #DDD6FE;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer">Detect origin field</button>'+
          '<button onclick="forceConfirmed(this)" style="background:#0F766E;color:#fff;border:none;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer" title="Re-push every confirmed student to the Portal and list any the Portal does not accept.">Re-push ALL confirmed</button>'+
          '<button onclick="syncDetailsAll(this)" style="background:#B45309;color:#fff;border:none;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer" title="Push every linked student\'s details (reference/enrollment no, DOB, mobile, email, name…) from the tracker to the Portal — fixes stale Portal data after tracker-side edits.">Sync details &rarr; Portal</button>'+
          '<button onclick="syncPortalNow(this)" style="background:#4F46E5;color:#fff;border:none;border-radius:8px;padding:5px 13px;font-size:12px;font-weight:700;cursor:pointer">Sync Portal now</button>'+
          '<button onclick="document.getElementById(&quot;reconcile-panel&quot;).style.display=&quot;none&quot;" style="background:var(--soft);border:1px solid var(--border);border-radius:8px;padding:5px 11px;font-size:12px;font-weight:700;cursor:pointer">&times; Close</button>'+
        '</span></div>'+
      '<div id="rp-syncline" style="display:none;margin:2px 0 10px;font-size:12.5px;font-weight:700;color:#4F46E5"></div>'+
      '<div id="rp-probe" style="display:none;margin:2px 0 10px;font-size:12px;color:var(--muted);background:var(--bg);border:1px solid var(--border);border-radius:9px;padding:10px 12px"></div>'+
      '<div id="rp-force" style="display:none;margin:2px 0 10px;font-size:12.5px;background:var(--bg);border:1px solid var(--border);border-radius:9px;padding:10px 12px"></div>'+
      '<div id="rp-details" style="display:none;margin:2px 0 10px;font-size:12.5px;background:var(--bg);border:1px solid var(--border);border-radius:9px;padding:10px 12px"></div>'+
      '<div id="rp-audit" style="display:none;margin:2px 0 10px;font-size:12px;background:var(--bg);border:1px solid var(--border);border-radius:9px;padding:10px 12px"></div>'+
      row("Total students in tracker", d.total_live)+
      row("Portal has, tracker skips", (d.deleted||0)+(d.no_reference||0), "#B45309",
          "deleted: "+(d.deleted||0)+" + no-reference pending: "+(d.no_reference||0))+
      row("Confirmed (tracker)", (c.by_flag||0), "#047857")+
      row("Confirmed but Portal not updated yet", at.confirmed_push_lag||0,
          (at.confirmed_push_lag>0?"#B91C1C":"#047857"),
          "auto-syncs every 30 min — or press Sync Portal now")+
      row("Confirmed with NO Portal link", at.confirmed_no_link||0,
          (at.confirmed_no_link>0?"#B45309":"#047857"),
          "tracker-only students — transfer to Portal to push")+
      row("Enrol vs Sheet split from Portal",
          "enrol "+(org.enrol||0)+" · legacy/sheet "+(org.sheet||0)+" · not sent "+(org.blank||0),
          "var(--text)",
          (org.blank>0?"Portal is not sending origin for "+org.blank+" students — they default to Enrol. ":""))+

      '<div style="margin-top:10px;font-size:12px;color:var(--muted)">'+
      'Simple rule: <b>tracker total + deleted + pending = Portal total</b>, and confirmed matches once "not updated yet" reaches 0.</div>';
  }catch(e){p.innerHTML='<div style="color:#B91C1C;font-size:13px">Could not load: '+e.message+'</div>';}
}
async function auditConfirmed(btn){
  var out=document.getElementById("rp-audit");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Comparing…";}
  if(out){out.style.display="block";out.textContent="Fetching the Portal list and comparing every confirmed student…";}
  try{
    const d=await api("/api/confirmed-audit");
    if(!d.ok){if(out)out.textContent=d.message;return;}
    var h='<b style="color:var(--text)">'+d.mismatch_count+' confirmed student(s) the Portal is NOT counting as confirmed</b> '+
          '<span style="color:var(--muted)">(of '+d.total_confirmed_linked+' linked). Portal sends a status field: '+(d.portal_sends_status?"yes":"no")+'.</span>';
    if(d.note)h+='<div style="margin-top:6px;color:#92400E">'+d.note+'</div>';
    if(d.mismatches&&d.mismatches.length){
      h+='<div style="overflow-x:auto;margin-top:8px"><table class="tbl" style="font-size:12px"><thead><tr>'+
         '<th>Student</th><th>Reference</th><th>Mobile</th><th>Session</th><th>Portal says</th><th>Tracker</th></tr></thead><tbody>';
      d.mismatches.forEach(function(m){
        h+='<tr><td><b>'+(m.name||"—")+'</b></td><td>'+(m.reference_no||"—")+'</td><td>'+(m.mobile||"—")+'</td>'+
           '<td>'+(m.session||"—")+'</td><td style="color:#B45309;font-weight:700">'+(m.portal_status||"—")+'</td>'+
           '<td style="color:#047857">'+(m.tracker_status||"—")+'</td></tr>';
      });
      h+='</tbody></table></div>';
    }else{
      h+='<div style="margin-top:6px;color:#047857">No per-student mismatch found — the two are aligned (any remaining difference is Portal-side display lag).</div>';
    }
    if(out)out.innerHTML=h;
  }catch(e){if(out)out.textContent=""+e.message;}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function forceConfirmed(btn){
  if(!confirm("Re-push EVERY confirmed student to the Portal? This re-sends 'Admission Confirmed' for all linked confirmed students and lists any the Portal does not accept. Safe to run anytime (not during a status run)."))return;
  var box=document.getElementById("rp-force");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/portal-force-confirmed","POST",{});
    if(box){box.style.display="block";box.textContent=r.message||"";}
    if(!r.started){showToast(r.message);if(btn){btn.disabled=false;btn.textContent=old;}return;}
    (async function poll(){
      try{
        const s=await api("/api/portal-force-status");
        var html='<b>'+(s.running?("Re-pushing… "+s.done+" / "+s.total+" ("+(s.percent||0)+"%) — accepted "+s.ok+", not accepted "+s.failed):s.message)+'</b>';
        if(!s.running && s.fail_list && s.fail_list.length){
          html+='<div style="margin-top:8px;font-weight:700;color:#B91C1C">Portal did NOT accept these — check them on the Portal:</div>';
          html+='<div style="max-height:220px;overflow:auto;margin-top:4px"><table class="tbl"><thead><tr><th>Name</th><th>Reference</th><th>Portal ID</th></tr></thead><tbody>'+
            s.fail_list.map(function(f){return '<tr><td>'+(f.name||"—")+'</td><td>'+(f.reference||"—")+'</td><td style="font-size:12px;color:var(--muted)">'+(f.student_id||"—")+'</td></tr>';}).join("")+
            '</tbody></table></div>';
        }
        if(box)box.innerHTML=html;
        if(s.running){setTimeout(poll,1500);}
        else{showToast(s.message||"Done");if(btn){btn.disabled=false;btn.textContent=old;}loadReconciliation(true);}
      }catch(e){setTimeout(poll,2500);}
    })();
  }catch(e){showToast(""+e.message);if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function syncDetailsAll(btn){
  if(!confirm("Push EVERY linked student's details (reference/enrollment no, DOB, mobile, alt no, email, name, class, session) from the TRACKER to the PORTAL?\n\nUse this to backfill students already corrected on the tracker (e.g. wrong references fixed after a run) whose Portal copy is still old. Empty tracker fields are skipped, so nothing on the Portal gets wiped."))return;
  var box=document.getElementById("rp-details");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/portal-sync-details","POST",{});
    if(box){box.style.display="block";box.textContent=r.message||"";}
    if(!r.started){showToast(r.message);if(btn){btn.disabled=false;btn.textContent=old;}return;}
    (async function poll(){
      try{
        const s=await api("/api/portal-sync-details-status");
        var html='<b>'+(s.running?("Syncing details… "+s.done+" / "+s.total+" ("+(s.percent||0)+"%) — synced "+s.ok+", failed "+s.failed):s.message)+'</b>';
        if(!s.running && s.fail_list && s.fail_list.length){
          html+='<div style="margin-top:8px;font-weight:700;color:#B91C1C">Portal did NOT accept these:</div>';
          html+='<div style="max-height:220px;overflow:auto;margin-top:4px"><table class="tbl"><thead><tr><th>Name</th><th>Reference</th><th>Reason</th></tr></thead><tbody>'+
            s.fail_list.map(function(f){return '<tr><td>'+(f.name||"—")+'</td><td>'+(f.reference||"—")+'</td><td style="font-size:12px;color:var(--muted)">'+(f.reason||"—")+'</td></tr>';}).join("")+
            '</tbody></table></div>';
        }
        if(box)box.innerHTML=html;
        if(s.running){setTimeout(poll,1500);}
        else{showToast(s.message||"Done");if(btn){btn.disabled=false;btn.textContent=old;}}
      }catch(e){setTimeout(poll,2500);}
    })();
  }catch(e){showToast(""+e.message);if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function probeOrigin(btn){
  var out=document.getElementById("rp-probe");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Checking Portal…";}
  if(out){out.style.display="block";out.textContent="Fetching a sample from the Portal…";}
  try{
    const d=await api("/api/portal-origin-probe");
    if(!d.ok){if(out)out.textContent=d.message;return;}
    var det=d.detected||{}; var found=((det["enrol"]||0)+(det["sheet"]||0))>0;
    var html='<b style="color:var(--text)">'+d.message+'</b>';
    if(found){
      html+='<div style="margin-top:8px"><button onclick="applyOrigin(this)" style="background:#16A34A;color:#fff;border:none;border-radius:8px;padding:6px 14px;font-size:12.5px;font-weight:700;cursor:pointer">Apply origin to ALL students now</button>'+
            ' <span style="font-size:11.5px;color:var(--muted)">— ek click, poori list save, cards turant Portal jaisa split</span></div>';
    }
    html+='<div style="margin-top:6px">Auto-detected in sample: '+Object.keys(d.detected).map(function(k){return k+" = "+d.detected[k];}).join(" · ")+'</div>';
    var ck=Object.keys(d.candidate_fields||{});
    if(ck.length){
      html+='<div style="margin-top:6px">Origin-jaise fields jo Portal bhej raha hai:</div><ul style="margin:4px 0 0 18px;padding:0">';
      ck.forEach(function(k){
        var vals=d.candidate_fields[k];
        html+='<li><b>'+k+'</b>: '+Object.keys(vals).map(function(v){return v+" ("+vals[v]+")";}).join(", ")+'</li>';
      });
      html+='</ul>';
    }else{
      html+='<div style="margin-top:6px;color:#B45309">Portal ke trackerList response mein koi origin-jaisa field NAHI mila — Portal side pe field add karni hogi (e.g. origin: "real_enrolment" / "bulk_imported").</div>';
    }
    if(out)out.innerHTML=html;
  }catch(e){if(out)out.textContent=""+e.message;}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function compareConfirmed(btn){
  var out=document.getElementById("rp-probe");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Comparing…";}
  if(out){out.style.display="block";out.textContent="Portal se poori list milakar student-by-student compare ho raha hai…";}
  try{
    const d=await api("/api/confirmed-compare");
    if(!d.ok){if(out)out.textContent=d.message;return;}
    var list=function(title,items,color,showPV){
      if(!items||!items.length)return "";
      var h='<div style="margin-top:9px"><b style="color:'+color+'">'+title+'</b><ul style="margin:4px 0 0 18px;padding:0">';
      items.forEach(function(s){
        h+='<li><b>'+(s.name||"—")+'</b> · '+(s.reference||"—")+' · '+(s.session||"—")+
           (showPV?(' — Portal says: <b style="color:'+color+'">'+(s.portal_value||"?")+'</b>'):'')+'</li>';
      });
      return h+'</ul></div>';
    };
    var html='<b style="color:var(--text)">Tracker confirmed: '+d.tracker_confirmed+
             (d.portal_confirmed_by_field!=null?(' · Portal confirmed (its own field "'+d.stage_field+'"): '+d.portal_confirmed_by_field):'')+
             ' · matched: '+d.matched+'</b>';
    if(d.differs_count){
      html+=list("Portal ke paas hain par stage 'confirmed' NAHI ("+d.differs_count+") — ye hi gap hai:",d.differs,"#B91C1C",true);
      var keys=d.differs.map(function(s){return s.row_key;});
      window.__cmpDiffKeys=keys;
      html+='<div style="margin-top:8px"><button onclick="forcePush(window.__cmpDiffKeys,this)" style="background:#DC2626;color:#fff;border:none;border-radius:8px;padding:6px 14px;font-size:12.5px;font-weight:700;cursor:pointer">Force re-push these '+d.differs_count+' now</button></div>';
    }
    if(d.missing_count) html+=list("Tracker linked hai par Portal list mein studentId NAHI mila ("+d.missing_count+") — Portal pe delete/merge hue honge:",d.missing,"#B45309",false);
    if(d.no_link_count) html+=list("No Portal link ("+d.no_link_count+") — Transfer Data se bhejo:",d.no_link,"#6B7280",false);
    if(!d.differs_count&&!d.missing_count&&!d.no_link_count) html+='<div style="margin-top:8px;color:#047857;font-weight:700">Sab match — koi gap nahi.</div>';
    if(d.stage_field==null) html+='<div style="margin-top:8px;color:#B45309">Note: Portal list mein koi stage/status field nahi mila, isliye sirf membership compare hua (stage compare nahi ho paya).</div>';
    if(out)out.innerHTML=html;
  }catch(e){if(out)out.textContent=""+e.message;}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function forcePush(keys,btn){
  if(!confirm("Force re-push "+keys.length+" student(s) ka current status Portal pe?"))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Pushing…";}
  try{const r=await api("/api/force-push","POST",{row_keys:keys});showToast(r.message||"Done");loadReconciliation(true);}
  catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function applyOrigin(btn){
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Applying…";}
  try{
    const r=await api("/api/portal-origin-apply","POST",{});
    showToast(r.message||"Done");
    if(r.ok){ loadReconciliation(true); try{loadDashboard();}catch(e){} }
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
function showUnlinked(list){
  var out=document.getElementById("rp-probe");
  if(!out||!list||!list.length)return;
  out.style.display="block";
  out.innerHTML='<b style="color:#B45309">Ye '+list.length+' confirmed student(s) Portal ki list mein NAHI mile (isliye push impossible) — Portal pe inka record banao/transfer karo, ya inka reference Portal se milao:</b>'+
    '<ul style="margin:6px 0 0 18px;padding:0">'+
    list.map(function(u){return '<li><b>'+(u.name||"—")+'</b> · '+(u.reference||"no ref")+' · '+(u.mobile||"no mobile")+'</li>';}).join("")+
    '</ul>';
}
async function syncPortalNow(btn){
  var line=document.getElementById("rp-syncline");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Checking…";}
  try{
    const r=await api("/api/portal-resync-now","POST",{});
    showUnlinked(r.unlinked);
    if(r.nothing || !r.started){
      // Nothing to push (or run active) — say it clearly, right in the panel.
      if(line){line.style.display="block";line.style.color=r.nothing?"#047857":"#B45309";line.textContent=r.message;}
      showToast(r.message);
      if(btn){btn.disabled=false;btn.textContent=old;}
      if(r.linked>0){ loadReconciliation(true); }
      return;
    }
    // Started — poll live progress until done, then refresh the numbers in place.
    if(line){line.style.display="block";line.style.color="#4F46E5";line.textContent=r.message;}
    (async function poll(){
      try{
        const s=await api("/api/portal-sync-status");
        if(line)line.textContent=(s.running?("Syncing… "+s.done+" / "+s.total+" ("+(s.percent||0)+"%) — pushed "+s.pushed):s.message);
        if(s.running){setTimeout(poll,1500);}
        else{
          showToast(s.message||"Portal sync done");
          if(btn){btn.disabled=false;btn.textContent=old;}
          loadReconciliation(true);   // refresh the numbers in place (panel stays open)
        }
      }catch(e){ setTimeout(poll,2500); }
    })();
  }catch(e){
    showToast(""+e.message);
    if(btn){btn.disabled=false;btn.textContent=old;}
  }
}
function statCard(lbl,val,svg,col){
  return '<div class="stat"><div class="ic" style="background:'+col+'1A;color:'+col+'">'+svg+
    '</div><div class="lbl">'+lbl+'</div><div class="val">'+val+
    '</div><div class="bar" style="background:'+col+'"></div></div>';
}
const SI={
  users:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87M16 3.13a4 4 0 0 1 0 7.75"/></svg>',
  activity:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="22 12 18 12 15 21 9 3 6 12 2 12"/></svg>',
  check:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M22 11.08V12a10 10 0 1 1-5.93-9.14"/><polyline points="22 4 12 14.01 9 11.01"/></svg>',
  shield:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><polyline points="9 12 11 14 15 10"/></svg>',
  file:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>',
  loader:'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="9"/><polyline points="12 7 12 12 15 14"/></svg>'
};
const DIST_COLOURS={
  "Pending":"#FBBF24","Documents Verification In Progress":"#FB923C","Document Required":"#F97316",
  "Verified":"#4ADE80","Approved":"#2DD4BF","Admission Confirmed":"#22C55E",
  "Rejected":"#EF4444","Fetch Error":"#94A3B8","Unknown":"#CBD5E1"};
function renderDistribution(dist,total){
  const el=document.getElementById("distribution");
  if(!dist||!dist.length){el.innerHTML='<div class="empty">No data yet. Upload Excel & run a check.</div>';return;}
  const max=Math.max.apply(null,dist.map(d=>d.cnt).concat([1]));
  el.innerHTML=dist.map(d=>{
    const col=DIST_COLOURS[d.current_status]||"#CBD5E1";
    const w=(d.cnt/max*100).toFixed(0);
    return '<div class="dist-row"><div class="nm">'+(d.current_status||"Unknown")+'</div>'+
      '<div class="track"><div class="fill" style="width:'+w+'%;background:'+col+'"></div></div>'+
      '<div class="ct">'+d.cnt+'</div></div>';
  }).join("");
}
function renderRuns(runs,id){
  const el=document.getElementById(id);
  if(!runs||!runs.length){el.innerHTML='<tr><td colspan="7" class="empty">No runs yet</td></tr>';return;}
  el.innerHTML=runs.map(r=>{
    var st;
    if(r.status==="running"){
      st='<span class="run-live">● running</span>'+
         ' <button class="btn-cancel" onclick="cancelRun('+r.id+')">Cancel</button>';
    }else if(r.status==="cancelled"){
      st='<span class="run-cancel">✕ cancelled</span>';
    }else if(r.status==="completed"){
      st='<span class="run-done">✓ completed</span>';
    }else{
      st='<span class="run-err">'+r.status+'</span>';
    }
    var act=(r.status==="running")?'<span style="color:var(--muted);font-size:12px">—</span>':
      '<button title="Delete this run log" onclick="deleteRunLog('+r.id+')" '+
      'style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:4px 9px;font-size:12px;font-weight:600;cursor:pointer">&#128465;</button>';
    return '<tr><td>'+r.run_at+'</td><td><span class="ref-tag">'+(r.group_type||"all")+
      '</span></td><td>'+r.total_checked+'</td><td style="color:var(--primary);font-weight:700">'+r.total_changed+
      '</td><td style="color:'+(r.total_failed?'var(--danger)':'inherit')+'">'+r.total_failed+
      '</td><td>'+st+'</td><td>'+act+'</td></tr>';
  }).join("");
}
async function deleteRunLog(id){
  if(!confirm("Delete this run log entry?"))return;
  try{
    await api("/api/run-log-delete?id="+id,"POST");
    showToast("Run log deleted");
    if(secActive("dashboard"))loadDashboard();
    if(secActive("runlogs"))loadRunLogs();
  }catch(e){showToast(""+e.message);}
}
async function clearRunLogs(){
  if(!confirm("Delete ALL run logs (except any currently running)?"))return;
  try{const r=await api("/api/run-logs-clear","POST");showToast("Cleared "+(r.deleted||0)+" run logs");
    if(secActive("dashboard"))loadDashboard();if(secActive("runlogs"))loadRunLogs();}
  catch(e){showToast(""+e.message);}
}

async function cancelRun(rid){
  if(!confirm("Cancel this run?")) return;
  try{
    const fd=new FormData();fd.append("run_id",rid);
    const res=await fetch("/api/cancel-run",{method:"POST",headers:{Authorization:"Bearer "+TOKEN},body:fd});
    const d=await res.json();
    if(!res.ok) throw new Error(d.detail||"failed");
    showToast("✓ "+(d.message||"Cancelled"));
    const dash=document.getElementById("sec-dashboard");
    const rlog=document.getElementById("sec-runlogs");
    if(dash&&dash.classList.contains("active")) loadDashboard();
    if(rlog&&rlog.classList.contains("active")) loadRunLogs();
  }catch(e){showToast(""+e.message);}
}

function toggleBell(e){e.stopPropagation();document.getElementById("bell-dropdown").classList.toggle("open");}
document.addEventListener("click",()=>document.getElementById("bell-dropdown").classList.remove("open"));
async function refreshBell(){try{const d=await api("/api/dashboard");renderBell(d.notifications||[],d.dup_notifications||[],d.unknown_notifications||[]);}catch(e){}}
function renderBell(notifs,dups,unknowns){
  dups=dups||[];unknowns=unknowns||[];
  const n=notifs.length, total=n+dups.length+unknowns.length;
  const badge=document.getElementById("bell-badge");
  const navB=document.getElementById("nav-required-badge");
  badge.textContent=total;badge.style.display=total?"flex":"none";
  navB.textContent=n;navB.style.display=n?"inline-block":"none";
  document.getElementById("bell-head-cnt").textContent=total;
  const list=document.getElementById("bell-list");
  let html="";
  if(n){html+=notifs.map(x=>'<div class="notif-item"><div class="dot"></div><div>'+
    '<div class="nm">'+(x.student_name||"—")+'</div>'+
    '<div class="rf">'+(x.reference_no||"No ref")+'</div>'+
    (x.remark?'<div class="rk">'+x.remark+'</div>':"")+'</div></div>').join("");}
  if(unknowns.length){
    html+='<div style="padding:7px 14px;font-size:11px;font-weight:800;color:#991B1B;background:#FEE2E2;border-top:1px solid var(--border)">STATUS NOT FOUND — re-check needed</div>';
    html+=unknowns.map(x=>'<div class="notif-item"><div class="dot" style="background:#DC2626"></div><div>'+
      '<div class="nm">'+(x.student_name||"—")+'</div>'+
      '<div class="rf">'+(x.reference_no||x.enrollment_no||"—")+'</div>'+
      '<div class="rk">'+(x.current_status||"Unknown")+' — '+(x.remark||"NIOS didn\'t return a readable status. Re-check this student.")+'</div></div></div>').join("");
  }
  if(dups.length){
    html+='<div style="padding:7px 14px;font-size:11px;font-weight:800;color:#5B21B6;background:#EDE9FE;border-top:1px solid var(--border)">DUPLICATE — kept as MVS Portal</div>';
    html+=dups.map(x=>'<div class="notif-item"><div class="dot" style="background:#7C3AED"></div><div>'+
      '<div class="nm">'+(x.student_name||"—")+'</div>'+
      '<div class="rf">'+(x.reference_no||x.enrollment_no||"—")+'</div>'+
      '<div class="rk">Same student in MVS Portal &amp; MVS Tracker → kept once as MVS Portal</div></div></div>').join("");
  }
  list.innerHTML=html||'<div class="notif-empty">No pending items</div>';
}

let stTimer,cTimer,rTimer;
function debounceStudents(){clearTimeout(stTimer);stTimer=setTimeout(()=>loadStudents(1),400);}
function debounceConfirmed(){clearTimeout(cTimer);cTimer=setTimeout(()=>loadConfirmed(1),400);}
function debounceRequired(){clearTimeout(rTimer);rTimer=setTimeout(()=>loadRequired(1),400);}

function badge(s){
  const m={"Pending":"b-pending","Documents Verification In Progress":"b-docs","Document Required":"b-required",
    "Verified":"b-verified","Approved":"b-approved","Admission Confirmed":"b-confirmed","Rejected":"b-rejected","SYC":"b-syc"};
  return '<span class="badge '+(m[s]||'b-error')+'">'+(s||'Unknown')+'</span>';
}
function srcBadge(s){
  var p=(s.source||"mvs_tracker")==="mvs_portal";
  return '<span style="display:inline-block;white-space:nowrap;font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px;'+
    (p?'background:#EDE9FE;color:#5B21B6':'background:#E0F2FE;color:#075985')+'">'+
    (p?'MVS Portal':'MVS Tracker')+'</span>';
}
function isLoginFailed(s){return (s.login_failed==1||s.login_failed===true);}
function fixBtn(s){
  return '<button onclick="editStudent(&quot;'+s.row_key+'&quot;)" '+
    'style="background:#e0e7ff;color:#3730a3;border:1px solid #c7d2fe;border-radius:8px;padding:4px 10px;font-size:11.5px;font-weight:600;cursor:pointer">Edit &amp; fix data</button>';
}
function dlLinks(s){
  if(isLoginFailed(s))
    return '<div style="font-size:11.5px;font-weight:600;color:#b91c1c;max-width:240px">Documents blocked — NIOS login failed with this data. '+
      'Links not shared (would open to an error). <div style="margin-top:5px">'+fixBtn(s)+'</div></div>';
  if(!s.reference_no||!s.dob) return '<span style="color:var(--warn);font-size:11px">ref/DOB missing</span>';
  const sess=(s.session||"").toLowerCase();
  const isStream2=sess.includes("stream 2")||sess.includes("stream2")||sess.includes("stream-2")||sess.includes("stream ii");
  const isOnDemand=sess.includes("on demand")||sess.includes("ondemand")||sess.includes("on-demand")||sess.includes("odes");
  const isPublic=!isStream2&&!isOnDemand;   // safe default: April/October/apr-27/unknown = public
  const toc=(s.toc_status||"").toLowerCase();
  const notoc=(toc==="no");
  // Show exactly the documents this student receives (matches the WhatsApp doc set):
  //   On Demand: id + hall (+ app form if TOC)   | Stream 2: id (+ app form if TOC)
  //   Public:    id (+ app form / registration summary if TOC=yes)
  let b=[dlBtn(s,"id_card","ID Card")];
  if(isOnDemand){
    if(!notoc) b.push(dlBtn(s,"app_form","App Form"));
    b.push(dlBtn(s,"hall_ticket","Hall Ticket"));
  }else if(isStream2){
    if(!notoc) b.push(dlBtn(s,"app_form","App Form"));
  }else{ // public — only TOC=yes also gets the Registration Summary (App Form)
    if(toc==="yes") b.push(dlBtn(s,"app_form","App Form"));
  }
  return '<div style="display:flex;flex-wrap:wrap;gap:5px">'+b.join("")+'</div>';
}
async function refreshSelectedDocs(btn){
  const keys=(typeof CONF_SEL!=="undefined")?[...CONF_SEL]:[];
  if(!keys.length){ showToast("Select one or more students first (tick the checkboxes)"); return; }
  if(!confirm("Re-fetch and overwrite the saved documents for "+keys.length+" selected student(s) from NIOS?")) return;
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/refresh-selected-docs","POST",{row_keys:keys});
    showToast((r&&r.message)||("Refreshed "+((r&&r.saved)||0)+" document(s)"));
    if(typeof loadConfirmed==="function")loadConfirmed();
  }catch(e){ showToast("Error: "+e.message); }
  finally{ if(btn){btn.disabled=false;btn.style.opacity="";} }
}
function dlBtn(s,kind,label){
  var saved=(s.cached_docs&&s.cached_docs.indexOf(kind)>=0);
  var mark=saved?' <span title="Saved in our database \u2014 opens without NIOS" style="color:#15803D;font-weight:700">&#9679;</span>':'';
  return '<button class="btn-dl" onclick="downloadDoc(this,&quot;'+s.reference_no+'&quot;,&quot;'+
    (s.dob||"")+'&quot;,&quot;'+kind+'&quot;,&quot;'+label+'&quot;)"'+(saved?' style="border-color:#86EFAC"':'')+'>'+DL_ICON+' '+label+mark+'</button>';
}
function waBtn(s){
  if(!s.row_key)return "";
  if(isLoginFailed(s))
    return '<div style="margin-top:5px;font-size:11.5px;font-weight:700;color:#b91c1c">WhatsApp: FAILED (login error — not sent)</div>';
  const sent=(s.whatsapp_sent==1);
  const dv=(s.whatsapp_delivery||"");
  const at=(s.whatsapp_sent_at||"");
  let badge="";
  if(dv==="delivered"){
    badge='<div style="margin-top:5px;font-size:11.5px;font-weight:700;color:#15803D">Delivered on WhatsApp'+(s.whatsapp_delivery_at?' &middot; '+s.whatsapp_delivery_at:'')+'</div>';
  }else if(dv==="failed"){
    badge='<div style="margin-top:5px;font-size:11.5px;font-weight:700;color:#b91c1c">Delivery FAILED — please resend</div>';
  }else if(sent){
    badge='<div style="margin-top:5px;font-size:11.5px;font-weight:600;color:#B45309">Sent to WhatsApp'+(at?' &middot; '+at:'')+' <span style="font-weight:400;color:var(--muted)">(delivery not yet confirmed)</span></div>';
    var info2=(s.whatsapp_info||"");
    if(info2 && info2.indexOf("via")>=0){badge+='<div style="font-size:10.5px;color:var(--muted);margin-top:1px">'+info2.replace(/</g,"&lt;").slice(0,90)+'</div>';}
  }else{
    var why=(s.whatsapp_info||"");
    if(why && why.toLowerCase().indexOf("accepted")>=0){
      badge='<div style="margin-top:5px;font-size:11.5px;font-weight:600;color:#B45309">Sent to WhatsApp'+(at?' &middot; '+at:'')+' <span style="font-weight:400;color:var(--muted)">(gateway accepted, delivery pending)</span></div>';
    }else if(why && why.indexOf("2 numbers")<0){
      badge='<div style="margin-top:5px;font-size:11.5px;font-weight:700;color:#b91c1c">Not sent &middot; '+why.replace(/</g,"&lt;").slice(0,150)+'</div>';
    }else{
      badge='<div style="margin-top:5px;font-size:11.5px;font-weight:600;color:var(--muted)">Not sent yet</div>';
    }
  }
  if((s.whatsapp_info||"").indexOf("2 numbers")>=0){
    badge+='<div style="margin-top:3px;font-size:11px;font-weight:700;color:#7C3AED">&#128241; Sent to 2 numbers (own + alternate)</div>';
  }
  const bc=(dv==="failed")?'#DC2626':'#16A34A';
  return badge+'<button class="btn-dl" style="background:'+bc+';color:#fff;border-color:'+bc+';margin-top:4px" '+
    'onclick="resendWa(&quot;'+s.row_key+'&quot;,this)">'+(sent?'Resend WhatsApp':'Send WhatsApp')+'</button>';
}
async function resendWa(rowKey,btn){
  if(btn&&btn.dataset.busy==="1")return;
  if(!confirm("Send the documents to this student on WhatsApp?"))return;
  let orig="",pct=1,fake=null;
  if(btn){btn.dataset.busy="1";orig=btn.innerHTML;
    btn.innerHTML='<span class="dl-spin"></span> '+pct+'%';
    fake=setInterval(()=>{ if(pct<90){pct+=Math.max(1,Math.floor((90-pct)/6));
      btn.innerHTML='<span class="dl-spin"></span> '+pct+'%';}},420);}
  try{
    const r=await api("/api/wa-resend","POST",{row_key:rowKey});
    if(fake){clearInterval(fake);fake=null;}
    if(btn)btn.innerHTML='100%';
    if(r.ok){ showToast("Sent \u2713 "+(r.info||"")); }
    else{ showToast("Not sent \u2014 "+(r.info||"failed")); }
    setTimeout(()=>{try{loadConfirmed(1);}catch(e){}try{loadStudents(1);}catch(e){}},1200);
  }catch(e){showToast("Error: "+e.message);}
  finally{if(fake)clearInterval(fake);if(btn){setTimeout(()=>{btn.innerHTML=orig;btn.dataset.busy="";},1600);}}
}
async function downloadDoc(btn,ref,dob,kind,name){
  if(btn&&btn.dataset.busy==="1")return;          // already running -> ignore extra clicks
  let orig="",pct=1,fake=null;
  if(btn){
    btn.dataset.busy="1";btn.classList.add("loading");orig=btn.innerHTML;
    btn.innerHTML='<span class="dl-spin"></span> '+pct+'%';
    fake=setInterval(()=>{ if(pct<92){pct+=Math.max(1,Math.floor((92-pct)/9));
      btn.innerHTML='<span class="dl-spin"></span> '+pct+'%';}},650);
  }
  const restore=(txt)=>{
    if(fake)clearInterval(fake);
    if(btn){ btn.innerHTML=txt||orig;
      setTimeout(()=>{btn.innerHTML=orig;btn.dataset.busy="";btn.classList.remove("loading");},1400); }
  };
  try{
    const q=new URLSearchParams({ref:ref,dob:dob,kind:kind});
    const r=await fetch(API+"/api/download-doc?"+q.toString(),{headers:{"Authorization":"Bearer "+TOKEN}});
    if(!r.ok){const e=await r.json().catch(()=>({detail:"failed"}));
      var msg=(e.detail||"download failed");
      showToast("Error: "+msg);
      // A login/DOB failure just moved this student to 'Failed to Run' on the server —
      // refresh the badges and current view so it disappears from here immediately.
      if(/login|rejected|dob/i.test(msg)){
        updateNavCounts();
        try{loadConfirmed(1);}catch(_e){}try{loadStudents(1);}catch(_e){}
        try{loadRequired(1);}catch(_e){}try{loadFailed(1);}catch(_e){}
        showToast("Moved to 'Failed to Run' — fix the data there & run again");
      }
      restore();return;}
    const ctype=r.headers.get("Content-Type")||"";
    const blob=await r.blob();
    const url=URL.createObjectURL(blob);
    if(ctype.includes("pdf")){
      const a=document.createElement("a");a.href=url;a.download=name.replace(/ /g,"_")+"_"+ref+".pdf";a.click();
      showToast(name+" downloaded");
    }else{
      const w=window.open(url,"_blank");
      if(!w){showToast("Popup blocked — please allow popups and click again");}
      else{showToast(name+" opened — tap 'Save as PDF' at the top");}
    }
    setTimeout(()=>URL.revokeObjectURL(url),120000);
    restore('100%');
  }catch(e){showToast("Error: "+e.message);restore();}
}
function fillSessions(arr){
  if(!arr)return;
  ["s-session","c-session","r-session","u-session","nt-session"].forEach(id=>{
    const sel=document.getElementById(id);
    if(!sel)return;
    const cur=sel.value;
    sel.length=1;  // keep only the first "All Sessions" option
    arr.forEach(s=>{if(s){const o=document.createElement("option");o.value=s;o.textContent=s;sel.appendChild(o);}});
    sel.value=cur; // restore previous selection
  });
}

async function loadStudents(page){
  page=page||1;
  const dr=dateRange("s");
  const q=new URLSearchParams({page:page,per_page:perPage,view:"normal",
    search:document.getElementById("s-search").value,
    status_filter:document.getElementById("s-status").value,
    session_filter:document.getElementById("s-session").value,
    class_filter:fval("s-class"),source_filter:fval("s-source"),date_from:dr.from,date_to:dr.to,
    toc_filter:fval("s-toc"),
    check_state:(document.getElementById("s-checkstate")?document.getElementById("s-checkstate").value:"")});
  try{
    const d=await api("/api/students?"+q.toString());
    fillSessions(d.sessions);
    document.getElementById("s-count").textContent="Showing "+d.students.length+" of "+d.total+" active students";
    const b=document.getElementById("s-body");
    b.innerHTML=d.students.length?d.students.map((s,i)=>'<tr>'+
      '<td><input type="checkbox" class="sel-cb" data-rk="'+s.row_key+'" '+(SELECTED.has(s.row_key)?'checked':'')+' onclick="toggleSel(&quot;'+s.row_key+'&quot;,this)"></td>'+
      '<td style="color:var(--muted)">'+((page-1)*perPage+i+1)+'</td>'+
      '<td><span class="ref-tag">'+(s.reference_no||"—")+'</span></td>'+
      '<td>'+(s.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(s)+'</div></td><td style="font-size:13px">'+(s.session||"—")+'</td>'+
      '<td>'+badge(s.current_status)+loginWarn(s)+'</td>'+
      '<td>'+tocCell(s)+'</td>'+
      '<td style="font-size:12px;color:var(--muted)">'+(s.last_checked||"—")+checkBadge(s)+'</td>'+
      '<td>'+delBtn(s)+'</td></tr>').join("")
      :'<tr><td colspan="9" class="empty">No active students found</td></tr>';
    renderPg("s-pg",page,d.pages,"loadStudents");
    updateSelBar();
    const sa=document.getElementById("sel-all");if(sa)sa.checked=false;
    loadCheckSummary();
  }catch(e){showToast(""+e.message);}
}
function tocCell(s){
  var t=((s.toc_status||"")+"").toLowerCase();
  var pill=function(txt,col,bg){return '<span style="font-size:11px;font-weight:700;color:'+col+';background:'+bg+';padding:2px 8px;border-radius:6px">'+txt+'</span>';};
  var base = t==="yes"?pill("yes","#047857","#D1FAE5"):(t==="no"?pill("no","#92400E","#FEF3C7"):'<span style="color:var(--muted);font-size:12px">—</span>');
  if(s.toc_mismatch==1 && s.toc_verified!=1){
    return base+'<div style="margin-top:4px"><span style="font-size:10.5px;font-weight:700;color:#B91C1C;background:#FEE2E2;padding:2px 7px;border-radius:6px" title="NIOS says '+((s.nios_toc||"").toUpperCase())+', Portal says '+t.toUpperCase()+'">TOC mismatch</span></div>';
  }
  if(s.toc_verified==1){
    return base+'<div style="margin-top:4px"><span style="font-size:10.5px;font-weight:700;color:#047857;background:#D1FAE5;padding:2px 7px;border-radius:6px">verified</span></div>';
  }
  return base;
}
function checkBadge(s){
  if(s.check_state==="stale")
    return '<div style="margin-top:4px"><span style="font-size:11px;font-weight:600;color:#B45309;background:#FEF3C7;padding:2px 8px;border-radius:6px">Not checked</span></div>';
  if(s.check_state==="new")
    return '<div style="margin-top:4px"><span style="font-size:11px;font-weight:600;color:#047857;background:#D1FAE5;padding:2px 8px;border-radius:6px">New check</span></div>';
  return "";
}
async function loadCheckSummary(){
  const el=document.getElementById("s-checksummary");if(!el)return;
  try{
    const q=new URLSearchParams({view:"normal",
      session_filter:document.getElementById("s-session").value,
      source_filter:fval("s-source")});
    const d=await api("/api/check-summary?"+q.toString());
    const sess=document.getElementById("s-session").value||"All sessions";
    el.style.display="block";
    el.innerHTML='<b>'+sess+'</b> &nbsp;·&nbsp; '+
      '<span style="color:#047857;font-weight:600">New check: '+(d.new||0)+'</span> &nbsp;·&nbsp; '+
      '<span style="color:#B45309;font-weight:600">Not checked: '+(d.stale||0)+'</span>';
  }catch(e){el.style.display="none";}
}

let _ntT;
function debounceNoToc(){clearTimeout(_ntT);_ntT=setTimeout(()=>loadNoToc(1),350);}
async function loadNoToc(page){
  page=page||1;
  const q=new URLSearchParams({page:page,per_page:perPage,view:"notoc",
    search:document.getElementById("nt-search").value,
    session_filter:document.getElementById("nt-session").value});
  try{
    const d=await api("/api/students?"+q.toString());
    fillSessions(d.sessions);
    document.getElementById("nt-count").textContent="Showing "+d.students.length+" of "+d.total+" no-TOC students";
    const b=document.getElementById("nt-body");
    b.innerHTML=d.students.length?d.students.map((s,i)=>'<tr>'+
      '<td style="color:var(--muted)">'+((page-1)*perPage+i+1)+'</td>'+
      '<td><span class="ref-tag">'+(s.reference_no||"\u2014")+'</span></td>'+
      '<td>'+(s.student_name||"\u2014")+'<div style="margin-top:4px">'+srcBadge(s)+'</div></td>'+
      '<td style="font-size:13px">'+(s.session||"\u2014")+'</td>'+
      '<td>'+badge(s.current_status)+'</td>'+
      '<td>'+(((s.toc_status||"").toLowerCase()==="yes")
        ?'<span style="font-size:11px;font-weight:700;padding:2px 9px;border-radius:20px;background:#DCFCE7;color:#15803D">yes</span>'
        :'<span style="font-size:11px;font-weight:700;padding:2px 9px;border-radius:20px;background:#FEF3C7;color:#92400E">no</span>')+'</td>'+
      '<td style="font-size:12px;color:var(--muted)">'+(s.last_checked||"\u2014")+'</td></tr>').join("")
      :'<tr><td colspan="7" class="empty">No no-TOC students found. If the Portal shows some, their tocStatus has not reached the Tracker yet \u2014 run a fetch, and if still empty the Portal is not sending tocStatus.</td></tr>';
    renderPg("nt-pg",page,d.pages,"loadNoToc");
    const nb=document.getElementById("nav-notoc-badge");
    if(nb){if(d.total>0){nb.textContent=d.total;nb.style.display="";}else{nb.style.display="none";}}
  }catch(e){showToast(""+e.message);}
}
async function syncToc(btn){
  var o=btn?btn.innerHTML:"";
  if(btn){btn.disabled=true;btn.style.opacity="0.6";btn.innerHTML="Syncing\u2026";}
  try{
    const r=await api("/api/sync-toc","POST");
    showToast((r&&r.message)?r.message:"Synced");
    loadNoToc(1);
  }catch(e){showToast("Error: "+e.message);}
  finally{if(btn){btn.disabled=false;btn.style.opacity="";btn.innerHTML=o;}}
}
async function diagnoseLogin(){
  var ref=(fval("dg-ref")||"").trim(), dob=(fval("dg-dob")||"").trim();
  var out=document.getElementById("dg-out");
  if(!ref||!dob){showToast("Enter Reference No and DOB");return;}
  out.style.display="block"; out.textContent="Running live NIOS login test\u2026 (this can take ~20-40s)";
  try{
    const q=new URLSearchParams({ref:ref,dob:dob});
    const r=await api("/api/diagnose-login?"+q.toString(),"GET");
    out.textContent=JSON.stringify(r,null,2);
  }catch(e){out.textContent="Error: "+e.message;}
}
async function runNotoc(btn){
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/run-now-notoc","POST");
    showToast(r.message||"Running\u2026");
  }catch(e){showToast("Error: "+e.message);}
  finally{setTimeout(()=>{if(btn){btn.disabled=false;btn.style.opacity="";}},2000);}
}
function exportNoToc(){
  const q=new URLSearchParams({view:"notoc",search:fval("nt-search"),session_filter:fval("nt-session")});
  showToast("Preparing Excel\u2026");
  fetch(API+"/api/export-students?"+q.toString(),{headers:{Authorization:"Bearer "+TOKEN}})
    .then(r=>r.ok?r.blob():Promise.reject())
    .then(blob=>{const u=URL.createObjectURL(blob);const a=document.createElement("a");a.href=u;a.download="no-toc-students.xlsx";a.click();URL.revokeObjectURL(u);})
    .catch(()=>showToast("Export failed"));
}
async function loadConfirmed(page){
  page=page||1;
  try{ if(!CACHE_POLL) startCachePoll(); }catch(e){}
  const dr=dateRange("c");
  const q=new URLSearchParams({page:page,per_page:perPage,view:"confirmed",
    search:document.getElementById("c-search").value,
    status_filter:(document.getElementById("c-status")?document.getElementById("c-status").value:""),
    wa_status:fval("c-wa"),
    session_filter:document.getElementById("c-session").value,
    class_filter:fval("c-class"),source_filter:fval("c-source"),
    toc_filter:fval("c-toc"),
    saved_filter:(document.getElementById("c-saved")?document.getElementById("c-saved").value:""),
    date_from:dr.from,date_to:dr.to});
  try{
    const d=await api("/api/students?"+q.toString());
    fillSessions(d.sessions);
    document.getElementById("c-count").textContent=d.total+" confirmed students";
    const b=document.getElementById("c-body");
    b.innerHTML=d.students.length?d.students.map((s,i)=>'<tr>'+
      '<td><input type="checkbox" class="c-selbox" value="'+s.row_key+'" onchange="onConfSel(this)"'+(CONF_SEL.has(s.row_key)?' checked':'')+'></td>'+
      '<td style="color:var(--muted)">'+((page-1)*perPage+i+1)+'</td>'+
      '<td><span class="ref-tag">'+(s.reference_no||"—")+'</span></td>'+
      '<td>'+(s.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(s)+'</div></td><td style="font-size:13px">'+(s.session||"—")+'</td>'+
      '<td>'+badge(s.current_status)+loginWarn(s)+'</td><td>'+tocCell(s)+'</td><td style="font-size:12px">'+dlLinks(s)+waBtn(s)+'</td>'+
      '<td style="font-size:12px;color:var(--muted)">'+(s.last_changed||"—")+'</td>'+
      '<td>'+delBtn(s)+'</td></tr>').join("")
      :'<tr><td colspan="10" class="empty">No confirmed students yet</td></tr>';
    const sa=document.getElementById("c-selall");if(sa)sa.checked=false;
    renderPg("c-pg",page,d.pages,"loadConfirmed");
    if(!WA_POLL){pollWaOnce().then(p=>{if(p&&p.running){renderWaProgress(p);startWaProgressPoll();}else{setWaBox(false);}});}
  }catch(e){showToast(""+e.message);}
}
const CONF_SEL=new Set();
function updateConfBar(){
  const bar=document.getElementById("c-bulkbar");const n=CONF_SEL.size;
  if(bar)bar.style.display=n?"flex":"none";
  const c=document.getElementById("c-selcount");if(c)c.textContent=n;
}
function onConfSel(cb){ if(cb.checked)CONF_SEL.add(cb.value); else CONF_SEL.delete(cb.value); updateConfBar(); }
function toggleConfAll(cb){
  document.querySelectorAll(".c-selbox").forEach(b=>{b.checked=cb.checked; if(cb.checked)CONF_SEL.add(b.value); else CONF_SEL.delete(b.value);});
  updateConfBar();
}
function clearConfSel(){ CONF_SEL.clear(); document.querySelectorAll(".c-selbox").forEach(b=>b.checked=false); const sa=document.getElementById("c-selall");if(sa)sa.checked=false; updateConfBar(); }
async function bulkDeleteConf(){
  const keys=Array.from(CONF_SEL);
  if(!keys.length)return;
  if(!confirm("Remove "+keys.length+" selected confirmed student(s)?\\n\\nThey move to Trash — restore from Settings → Deleted Students."))return;
  try{
    const r=await api("/api/students-delete-bulk","POST",{row_keys:keys});
    showToast("Moved "+(r.deleted||keys.length)+" student(s) to Trash");
    clearConfSel();
    refreshAllTables();
  }catch(e){showToast("Error: "+e.message);}
}
async function resendSelected(btn){
  if(!CONF_SEL.size)return;
  const keys=Array.from(CONF_SEL);
  if(!confirm("Resend WhatsApp documents to "+keys.length+" selected student(s)?"))return;
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/wa-resend-bulk","POST",{row_keys:keys});
    showToast(r.message||("Resending "+keys.length+" student(s)"));
    clearConfSel();
    setTimeout(()=>{try{loadConfirmed(1);}catch(e){}},6000);
  }catch(e){showToast("Error: "+e.message);}
  finally{if(btn){btn.disabled=false;btn.style.opacity="";}}
}
async function autoSendNow(btn){
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/wa-autosend-now","POST");
    showToast(r.already_running?"A send is already running":"Sending to all pending students…");
    startWaProgressPoll();
  }catch(e){showToast("Error: "+e.message);}
  finally{setTimeout(()=>{if(btn){btn.disabled=false;btn.style.opacity="";}},2000);}
}
async function resendNoTocConfirmed(btn){
  if(!confirm("Re-send corrected documents to all CONFIRMED no-TOC students? They previously received the wrong documents \u2014 this sends the right ones via the no-TOC campaign, and their old links stop working.")) return;
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/wa-resend-notoc","POST");
    if(r&&r.already_running){showToast("A send is already running");}
    else if(r&&r.queued===0){showToast(r.message||"No no-TOC confirmed students to resend");}
    else{showToast(r.message||("Resending to "+((r&&r.queued)||0)+" students\u2026"));startWaProgressPoll();}
  }catch(e){showToast("Error: "+e.message);}
  finally{setTimeout(()=>{if(btn){btn.disabled=false;btn.style.opacity="";}},2000);}
}
let WA_POLL=null, WA_SEEN_RUNNING=false;
async function cacheDocsNow(btn,force){
  var msg=force?"Re-fetch and OVERWRITE all saved documents from NIOS? Use this if documents changed on NIOS. Runs in the background.":"Save every confirmed student document into our database now? This fetches each document from NIOS once and stores our own copy, so WhatsApp links open straight from our copy \u2014 no NIOS/CapSolver, and they keep working even if NIOS is down. Runs in the background; may take a while for many students.";
  if(!confirm(msg)) return;
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/cache-docs","POST",{force:!!force});
    if(r&&r.ok){ startCachePoll(); }
    else { showToast((r&&r.message)||"Already saving\u2026"); startCachePoll(); }
  }catch(e){ showToast("Error: "+e.message); if(btn){btn.disabled=false;btn.style.opacity="";} }
}
async function stopCacheDocs(btn){
  if(!confirm("Stop the auto-save? Documents already saved are kept. You can start it again anytime.")) return;
  try{ await api("/api/cache-docs-stop","POST"); showToast("Auto-save stopped."); }
  catch(e){ showToast("Error: "+e.message); }
}
let CACHE_POLL=null;
function startCachePoll(){
  if(CACHE_POLL)clearInterval(CACHE_POLL);
  const el=document.getElementById("cache-docs-status");
  const btn=document.getElementById("cache-docs-btn");
  const box=document.getElementById("cache-progress");
  const setTxt=(id,v)=>{var e=document.getElementById(id);if(e)e.textContent=v;};
  const tick=async()=>{
    let d; try{ d=await api("/api/cache-docs-progress"); }catch(e){ return; }
    const p=(d&&d.progress)||{}; const active=!!(d&&d.auto_resume);
    const pct=p.total?Math.round((p.done/p.total)*100):0;
    // small persistent summary line
    if(el){ el.textContent="Saved in database: "+((d&&d.students_cached)||0)+" students \u00b7 "+((d&&d.total_cached)||0)+" documents \u00b7 Fully saved: "+((d&&d.fully_saved)||0)+" / "+((d&&d.confirmed_total)||0)+((d&&d.pending_docs)?(" \u00b7 "+d.pending_docs+" pending"):"")+(((d&&d.genuine_failed)||0)?(" \u00b7 \u26a0 "+d.genuine_failed+" need data fix"):""); }
    // the progress bar box
    if(box){
      if(p.running||active){
        box.style.display="block";
        setTxt("cap-title", p.running?(p.phase==="saving"?"Saving documents to database":"Preparing\u2026"):"Auto-resume ON \u2014 will retry the rest shortly");
        setTxt("cap-pct", p.running?pct:0);
        var bar=document.getElementById("cap-bar"); if(bar)bar.style.width=(p.running?pct:100)+"%";
        setTxt("cap-done", p.done||0); setTxt("cap-total", p.total||0);
        setTxt("cap-saved", p.saved||0);
        setTxt("cap-retry", (d&&d.retrying)||0);
        setTxt("cap-genfail", (d&&d.genuine_failed)||0);
        setTxt("cap-resume", active?"\u21bb auto-resume ON":"");
        var ce=document.getElementById("cap-err");
        if(ce){ if(p.last_error){ce.style.display="block";ce.textContent="Last issue: "+p.last_error;} else {ce.style.display="none";} }
      } else {
        box.style.display="none";
      }
    }
    if(btn){ btn.disabled=(p.running||active); btn.style.opacity=(p.running||active)?"0.6":""; }
    if(!p.running && !active){
      clearInterval(CACHE_POLL); CACHE_POLL=null;
      if(p.total){ showToast("All done \u2014 "+(p.saved||0)+" document(s) saved"); }
    }
  };
  tick(); CACHE_POLL=setInterval(tick,2500);
}
function renderWaProgress(p){
  document.getElementById("wap-title").textContent=(p.label||"Sending WhatsApp")+(p.running?"":" — done");
  document.getElementById("wap-pct").textContent=p.pct||0;
  document.getElementById("wap-bar").style.width=(p.pct||0)+"%";
  document.getElementById("wap-total").textContent=p.total||0;
  document.getElementById("wap-sent").textContent=p.sent||0;
  document.getElementById("wap-remaining").textContent=p.remaining||0;
  document.getElementById("wap-failed").textContent=p.failed||0;
}
function setWaBox(show){const b=document.getElementById("wa-progress");if(b)b.style.display=show?"block":"none";}
async function pollWaOnce(){
  try{return await api("/api/wa-progress");}catch(e){return null;}
}
function startWaProgressPoll(){
  if(WA_POLL)clearInterval(WA_POLL);
  WA_SEEN_RUNNING=false;
  let ticks=0;
  setWaBox(true);
  const tick=async()=>{
    ticks++;
    const p=await pollWaOnce();
    if(!p)return;
    renderWaProgress(p);
    if(p.running){WA_SEEN_RUNNING=true;setWaBox(true);return;}
    if(!WA_SEEN_RUNNING && ticks<6){return;}        // give the background task a moment to start
    clearInterval(WA_POLL);WA_POLL=null;
    setWaBox(true);                                  // show the final 100% state…
    try{loadConfirmed(1);}catch(e){}
    setTimeout(()=>{if(!WA_POLL)setWaBox(false);},6000);   // …then auto-hide it
  };
  tick();
  WA_POLL=setInterval(tick,1500);
}

let sycTimer;
function debounceSyc(){clearTimeout(sycTimer);sycTimer=setTimeout(()=>loadSyc(1),400);}
async function loadSyc(page){
  page=page||1;
  const search=fval("syc-search");
  try{
    const q=new URLSearchParams({page:page,search:search});
    const d=await api("/api/syc?"+q.toString());
    document.getElementById("syc-count").textContent=(d.total||0)+" SYC student(s)";
    const tb=document.getElementById("syc-body");
    if(!d.students||!d.students.length){tb.innerHTML='<tr><td colspan="6" style="text-align:center;color:var(--muted);padding:24px">No SYC students yet. Upload a sheet with ADMISSION SESSION = SYC and click Run Now.</td></tr>';document.getElementById("syc-pg").innerHTML="";return;}
    tb.innerHTML=d.students.map((s,i)=>sycRow(s,(page-1)*20+i+1)).join("");
    renderPg("syc-pg",page,d.pages,"loadSyc");
  }catch(e){showToast(""+e.message);}
}
function sycRow(s,i){
  var nm=(s.student_name||'').replace(/"/g,'&quot;');
  return '<tr><td>'+i+'</td><td>'+(s.enrollment_no||'—')+'</td><td>'+(s.student_name||'')+'</td>'+
    '<td>'+(s.mobile||'')+'</td><td>'+(s.dob||'')+'</td>'+
    '<td><span class="badge b-syc">SYC &#10003; Ready</span></td><td>'+
    '<button class="btn-dl" onclick="downloadSycHall(&quot;'+s.row_key+'&quot;,this)">'+DL_ICON+' Download Hall Ticket</button>'+
    '<div class="syc-prog" style="display:none;height:7px;background:#E2E8F0;border-radius:4px;margin-top:7px;overflow:hidden;max-width:220px"><div class="syc-bar" style="height:100%;width:0;background:#4F46E5;transition:width .35s"></div></div>'+
    '<div class="syc-msg" style="font-size:11px;color:var(--muted);margin-top:3px"></div></td>'+
    '<td><button title="Delete this SYC student" onclick="deleteSyc(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" '+
    'style="background:rgba(239,68,68,.12);color:#DC2626;border:1px solid rgba(239,68,68,.3);border-radius:8px;padding:6px 9px;cursor:pointer;font-weight:600">'+
    '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" width="14" height="14" style="vertical-align:-2px"><polyline points="3 6 5 6 21 6"/><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/></svg></button></td></tr>';
}
async function deleteSyc(rowKey,name){
  if(!confirm("Remove SYC student "+name+"?\\n\\nThis moves the student to Trash. You can restore it from Settings → Deleted Students if removed by mistake."))return;
  try{
    await api("/api/syc-delete?row_key="+encodeURIComponent(rowKey),"POST");
    showToast("Moved "+name+" to Trash");
    loadSyc(1);
    try{loadDashboard();}catch(e){}
  }catch(e){showToast(""+e.message);}
}
function loginWarn(s){
  if(!(s.login_failed==1||s.login_failed===true))return "";
  var rm=(s.login_remark||"NIOS login failed — check Reference/Enrollment No & DOB").replace(/</g,"&lt;");
  return '<div style="margin-top:5px;font-size:11.5px;font-weight:600;color:#b91c1c;background:#fee2e2;border:1px solid #fecaca;border-radius:7px;padding:4px 8px;white-space:normal;max-width:260px">'+
    '&#9888; Not sent — '+rm+'</div>';
}
function delBtn(s){
  var nm=(s.student_name||"this student").replace(/[\\"']/g," ");
  var edit='<button title="Edit details (DOB / Reference No etc.) and re-run" '+
    'style="background:#e0e7ff;color:#3730a3;border:1px solid #c7d2fe;border-radius:8px;padding:5px 10px;font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap;margin-right:6px" '+
    'onclick="editStudent(&quot;'+s.row_key+'&quot;)">&#9998; Edit</button>';
  var rem='<button title="Remove this student (moves to Trash — restore from Settings)" '+
    'style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:5px 10px;font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap" '+
    'onclick="deleteStudent(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)">&#128465; Remove</button>';
  return '<div style="display:flex;flex-wrap:wrap;gap:4px">'+edit+rem+'</div>';
}
async function deleteStudent(rowKey,name){
  if(!confirm("Remove "+name+"?\\n\\nThis hides the student from the portal and moves it to Trash. You can restore it from Settings → Deleted Students if removed by mistake."))return;
  try{
    await api("/api/student-delete?row_key="+encodeURIComponent(rowKey),"POST");
    showToast("Moved "+name+" to Trash");
    try{loadStudents(1);}catch(e){}
    try{loadConfirmed(1);}catch(e){}
    try{loadRequired(1);}catch(e){}
    try{loadFailed(1);}catch(e){}
    updateFailedBadge();
    try{loadDashboard();}catch(e){}
  }catch(e){showToast(""+e.message);}
}
// ── Reusable multi-select + bulk delete (works on any table) ──
const SEL={};
function selSet(k){ if(!SEL[k])SEL[k]=new Set(); return SEL[k]; }
function selBox(k,rowKey){
  return '<input type="checkbox" class="selbox-'+k+'" value="'+rowKey+'" onchange="onSel(&quot;'+k+'&quot;,this)"'+(selSet(k).has(rowKey)?' checked':'')+'>';
}
function selHead(k){
  return '<input type="checkbox" id="'+k+'-selall" onclick="selAll(&quot;'+k+'&quot;,this)" title="Select all on this page">';
}
function onSel(k,cb){ const s=selSet(k); if(cb.checked)s.add(cb.value); else s.delete(cb.value); selBar(k); }
function selAll(k,cb){ const s=selSet(k); document.querySelectorAll(".selbox-"+k).forEach(b=>{b.checked=cb.checked; if(cb.checked)s.add(b.value); else s.delete(b.value);}); selBar(k); }
function selClear(k){ selSet(k).clear(); document.querySelectorAll(".selbox-"+k).forEach(b=>b.checked=false); const a=document.getElementById(k+"-selall"); if(a)a.checked=false; selBar(k); }
function selBar(k){ const bar=document.getElementById(k+"-bulkbar"); const n=selSet(k).size; if(bar)bar.style.display=n?"flex":"none"; const c=document.getElementById(k+"-selcount"); if(c)c.textContent=n; }
function bulkBarHtml(k){
  return '<div id="'+k+'-bulkbar" style="display:none;align-items:center;gap:12px;flex-wrap:wrap;background:#FEF2F2;border:1px solid #FECACA;border-radius:10px;padding:10px 14px;margin-bottom:12px">'+
    '<span style="font-weight:700;font-size:13px"><span id="'+k+'-selcount">0</span> selected</span>'+
    '<button class="btn btn-sm" style="background:#DC2626;color:#fff" onclick="bulkDelete(&quot;'+k+'&quot;)">&#128465; Delete selected</button>'+
    '<button class="btn btn-sm" style="background:var(--soft);color:var(--text)" onclick="selClear(&quot;'+k+'&quot;)">Clear</button></div>';
}
async function runSelectedFrom(k){
  const keys=Array.from(selSet(k));
  if(!keys.length){showToast("Select at least one student first");return;}
  if(!confirm("Run a fresh NIOS check for "+keys.length+" selected student(s)? Status + TOC both get checked; any change is pushed to the Portal."))return;
  try{
    const r=await api("/api/run-selected","POST",{row_keys:keys});
    showToast(r.message||"Run started");
    selClear(k);
  }catch(e){showToast(""+e.message);}
}
async function bulkDelete(k){
  const keys=Array.from(selSet(k));
  if(!keys.length)return;
  if(!confirm("Remove "+keys.length+" selected student(s)?\\n\\nThey move to Trash — you can restore from Settings → Deleted Students."))return;
  try{
    const r=await api("/api/students-delete-bulk","POST",{row_keys:keys});
    showToast("Moved "+(r.deleted||keys.length)+" student(s) to Trash");
    selClear(k);
    refreshAllTables();
  }catch(e){showToast("Error: "+e.message);}
}
function refreshAllTables(){
  ["loadStudents","loadConfirmed","loadRequired","loadFailed"].forEach(fn=>{try{window[fn](1);}catch(e){}});
  try{updateFailedBadge();}catch(e){}
  try{loadDashboard();}catch(e){}
}
const EDIT_FIELDS=["student_name","reference_no","enrollment_no","dob","mobile","alt_mobile","email","class_level","session"];
async function editStudent(rowKey){
  try{
    const s=await api("/api/student-get?row_key="+encodeURIComponent(rowKey));
    document.getElementById("edit-rowkey").value=s.row_key;
    EDIT_FIELDS.forEach(f=>{const el=document.getElementById("edit-"+f);if(el)el.value=s[f]||"";});
    const w=document.getElementById("edit-warn");
    if(s.login_failed==1||s.login_failed===true){
      w.style.display="block";
      w.innerHTML="&#9888; NIOS login failed — "+((s.login_remark||"check Reference/Enrollment No & DOB").replace(/</g,"&lt;"));
    }else{w.style.display="none";}
    var sy=document.getElementById("edit-syncline");
    if(sy){
      if(!s.student_id){sy.innerHTML='<span style="color:var(--muted)">Not linked to a Portal student — edits stay tracker-only.</span>';}
      else if(!s.edit_sync){sy.innerHTML='<span style="color:var(--muted)">Portal sync: details not pushed yet (will push on Save).</span>';}
      else if(s.edit_sync.indexOf("synced")===0){sy.innerHTML='<span style="color:var(--success)">Portal sync: '+s.edit_sync.replace(/</g,"&lt;")+' \u2713</span>';}
      else{sy.innerHTML='<span style="color:#B45309">\u26A0 Portal sync: '+s.edit_sync.replace(/</g,"&lt;")+'</span>';}
    }
    document.getElementById("edit-status").textContent="";
    document.getElementById("edit-overlay").style.display="flex";
  }catch(e){showToast(""+e.message);}
}
function closeEdit(){document.getElementById("edit-overlay").style.display="none";}
async function saveEdit(thenRun){
  const rk=document.getElementById("edit-rowkey").value;
  const body={row_key:rk};
  EDIT_FIELDS.forEach(f=>{const el=document.getElementById("edit-"+f);if(el)body[f]=el.value;});
  const st=document.getElementById("edit-status");
  try{
    st.style.color="var(--muted)";st.textContent="Saving…";
    const r=await api("/api/student-edit","POST",body);
    var pLine=r.portal_synced?" · Portal updated \u2713":
      " · \u26A0 Portal NOT updated ("+((r.portal_message||"").replace(/</g,"&lt;"))+")";
    if(thenRun){
      st.innerHTML="Saved"+pLine+". Re-checking on NIOS (status + login)… this can take ~20–40 sec.";
      await api("/api/student-recheck?row_key="+encodeURIComponent(rk),"POST");
      pollAfterRecheck(rk,0);
    }else{
      st.style.color=r.portal_synced?"var(--success)":"#B45309";
      st.innerHTML="Saved!"+pLine;
      setTimeout(closeEdit,r.portal_synced?900:3200);
      try{loadStudents(1);}catch(e){}try{loadConfirmed(1);}catch(e){}try{loadRequired(1);}catch(e){}
    }
  }catch(e){st.style.color="var(--danger)";st.textContent=e.message;}
}
async function pollAfterRecheck(rk,n){
  if(n>20){document.getElementById("edit-status").textContent="Still running… you can close this; the row will update shortly.";return;}
  try{
    const s=await api("/api/student-get?row_key="+encodeURIComponent(rk));
    const st=document.getElementById("edit-status");
    const failed=(s.login_failed==1||s.login_failed===true||s.check_failed==1||s.check_failed===true);
    if(failed){
      st.style.color="var(--danger)";
      st.innerHTML="&#9888; Still failing — "+((s.login_remark||"").replace(/</g,"&lt;"))+" Fix the field and run again.";
      const w=document.getElementById("edit-warn");w.style.display="block";w.innerHTML=st.innerHTML;
      try{loadStudents(1);}catch(e){}try{loadConfirmed(1);}catch(e){}try{loadDashboard();}catch(e){}
      try{loadFailed(1);}catch(e){}updateFailedBadge();
      return;
    }
    if((s.current_status||"")&&n>=1){
      st.style.color="var(--success)";
      st.textContent="✓ Done — status: "+s.current_status+(s.current_status==="Admission Confirmed"?" · login OK, documents sent.":".");
      try{loadStudents(1);}catch(e){}try{loadConfirmed(1);}catch(e){}try{loadRequired(1);}catch(e){}try{loadDashboard();}catch(e){}
      try{loadFailed(1);}catch(e){}updateFailedBadge();
      return;
    }
  }catch(e){}
  setTimeout(()=>pollAfterRecheck(rk,n+1),3000);
}
async function downloadSycHall(rowKey,btn){
  if(btn.dataset.busy==="1")return;
  btn.dataset.busy="1";
  const cell=btn.parentElement;
  const prog=cell.querySelector(".syc-prog"),bar=cell.querySelector(".syc-bar"),msg=cell.querySelector(".syc-msg");
  prog.style.display="block";bar.style.background="#4F46E5";bar.style.width="3%";msg.textContent="Fetching from NIOS... (~15 sec)";
  let p=3;const tick=setInterval(function(){p=Math.min(p+(p<70?7:2),92);bar.style.width=p+"%";},900);
  try{
    const r=await fetch(API+"/api/syc-doc?"+new URLSearchParams({row_key:rowKey}).toString(),{headers:{"Authorization":"Bearer "+TOKEN}});
    clearInterval(tick);
    if(!r.ok){let dt="failed";try{dt=(await r.json()).detail||dt;}catch(e){}throw new Error(dt);}
    bar.style.width="100%";msg.textContent="Opening...";
    const blob=await r.blob();
    window.open(URL.createObjectURL(blob),"_blank");
    msg.innerHTML='<span style="color:var(--success)">Opened — use "Save as PDF" to download</span>';
  }catch(e){clearInterval(tick);bar.style.background="var(--danger)";bar.style.width="100%";msg.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
  finally{btn.dataset.busy="";setTimeout(function(){prog.style.display="none";},5000);}
}

async function loadRequired(page){
  page=page||1;
  const q=new URLSearchParams({page:page,per_page:perPage,view:"required",
    search:document.getElementById("r-search").value,
    status_filter:(document.getElementById("r-status")?document.getElementById("r-status").value:""),
    session_filter:document.getElementById("r-session").value,
    source_filter:fval("r-source")});
  try{
    const d=await api("/api/students?"+q.toString());
    fillSessions(d.sessions);
    document.getElementById("r-count").textContent=d.total+" students need documents";
    const b=document.getElementById("r-body");
    b.innerHTML=d.students.length?d.students.map((s,i)=>'<tr>'+
      '<td>'+selBox("r",s.row_key)+'</td>'+
      '<td style="color:var(--muted)">'+((page-1)*perPage+i+1)+'</td>'+
      '<td><span class="ref-tag">'+(s.reference_no||"—")+'</span></td>'+
      '<td>'+(s.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(s)+'</div></td><td style="font-size:13px">'+(s.session||"—")+'</td>'+
      '<td style="font-size:13px;color:var(--warn);max-width:420px">'+(s.remark||"(no comment captured)")+'</td>'+
      '<td><button onclick="editStudent(&quot;'+s.row_key+'&quot;)" style="background:#e0e7ff;color:#3730a3;border:1px solid #c7d2fe;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer;white-space:nowrap">&#9998; Edit</button></td></tr>').join("")
      :'<tr><td colspan="7" class="empty">No pending documents</td></tr>';
    renderPg("r-pg",page,d.pages,"loadRequired");
    const sa=document.getElementById("r-selall");if(sa)sa.checked=false;
    selBar("r");
  }catch(e){showToast(""+e.message);}
}

let fTimer=null;
function debounceFailed(){clearTimeout(fTimer);fTimer=setTimeout(()=>loadFailed(1),400);}
async function diagnoseLogin(){
  var NL=String.fromCharCode(10);
  var ref=prompt("To test a REAL login, enter a confirmed student REFERENCE NO (or leave blank to only check reachability):","");
  if(ref===null) return;
  ref=(ref||"").trim();
  var dob="";
  if(ref){ dob=prompt("Enter that student DATE OF BIRTH (DD-MM-YYYY):",""); if(dob===null) return; dob=(dob||"").trim(); }
  try{
    var q=new URLSearchParams(); if(ref)q.set("ref",ref); if(dob)q.set("dob",dob);
    var resp=await fetch(API+"/api/nios-reach?"+q.toString(),{headers:{"Authorization":"Bearer "+TOKEN}});
    var txt=await resp.text();
    var r={}; try{r=JSON.parse(txt);}catch(e){}
    if(!resp.ok){ alert("Diagnose returned HTTP "+resp.status+NL+NL+txt.slice(0,500)); return; }
    var msg="NIOS DIAGNOSIS  (code version: "+(r.endpoint_version||"OLD")+")"+NL+NL+
      "Page HTTP status : "+(r.page_status)+NL+
      "Looks blocked    : "+(r.looks_blocked?"YES":"no")+NL+
      "CSRF found       : "+(r.csrf_found?"yes":"NO")+NL+
      "Site key OK      : "+((r.live_sitekey&&r.live_sitekey===r.built_in_sitekey)?"yes (same)":(r.live_sitekey?"CHANGED":"(none)"))+NL+
      "reCAPTCHA action : "+(r.recaptcha_action||"(none)")+NL+
      "reCAPTCHA TYPE   : "+(r.recaptcha_type||"?")+NL+
      "Action from JS   : "+(r.action_from_js||"?")+NL+
      "FORM FIELDS      : "+(r.form_fields||"(none)")+NL+
      "reCAPTCHA/Login fields: "+(r.recaptcha_fields||"(none)")+NL+NL+
      "--- LOGIN TEST ---"+NL+
      "Proxy set     : "+(r.proxy_set||"?")+NL+
      "Proxy DIRECT test: "+(r.proxy_direct_test||"?")+NL+
      "CapSolver create: "+(r.capsolver_create||"?")+NL+
      "CapSolver result: "+(r.capsolver_result||"?")+NL+
      "Captcha token : "+(r.captcha_token||"(skipped)")+NL+
      "DOB format test: "+(r.dob_format_test||"(skipped)")+NL+
      "verify_login (real path): "+(r.verify_login_test||"(skipped)")+NL+
      "Login result  : "+(r.login_result||"")+NL+
      (r.final_url?("Final URL: "+r.final_url+NL):"")+
      (r.snippet?(NL+"After login page says: "+r.snippet):"");
    alert(msg);
  }catch(e){ alert("Diagnose failed: "+e); }
}
async function loadFailed(page){
  page=page||1;
  const q=new URLSearchParams({page:page,per_page:perPage,
    search:(document.getElementById("f-search")?document.getElementById("f-search").value:""),
    source:fval("f-source")});
  try{
    const d=await api("/api/failed-students?"+q.toString());
    document.getElementById("f-count").textContent=d.total+" student(s) need data correction";
    const b=document.getElementById("f-body");
    b.innerHTML=d.students.length?d.students.map(s=>{
      var ref=(s.reference_no||s.enrollment_no||"—");
      var nm=(s.student_name||"this student").replace(/[\\"']/g," ");
      var rm=(s.login_remark||"NIOS login failed — check Reference/Enrollment No & DOB").replace(/</g,"&lt;");
      return '<tr>'+
        '<td>'+selBox("f",s.row_key)+'</td>'+
        '<td>'+(s.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(s)+
          ((s.check_count&&s.check_count>1)?' <span style="font-size:10.5px;color:#6b7280;background:#f3f4f6;border-radius:6px;padding:1px 6px;font-weight:600">checked '+s.check_count+'x</span>':'')+
          '</div></td>'+
        '<td><span class="ref-tag">'+ref+'</span></td>'+
        '<td style="font-size:13px">'+(s.session||"—")+'</td>'+
        '<td style="font-size:12px;color:#b91c1c;font-weight:600;max-width:300px">'+rm+'</td>'+
        '<td><div style="display:flex;gap:6px;flex-wrap:wrap">'+
          ((rm.indexOf("DIFFERENT student")>=0)?'<button onclick="markNameOk(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" style="background:#d1fae5;color:#065f46;border:1px solid #a7f3d0;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer" title="The name/Reference is actually correct — dismiss this warning">Name correct</button>':'')+
          '<button onclick="editStudent(&quot;'+s.row_key+'&quot;)" style="background:#e0e7ff;color:#3730a3;border:1px solid #c7d2fe;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer">Edit &amp; fix</button>'+
          '<button onclick="deleteStudent(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer">Remove</button>'+
        '</div></td></tr>';
    }).join("")
      :'<tr><td colspan="6" class="empty" style="color:var(--success)">No failed students — all clear!</td></tr>';
    renderPg("f-pg",page,d.pages,"loadFailed");
    const sa=document.getElementById("f-selall");if(sa)sa.checked=false;
    selBar("f");
  }catch(e){showToast(""+e.message);}
}
async function markNameOk(rk,nm){
  if(!confirm("Mark this student's name/Reference as CORRECT?\\n\\nOnly do this after checking on the portal that the Reference No genuinely belongs to "+nm+". This dismisses the wrong-reference warning."))return;
  try{
    const fd=new FormData();fd.append("row_key",rk);
    const r=await fetch("/api/mark-name-verified",{method:"POST",headers:{Authorization:"Bearer "+TOKEN},body:fd});
    const d=await r.json();
    if(!r.ok)throw new Error(d.detail||"Failed");
    showToast(d.confirmed?"Marked correct \u2014 moved to Confirmed":"Marked correct");
    loadFailed(1);updateNavCounts();
  }catch(e){showToast(""+e.message);}
}
async function runAllFailed(btn){
  if(!confirm("Re-check ALL failed students now?\\n\\nEach one is re-run with auto-retry (transient errors) and an automatic DOB date/month swap (formatting fixes). This uses CapSolver credits for every student in the list."))return;
  const old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/run-now-failed","POST");
    showToast(r.message||"Re-checking failed students…");
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
let uTimer=null;
function debounceUnknown(){clearTimeout(uTimer);uTimer=setTimeout(()=>loadUnknown(1),400);}
async function loadUnknown(page){
  page=page||1;
  const q=new URLSearchParams({page:page,per_page:perPage,
    search:(document.getElementById("u-search")?document.getElementById("u-search").value:""),
    session_filter:fval("u-session")});
  try{
    const d=await api("/api/unknown-students?"+q.toString());
    if(d.sessions)fillSessions(d.sessions);
    document.getElementById("u-count").textContent=d.total+" student(s) with Unknown status";
    const b=document.getElementById("u-body");
    b.innerHTML=d.students.length?d.students.map(s=>{
      var ref=(s.reference_no||s.enrollment_no||"\u2014");
      var nm=(s.student_name||"this student").replace(/[\\"']/g," ");
      var rm=(s.login_remark||"NIOS returned no recognisable status.").replace(/</g,"&lt;");
      return '<tr>'+
        '<td>'+(s.student_name||"\u2014")+'<div style="margin-top:4px">'+srcBadge(s)+
          ((s.check_count&&s.check_count>1)?' <span style="font-size:10.5px;color:#6b7280;background:#f3f4f6;border-radius:6px;padding:1px 6px;font-weight:600">checked '+s.check_count+'x</span>':'')+
          '</div></td>'+
        '<td><span class="ref-tag">'+ref+'</span></td>'+
        '<td style="font-size:13px">'+(s.session||"\u2014")+'</td>'+
        '<td style="font-size:12px;color:#b45309;font-weight:500;max-width:320px">'+rm+'</td>'+
        '<td style="font-size:12px;color:var(--muted)">'+(s.last_checked||"\u2014")+'</td>'+
        '<td><div style="display:flex;gap:6px;flex-wrap:wrap">'+
          '<button onclick="editStudent(&quot;'+s.row_key+'&quot;)" style="background:#e0e7ff;color:#3730a3;border:1px solid #c7d2fe;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer">Edit &amp; fix</button>'+
          '<button onclick="deleteStudent(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;cursor:pointer">Remove</button>'+
        '</div></td></tr>';
    }).join("")
      :'<tr><td colspan="6" class="empty" style="color:var(--success)">No Unknown students \u2014 all clear!</td></tr>';
    renderUnknownPg(page,d.pages,d.total);
  }catch(e){showToast(""+e.message);}
}
function renderUnknownPg(page,total,totalRows){
  const el=document.getElementById("u-pg");if(!el)return;
  if(!total||total<=0){el.innerHTML="";return;}
  let ctrl='<div class="pg-controls">';
  ctrl+='<button onclick="loadUnknown('+(page-1)+')" '+(page<=1?"disabled":"")+'>\u2039 Prev</button>';
  const start=Math.max(1,page-2),end=Math.min(total,page+2);
  for(let i=start;i<=end;i++)ctrl+='<button class="'+(i===page?'active':'')+'" onclick="loadUnknown('+i+')">'+i+'</button>';
  ctrl+='<button onclick="loadUnknown('+(page+1)+')" '+(page>=total?"disabled":"")+'>Next \u203a</button></div>';
  const sel='<div class="perpage">'+(totalRows!=null?'<span>'+totalRows+' students</span> \u00b7 ':'')+
    'Per page: <select onchange="perPage=parseInt(this.value);loadUnknown(1)">'+
    [10,20,50,100].map(n=>'<option value="'+n+'" '+(n===perPage?"selected":"")+'>'+n+'</option>').join("")+
    '</select></div>';
  el.innerHTML=ctrl+sel;
}
async function loadPending(force){
  var tb=document.getElementById("pend-body");
  if(!tb)return;
  tb.innerHTML='<tr><td colspan="6" class="empty">Fetching live from the Portal…</td></tr>';
  try{
    const d=await api("/api/pending-students");
    var badge=document.getElementById("nav-pending-badge");
    if(badge){badge.style.display=(d.total>0)?"inline-block":"none";badge.textContent=d.total||0;}
    var cnt=document.getElementById("pend-count");
    var c=(d.counts||{});
    if(cnt){
      cnt.innerHTML=(d.total||0)+" pending student(s) — "+(c.no_ref||0)+" with a blank reference, "+
        (c.bad_ref||0)+" with an invalid one"+
        '<div style="margin-top:4px;font-size:12px">The Portal API sent <b>'+(d.portal_total||0)+
        '</b> students in total. If that is lower than the Total Students on your Portal dashboard, '+
        'the missing ones are students the Portal API does not send at all (usually blank-reference ones) '+
        '— they cannot appear here until the Portal includes them.</div>';
    }
    if(!d.ok){tb.innerHTML='<tr><td colspan="8" class="empty">'+(d.message||"Could not load")+'</td></tr>';return;}
    if(!d.students.length){tb.innerHTML='<tr><td colspan="8" class="empty">No pending students — every Portal student has a valid reference. 🎉</td></tr>';return;}
    tb.innerHTML=d.students.map(function(s,i){
      var bad=(s.reason||"").indexOf("No reference")!==0;
      return '<tr><td>'+(i+1)+'</td>'+
        '<td><b>'+(s.name||"—")+'</b></td>'+
        '<td>'+(s.mobile||"—")+'</td>'+
        '<td>'+(s.session||"—")+'</td>'+
        '<td>'+(s.class_level||"—")+'</td>'+
        '<td style="font-size:12px">'+(s.reference?('<code style="background:var(--soft);padding:2px 6px;border-radius:5px">'+s.reference+'</code>'):'<span style="color:var(--muted)">—</span>')+'</td>'+
        '<td style="font-size:12px;color:'+(bad?"#B91C1C":"var(--muted)")+'">'+(s.reason||"—")+'</td>'+
        '<td style="font-size:12px;color:var(--muted)">'+(s.student_id||"—")+'</td></tr>';
    }).join("");
  }catch(e){tb.innerHTML='<tr><td colspan="8" class="empty">'+e.message+'</td></tr>';}
}
let TE_PAGE=1, TE_TIMER=null;
function teSearch(){ clearTimeout(TE_TIMER); TE_TIMER=setTimeout(function(){loadTocErrors(1);},400); }
async function loadTocErrors(page){
  if(page)TE_PAGE=page;
  var tb=document.getElementById("te-body");
  if(!tb)return;
  try{ pollTocProgress(); }catch(e){}
  var badge=document.getElementById("nav-tocerror-badge");
  tb.innerHTML='<tr><td colspan="10" class="empty">Loading…</td></tr>';
  try{
    const q=new URLSearchParams({search:(document.getElementById("te-search")?document.getElementById("te-search").value:""),
                                 page:TE_PAGE,per_page:20});
    const d=await api("/api/toc-fixes?"+q.toString());
    if(badge){badge.style.display=(d.total>0)?"inline-block":"none";badge.textContent=d.total;}
    var cnt=document.getElementById("te-count");
    if(cnt)cnt.textContent=d.total+" auto-corrected student(s)";
    if(!d.fixes||!d.fixes.length){
      tb.innerHTML='<tr><td colspan="10" class="empty">No auto-corrections yet. Mismatches found during any check or run are fixed automatically and listed here.</td></tr>';
      renderTePg(1,1);return;
    }
    tb.innerHTML=d.fixes.map(function(f){
      var rk=(f.row_key||"").replace(/'/g,"\\'");
      return '<tr>'+
        '<td><b>'+(f.student_name||"—")+'</b></td>'+
        '<td>'+(f.reference_no||"—")+'</td>'+
        '<td>'+(f.session||"—")+'</td>'+
        '<td style="font-size:12px">'+(f.current_status||"—")+'</td>'+
        '<td><span style="font-weight:700;color:#B45309">'+((f.old_toc||"?")+"").toUpperCase()+'</span> '+
            '<span style="color:var(--muted)">&rarr;</span> '+
            '<span style="font-weight:700;color:#047857">'+((f.new_toc||"?")+"").toUpperCase()+'</span></td>'+
        '<td style="font-size:12px">'+(f.subjects||"—")+'</td>'+
        '<td style="font-size:12px;color:var(--muted)">'+(f.fixed_at||"—")+'</td>'+
        '<td>'+(f.pushed==1?'<span style="font-size:11.5px;font-weight:700;color:#047857;background:#D1FAE5;padding:2px 8px;border-radius:6px">pushed</span>'
                           :'<span style="font-size:11.5px;font-weight:700;color:#B45309;background:#FEF3C7;padding:2px 8px;border-radius:6px" title="Push will be retried automatically">pending</span>')+'</td>'+
        '<td>'+(f.wa_sent_before==1?'<span style="font-size:11.5px;font-weight:700;color:#B91C1C;background:#FEE2E2;padding:2px 8px;border-radius:6px" title="Documents had ALREADY been sent using the OLD (wrong) TOC campaign — resend the correct ones.">wrong docs sent</span>':'<span style="font-size:11.5px;color:var(--muted)">not sent before</span>')+'</td>'+
        '<td style="white-space:nowrap">'+
          (f.wa_sent_before==1?('<button class="btn btn-sm" style="background:#16A34A;color:#fff" onclick="resendWa(&quot;'+rk+'&quot;,this)">Resend correct docs</button> '):'')+
          '<button class="btn btn-outline btn-sm" onclick="tocEdit(&quot;'+rk+'&quot;)">Edit</button></td></tr>';
    }).join("");
    renderTePg(d.page,d.pages);
  }catch(e){tb.innerHTML='<tr><td colspan="10" class="empty">'+e.message+'</td></tr>';}
}
function renderTePg(page,total){
  var el=document.getElementById("te-pg");
  if(!el)return;
  if(!total||total<=1){el.innerHTML="";return;}
  el.innerHTML='<div class="pg-controls">'+
    '<button class="btn btn-outline btn-sm" '+(page<=1?"disabled":"")+' onclick="loadTocErrors('+(page-1)+')">&lsaquo; Prev</button>'+
    '<span style="font-size:12.5px;color:var(--muted);padding:0 10px">Page '+page+' of '+total+'</span>'+
    '<button class="btn btn-outline btn-sm" '+(page>=total?"disabled":"")+' onclick="loadTocErrors('+(page+1)+')">Next &rsaquo;</button></div>';
}
async function tocVerify(rk,status,subs){
  if(!confirm("Verify this student as TOC = "+status.toUpperCase()+"? The correct TOC is pushed to the Portal and the WhatsApp goes with the right campaign."))return;
  try{
    const r=await api("/api/toc-verify","POST",{row_key:rk,toc_status:status,toc_subjects:subs||[]});
    showToast(r.message||"Verified"); loadTocErrors();
  }catch(e){showToast(""+e.message);}
}
async function tocEdit(rk){
  var choice=prompt("Set TOC status - type 'yes' or 'no':","yes");
  if(!choice)return; choice=choice.trim().toLowerCase();
  if(choice!=="yes"&&choice!=="no"){showToast("Type yes or no");return;}
  var subs=[];
  if(choice==="yes"){
    var s=prompt("Which subjects are TOC (Yes)? Comma-separated (optional):","");
    if(s)subs=s.split(",").map(function(x){return x.trim();}).filter(Boolean);
  }
  try{
    const r=await api("/api/toc-verify","POST",{row_key:rk,toc_status:choice,toc_subjects:subs});
    showToast(r.message||"Saved"); loadTocErrors();
  }catch(e){showToast(""+e.message);}
}
async function runTocCheck(btn){
  if(!confirm("Read the real TOC from NIOS for all unverified no-TOC students? This uses CapSolver per student and flags mismatches here. It won't run while a status run is active."))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{ const r=await api("/api/toc-check","POST",{}); showToast(r.message||"TOC check started"); }
  catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function runTocCheckConfirmed(btn){
  if(!confirm("One-time: read the real NIOS TOC for every confirmed student not yet checked. This uses CapSolver per student and can take a while for many students — it runs in the background and you can watch the progress. Safe to re-click later to finish any that remain."))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/toc-check-confirmed","POST",{});
    showToast(r.message||"Started");
    if(r.count>0) pollTocProgress();
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function forceRecheckAll(btn){
  if(!confirm("Re-read the NIOS TOC for EVERY confirmed student (even already-checked / verified ones)? Any wrong TOC is corrected automatically, pushed to the Portal, and listed below — including students whose documents already went out with the wrong campaign. Uses one CapSolver credit per student."))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/toc-check-confirmed","POST",{force:true});
    showToast(r.message||"Started");
    if(r.count>0) pollTocProgress();
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function recheckFlagged(btn){
  if(!confirm("Fresh NIOS read for every still-flagged mismatch? Genuine mismatches are auto-corrected (NIOS value applied + pushed to the Portal) and listed below; wrong flags clear."))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/toc-check-confirmed","POST",{flagged:true});
    showToast(r.message||"Started");
    if(r.count>0) pollTocProgress();
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function tocCheckSelected(which,btn){
  var keys = (which==="confirmed") ? Array.from((typeof CONF_SEL!=="undefined")?CONF_SEL:[])
                                   : Array.from((typeof SELECTED!=="undefined")?SELECTED:[]);
  if(!keys.length){showToast("Select at least one student first");return;}
  if(!confirm("Read the real NIOS TOC for the "+keys.length+" selected student(s)? Uses CapSolver per student. Any mismatch with the Portal appears in TOC Status Error (WhatsApp held until verified)."))return;
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting…";}
  try{
    const r=await api("/api/toc-check-confirmed","POST",{row_keys:keys});
    showToast(r.message||"Started");
    if(r.count>0){ nav("tocerror"); pollTocProgress(); }
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
let TOC_POLL_WAS_RUNNING=false, TOC_POLL_TIMER=null;
async function pollTocProgress(){
  const box=document.getElementById("toc-progress");
  if(!box)return;
  if(TOC_POLL_TIMER){clearTimeout(TOC_POLL_TIMER);TOC_POLL_TIMER=null;}
  try{
    const d=await api("/api/toc-check-progress");
    if((d.total||0)===0 && !d.running){box.style.display="none";TOC_POLL_WAS_RUNNING=false;return;}
    box.style.display="block";
    var pct=d.percent||0;
    var cell=function(lbl,val,col){return '<span style="font-weight:700;color:'+col+'">'+(val||0)+'</span> <span style="color:var(--muted);font-size:12px">'+lbl+'</span>';};
    box.innerHTML=
      '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:9px">'+
        '<span style="font-weight:700;font-size:13.5px">'+(d.running?"Checking students\u2019 TOC\u2026":"Confirmed TOC check")+'</span>'+
        '<span style="font-size:13px;color:var(--muted)"><b style="color:var(--text)">'+(d.done||0)+'</b> / '+(d.total||0)+' &nbsp;('+pct+'%)</span>'+
      '</div>'+
      '<div style="height:9px;background:var(--border);border-radius:6px;overflow:hidden;margin-bottom:11px"><div style="height:100%;width:'+pct+'%;background:#4F46E5;transition:width .4s"></div></div>'+
      '<div style="display:flex;gap:18px;flex-wrap:wrap;font-size:13px">'+
        cell("checked",d.done,"#111827")+cell("yes-TOC",d.yes_toc,"#047857")+
        cell("no-TOC",d.no_toc,"#B45309")+cell("auto-fixed",d.errors,"#B91C1C")+
        cell("could not read",Math.max(0,(d.done||0)-((d.yes_toc||0)+(d.no_toc||0))),"#6B7280")+
        (d.running?'<button onclick="cancelTocCheck()" style="margin-left:auto;background:#FEE2E2;color:#B91C1C;border:1px solid #FECACA;border-radius:7px;padding:3px 12px;font-size:12px;font-weight:700;cursor:pointer">Stop</button>':'')+
      '</div>'+
      (d.message?('<div style="margin-top:10px;font-size:12px;color:var(--muted)">'+d.message+'</div>'):'');
    if(d.running){
      TOC_POLL_WAS_RUNNING=true;
      TOC_POLL_TIMER=setTimeout(pollTocProgress,2000);
    } else if(TOC_POLL_WAS_RUNNING){
      // Just transitioned running -> done: refresh the audit list ONCE, show the final summary
      // briefly, then hide the panel (it no longer sticks around forever).
      TOC_POLL_WAS_RUNNING=false;
      loadTocErrors();
      setTimeout(function(){var b=document.getElementById("toc-progress");if(b)b.style.display="none";},12000);
    } else {
      // Idle visit with a stale completed state — don't resurrect the old panel.
      box.style.display="none";
    }
  }catch(e){ if(TOC_POLL_WAS_RUNNING){TOC_POLL_TIMER=setTimeout(pollTocProgress,3000);} }
}
async function cancelTocCheck(){
  try{ await api("/api/toc-check-cancel","POST",{}); showToast("Stopping…"); }catch(e){}
}
async function runNotChecked(){
  var sess=document.getElementById("s-session").value;
  var scope=sess?("session "+sess):"ALL sessions";
  if(!confirm("Re-check every not-checked student in "+scope+"?\\n\\nThis re-runs only the students the latest run couldn't read (captcha/proxy). Confirmed students are skipped. Uses CapSolver credits for each."))return;
  var btn=document.getElementById("s-run-nc-btn");
  var old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting\u2026";}
  try{
    const r=await api("/api/run-now-notchecked","POST",{session_filter:sess,source_filter:fval("s-source")});
    showToast(r.message||"Re-checking not-checked students\u2026");
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function runAllUnknown(btn){
  if(!confirm("Re-check ALL Unknown students now?\\n\\nEach is re-run with auto-retry. Uses CapSolver credits for every student in this list."))return;
  const old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Starting\u2026";}
  try{
    const r=await api("/api/run-now-unknown","POST");
    showToast(r.message||"Re-checking Unknown students\u2026");
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
function _setNavBadge(id,n){
  const b=document.getElementById(id);if(!b)return;
  if(n>0){b.textContent=n;b.style.display="inline-block";b.classList.add("has");}
  else{b.style.display="none";b.classList.remove("has");}
}
async function updateNavCounts(){
  try{
    const d=await api("/api/nav-count");
    _setNavBadge("nav-students-badge",d.students);
    _setNavBadge("nav-confirmed-badge",d.confirmed);
    _setNavBadge("nav-required-badge",d.required);
    _setNavBadge("nav-syc-badge",d.syc);
    _setNavBadge("nav-failed-badge",d.failed);
    _setNavBadge("nav-unknown-badge",d.unknown);
  }catch(e){}
}
async function updateFailedBadge(){return updateNavCounts();}   // back-compat
function renderPg(id,page,total,fnName){
  const el=document.getElementById(id);
  let ctrl='<div class="pg-controls">';
  ctrl+='<button onclick="'+fnName+'('+(page-1)+')" '+(page<=1?"disabled":"")+'>‹ Prev</button>';
  const start=Math.max(1,page-2),end=Math.min(total,page+2);
  for(let i=start;i<=end;i++)ctrl+='<button class="'+(i===page?'active':'')+'" onclick="'+fnName+'('+i+')">'+i+'</button>';
  ctrl+='<button onclick="'+fnName+'('+(page+1)+')" '+(page>=total?"disabled":"")+'>Next ›</button></div>';
  const sel='<div class="perpage">Per page: <select onchange="setPerPage(this.value,&quot;'+fnName+'&quot;)">'+
    [10,20,50,100].map(n=>'<option value="'+n+'" '+(n===perPage?"selected":"")+'>'+n+'</option>').join("")+
    '</select></div>';
  el.innerHTML=ctrl+sel;
}
function setPerPage(v,fnName){perPage=parseInt(v);window[fnName](1);}

let histPerPage=10;
function fmtDT(d){
  const p=n=>String(n).padStart(2,"0");
  return d.getFullYear()+"-"+p(d.getMonth()+1)+"-"+p(d.getDate())+" "+p(d.getHours())+":"+p(d.getMinutes())+":"+p(d.getSeconds());
}
function dtLocalToStr(v){ // "2026-06-11T18:30" -> "2026-06-11 18:30:00"
  if(!v)return "";
  v=v.replace("T"," ");
  return v.length===16?v+":00":v;
}
function onHistPreset(){
  const custom=document.getElementById("h-preset").value==="custom";
  document.getElementById("h-custom").style.display=custom?"flex":"none";
  if(custom){
    const f=document.getElementById("h-from"),t=document.getElementById("h-to");
    if(f&&!f.value){const y=new Date();y.setDate(y.getDate()-1);y.setHours(18,30,0,0);f.value=toLocalInput(y);}
    if(t&&!t.value){t.value=toLocalInput(new Date());}
  }else loadHistory(1);
}
function toLocalInput(d){const p=n=>String(n).padStart(2,"0");
  return d.getFullYear()+"-"+p(d.getMonth()+1)+"-"+p(d.getDate())+"T"+p(d.getHours())+":"+p(d.getMinutes());}
function histRange(){
  const preset=fval("h-preset");
  const now=new Date();
  let from="",to="";
  if(preset==="today"){const s=new Date(now);s.setHours(0,0,0,0);from=fmtDT(s);to=fmtDT(now);}
  else if(preset==="yesterday"){const s=new Date(now);s.setDate(s.getDate()-1);s.setHours(0,0,0,0);
    const e=new Date(s);e.setHours(23,59,59,0);from=fmtDT(s);to=fmtDT(e);}
  else if(preset==="24h"){const s=new Date(now.getTime()-86400000);from=fmtDT(s);to=fmtDT(now);}
  else if(preset==="custom"){from=dtLocalToStr(fval("h-from"));to=dtLocalToStr(fval("h-to"));}
  return {from,to};
}
let _histSearchT=null;
function histSearchDebounced(){clearTimeout(_histSearchT);_histSearchT=setTimeout(()=>loadHistory(1),350);}
async function loadHistory(page){
  page=page||1;
  const rg=histRange();
  const q=new URLSearchParams({page:page,per_page:histPerPage,from_dt:rg.from,to_dt:rg.to,
    status:fval("h-status"),source:fval("h-source"),search:fval("h-search")});
  try{
    const d=await api("/api/history?"+q.toString());
    const items=d.items||[];
    const cnt=document.getElementById("h-count");
    if(cnt)cnt.textContent=d.total+" change(s)"+(rg.from?" in selected range":"");
    document.getElementById("h-body").innerHTML=items.length?items.map(x=>'<tr>'+
      '<td><span class="ref-tag">'+(x.reference_no||"—")+'</span></td>'+
      '<td>'+(x.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(x)+'</div></td>'+
      '<td>'+badge(x.old_status)+'</td><td>'+badge(x.new_status)+'</td>'+
      '<td style="font-size:12px;color:var(--muted)">'+x.changed_at+'</td>'+
      '<td><button title="Delete this history entry" style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:4px 9px;font-size:11px;font-weight:600;cursor:pointer" onclick="deleteHistory('+x.id+')">&#128465;</button></td></tr>').join("")
      :'<tr><td colspan="6" class="empty">No changes in this range</td></tr>';
    renderHistPg(d.page,d.pages,d.total);
  }catch(e){showToast(""+e.message);}
}
async function deleteHistory(id){
  if(!confirm("Delete this history entry?"))return;
  try{await api("/api/history-delete?id="+id,"POST");showToast("Deleted");loadHistory(1);}
  catch(e){showToast(""+e.message);}
}
async function clearHistory(){
  if(!confirm("Clear ALL status-change history?\\n\\nThis permanently deletes every history entry. Students are NOT affected."))return;
  try{const r=await api("/api/history-clear","POST");showToast("Cleared "+(r.deleted||0)+" entries");loadHistory(1);}
  catch(e){showToast(""+e.message);}
}
async function exportHistory(){
  const rg=histRange();
  const q=new URLSearchParams({from_dt:rg.from,to_dt:rg.to,status:fval("h-status"),source:fval("h-source")});
  try{
    showToast("Preparing Excel...");
    const r=await fetch(API+"/api/export-history?"+q.toString(),{headers:{Authorization:"Bearer "+TOKEN}});
    if(!r.ok){showToast("Export failed");return;}
    const blob=await r.blob();const url=URL.createObjectURL(blob);
    const a=document.createElement("a");a.href=url;a.download="nios_change_history.xlsx";
    document.body.appendChild(a);a.click();a.remove();
    setTimeout(()=>URL.revokeObjectURL(url),1500);showToast("Excel downloaded");
  }catch(e){showToast("Error: "+e.message);}
}
function renderHistPg(page,total,totalRows){
  const el=document.getElementById("h-pg");if(!el)return;
  if(!total||total<=0){el.innerHTML="";return;}
  let ctrl='<div class="pg-controls">';
  ctrl+='<button onclick="loadHistory('+(page-1)+')" '+(page<=1?"disabled":"")+'>‹ Prev</button>';
  const start=Math.max(1,page-2),end=Math.min(total,page+2);
  for(let i=start;i<=end;i++)ctrl+='<button class="'+(i===page?'active':'')+'" onclick="loadHistory('+i+')">'+i+'</button>';
  ctrl+='<button onclick="loadHistory('+(page+1)+')" '+(page>=total?"disabled":"")+'>Next ›</button></div>';
  const sel='<div class="perpage">'+(totalRows!=null?'<span>'+totalRows+' changes</span> · ':'')+
    'Per page: <select onchange="histPerPage=parseInt(this.value);loadHistory(1)">'+
    [10,20,30,50,100].map(n=>'<option value="'+n+'" '+(n===histPerPage?"selected":"")+'>'+n+'</option>').join("")+
    '</select></div>';
  el.innerHTML=ctrl+sel;
}
async function loadRunLogs(){
  const sel=document.getElementById("rl-limit");
  const lim=sel?parseInt(sel.value):50;
  try{const l=await api("/api/run-logs?limit="+lim);renderRuns(l,"rl-body");}catch(e){showToast(""+e.message);}
}
let transPerPage=10,_trSearchT=null;
function trSearchDebounced(){clearTimeout(_trSearchT);_trSearchT=setTimeout(()=>loadTransfers(1),350);}
async function clearTransferLog(){
  if(!confirm("Clear the transfer history list? Students already transferred to the Portal stay transferred \u2014 only this log below is reset.")) return;
  try{
    const r=await api("/api/transfers-clear","POST");
    showToast(((r&&r.deleted!=null)?r.deleted:0)+" log record(s) cleared");
  }catch(e){showToast("Could not clear log");}
  loadTransfers(1);
}
async function loadTransfers(page){
  page=page||1;
  loadLastMatch();
  const q=new URLSearchParams({page:page,per_page:transPerPage,
    search:fval("tr-search"),mode:fval("tr-mode")});
  try{
    const d=await api("/api/transfers?"+q.toString());
    const items=d.items||[];
    const cnt=document.getElementById("tr-count");
    if(cnt)cnt.textContent=d.total+" record(s) transferred Tracker \u2192 Portal";
    document.getElementById("tr-body").innerHTML=items.length?items.map(x=>{
      var mode=(x.mode==="manual")
        ?'<span style="font-size:11px;font-weight:600;color:#3730a3;background:#e0e7ff;border-radius:6px;padding:2px 8px">Manual</span>'
        :'<span style="font-size:11px;font-weight:600;color:#065f46;background:#d1fae5;border-radius:6px;padding:2px 8px">Auto</span>';
      return '<tr>'+
        '<td>'+(x.student_name||"\u2014")+'<div style="margin-top:3px"><span style="font-size:10.5px;color:#6b7280">'+(x.mobile||"")+'</span></div></td>'+
        '<td><span class="ref-tag">'+(x.reference_no||x.enrollment_no||"\u2014")+'</span></td>'+
        '<td style="font-size:13px">'+(x.session||"\u2014")+'</td>'+
        '<td>'+badge(x.old_status||"\u2014")+'</td>'+
        '<td>'+badge(x.new_status||"\u2014")+'</td>'+
        '<td style="font-size:12px;color:var(--muted)">'+(x.transferred_at||"")+'</td>'+
        '<td>'+mode+'</td></tr>';
    }).join("")
      :'<tr><td colspan="7" class="empty">No transfers yet. When a Tracker student is matched to Portal during a run, it appears here.</td></tr>';
    renderTransPg(d.page,d.pages,d.total);
  }catch(e){showToast(""+e.message);}
}
function renderTransPg(page,total,totalRows){
  const el=document.getElementById("tr-pg");if(!el)return;
  if(!total||total<=0){el.innerHTML="";return;}
  let ctrl='<div class="pg-controls">';
  ctrl+='<button onclick="loadTransfers('+(page-1)+')" '+(page<=1?"disabled":"")+'>\u2039 Prev</button>';
  const start=Math.max(1,page-2),end=Math.min(total,page+2);
  for(let i=start;i<=end;i++)ctrl+='<button class="'+(i===page?'active':'')+'" onclick="loadTransfers('+i+')">'+i+'</button>';
  ctrl+='<button onclick="loadTransfers('+(page+1)+')" '+(page>=total?"disabled":"")+'>Next \u203a</button></div>';
  const sel='<div class="perpage">'+(totalRows!=null?'<span>'+totalRows+' records</span> \u00b7 ':'')+
    'Per page: <select onchange="transPerPage=parseInt(this.value);loadTransfers(1)">'+
    [10,20,30,50,100].map(n=>'<option value="'+n+'" '+(n===transPerPage?"selected":"")+'>'+n+'</option>').join("")+
    '</select></div>';
  el.innerHTML=ctrl+sel;
}
let staleKeys=[];
async function findStalePortal(btn){
  const box=document.getElementById("stale-box");
  if(btn){btn.disabled=true;btn.textContent="Checking\u2026";}
  if(box)box.innerHTML='<div style="color:var(--muted);font-size:13px;padding:8px 0">Fetching the live portal and comparing\u2026</div>';
  try{
    const r=await api("/api/portal-stale");
    if(r.ok===false){box.innerHTML='<div style="color:var(--warn);font-size:13px">'+(r.message||"Could not check right now.")+'</div>';return;}
    staleKeys=(r.stale||[]).map(s=>s.row_key);
    if(!r.stale||!r.stale.length){
      box.innerHTML='<div style="color:var(--success);font-size:13.5px;font-weight:600;padding:8px 0">All clean — every tracked MVS Portal student is still on the live portal. (Compared against '+r.fetched+' live records.)</div>';
      return;
    }
    var rows=r.stale.map((s,i)=>'<tr>'+
      '<td><input type="checkbox" class="stale-cb" value="'+s.row_key+'" checked onchange="updateStaleCount()"></td>'+
      '<td style="font-size:12px;color:var(--muted)">'+(i+1)+'</td>'+
      '<td>'+(s.student_name||"\u2014")+'</td>'+
      '<td><span class="ref-tag">'+(s.reference_no||"\u2014")+'</span></td>'+
      '<td style="font-size:12.5px">'+(s.session||"\u2014")+'</td>'+
      '<td style="font-size:12.5px">'+(s.status||"\u2014")+'</td>'+
      '<td style="font-size:12px;color:var(--muted)">'+(s.last_checked||"\u2014")+'</td>'+
      '</tr>').join("");
    box.innerHTML='<div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:9px;padding:10px 13px;margin-bottom:10px;font-size:13.5px;color:#9a3412">'+
        'Found <b>'+r.stale.length+'</b> old portal student(s) no longer on the live portal (compared against '+r.fetched+' live records). '+
        '<b>Tick the ones you want</b>, then either move them to the Tracker (keeps the data) or to Trash.</div>'+
      '<div style="overflow-x:auto"><table><thead><tr>'+
        '<th><input type="checkbox" id="stale-all" checked onclick="toggleStaleAll(this)"></th>'+
        '<th>#</th><th>Student</th><th>Reference / Enroll</th><th>Session</th><th>Last Status</th><th>Last Checked</th>'+
        '</tr></thead><tbody>'+rows+'</tbody></table></div>'+
      '<div style="margin-top:13px;display:flex;gap:10px;flex-wrap:wrap;align-items:center">'+
        '<button class="btn btn-sm" style="background:#2563eb;color:#fff" onclick="transferStaleToTracker(this)">&#8594; Transfer selected to Tracker</button>'+
        '<button class="btn btn-sm" style="background:#dc2626;color:#fff" onclick="removeStalePortal(this)">&#128465; Move selected to Trash</button>'+
        '<span id="stale-count" style="font-size:12.5px;color:var(--muted)"></span>'+
      '</div>'+
      '<div style="font-size:12px;color:var(--muted);margin-top:7px">&#8594; <b>Transfer to Tracker</b> keeps the record (re-checked as Tracker data, no longer linked to the portal). &#128465; <b>Trash</b> is recoverable from Settings.</div>';
    updateStaleCount();
  }catch(e){box.innerHTML='<div style="color:var(--danger);font-size:13px">'+e.message+'</div>';}
  finally{if(btn){btn.disabled=false;btn.textContent="Find old portal data";}}
}
function getStaleSelected(){
  return Array.from(document.querySelectorAll(".stale-cb:checked")).map(c=>c.value);
}
function updateStaleCount(){
  const el=document.getElementById("stale-count");if(!el)return;
  el.textContent=getStaleSelected().length+" selected";
}
function toggleStaleAll(cb){
  document.querySelectorAll(".stale-cb").forEach(c=>{c.checked=cb.checked;});
  updateStaleCount();
}
async function transferStaleToTracker(btn){
  const keys=getStaleSelected();
  if(!keys.length){showToast("Select at least one student first");return;}
  if(!confirm("Move these "+keys.length+" student(s) from MVS Portal to MVS Tracker?\\n\\nThe records are kept and re-checked as Tracker data — they're just no longer linked to the live portal."))return;
  if(btn){btn.disabled=true;btn.textContent="Moving\u2026";}
  try{
    const r=await api("/api/portal-to-tracker","POST",{row_keys:keys});
    showToast("Moved "+(r.moved||keys.length)+" student(s) to MVS Tracker");
    document.getElementById("stale-box").innerHTML='<div style="color:var(--success);font-size:13.5px;font-weight:600;padding:8px 0">Done — '+(r.moved||0)+' record(s) moved to MVS Tracker. They are kept and will be re-checked as Tracker data.</div>';
    try{loadDashboard();}catch(e){}
  }catch(e){showToast("Error: "+e.message);if(btn){btn.disabled=false;btn.textContent="\u2192 Transfer selected to Tracker";}}
}
async function removeStalePortal(btn){
  const keys=getStaleSelected();
  if(!keys.length){showToast("Select at least one student first");return;}
  if(!confirm("Move these "+keys.length+" student(s) to Trash?\\n\\nThey can be restored anytime from Settings \u2192 Trash."))return;
  if(btn){btn.disabled=true;btn.textContent="Removing\u2026";}
  try{
    const r=await api("/api/students-delete-bulk","POST",{row_keys:keys});
    showToast("Moved "+(r.deleted||keys.length)+" student(s) to Trash");
    document.getElementById("stale-box").innerHTML='<div style="color:var(--success);font-size:13.5px;font-weight:600;padding:8px 0">Done — '+(r.deleted||0)+' record(s) moved to Trash.</div>';
    try{loadDashboard();}catch(e){}
  }catch(e){showToast("Error: "+e.message);if(btn){btn.disabled=false;btn.textContent="Move selected to Trash";}}
}
async function syncTransfers(btn){
  if(!confirm("Push the current status of every transferred Portal student to MVS Portal now?\\n\\nThis re-sends each transferred student's latest NIOS status + document links to the portal and logs them as Manual transfers."))return;
  const old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Syncing\u2026";}
  try{
    const r=await api("/api/transfer-sync","POST");
    showToast(r.message||"Synced to Portal");
    loadTransfers(1);
  }catch(e){showToast(""+e.message);}
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
function renderMatchResult(r){
  const box=document.getElementById("tr-match-result");
  if(!box||!r||!r.ok)return;
  const esc=(v)=>(""+(v==null?"":v)).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
  const scanned=(r.transferred||0)+(r.new_fetch||0);
  let html='<div style="background:var(--primary-light);border:1px solid var(--primary);border-radius:10px;padding:12px 14px">'+
    '<div style="font-weight:700;color:var(--primary-dark)">\u2713 '+(r.transferred||0)+' matched & transferred to Portal (no CapSolver used)'+(r.at?' \u00b7 <span style="font-weight:500;color:var(--muted);font-size:12px">'+esc(r.at)+'</span>':'')+'</div>'+
    '<div style="font-size:13px;color:var(--muted);margin-top:2px">Scanned '+scanned+' portal students \u00b7 '+(r.new_fetch||0)+' not matched (left for New Fetch).</div></div>';
  const nm=r.not_matched||[];
  if(nm.length){
    html+='<div style="margin-top:12px;font-weight:600;font-size:13.5px">Not transferred yet \u2014 reason ('+nm.length+'):</div>'+
      '<div style="overflow-x:auto;margin-top:6px;max-height:340px;overflow-y:auto"><table><thead><tr><th>Student</th><th>Reference No</th><th>Why not checked</th></tr></thead><tbody>';
    nm.forEach(function(x){
      html+='<tr><td>'+esc(x.student_name||"\u2014")+'</td><td>'+esc(x.reference_no||"\u2014")+'</td>'+
        '<td style="color:var(--warn);font-size:12.5px">'+esc(x.reason||"")+'</td></tr>';
    });
    html+='</tbody></table></div>';
  }
  box.innerHTML=html;
}
async function loadLastMatch(){
  try{const r=await api("/api/transfer-match-last");if(r&&r.ok)renderMatchResult(r);}catch(e){}
}
function _matchBarHTML(){
  return '<div style="background:var(--soft);border:1px solid var(--border);border-radius:10px;padding:12px 14px">'+
    '<div style="font-weight:600;font-size:13.5px">Match &amp; Transfer running\u2026 <span id="tm-pct">\u2026</span></div>'+
    '<div style="height:9px;background:var(--border);border-radius:6px;overflow:hidden;margin-top:8px"><div id="tm-fill" style="height:100%;width:0%;background:var(--primary);transition:width .3s"></div></div>'+
    '<div id="tm-sub" style="font-size:12px;color:var(--muted);margin-top:6px">Starting\u2026</div></div>';
}
function _updMatchBar(p){
  const pct=(p.total>0)?Math.round(p.done*100/p.total):0;
  const fill=document.getElementById("tm-fill");if(fill)fill.style.width=pct+"%";
  const showPct=(p.phase==="Pushing"||p.phase==="Matching");
  const pe=document.getElementById("tm-pct");if(pe)pe.textContent=showPct?(pct+"%"):"\u2026";
  let txt;
  if(p.phase==="Pushing")txt=p.done+" / "+p.total+" matched students transferred to Portal";
  else if(p.phase==="Matching")txt="Scanning "+p.total+" portal students\u2026";
  else if(p.phase==="Loading Tracker data")txt="Loading Tracker data\u2026";
  else if(p.phase==="Saving")txt="Saving results\u2026";
  else txt=(p.phase||"Working")+"\u2026";
  const sub=document.getElementById("tm-sub");if(sub)sub.textContent=txt;
}
async function matchTransfers(btn){
  if(!confirm("Match live MVS Portal students to your already-checked Tracker data by Reference No, and push their status to the Portal WITHOUT using CapSolver?\\n\\nMatched = status pushed now + moved to Portal data (still re-checked normally). Unmatched = left for New Fetch."))return;
  const old=btn?btn.textContent:"";
  if(btn){btn.disabled=true;btn.textContent="Matching\u2026";}
  const box=document.getElementById("tr-match-result");
  if(box)box.innerHTML=_matchBarHTML();
  try{
    const start=await api("/api/transfer-match","POST");
    if(start&&start.started===false){showToast(start.message||"Already running");}
    let guard=0;
    while(guard++<4000){
      await new Promise(r=>setTimeout(r,800));
      let p;
      try{p=await api("/api/transfer-match-progress");}catch(e){continue;}
      _updMatchBar(p);
      if(p.finished){
        if(p.error){if(box)box.innerHTML='<div style="color:var(--danger);font-size:13px">Error: '+(""+p.error).replace(/</g,"&lt;")+'</div>';showToast("Error: "+p.error);}
        else if(p.result){showToast(p.result.message||"Done");renderMatchResult(p.result);loadTransfers(1);}
        break;
      }
      if(!p.running&&!p.finished&&guard>2){break;}
    }
  }catch(e){
    showToast(""+e.message);
    if(box)box.innerHTML='<div style="color:var(--danger);font-size:13px">Error: '+(""+e.message).replace(/</g,"&lt;")+'</div>';
  }
  finally{if(btn){btn.disabled=false;btn.textContent=old;}}
}
async function downloadTransfers(){
  const q=new URLSearchParams();
  const s=document.getElementById("tr-search");if(s&&s.value)q.set("search",s.value);
  const m=document.getElementById("tr-mode");if(m&&m.value)q.set("mode",m.value);
  try{
    const r=await fetch(API+"/api/transfers-download?"+q.toString(),{headers:{Authorization:"Bearer "+TOKEN}});
    if(!r.ok){showToast("Download failed");return;}
    const blob=await r.blob();const url=URL.createObjectURL(blob);
    const a=document.createElement("a");a.href=url;a.download="mvs_transfer_data.xlsx";a.click();
    URL.revokeObjectURL(url);
  }catch(e){showToast(""+e.message);}
}

const drop=document.getElementById("drop");
["dragover","dragenter"].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.add("drag");}));
["dragleave","drop"].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.remove("drag");}));
drop.addEventListener("drop",e=>{if(e.dataTransfer.files[0])handleFile(e.dataTransfer.files[0]);});
function handleFile(file){
  if(!file)return;
  const st=document.getElementById("upload-status");
  const sm=document.getElementById("upload-summary");
  const pv=document.getElementById("upload-preview");
  sm.innerHTML="";pv.innerHTML="";
  st.innerHTML='<div class="up-prog"><div class="up-prog-top"><span>Uploading…</span>'+
    '<span id="up-pct">0%</span></div><div class="up-track"><div class="up-fill" id="up-fill"></div></div>'+
    '<div class="up-eta" id="up-eta"></div></div>';
  const fd=new FormData();fd.append("file",file);
  const xhr=new XMLHttpRequest();
  const startT=Date.now();
  xhr.open("POST",API+"/api/upload-excel");
  xhr.setRequestHeader("Authorization","Bearer "+TOKEN);
  xhr.upload.onprogress=function(e){
    if(!e.lengthComputable)return;
    const pct=Math.round(e.loaded*100/e.total);
    const pe=document.getElementById("up-pct"),pf=document.getElementById("up-fill"),pet=document.getElementById("up-eta");
    if(pe)pe.textContent=pct+"%";
    if(pf)pf.style.width=pct+"%";
    const elapsed=(Date.now()-startT)/1000;
    const speed=e.loaded/Math.max(elapsed,0.001);
    const eta=speed>0?(e.total-e.loaded)/speed:0;
    if(pet)pet.textContent=pct<100?("~"+fmtEta(eta)+" remaining · "+fmtBytes(e.loaded)+" / "+fmtBytes(e.total)):"Processing file on server…";
  };
  xhr.onload=function(){
    let d={};try{d=JSON.parse(xhr.responseText);}catch(e){}
    if(xhr.status<200||xhr.status>=300){st.innerHTML='<div style="color:var(--danger)">'+(d.detail||"Upload failed")+'</div>';return;}
    st.innerHTML='<div style="color:var(--success);font-weight:600">'+(d.message||"Uploaded")+'</div>'+
      '<div style="margin-top:12px"><button class="btn btn-success btn-sm" onclick="runUploaded(this)">Run Check Now (uploaded only)</button>'+
      '<div style="font-size:11.5px;color:var(--muted);margin-top:7px">This checks <b>only the '+(d.total||0)+' student(s) in this sheet</b>. To run all data (MVS Tracker + MVS Portal), use <b>Run Now</b> on the Dashboard.</div></div>';
    if(d.parse_error){sm.innerHTML='<div style="color:var(--danger);font-size:13px">Preview error: '+d.parse_error+'</div>';}
    else{renderUploadSummary(d);renderUploadPreview(d.preview||[]);}
    showToast("Excel uploaded — "+(d.unique||0)+" new students");
  };
  xhr.onerror=function(){st.innerHTML='<div style="color:var(--danger)">Upload error — try again</div>';};
  xhr.send(fd);
}
const altDrop=document.getElementById("alt-drop");
if(altDrop){
  ["dragover","dragenter"].forEach(ev=>altDrop.addEventListener(ev,e=>{e.preventDefault();altDrop.classList.add("drag");}));
  ["dragleave","drop"].forEach(ev=>altDrop.addEventListener(ev,e=>{e.preventDefault();altDrop.classList.remove("drag");}));
  altDrop.addEventListener("drop",e=>{if(e.dataTransfer.files[0])handleAltFile(e.dataTransfer.files[0]);});
}
function handleAltFile(file){
  if(!file)return;
  const st=document.getElementById("alt-status");
  st.innerHTML='<div class="up-prog"><div class="up-prog-top"><span>Uploading alternate numbers…</span>'+
    '<span id="alt-pct">0%</span></div><div class="up-track"><div class="up-fill" id="alt-fill"></div></div>'+
    '<div class="up-eta" id="alt-eta"></div></div>';
  const fd=new FormData();fd.append("file",file);
  const xhr=new XMLHttpRequest();
  const startT=Date.now();
  xhr.open("POST",API+"/api/upload-alt-numbers");
  xhr.setRequestHeader("Authorization","Bearer "+TOKEN);
  xhr.upload.onprogress=function(e){
    if(!e.lengthComputable)return;
    const pct=Math.round(e.loaded*100/e.total);
    const pe=document.getElementById("alt-pct"),pf=document.getElementById("alt-fill"),pet=document.getElementById("alt-eta");
    if(pe)pe.textContent=pct+"%";
    if(pf)pf.style.width=pct+"%";
    const elapsed=(Date.now()-startT)/1000;
    const speed=e.loaded/Math.max(elapsed,0.001);
    const eta=speed>0?(e.total-e.loaded)/speed:0;
    if(pet)pet.textContent=pct<100?("~"+fmtEta(eta)+" remaining · "+fmtBytes(e.loaded)+" / "+fmtBytes(e.total)):"Matching references & updating numbers on the server…";
  };
  xhr.onload=function(){
    let d={};try{d=JSON.parse(xhr.responseText);}catch(e){}
    if(xhr.status<200||xhr.status>=300){st.innerHTML='<div style="color:var(--danger)">'+(d.detail||"Upload failed")+'</div>';return;}
    st.innerHTML='<div style="color:var(--success);font-weight:600">'+(d.message||"Done")+'</div>'+
      '<div class="up-stats" style="margin-top:12px">'+
        '<div class="up-stat new"><div class="us-lbl">Updated</div><div class="us-val">'+(d.updated||0)+'</div></div>'+
        '<div class="up-stat sim"><div class="us-lbl">Sent to alt (confirmed)</div><div class="us-val">'+(d.to_send||0)+'</div></div>'+
        '<div class="up-stat dup"><div class="us-lbl">Not found</div><div class="us-val">'+(d.not_found||0)+'</div></div>'+
      '</div>';
    showToast("Alternate numbers — updated "+(d.updated||0));
    const fi=document.getElementById("alt-file-input");if(fi)fi.value="";
  };
  xhr.onerror=function(){st.innerHTML='<div style="color:var(--danger)">Upload error — try again</div>';};
  xhr.send(fd);
}
async function downloadAltSample(){
  try{
    const r=await fetch(API+"/api/sample-alt-sheet",{headers:{Authorization:"Bearer "+TOKEN}});
    if(!r.ok){showToast("Sample download failed ("+r.status+")");return;}
    const blob=await r.blob();const url=URL.createObjectURL(blob);
    const a=document.createElement("a");a.href=url;a.download="MVS_alternate_numbers_sample.xlsx";
    document.body.appendChild(a);a.click();a.remove();
    setTimeout(()=>URL.revokeObjectURL(url),1500);
    showToast("Sample sheet downloaded");
  }catch(e){showToast("Error: "+e.message);}
}
function fmtBytes(b){if(b<1024)return b+" B";if(b<1048576)return (b/1024).toFixed(1)+" KB";return (b/1048576).toFixed(1)+" MB";}
function fmtEta(s){s=Math.ceil(s);if(s<1)return "0s";if(s<60)return s+"s";const m=Math.floor(s/60),ss=s%60;return m+"m "+ss+"s";}
function renderUploadSummary(d){
  const sm=document.getElementById("upload-summary");
  const sim=d.similar||0;
  sm.innerHTML='<div class="up-stats">'+
    '<div class="up-stat"><div class="us-lbl">Total in list</div><div class="us-val">'+(d.total||0)+'</div></div>'+
    '<div class="up-stat new"><div class="us-lbl">New students</div><div class="us-val">'+(d.unique||0)+'</div></div>'+
    '<div class="up-stat dup"><div class="us-lbl">Already tracked</div><div class="us-val">'+(d.duplicates||0)+'</div></div>'+
    (sim>0?'<div class="up-stat sim"><div class="us-lbl">Similar (will run)</div><div class="us-val">'+sim+'</div></div>':'')+
    '</div>'+
    '<div style="font-size:12px;color:var(--muted);margin-top:8px;line-height:1.6">'+
      '<b>'+(d.unique||0)+'</b> new student(s) will be added. '+
      ((d.duplicates||0)>0?'<b>'+d.duplicates+'</b> true duplicate(s) (<b>same reference number</b>) — merged & re-checked, <b>not</b> removed. ':'')+
      (sim>0?'<b>'+sim+'</b> share a phone/email with another record but have a <b>different reference</b> — so they are separate students (e.g. siblings) and <b>WILL run</b>. ':'')+
      'Nothing is deleted on upload.</div>';
}
function renderUploadPreview(rows){
  const pv=document.getElementById("upload-preview");
  if(!rows.length){pv.innerHTML="";return;}
  const head='<tr><th>#</th><th>Name</th><th>Reference</th><th>Mobile</th><th>Session</th><th>Status / Matches</th></tr>';
  const body=rows.map((s,i)=>{
    var matchLine="";
    if((s.dup||s.similar)&&(s.match_name||s.match_ref)){
      matchLine='<div style="font-size:11px;color:var(--muted);margin-top:3px">matches: <b>'+(s.match_name||"—")+'</b>'+
        (s.match_ref?' &middot; Ref <span style="font-family:monospace">'+s.match_ref+'</span>':'')+'</div>';
    }
    var tag;
    if(s.dup){
      tag='<span class="dup-tag" title="Same reference number as an existing student — merged & re-checked (not removed).">Duplicate &middot; '+(s.dup_basis||"")+'</span>';
    }else if(s.similar){
      tag='<span class="sim-tag" title="Different reference but same '+s.similar+'. Treated as a SEPARATE student and WILL run. Check the match to confirm.">Similar &middot; '+s.similar+'</span>';
    }else{
      tag='<span class="new-tag">New</span>';
    }
    return '<tr'+(s.dup?' style="background:var(--dup-bg)"':(s.similar?' style="background:#FEF9E7"':''))+'>'+
    '<td style="color:var(--muted)">'+(i+1)+'</td>'+
    '<td>'+(s.student_name||"—")+'</td>'+
    '<td><span class="ref-tag">'+(s.reference_no||s.enrollment_no||"—")+'</span></td>'+
    '<td style="font-size:12px">'+(s.mobile||"—")+'</td>'+
    '<td style="font-size:12px">'+(s.session||"—")+'</td>'+
    '<td>'+tag+matchLine+'</td></tr>';
  }).join("");
  pv.innerHTML='<div class="prev-head">Preview — '+rows.length+' rows (scroll to see all)</div>'+
    '<div style="font-size:11.5px;color:var(--muted);margin:2px 0 8px;line-height:1.5"><b>Similar</b> rows share a phone/email with another student but have a <b>different reference number</b>, so they are <b>NOT</b> duplicates — they run as separate students (e.g. siblings). Check &ldquo;matches&rdquo; to confirm. A true <b>Duplicate</b> has the <b>same reference number</b>. To drop a real duplicate, delete it from the Students list after upload (select &rarr; Delete).</div>'+
    '<div class="prev-box"><table>'+head+body+'</table></div>';
}
async function downloadExcel(){
  const r=await fetch(API+"/api/download-excel",{headers:{"Authorization":"Bearer "+TOKEN}});
  if(!r.ok){showToast("No Excel found. Run a check first.");return;}
  const blob=await r.blob();const url=URL.createObjectURL(blob);
  const a=document.createElement("a");a.href=url;a.download="nios_status_updated.xlsx";a.click();
  URL.revokeObjectURL(url);
}
async function saveUploadSource(){
  try{await api("/api/source-override?value="+encodeURIComponent(fval("upload-source")),"POST");
    showToast("Data source set for next run");}catch(e){showToast(""+e.message);}
}
async function loadUploadSource(){
  try{const d=await api("/api/source-override");const sel=document.getElementById("upload-source");if(sel)sel.value=d.value||"";}catch(e){}
}
async function loadSourceCounts(){
  try{
    const d=await api("/api/source-counts");
    const el=document.getElementById("src-counts");
    if(el)el.innerHTML='Current: <b>MVS Portal</b> &mdash; '+(d.mvs_portal||0)+' &middot; <b>MVS Tracker</b> &mdash; '+(d.mvs_tracker||0);
  }catch(e){}
}
function srcLabel(s){return s==="mvs_portal"?"MVS Portal":"MVS Tracker";}
async function moveAllSource(btn){
  const frm=fval("src-from"), to=fval("src-to");
  if(frm===to){showToast("Pick two different data types");return;}
  if(!confirm("Move ALL students from "+srcLabel(frm)+" to "+srcLabel(to)+"?\\n\\nOnly the data-type tag changes — no student is deleted or lost."))return;
  const st=document.getElementById("src-status");
  if(btn){btn.disabled=true;btn.textContent="Moving…";}
  try{
    const r=await api("/api/change-source-bulk","POST",{from_source:frm,to_source:to});
    if(st)st.innerHTML='<span style="color:var(--success)">Moved '+(r.moved||0)+' student(s) to '+srcLabel(to)+'.</span>';
    loadSourceCounts();
    refreshAllTables();
  }catch(e){if(st)st.innerHTML='<span style="color:var(--danger)">Error: '+e.message+'</span>';}
  finally{if(btn){btn.disabled=false;btn.textContent="Move all";}}
}
async function changeSelectedSource(sel){
  const to=sel.value; if(!to){return;}
  const keys=[...SELECTED];
  sel.value="";
  if(!keys.length){showToast("Select at least 1 student first");return;}
  if(!confirm("Move "+keys.length+" selected student(s) to "+srcLabel(to)+"?\\n\\nOnly the data-type tag changes — nothing is deleted."))return;
  try{
    const r=await api("/api/change-source-selected","POST",{row_keys:keys,to_source:to});
    showToast("Moved "+(r.moved||keys.length)+" student(s) to "+srcLabel(to));
    clearSelection();
    refreshAllTables();
    try{loadSourceCounts();}catch(e){}
  }catch(e){showToast("Error: "+e.message);}
}
async function downloadSample(type){
  try{
    const r=await fetch(API+"/api/sample-sheet?type="+type,{headers:{"Authorization":"Bearer "+TOKEN}});
    if(!r.ok){showToast("Sample download failed ("+r.status+")");return;}
    const blob=await r.blob();const url=URL.createObjectURL(blob);
    const a=document.createElement("a");a.href=url;
    a.download="MVS_sample_"+type+".xlsx";
    document.body.appendChild(a);a.click();a.remove();
    setTimeout(()=>URL.revokeObjectURL(url),1500);
    showToast("Sample sheet downloaded");
  }catch(e){showToast("Error: "+e.message);}
}
function fval(id){const e=document.getElementById(id);return e?e.value:"";}
function dateRange(prefix){
  const preset=fval(prefix+"-datepreset");
  const now=new Date();
  let from="",to="";
  if(preset==="today"){const s=new Date(now);s.setHours(0,0,0,0);from=fmtDT(s);to=fmtDT(now);}
  else if(preset==="yesterday"){const s=new Date(now);s.setDate(s.getDate()-1);s.setHours(0,0,0,0);
    const e=new Date(s);e.setHours(23,59,59,0);from=fmtDT(s);to=fmtDT(e);}
  else if(preset==="7d"){const s=new Date(now);s.setDate(s.getDate()-6);s.setHours(0,0,0,0);from=fmtDT(s);to=fmtDT(now);}
  else if(preset==="custom"){from=dtLocalToStr(fval(prefix+"-from"));to=dtLocalToStr(fval(prefix+"-to"));}
  return {from,to};
}
function onDatePreset(prefix,reload){
  const custom=fval(prefix+"-datepreset")==="custom";
  const row=document.getElementById(prefix+"-daterow");
  if(row)row.style.display=custom?"flex":"none";
  if(!custom&&reload)reload();
}
async function exportStudents(view){
  let search="",status="",session="",cls="",from="",to="",source="",toc="";
  if(view==="confirmed"){search=fval("c-search");status=fval("c-status");session=fval("c-session");cls=fval("c-class");source=fval("c-source");toc=fval("c-toc");const dr=dateRange("c");from=dr.from;to=dr.to;}
  else if(view==="required"){search=fval("r-search");status=fval("r-status");session=fval("r-session");source=fval("r-source");}
  else{search=fval("s-search");status=fval("s-status");session=fval("s-session");cls=fval("s-class");source=fval("s-source");toc=fval("s-toc");const dr=dateRange("s");from=dr.from;to=dr.to;}
  const q=new URLSearchParams({view:view,search:search,status_filter:status,session_filter:session,
    class_filter:cls,source_filter:source,toc_filter:toc,date_from:from,date_to:to});
  try{
    showToast("Preparing Excel...");
    const r=await fetch(API+"/api/export-students?"+q.toString(),{headers:{Authorization:"Bearer "+TOKEN}});
    if(!r.ok){showToast("Export failed");return;}
    const blob=await r.blob();const url=URL.createObjectURL(blob);
    const a=document.createElement("a");a.href=url;
    a.download="nios_"+(view==="normal"?"active":view)+"_students.xlsx";
    document.body.appendChild(a);a.click();a.remove();
    setTimeout(()=>URL.revokeObjectURL(url),1500);
    showToast("Excel downloaded");
  }catch(e){showToast("Error: "+e.message);}
}
function toggleRunMenu(e){
  if(e)e.stopPropagation();
  const m=document.getElementById("run-menu");
  if(!m)return;
  const open=(m.style.display==="none"||!m.style.display);
  m.style.display=open?"block":"none";
  const b=document.getElementById("run-now-btn");
  if(b)b.classList.toggle("open",open);
}
document.addEventListener("click",function(e){
  const w=e.target.closest&&e.target.closest(".run-menu-wrap");
  if(!w){const m=document.getElementById("run-menu");if(m)m.style.display="none";
    const b=document.getElementById("run-now-btn");if(b)b.classList.remove("open");}
});
const RUN_EP={all:"/api/run-now",tracker:"/api/run-now-tracker",portal:"/api/run-now-portal",portalnew:"/api/run-now-portal-new"};
const RUN_LBL={all:"all students",tracker:"MVS Tracker students",portal:"MVS Portal students",portalnew:"MVS Portal — new data only"};
const PORTAL_GRP_LBL={ondemand:"On Demand",stream2:"Stream 2",public:"Public",all:"all MVS Portal"};
async function runRequired(btn){
  if(btn&&btn.dataset.busy==="1")return;
  if(btn){btn.dataset.busy="1";btn.style.opacity="0.6";btn.style.pointerEvents="none";}
  try{
    const r=await api("/api/run-required","POST");
    if((r.count||0)===0){showToast("No Document Required students to run.");}
    else{
      showToast(r.message+" — running in the background");
      const box=document.getElementById("run-progress");
      if(box){box.style.display="block";document.getElementById("pb-pct").textContent="0%";
        document.getElementById("pb-fill").style.width="0%";
        document.getElementById("pb-sub").textContent="Starting…";
        document.getElementById("pb-label").textContent="Re-checking Document Required students…";}
      startProgressPoll();
    }
  }catch(e){showToast("Error: "+e.message);}
  finally{ setTimeout(()=>{if(btn){btn.dataset.busy="";btn.style.opacity="";btn.style.pointerEvents="";}},4000); }
}
async function runChoice(kind,group){
  const m=document.getElementById("run-menu");if(m)m.style.display="none";
  const btn=document.getElementById("run-now-btn");if(btn)btn.classList.remove("open");
  if(btn&&btn.dataset.busy==="1")return;
  if(btn){btn.dataset.busy="1";btn.style.opacity="0.6";btn.style.pointerEvents="none";}
  try{
    let ep=RUN_EP[kind]||RUN_EP.all;
    if((kind==="portal"||kind==="tracker")&&group)ep=RUN_EP[kind]+"?group="+encodeURIComponent(group);
    const r=await api(ep,"POST");
    showToast(r.message+" — running in the background");
    let lbl=RUN_LBL[kind]||"students";
    if((kind==="portal"||kind==="tracker")&&group){
      lbl=(kind==="portal"?"MVS Portal":"MVS Tracker")+" — "+(PORTAL_GRP_LBL[group]||group)+" students";
    }
    const box=document.getElementById("run-progress");
    if(box){box.style.display="block";document.getElementById("pb-pct").textContent="0%";
      document.getElementById("pb-fill").style.width="0%";
      document.getElementById("pb-sub").textContent="Starting…";
      document.getElementById("pb-label").textContent="Checking "+lbl+"…";}
    startProgressPoll();
  }catch(e){showToast("Error: "+e.message);}
  finally{ setTimeout(()=>{if(btn){btn.dataset.busy="";btn.style.opacity="";btn.style.pointerEvents="";}},4000); }
}
async function runNow(){return runChoice('all');}   // back-compat
async function runUploaded(btn){
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  try{
    const r=await api("/api/run-now-upload","POST");
    showToast(r.message+" — running in the background");
    // Clear the upload preview/summary now that the run has started.
    ["upload-status","upload-summary","upload-preview"].forEach(id=>{
      const el=document.getElementById(id);if(el)el.innerHTML="";});
    const fi=document.getElementById("file-input");if(fi)fi.value="";
    const box=document.getElementById("run-progress");
    if(box){box.style.display="block";document.getElementById("pb-pct").textContent="0%";
      document.getElementById("pb-fill").style.width="0%";
      document.getElementById("pb-sub").textContent="Starting…";
      document.getElementById("pb-label").textContent="Checking uploaded students…";}
    startProgressPoll();
    showToast("Upload preview cleared — run started");
  }catch(e){showToast("Error: "+e.message);}
  finally{ setTimeout(()=>{if(btn){btn.disabled=false;btn.style.opacity="";}},4000); }
}

function setIvField(which,mins){
  const inp=document.getElementById("iv-"+which),unit=document.getElementById("iv-"+which+"-unit");
  if(!inp||!unit)return;
  if(mins>=1440 && mins%1440===0){inp.value=mins/1440;unit.value="days";}
  else if(mins>=60 && mins%60===0){inp.value=mins/60;unit.value="hours";}
  else{inp.value=mins;unit.value="minutes";}
}
function ivToMin(which){
  const v=parseInt(document.getElementById("iv-"+which).value)||0;
  const unit=document.getElementById("iv-"+which+"-unit").value;
  if(unit==="days")return v*1440;
  if(unit==="hours")return v*60;
  return v;
}
async function loadIntervals(){
  try{const r=await api("/api/intervals");
    setIvField("ondemand",r.ondemand_min);
    setIvField("stream2",r.stream2_min);
    setIvField("public",r.public_min);
    setIvField("portalnew",r.portalnew_min);
    var sm={ondemand:r.ondemand_src,stream2:r.stream2_src,public:r.public_src};
    for(var g in sm){var el=document.getElementById("iv-"+g+"-src");if(el)el.value=sm[g]||"both";}
  }catch(e){}
}
async function loadTrash(){
  const b=document.getElementById("trash-body");if(!b)return;
  try{
    const rows=await api("/api/deleted-students");
    if(!rows.length){b.innerHTML='<tr><td colspan="6" class="empty">Trash is empty</td></tr>';selBar("trash");return;}
    b.innerHTML=rows.map(s=>{
      var nm=(s.student_name||"this student").replace(/[\\"']/g," ");
      var ref=(s.reference_no||s.enrollment_no||"—");
      return '<tr><td>'+selBox("trash",s.row_key)+'</td><td>'+(s.student_name||"—")+'<div style="margin-top:4px">'+srcBadge(s)+'</div></td>'+
        '<td><span class="ref-tag">'+ref+'</span></td>'+
        '<td style="font-size:13px">'+(s.session||"—")+'</td>'+
        '<td style="font-size:12px;color:var(--muted)">'+(s.deleted_at||"—")+'</td>'+
        '<td><div style="display:flex;gap:6px;flex-wrap:wrap">'+
        '<button onclick="restoreStudent(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" style="background:#dcfce7;color:#166534;border:1px solid #bbf7d0;border-radius:8px;padding:5px 10px;font-size:12px;font-weight:600;cursor:pointer">&#8617; Restore</button>'+
        '<button onclick="purgeStudent(&quot;'+s.row_key+'&quot;,&quot;'+nm+'&quot;)" style="background:#fee2e2;color:#b91c1c;border:1px solid #fecaca;border-radius:8px;padding:5px 10px;font-size:12px;font-weight:600;cursor:pointer">Delete permanently</button>'+
        '</div></td></tr>';
    }).join("");
    const sa=document.getElementById("trash-selall");if(sa)sa.checked=false;
    selBar("trash");
  }catch(e){b.innerHTML='<tr><td colspan="6" class="empty">'+e.message+'</td></tr>';}
}
async function restoreSelected(){
  const keys=Array.from(selSet("trash"));
  if(!keys.length)return;
  try{
    const r=await api("/api/students-restore-bulk","POST",{row_keys:keys});
    showToast("Restored "+(r.restored||keys.length)+" student(s)");
    selClear("trash");loadTrash();try{loadDashboard();}catch(e){}
  }catch(e){showToast("Error: "+e.message);}
}
async function purgeSelected(){
  const keys=Array.from(selSet("trash"));
  if(!keys.length)return;
  if(!confirm("Permanently delete "+keys.length+" student(s)?\\n\\nThis CANNOT be undone — they will be gone forever."))return;
  try{
    const r=await api("/api/students-purge-bulk","POST",{row_keys:keys});
    showToast("Permanently deleted "+(r.purged||keys.length)+" student(s)");
    selClear("trash");loadTrash();
  }catch(e){showToast("Error: "+e.message);}
}
async function restoreStudent(rk,name){
  try{await api("/api/student-restore?row_key="+encodeURIComponent(rk),"POST");
    showToast("Restored "+name);loadTrash();try{loadDashboard();}catch(e){}}
  catch(e){showToast(""+e.message);}
}
async function purgeStudent(rk,name){
  if(!confirm("Permanently delete "+name+"?\\n\\nThis CANNOT be undone."))return;
  try{await api("/api/student-purge?row_key="+encodeURIComponent(rk),"POST");
    showToast("Permanently deleted "+name);loadTrash();}
  catch(e){showToast(""+e.message);}
}
async function loadReportSettings(){
  try{
    const d=await api("/api/report-settings");
    const cb=document.getElementById("rep-enabled");if(cb)cb.checked=!!d.enabled;
    const ta=document.getElementById("rep-numbers");if(ta)ta.value=(d.numbers||"").split(",").filter(Boolean).join(String.fromCharCode(10));
    const last=document.getElementById("rep-last");
    if(last)last.innerHTML=d.last_status?('Last report: '+d.last_status):'';
    if(!d.campaign_set){const s=document.getElementById("rep-status");if(s)s.innerHTML='<span style="color:var(--warn)">Note: AISENSY_CAMPAIGN_REPORT env var not set on Railway — report will not send until it is added.</span>';}
  }catch(e){}
}
async function sendLatestReport(btn){
  const s=document.getElementById("rep-status");
  s.innerHTML='<span style="color:var(--muted)">Loading recent runs…</span>';
  let runs=[];
  try{ runs=await api("/api/recent-runs"); }catch(e){}
  let h='<div style="border:1px solid var(--border);border-radius:12px;padding:14px;margin-top:8px;background:var(--card)">';
  h+='<div style="font-weight:700;font-size:13.5px;margin-bottom:4px">Which report do you want to send to all management numbers?</div>';
  h+='<div style="color:var(--muted);font-size:12px;margin-bottom:10px">Pick a run below. (Automatic reports always send the latest run on their own.)</div>';
  h+='<div style="display:flex;flex-direction:column;gap:8px">';
  h+='<div style="display:flex;justify-content:space-between;align-items:center;gap:10px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:9px;padding:9px 11px">'
    +'<span style="font-size:12.5px"><b>Current live snapshot</b> <span style="color:var(--muted)">— totals right now</span></span>'
    +'<button class="btn btn-sm btn-primary" onclick="sendSnapshotReport(this)">Send</button></div>';
  (runs||[]).forEach(r=>{
    h+='<div style="display:flex;justify-content:space-between;align-items:center;gap:10px;border:1px solid var(--border);border-radius:9px;padding:9px 11px">'
      +'<span style="font-size:12.5px">'+(r.label||"Run")+'<br><span style="color:var(--muted);font-size:11.5px">checked '+(r.checked||0)+' &middot; confirmed '+(r.confirmed||0)+' &middot; required '+(r.required||0)+'</span></span>'
      +'<button class="btn btn-sm" style="background:#16A34A;color:#fff" onclick="sendPickedReport('+r.id+',this)">Send</button></div>';
  });
  if(!runs||!runs.length) h+='<div style="color:var(--muted);font-size:12px">No saved past runs yet — use Current snapshot. New runs will appear here automatically.</div>';
  h+='</div></div>';
  s.innerHTML=h;
}
async function _doReportSend(promise,btn){
  const s=document.getElementById("rep-status");
  if(btn){btn.disabled=true;btn.dataset.t=btn.textContent;btn.textContent="Sending…";}
  try{
    const r=await promise;
    s.innerHTML=r.sent?'<span style="color:var(--success)">Report sent to '+r.sent+'/'+r.total+' management number(s).</span>'
                      :'<span style="color:var(--danger)">Could not send to anyone.</span>';
    if(r.errors&&r.errors.length)s.innerHTML+='<div style="color:var(--danger);font-size:12px;margin-top:6px">'+r.errors.join("<br>")+'</div>';
    try{loadReportSettings();}catch(e){}
  }catch(e){s.innerHTML='<span style="color:var(--danger)">Error: '+e.message+'</span>';
    if(btn){btn.disabled=false;btn.textContent=btn.dataset.t||"Send";}}
}
function sendSnapshotReport(btn){ _doReportSend(api("/api/send-latest-report","POST"),btn); }
function sendPickedReport(runId,btn){ _doReportSend(api("/api/send-run-report","POST",{run_id:runId}),btn); }
async function saveReportSettings(){
  const enabled=document.getElementById("rep-enabled").checked;
  const numbers=document.getElementById("rep-numbers").value;
  const s=document.getElementById("rep-status");
  try{
    const r=await api("/api/report-settings","POST",{enabled:enabled,numbers:numbers});
    s.innerHTML='<span style="color:var(--success)">Saved — '+r.count+' number(s). '+(r.enabled?"Reports ON.":"Reports OFF.")+'</span>';
    const ta=document.getElementById("rep-numbers");if(ta)ta.value=(r.numbers||"").split(",").filter(Boolean).join(String.fromCharCode(10));
  }catch(e){s.innerHTML='<span style="color:var(--danger)">Error: '+e.message+'</span>';}
}
async function testReport(btn){
  const s=document.getElementById("rep-status");
  if(btn){btn.disabled=true;btn.style.opacity="0.6";}
  s.innerHTML='<span style="color:var(--muted)">Sending test report…</span>';
  try{
    const r=await api("/api/report-test","POST");
    s.innerHTML='<span style="color:var(--success)">Test sent to '+r.sent+'/'+r.total+' number(s).</span>'+
      (r.errors&&r.errors.length?'<div style="color:var(--danger);font-size:12px;margin-top:6px">'+r.errors.join("<br>")+'</div>':"");
  }catch(e){s.innerHTML='<span style="color:var(--danger)">Error: '+e.message+'</span>';}
  finally{if(btn){btn.disabled=false;btn.style.opacity="";}}
}
async function saveIntervals(){
  const od=ivToMin("ondemand"),st=ivToMin("stream2"),pm=ivToMin("public"),pn=ivToMin("portalnew");
  const S=document.getElementById("iv-status");
  if(od<15||st<15||pm<15||pn<15){S.innerHTML='<span style="color:var(--danger)">Minimum interval is 15 minutes</span>';return;}
  if(od>43200||st>43200||pm>43200||pn>43200){S.innerHTML='<span style="color:var(--danger)">Maximum interval is 30 days</span>';return;}
  var ivsrc=function(g){var el=document.getElementById("iv-"+g+"-src");return el?el.value:"both";};
  try{const r=await api("/api/intervals","POST",{ondemand_min:od,stream2_min:st,public_min:pm,portalnew_min:pn,
        ondemand_src:ivsrc("ondemand"),stream2_src:ivsrc("stream2"),public_src:ivsrc("public")});
    S.innerHTML='<span style="color:var(--success)">'+r.message+'</span>';
    showToast("Intervals saved!");try{loadNextRuns();}catch(e){}}
  catch(e){S.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}

async function loadWa(){
  let r=null;
  for(let attempt=0;attempt<2;attempt++){
    try{r=await api("/api/wa-settings");break;}
    catch(e){
      if(attempt>=1){const cfg=document.getElementById("wa-config");if(cfg)cfg.innerHTML='<span style="color:var(--danger)">&#10007; Could not load WhatsApp settings: '+(e.message||e)+'. <a href="#" onclick="loadWa();return false;" style="color:#2563eb;font-weight:600">Retry</a></span>';break;}
      await new Promise(s=>setTimeout(s,900));
    }
  }
  if(r){try{
    document.getElementById("wa-enabled").checked=r.enabled;
    var re=document.getElementById("wa-required-enabled");if(re)re.checked=!!r.required_enabled;
    const cfg=document.getElementById("wa-config");
    if(!r.configured){
      cfg.innerHTML='<span style="color:var(--danger)">&#10007; AISENSY_API_KEY is not set in Railway environment variables</span>';
    }else{
      const c=r.campaigns||{};const rc=r.required_campaigns||{};const ce=r.campaigns_env||{};
      const esc=(v)=>(v||"").replace(/"/g,"&quot;");
      const row=(lbl,v)=>'<div style="margin:2px 0">'+lbl+': '+
        (v?'<b style="color:var(--success)">'+v+'</b>':'<span style="color:var(--warn)">not set</span>')+'</div>';
      const camp=(lbl,key,v)=>'<div style="display:flex;align-items:center;gap:8px;margin:5px 0;flex-wrap:wrap">'+
        '<span style="width:96px;font-size:12.5px;color:var(--muted)">'+lbl+'</span>'+
        '<input class="wa-camp" data-g="'+key+'" value="'+esc(v)+'" placeholder="AiSensy campaign name" '+
        'style="flex:1;max-width:300px;padding:6px 9px;border:1px solid '+(v?'var(--border)':'#fca5a5')+';border-radius:7px;font-size:13px">'+
        '<span style="font-size:11px;font-weight:600;color:'+(ce[key]?'#15803d':'#b91c1c')+'">Railway: '+(ce[key]?'seen &#10003;':'empty &#10007;')+'</span>'+
        '</div>';
      cfg.innerHTML='<div style="color:var(--success);margin-bottom:8px">API key configured</div>'+
        '<div style="font-size:12px;color:var(--muted);margin-bottom:4px"><b>Confirmed-send campaigns</b> — type the AiSensy campaign name for each and Save (works even if Railway shows empty):</div>'+
        camp("On Demand","ondemand",c.ondemand)+camp("Stream 2","stream2",c.stream2)+
        camp("Public","public",c.public)+camp("SYC","syc",c.syc)+
        '<div style="font-size:12px;color:var(--muted);margin:8px 0 4px"><b>No-TOC campaigns</b> &middot; for confirmed students whose tocStatus is <b>no</b> (shorter document set):</div>'+
        camp("On Demand no-TOC","ondemand_notoc",c.ondemand_notoc)+camp("Stream 2 no-TOC","stream2_notoc",c.stream2_notoc)+
        '<div style="font-size:12px;color:var(--muted);margin:8px 0 4px"><b>Public yes-TOC</b> &middot; Public students whose tocStatus is <b>yes</b> (5 variables: name, reference, id card, registration summary, regional centre address):</div>'+
        camp("Public yes-TOC","public_yestoc",c.public_yestoc)+
        '<button class="btn btn-sm" style="background:var(--primary);color:#fff;margin-top:6px" onclick="saveCampaigns(this)">Save campaigns</button>'+
        '<div style="margin-top:10px;padding-top:8px;border-top:1px solid var(--border)"><b style="font-size:12px;color:var(--muted)">Document-Required reminder</b></div>'+
        row("Reminder &middot; main (On Demand/Stream 2)",rc.main)+row("Reminder &middot; public",rc.public);
    }
  }catch(e){}}
  try{const w=await api("/api/webhook-url");const el=document.getElementById("wh-url");if(el)el.value=w.url||"";}catch(e){}
}
function copyWebhook(btn){
  const el=document.getElementById("wh-url");if(!el)return;
  el.select();el.setSelectionRange(0,99999);
  try{navigator.clipboard.writeText(el.value);}catch(e){try{document.execCommand("copy");}catch(_){}}
  if(btn){const o=btn.textContent;btn.textContent="Copied!";setTimeout(()=>{btn.textContent=o;},1500);}
}
async function saveCampaigns(btn){
  const body={};
  document.querySelectorAll(".wa-camp").forEach(i=>{body[i.dataset.g]=i.value.trim();});
  if(btn)btn.disabled=true;
  try{await api("/api/wa-campaigns","POST",body);showToast("Campaigns saved \u2713");loadWa();}
  catch(e){showToast("Error: "+e.message);}
  finally{if(btn)btn.disabled=false;}
}
async function saveWa(){
  const en=document.getElementById("wa-enabled").checked;
  var reEl=document.getElementById("wa-required-enabled");
  const ren=reEl?reEl.checked:false;
  try{await api("/api/wa-settings","POST",{enabled:en,required_enabled:ren});
    showToast("WhatsApp settings saved");}
  catch(e){showToast("Error: "+e.message);}
}
async function testWa(){
  const num=document.getElementById("wa-num").value.trim();
  const group=document.getElementById("wa-group").value;
  const s=document.getElementById("wa-status");
  if(!num){s.innerHTML='<span style="color:var(--danger)">Please enter a number</span>';return;}
  s.innerHTML='Sending...';
  try{const r=await api("/api/wa-test","POST",{number:num,group:group});
    if(r.ok)s.innerHTML='<span style="color:var(--success)">Test message sent. Please check your WhatsApp.</span>';
    else s.innerHTML='<span style="color:var(--danger)">&#10007; '+(r.info||"failed")+'</span>';}
  catch(e){s.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}

function showToast(msg){
  const t=document.getElementById("toast");t.textContent=msg;t.classList.add("show");
  setTimeout(()=>t.classList.remove("show"),3500);
}

async function testLogin(){
  const ref=document.getElementById("dbg-ref").value.trim();
  const dob=document.getElementById("dbg-dob").value.trim();
  if(!ref||!dob){showToast("Please enter both Reference No and DOB");return;}
  const st=document.getElementById("dbg-status");
  const pre=document.getElementById("dbg-result");
  st.innerHTML='<span style="color:var(--muted)">Logging in (solving captcha, ~15 sec)...</span>';
  pre.style.display="none";
  try{
    const q=new URLSearchParams({ref:ref,dob:dob});
    const d=await api("/api/debug-login?"+q.toString());
    if(d.error){st.innerHTML='<span style="color:var(--danger)">'+d.error+'</span>';return;}
    const ok=d.logged_in_guess;
    st.innerHTML=ok?'<span style="color:var(--success)">Login successful! See the links below.</span>'
      :'<span style="color:var(--warn)">Login may have failed (or the page is different). Details below.</span>';
    pre.style.display="block";
    pre.textContent=JSON.stringify(d,null,2);
  }catch(e){st.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}

async function inspectDoc(){
  const ref=document.getElementById("dbg-ref").value.trim();
  const dob=document.getElementById("dbg-dob").value.trim();
  const kind=document.getElementById("dbg-kind").value;
  if(!ref||!dob){showToast("Please enter both Reference No and DOB");return;}
  const st=document.getElementById("dbg-status");
  const pre=document.getElementById("dbg-result");
  st.innerHTML='<span style="color:var(--muted)">Inspecting the '+kind+' page (~15 sec)...</span>';
  pre.style.display="none";
  try{
    const q=new URLSearchParams({ref:ref,dob:dob,kind:kind});
    const d=await api("/api/debug-doc?"+q.toString());
    if(d.error){st.innerHTML='<span style="color:var(--danger)">'+d.error+'</span>';return;}
    st.innerHTML='<span style="color:var(--success)">Inspection complete — structure below</span>';
    pre.style.display="block";
    pre.textContent=JSON.stringify(d,null,2);
  }catch(e){st.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}

async function findAddr(){
  const ref=document.getElementById("dbg-ref").value.trim();
  const dob=document.getElementById("dbg-dob").value.trim();
  if(!ref||!dob){showToast("Please enter both Reference No and DOB");return;}
  const st=document.getElementById("dbg-status");
  const pre=document.getElementById("dbg-result");
  st.innerHTML='<span style="color:var(--muted)">Finding the Regional Centre address from the ID card (~15 sec)...</span>';
  pre.style.display="none";
  try{
    const q=new URLSearchParams({ref:ref,dob:dob});
    const d=await api("/api/debug-idcard?"+q.toString());
    if(d.error){st.innerHTML='<span style="color:var(--danger)">'+d.error+'</span>';return;}
    st.innerHTML='<span style="color:var(--success)">Extracted address: <b>'+(d.extracted_address||"(blank — please share the text below)")+'</b></span>';
    pre.style.display="block";
    pre.textContent=JSON.stringify(d,null,2);
  }catch(e){st.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}

async function resetData(){
  if(!confirm("Are you sure? All students, statuses, history and the sheet will be DELETED. This cannot be undone."))return;
  if(!confirm("Final confirmation — do you really want to clear all data?"))return;
  const s=document.getElementById("reset-status");
  s.innerHTML="Clearing...";
  try{const r=await api("/api/reset-data","POST",{});
    s.innerHTML='<span style="color:var(--success)">'+r.message+' You can now upload a new sheet.</span>';
    showToast("All data cleared");}
  catch(e){s.innerHTML='<span style="color:var(--danger)">'+e.message+'</span>';}
}
</script>
</body>
</html>
"""

def create_token(username):
    expire = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return jwt.encode({"sub": username, "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(creds: HTTPAuthorizationCredentials = Depends(bearer)):
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("sub") != PORTAL_USER:
            raise HTTPException(status_code=401, detail="Invalid token")
        return payload["sub"]
    except JWTError:
        raise HTTPException(status_code=401, detail="Token expired or invalid")

scheduler = BackgroundScheduler()

def _last_run_time(group_type):
    """Most recent completed run for this group (or a manual 'All' run)."""
    pat = {"ondemand": "%On Demand%", "stream2": "%Stream 2%",
           "public": "%Public%"}.get(group_type, f"%{group_type}%")
    conn = get_db()
    row = conn.execute(
        "SELECT run_at FROM run_logs WHERE (group_type LIKE ? OR group_type IN ('All','all')) "
        "AND status LIKE 'completed%' ORDER BY id DESC LIMIT 1", (pat,)).fetchone()
    conn.close()
    if row and row["run_at"]:
        try:
            return datetime.strptime(row["run_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None
    return None

def _interval_minutes(grp, default_h):
    """Interval in MINUTES. Prefers the group's own *_min setting; On Demand & Stream 2
    fall back to the old combined 'regular' interval (so existing setups keep working
    until they're set separately). Legacy hour setting (×60) honoured last."""
    m = get_setting(f"interval_{grp}_min", "")
    if m:
        try:
            return max(15, int(m))
        except Exception:
            pass
    if grp in ("ondemand", "stream2"):
        rm = get_setting("interval_regular_min", "")
        if rm:
            try:
                return max(15, int(rm))
            except Exception:
                pass
    try:
        return int(get_setting(f"interval_{grp}", str(default_h))) * 60
    except Exception:
        return default_h * 60

def _mvs_enabled():
    try:
        import mvs_sync
        return mvs_sync.enabled()
    except Exception:
        return False

# Run groups — each has its own interval, dashboard timer, pause/resume + manual run.
# (job id, default hours)
RUN_GROUPS = [("ondemand", "job_ondemand", 6), ("stream2", "job_stream2", 6),
              ("public", "job_public", 12)]

def _group_source(grp):
    """Per-group data-source preference for AUTO runs (set in Recheck Intervals):
    'both' (default) -> None = both sources run together; otherwise restrict the
    scheduled run to just MVS Tracker or just MVS Portal data. This is what lets a
    counsellor say e.g. 'auto-run Stream 2 from MVS Tracker only'."""
    v = get_setting(f"runsrc_{grp}", "both")
    return v if v in ("mvs_tracker", "mvs_portal") else None

def reschedule_jobs():
    """One interval job per group (On Demand / Stream 2 / Public). Both data sources
    (MVS Portal + MVS Tracker) run together; the split is by session group only.
    Automatic runs happen ONLY at the set interval — never a 'catch-up' burst after a
    restart/redeploy/upload (next run is always at least one full interval away)."""
    now = datetime.now()
    for old in ("job_mvs", "job_regular"):     # drop legacy combined jobs
        try:
            scheduler.remove_job(old)
        except Exception:
            pass
    for grp, jid, dh in RUN_GROUPS:
        mins = _interval_minutes(grp, dh)
        # If THIS group is paused, keep it off (across restarts too) and skip scheduling.
        if get_setting(f"paused_{grp}", "") == "1":
            try:
                scheduler.remove_job(jid)
            except Exception:
                pass
            logger.info(f"{jid}: PAUSED — not scheduled.")
            continue
        last = _last_run_time(grp)
        nxt = last + timedelta(minutes=mins) if last else now + timedelta(minutes=mins)
        if nxt <= now:
            nxt = now + timedelta(minutes=mins)
        scheduler.add_job(lambda g=grp: run_status_check(g, _group_source(g), is_auto=True),
                          trigger=IntervalTrigger(minutes=mins),
                          id=jid, replace_existing=True, next_run_time=nxt)
        logger.info(f"{jid}: every {mins}min | last_run={last} | next_run={nxt}")

    # MVS Portal — NEW data only: a cheap, frequent run that checks just the students
    # newly arrived on the portal (skips already-tracked/active data) to save CapSolver.
    if get_setting("paused_portalnew", "") == "1":
        try:
            scheduler.remove_job("job_portalnew")
        except Exception:
            pass
        logger.info("job_portalnew: PAUSED — not scheduled.")
    else:
        pn_mins = _interval_minutes("portalnew", 3)
        pn_nxt = now + timedelta(minutes=pn_mins)
        scheduler.add_job(lambda: run_status_check("all", "mvs_portal", "new", is_auto=True),
                          trigger=IntervalTrigger(minutes=pn_mins),
                          id="job_portalnew", replace_existing=True, next_run_time=pn_nxt)
        logger.info(f"job_portalnew: every {pn_mins}min | next_run={pn_nxt}")

@app.on_event("startup")
async def startup():
    init_db()
    reschedule_jobs()
    # More worker threads for sync endpoints (default is 40). Students opening
    # documents + admin using the portal + background jobs all share this pool;
    # 100 keeps the portal responsive even when many students click at once.
    try:
        import anyio
        anyio.to_thread.current_default_thread_limiter().total_tokens = 100
    except Exception as e:
        logger.warning(f"Threadpool bump failed (using default): {e}")
    # Document auto-save now RESUMES across restarts/deploys: if "Save all documents
    # to DB" was active before the restart, the watchdog picks it up ~2 minutes after
    # boot (app serves links first) and continues where it left off — no need to press
    # the button again after every deploy. It also waits automatically while a run is
    # in progress, so it never competes for NIOS/CapSolver.
    # Auto-send WhatsApp documents to confirmed students that still haven't received
    # them — every 10 minutes, in small batches, so nothing stays "Not sent".
    try:
        from apscheduler.triggers.interval import IntervalTrigger as _IT
        scheduler.add_job(auto_send_pending_whatsapp, trigger=_IT(minutes=10),
                          id="job_wa_autosend", replace_existing=True,
                          next_run_time=datetime.now() + timedelta(minutes=2))
    except Exception as e:
        logger.warning(f"WhatsApp auto-send job not scheduled: {e}")
    # Auto-retry failed/unknown students — every 90 minutes, max 10 students per
    # sweep, each student max ~4 auto-retries/day, and it never overlaps a run.
    # Students whose data is fine but the check failed (captcha/NIOS hiccup) now
    # clear on their own instead of needing a manual "Run now".
    try:
        from job_runner import auto_retry_failed_sweep
        from apscheduler.triggers.interval import IntervalTrigger as _IT2
        scheduler.add_job(auto_retry_failed_sweep, trigger=_IT2(minutes=90),
                          id="job_auto_retry_failed", replace_existing=True,
                          next_run_time=datetime.now() + timedelta(minutes=20))
    except Exception as e:
        logger.warning(f"Auto-retry job not scheduled: {e}")
    # Portal resync — every 30 min re-push any confirmed/verified/in-progress student whose
    # status didn't reach the Portal (failed push / Portal briefly down), so the two dashboards
    # stay in agreement. Skips while a run is active; bounded per sweep.
    try:
        from job_runner import portal_resync_sweep
        from apscheduler.triggers.interval import IntervalTrigger as _IT3
        scheduler.add_job(portal_resync_sweep, trigger=_IT3(minutes=30),
                          id="job_portal_resync", replace_existing=True,
                          next_run_time=datetime.now() + timedelta(minutes=8))
    except Exception as e:
        logger.warning(f"Portal resync job not scheduled: {e}")
    scheduler.start()
    logger.info("Scheduler started")

@app.on_event("shutdown")
async def shutdown():
    scheduler.shutdown(wait=False)

@app.get("/", response_class=HTMLResponse)
async def serve_portal():
    return PORTAL_HTML

@app.get("/health")
async def health():
    return {"status": "ok", "captcha_key_set": bool(os.environ.get("CAPTCHA_API_KEY", ""))}

@app.post("/api/login")
async def login(body: dict):
    import asyncio
    u = str(body.get("username") or "")
    p = str(body.get("password") or "")
    # Constant-time comparison avoids leaking the password via response timing.
    ok = (_secrets.compare_digest(u, PORTAL_USER) and _secrets.compare_digest(p, PORTAL_PASS))
    if not ok:
        await asyncio.sleep(1.0)        # slow down brute-force guessing
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"token": create_token(u), "username": u}

def _dist_with_confirmed(dist_rows, confirmed_cnt):
    """Status distribution for the chart, but with 'Admission Confirmed' shown on the
    SAME definition as the card / sidebar / reports (the is_confirmed flag), so the
    dashboard never shows two different confirmed numbers. The raw-status confirmed rows
    that are actually login/check-failed are surfaced under 'Failed to Run' instead."""
    out = []
    for r in dist_rows:
        d = dict(r)
        if d.get("current_status") == "Admission Confirmed":
            d["cnt"] = confirmed_cnt
        out.append(d)
    return out

@app.get("/api/dashboard")
def dashboard(user=Depends(verify_token)):
    conn = get_db()
    ND = "COALESCE(deleted,0)=0"          # not-deleted guard for every count
    total = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND}").fetchone()[0]
    runs = conn.execute("SELECT * FROM run_logs ORDER BY id DESC LIMIT 10").fetchall()
    dist = conn.execute(f"SELECT current_status, COUNT(*) as cnt FROM student_status WHERE {ND} GROUP BY current_status").fetchall()

    # Counts by status
    NF = "COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0"   # exclude failed-to-run
    def count_status(s):
        return conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status=?", (s,)).fetchone()[0]
    # "Confirmed" everywhere else (sidebar badge, session-wise card, exported reports)
    # means the is_confirmed FLAG, not the raw NIOS status text. Keep this dashboard
    # card on the SAME definition so the numbers never disagree across the screen.
    confirmed_cnt = conn.execute(
        f"SELECT COUNT(*) FROM student_status WHERE {ND} AND {NF} AND is_confirmed=1").fetchone()[0]
    verified_cnt  = count_status("Verified")
    docreq_cnt    = count_status("Document Required")
    docverif_cnt  = count_status("Documents Verification In Progress")
    syc_cnt = conn.execute(
        f"SELECT COUNT(*) FROM student_status WHERE {ND} AND (session LIKE '%syc%' OR current_status='SYC')"
    ).fetchone()[0]

    # Changes today
    today = datetime.now().strftime("%Y-%m-%d")
    changes_today = conn.execute(
        "SELECT COUNT(*) FROM status_history WHERE changed_at LIKE ?", (f"{today}%",)).fetchone()[0]

    # Notifications: document required students (name + ref)
    notifs = conn.execute(
        f"SELECT student_name, reference_no, remark FROM student_status WHERE {ND} AND current_status='Document Required' ORDER BY last_changed DESC LIMIT 50"
    ).fetchall()

    # "Both" (cross-source) system removed: a student now belongs to ONE data source
    # (transferred students are Portal). No cross-source duplicate notification is shown.
    dup_notifs = []

    # Failed-to-run students: NIOS login failed OR status check failed -> need edit/re-run
    login_issues = conn.execute(
        f"SELECT student_name, reference_no, enrollment_no, login_remark, row_key FROM student_status "
        f"WHERE {ND} AND (COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1) "
        f"ORDER BY student_name LIMIT 100"
    ).fetchall()

    # Status not found (Unknown / Fetch Error): NIOS didn't return a readable status —
    # likely an error or wrong reference. Surfaced in the bell with a remark so the
    # counsellor can re-check, since these are easy to miss in Active Students.
    unknown_notifs = conn.execute(
        f"SELECT student_name, reference_no, enrollment_no, current_status, "
        f"COALESCE(NULLIF(login_remark,''), NULLIF(remark,''), '') as remark "
        f"FROM student_status WHERE {ND} AND current_status IN ('Unknown','Fetch Error') "
        f"ORDER BY last_changed DESC LIMIT 100"
    ).fetchall()

    # Session-wise totals WITH a per-session status breakdown (same definitions the
    # nav tabs use), so a counsellor can see, per session, how many are still pending.
    NFAIL = "COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0"
    sess_counts = conn.execute(
        f"SELECT session, COUNT(*) as cnt, "
        f"SUM(CASE WHEN is_confirmed=1 THEN 1 ELSE 0 END) as confirmed, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND current_status='Verified' THEN 1 ELSE 0 END) as verified, "
        f"SUM(CASE WHEN current_status='Document Required' THEN 1 ELSE 0 END) as required, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND COALESCE(current_status,'')!='SYC' "
        f"         AND (session IS NULL OR session NOT LIKE '%syc%') THEN 1 ELSE 0 END) as active "
        f"FROM student_status WHERE {ND} AND {NFAIL} AND session != '' GROUP BY session ORDER BY cnt DESC"
    ).fetchall()

    # Three dashboard cards, using the flags we have:
    #   • "Enrol. MVS Portal" = came from the Portal enrollment form
    #     -> source='mvs_portal' AND cross_dup=0 AND portal_origin != 'sheet'
    #   • "MVS Portal"        = the rest of the Portal data (sheet-uploaded on the Portal +
    #     students the tracker checked & pushed to the Portal)
    #     -> source='mvs_portal' AND (cross_dup=1 OR portal_origin='sheet')
    #   • "MVS Tracker"       = uploaded straight into the tracker, not (yet) on the Portal
    #     -> source='mvs_tracker'
    _bucket = ("CASE "
               "WHEN COALESCE(source,'mvs_tracker')='mvs_tracker' THEN 'mvs_tracker' "
               "WHEN COALESCE(portal_origin,'')='enrol' THEN 'enrol_portal' "
               "WHEN COALESCE(portal_origin,'')='sheet' THEN 'mvs_portal' "
               "WHEN COALESCE(cross_dup,0)=1 THEN 'mvs_portal' "
               "ELSE 'enrol_portal' END")
    src_counts = conn.execute(
        f"SELECT {_bucket} AS bucket, COUNT(*) as cnt, "
        f"SUM(CASE WHEN is_confirmed=1 THEN 1 ELSE 0 END) as confirmed, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND current_status='Verified' THEN 1 ELSE 0 END) as verified, "
        f"SUM(CASE WHEN current_status='Document Required' THEN 1 ELSE 0 END) as required, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND COALESCE(current_status,'')!='SYC' "
        f"         AND (session IS NULL OR session NOT LIKE '%syc%') THEN 1 ELSE 0 END) as active "
        f"FROM student_status WHERE {ND} AND {NFAIL} GROUP BY bucket"
    ).fetchall()

    # How many students have been transferred Tracker -> Portal (distinct, no duplicates) —
    # surfaced as a small card on the dashboard so the count is visible at a glance.
    try:
        transfer_count = conn.execute("SELECT COUNT(DISTINCT row_key) FROM transfer_log").fetchone()[0] or 0
    except Exception:
        transfer_count = 0

    conn.close()
    _nexts = [j.next_run_time for j in (scheduler.get_job(jid) for _g, jid, _d in RUN_GROUPS)
              if j and j.next_run_time]
    next_run = str(min(_nexts)) if _nexts else "Not scheduled"
    return {
        "total_students": total, "next_run": next_run,
        "status_distribution": _dist_with_confirmed(dist, confirmed_cnt),
        "recent_runs": [dict(r) for r in runs],
        "counts": {
            "confirmed": confirmed_cnt, "verified": verified_cnt,
            "document_required": docreq_cnt, "doc_verification": docverif_cnt,
            "changes_today": changes_today, "syc": syc_cnt,
        },
        "notifications": [dict(n) for n in notifs],
        "dup_notifications": [dict(d) for d in dup_notifs],
        "unknown_notifications": [dict(u) for u in unknown_notifs],
        "login_issues": [dict(li) for li in login_issues],
        "session_counts": _normalized_session_counts(sess_counts),
        "source_counts": _normalized_source_counts(src_counts),
        "transfer_count": transfer_count,
    }

def _normalized_source_counts(raw_rows):
    """Three dashboard cards:
      • 'Enrol. MVS Portal' — students from the Portal enrollment form.
      • 'MVS Portal'        — the rest of the Portal data (sheet uploads on the Portal + transfers).
      • 'MVS Tracker'       — uploaded straight into the tracker.
    Order: Enrol, MVS Portal, MVS Tracker. The dashboard's Combine button merges Enrol + MVS
    Portal in place (client-side), so no separate combined card is returned here."""
    label = {"enrol_portal": "Enrol. MVS Portal", "mvs_portal": "MVS Portal",
             "mvs_tracker": "MVS Tracker"}
    order = {"enrol_portal": 0, "mvs_portal": 1, "mvs_tracker": 2}
    out = []
    for r in raw_rows:
        b = (r["bucket"] if "bucket" in r.keys() else "") or "mvs_tracker"
        out.append({"source": label.get(b, b), "key": b,
                    "cnt": r["cnt"] or 0, "confirmed": r["confirmed"] or 0,
                    "verified": r["verified"] or 0, "active": r["active"] or 0,
                    "required": r["required"] or 0})
    out.sort(key=lambda x: order.get(x["key"], 9))
    return out

def _normalized_session_counts(raw_rows):
    """Collapse raw session rows into the normalized categories (On Demand, Stream 2,
    Public, SYC) so the dashboard shows ONE row per real category, each with a
    per-session breakdown (total / confirmed / verified / active / required)."""
    agg = {}
    for r in raw_rows:
        norm = normalize_session(r["session"])
        if not norm:
            continue
        d = agg.setdefault(norm, {"session": norm, "cnt": 0, "confirmed": 0,
                                  "verified": 0, "active": 0, "required": 0})
        d["cnt"] += (r["cnt"] or 0)
        for k in ("confirmed", "verified", "active", "required"):
            d[k] += (r[k] or 0) if k in r.keys() else 0
    out = list(agg.values())
    out.sort(key=lambda x: x["cnt"], reverse=True)
    return out

def normalize_session(s):
    """Merge raw session text into the real filter categories so there is ONE entry
    per category, no matter how it's typed. Spelling tolerance lives in the shared
    excel_handler.session_category() (used by the run grouping and document rules too),
    so 'str-2'/'STREAM 2', 'ode'/'ODE', 'apr-27'/'APRIL 2027' all land in the right
    bucket. April / October / abbreviations / unknown all collapse to 'Public'."""
    from excel_handler import session_category
    return {"": "", "syc": "SYC", "ondemand": "On Demand", "stream2": "Stream 2",
            "stream1": "Stream 1", "public": "Public"}.get(session_category(s), "Public")

def _session_clause(cat):
    """SQL clause + params matching ALL raw variants of a normalized session category,
    so selecting 'On Demand' catches 'On Demand1' etc. 'Public' is defined as EVERYTHING
    that is not On Demand / Stream 2 / Stream 1 / SYC — so it catches every April / October
    spelling AND abbreviations like 'apr-27' / 'oct-26', exactly like the run grouping."""
    t = (cat or "").strip().lower()
    pats = {
        "on demand": ["%on demand%", "%ondemand%", "%on-demand%", "%on_demand%", "%odes%", "%ode%"],
        "stream 2":  ["%stream 2%", "%stream2%", "%stream-2%", "%stream ii%", "%stream_2%",
                      "%str-2%", "%str2%", "%str 2%"],
        "stream 1":  ["%stream 1%", "%stream1%", "%stream-1%", "%str_1%",
                      "%str-1%", "%str1%", "%str 1%"],
        "april":     ["%april%", "%apr-%", "%apr %"],
        "october":   ["%october%", "%oct-%", "%oct %"],
        "syc":       ["%syc%"],
    }
    if t == "public":
        # Everything that is NOT On Demand / Stream 2 / Stream 1 / SYC.
        neg = (pats["on demand"] + pats["stream 2"] + pats["stream 1"] + pats["syc"])
        clause = ("(TRIM(COALESCE(session,'')) != '' AND "
                  + " AND ".join(["LOWER(session) NOT LIKE ?"] * len(neg)) + ")")
        return clause, neg
    if t in pats:
        p = pats[t]
        return "(" + " OR ".join(["LOWER(session) LIKE ?"] * len(p)) + ")", p
    return "session = ?", [cat]

# Sessions that are On Demand or Stream 2 (everything else, except SYC, is "public").
_OD_S2_SQL = ("(LOWER(COALESCE(session,'')) LIKE '%on demand%' OR LOWER(COALESCE(session,'')) LIKE '%ondemand%' "
              "OR LOWER(COALESCE(session,'')) LIKE '%on-demand%' OR LOWER(COALESCE(session,'')) LIKE '%odes%' "
              "OR LOWER(COALESCE(session,'')) LIKE '%stream 2%' OR LOWER(COALESCE(session,'')) LIKE '%stream2%' "
              "OR LOWER(COALESCE(session,'')) LIKE '%stream-2%')")
_IS_SYC_SQL = "LOWER(COALESCE(session,'')) LIKE '%syc%'"
# Students who need SPECIAL document handling (the "No-TOC" page / run / resend):
#   - On Demand / Stream 2 with tocStatus = 'no'   (shorter doc set)
#   - Public with tocStatus = 'yes'                (gets the Application Form too)
# Public 'no' is the normal Public case, so it is intentionally excluded.
_SPECIAL_TOC_CLAUSE = (
    "( ( " + _OD_S2_SQL + " AND LOWER(COALESCE(toc_status,''))='no' ) "
    "OR ( NOT " + _OD_S2_SQL + " AND NOT " + _IS_SYC_SQL + " AND LOWER(COALESCE(toc_status,''))='yes' ) )"
)

def _fully_saved_keys():
    """Return the set of confirmed row_keys whose documents are FULLY saved in our DB — using the
    exact same per-session/TOC expected-doc rule as the 'Fully saved' header count, so the
    'Documents saved' / 'Not fully saved' filters always agree with that number. Only counts
    students that actually appear on the Confirmed page (i.e. NOT login-failed / check-failed —
    those live under 'Failed to Run'), so the filter and the header never disagree. Returns None
    only if it truly can't be computed (then the caller leaves the filter as a no-op)."""
    try:
        import whatsapp
        conn = get_db()
        confirmed = conn.execute(
            "SELECT row_key, session, toc_status FROM student_status "
            "WHERE is_confirmed=1 AND COALESCE(deleted,0)=0 "
            "AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0").fetchall()
        have = {}
        for x in conn.execute("SELECT row_key, kind FROM document_cache").fetchall():
            have.setdefault(x["row_key"], set()).add(x["kind"])
        conn.close()
        full = set()
        for s in confirmed:
            allowed = whatsapp.allowed_docs(s["session"], (s["toc_status"] or ""))
            if not allowed or allowed.issubset(have.get(s["row_key"], set())):
                full.add(s["row_key"])
        return full
    except Exception as e:
        logger.warning(f"_fully_saved_keys error: {e}")
        return None


def _build_student_where(view, search, status_filter, session_filter,
                         class_filter="", date_from="", date_to="", source_filter="",
                         wa_status="", saved_filter="", check_state="", check_boundary=None,
                         toc_filter=""):
    """Shared WHERE builder so the table and its Excel export stay perfectly in sync.
    NULL-safe so students with missing status/date are never silently hidden."""
    wc, params = [], []
    wc.append("COALESCE(deleted,0) = 0")          # never show soft-deleted (in Trash)
    wc.append("COALESCE(login_failed,0) = 0")     # login-failed -> only in 'Failed to Run'
    wc.append("COALESCE(check_failed,0) = 0")     # status-check-failed -> only in 'Failed to Run'
    # tocStatus filter: yes / no / blank (not set) / mismatch (unverified TOC error)
    tf = (toc_filter or "").strip().lower()
    if tf == "yes":
        wc.append("LOWER(COALESCE(toc_status,'')) = 'yes'")
    elif tf == "no":
        wc.append("LOWER(COALESCE(toc_status,'')) = 'no'")
    elif tf == "blank":
        wc.append("COALESCE(toc_status,'') = ''")
    elif tf == "mismatch":
        wc.append("COALESCE(toc_mismatch,0) = 1 AND COALESCE(toc_verified,0) = 0")
    if view == "confirmed":
        wc.append("is_confirmed = 1")
    elif view == "required":
        wc.append("current_status = 'Document Required'")
    elif view == "notoc":
        # Special-TOC page: On Demand/Stream 2 no-TOC + Public yes-TOC (any status), never SYC.
        wc.append(_SPECIAL_TOC_CLAUSE)
        wc.append("(session IS NULL OR session NOT LIKE '%syc%')")
    else:  # normal = active students, exclude confirmed and SYC (NULL-safe)
        wc.append("COALESCE(is_confirmed,0) = 0")
        wc.append("COALESCE(current_status,'') != 'SYC'")
        wc.append("(session IS NULL OR session NOT LIKE '%syc%')")
    if search:
        wc.append("(reference_no LIKE ? OR student_name LIKE ? OR email LIKE ?)")
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    if status_filter:
        wc.append("current_status = ?"); params.append(status_filter)
    if session_filter:
        clause, sp = _session_clause(session_filter)
        wc.append(clause); params += sp
    if source_filter:                      # mvs_portal | mvs_tracker | both (Data Type)
        if source_filter == "both":
            wc.append("COALESCE(cross_dup,0) = 1")   # exists in BOTH sources, merged
        else:
            wc.append("COALESCE(source,'mvs_tracker') = ?"); params.append(source_filter)
    if wa_status:                          # WhatsApp delivery filter (Confirmed view)
        if wa_status == "delivered":
            wc.append("whatsapp_delivery = 'delivered'")
        elif wa_status == "failed":
            wc.append("(whatsapp_delivery = 'failed' OR COALESCE(login_failed,0)=1)")
        elif wa_status == "sent":
            wc.append("COALESCE(whatsapp_sent,0)=1 AND COALESCE(whatsapp_delivery,'')=''")
        elif wa_status == "pending24":
            # Sent 24h+ ago and STILL no delivery confirmation — the ones most likely to have
            # silently not reached the student (like the June-29 case). Bulk-resend candidates.
            wc.append("COALESCE(whatsapp_sent,0)=1 AND COALESCE(whatsapp_delivery,'')='' "
                      "AND COALESCE(whatsapp_sent_at,'') != '' "
                      "AND whatsapp_sent_at <= datetime('now','localtime','-24 hours')")
        elif wa_status == "notsent":
            wc.append("COALESCE(whatsapp_sent,0)=0")
    if class_filter:                       # "10" matches 10/10TH, "12" matches 12/12TH
        wc.append("class_level LIKE ?"); params.append(f"{class_filter}%")
    # Date/time filter on when the status last changed; fall back to last_checked so a
    # student with an empty last_changed is never dropped. Accepts either a full
    # "YYYY-MM-DD HH:MM:SS" (custom range with time) or a plain "YYYY-MM-DD".
    def _dtval(v, end):
        v = (v or "").strip()
        if not v:
            return ""
        return v if len(v) > 10 else (v + (" 23:59:59" if end else " 00:00:00"))
    _dt = "COALESCE(NULLIF(last_changed,''), last_checked)"
    df, dtv = _dtval(date_from, False), _dtval(date_to, True)
    if df:
        wc.append(f"{_dt} >= ?"); params.append(df)
    if dtv:
        wc.append(f"{_dt} <= ?"); params.append(dtv)
    # Documents-saved filter. "Fully saved" = we have EVERY document this student's WhatsApp
    # message needs (id card / app form / hall ticket, per session + TOC). A student with 2 of 3
    # docs is NOT fully saved — they belong under "Not saved yet" so they're findable and fixable.
    if saved_filter in ("saved", "notsaved"):
        fully_keys = list(_fully_saved_keys() or [])
        key_list = ",".join("?" for _ in fully_keys)
        if saved_filter == "saved":
            if fully_keys:
                wc.append(f"student_status.row_key IN ({key_list})")
                params.extend(fully_keys)
            else:
                wc.append("1=0")   # nothing fully saved yet
        else:  # notsaved = NOT fully saved (missing at least one doc, or none)
            if fully_keys:
                wc.append(f"student_status.row_key NOT IN ({key_list})")
                params.extend(fully_keys)
            # else: no exclusions -> everyone is "not saved yet"
    # "New check / Not checked (this run)" — based on last_verified (only set when a run
    # actually READ a real NIOS status; a captcha/proxy fallback does NOT advance it). The
    # boundary is the most recent verify time WITHIN the current filter scope (session-aware),
    # minus a safety window that comfortably covers a single run but is well under the gap
    # between two auto-runs of the same session. So "new" = verified in the latest run batch;
    # "stale" = missed it (or never verified).
    if check_state in ("new", "stale") and check_boundary is not None:
        if check_state == "new":
            wc.append("(COALESCE(last_verified,'') != '' AND last_verified >= ?)")
            params.append(check_boundary)
        else:  # stale = not verified in the latest batch (or never)
            wc.append("(COALESCE(last_verified,'') = '' OR last_verified < ?)")
            params.append(check_boundary)
    return (("WHERE " + " AND ".join(wc)) if wc else ""), params

_CHECK_WINDOW_HOURS = 8   # a single run finishes well within this; auto-runs are >=21h apart

def _verify_boundary(conn, view, session_filter, source_filter):
    """Most-recent verify time within the current scope, minus a safety window. Students
    verified at/after this are 'new' (this run); older/never are 'not checked'. Session-aware
    so a run of one session doesn't mark another session's students as stale."""
    wc = ["COALESCE(deleted,0)=0", "COALESCE(last_verified,'') != ''"]
    params = []
    if view == "confirmed":
        wc.append("is_confirmed = 1")
    elif view == "required":
        wc.append("current_status = 'Document Required'")
    else:
        wc.append("COALESCE(is_confirmed,0) = 0")
        wc.append("COALESCE(current_status,'') != 'SYC'")
    if session_filter:
        clause, sp = _session_clause(session_filter)
        wc.append(clause); params += sp
    if source_filter and source_filter != "both":
        wc.append("COALESCE(source,'mvs_tracker') = ?"); params.append(source_filter)
    row = conn.execute(f"SELECT MAX(last_verified) AS m FROM student_status WHERE {' AND '.join(wc)}",
                       params).fetchone()
    if not row or not row["m"]:
        return None
    try:
        peak = datetime.strptime(row["m"], "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None
    return (peak - timedelta(hours=_CHECK_WINDOW_HOURS)).strftime("%Y-%m-%d %H:%M:%S")

def _row_check_state(last_verified, boundary):
    """'new' if verified in the latest batch, else 'stale'. Used for the per-row badge."""
    if not boundary:
        return "new" if (last_verified or "") else "stale"
    if last_verified and last_verified >= boundary:
        return "new"
    return "stale"

@app.get("/api/students")
def get_students(page: int=1, per_page: int=50, search: str="",
                       status_filter: str="", session_filter: str="",
                       class_filter: str="", date_from: str="", date_to: str="",
                       source_filter: str="", view: str="normal", wa_status: str="",
                       saved_filter: str="", check_state: str="", toc_filter: str="",
                       user=Depends(verify_token)):
    conn = get_db()
    offset = (page - 1) * per_page
    boundary = _verify_boundary(conn, view, session_filter, source_filter)
    where, params = _build_student_where(view, search, status_filter, session_filter,
                                         class_filter, date_from, date_to, source_filter,
                                         wa_status, saved_filter, check_state, boundary,
                                         toc_filter)
    total = conn.execute(f"SELECT COUNT(*) FROM student_status {where}", params).fetchone()[0]
    students = conn.execute(
        f"SELECT * FROM student_status {where} ORDER BY student_name LIMIT ? OFFSET ?",
        params + [per_page, offset]).fetchall()
    raw_sessions = conn.execute("SELECT DISTINCT session FROM student_status WHERE session != ''").fetchall()
    norm_sessions = sorted({normalize_session(r["session"]) for r in raw_sessions})
    norm_sessions = [x for x in norm_sessions if x and x != "SYC"]   # SYC has its own page
    slist = [dict(s) for s in students]
    # attach which documents we have SAVED in our DB for each student (for the "saved" badge)
    rks = [s.get("row_key") for s in slist if s.get("row_key")]
    if rks:
        ph = ",".join("?" * len(rks))
        saved = {}
        for cr in conn.execute(f"SELECT row_key, kind FROM document_cache WHERE row_key IN ({ph})", rks).fetchall():
            saved.setdefault(cr["row_key"], []).append(cr["kind"])
        for s in slist:
            s["cached_docs"] = saved.get(s.get("row_key"), [])
    for s in slist:
        s["check_state"] = _row_check_state(s.get("last_verified"), boundary)
    conn.close()
    return {"students": slist, "total": total, "page": page,
            "per_page": per_page, "pages": max(1, (total+per_page-1)//per_page),
            "sessions": norm_sessions}

@app.get("/api/failed-students")
def failed_students(page: int = 1, per_page: int = 50, search: str = "",
                          source: str = "", user=Depends(verify_token)):
    """Students whose NIOS login failed (wrong data) — the 'Failed to Run' list."""
    conn = get_db()
    wc = ["COALESCE(deleted,0)=0", "(COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1)",
          "COALESCE(current_status,'')!='Unknown'"]
    params = []
    if search:
        like = f"%{search.strip()}%"
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR email LIKE ? OR enrollment_no LIKE ?)")
        params += [like, like, like, like]
    if source:
        if source == "both":
            wc.append("COALESCE(cross_dup,0)=1")
        else:
            wc.append("COALESCE(source,'mvs_tracker')=?"); params.append(source)
    where = "WHERE " + " AND ".join(wc)
    total = conn.execute(f"SELECT COUNT(*) FROM student_status {where}", params).fetchone()[0]
    per_page = max(1, min(per_page, 200)); page = max(1, page)
    offset = (page - 1) * per_page
    rows = conn.execute(f"SELECT * FROM student_status {where} ORDER BY last_checked DESC LIMIT ? OFFSET ?",
                        params + [per_page, offset]).fetchall()
    conn.close()
    return {"students": [dict(r) for r in rows], "total": total, "page": page,
            "per_page": per_page, "pages": max(1, (total + per_page - 1) // per_page)}

@app.get("/api/unknown-students")
def unknown_students(page: int = 1, per_page: int = 50, search: str = "",
                           session_filter: str = "", user=Depends(verify_token)):
    """Students stuck at the 'Unknown' NIOS status — their own tab so they're easy to
    find, filter (by session) and re-run. Also returns the session list for the dropdown."""
    conn = get_db()
    wc = ["COALESCE(deleted,0)=0", "current_status='Unknown'"]
    params = []
    if search:
        like = f"%{search.strip()}%"
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR email LIKE ? OR enrollment_no LIKE ?)")
        params += [like, like, like, like]
    if session_filter:
        clause, sp = _session_clause(session_filter)
        wc.append(clause); params += sp
    where = "WHERE " + " AND ".join(wc)
    total = conn.execute(f"SELECT COUNT(*) FROM student_status {where}", params).fetchone()[0]
    per_page = max(1, min(per_page, 200)); page = max(1, page)
    offset = (page - 1) * per_page
    rows = conn.execute(f"SELECT * FROM student_status {where} ORDER BY last_checked DESC LIMIT ? OFFSET ?",
                        params + [per_page, offset]).fetchall()
    raw_sessions = conn.execute("SELECT DISTINCT session FROM student_status "
                                "WHERE COALESCE(deleted,0)=0 AND current_status='Unknown'").fetchall()
    norm_sessions = sorted({normalize_session(r["session"]) for r in raw_sessions})
    norm_sessions = [x for x in norm_sessions if x and x != "SYC"]
    conn.close()
    return {"students": [dict(r) for r in rows], "total": total, "page": page,
            "per_page": per_page, "pages": max(1, (total + per_page - 1) // per_page),
            "sessions": norm_sessions}

@app.post("/api/sync-toc")
async def sync_toc(user=Depends(verify_token)):
    """Pull the latest tocStatus from the Portal and update it on the matching Tracker students
    WITHOUT any CapSolver / NIOS status check — only the toc_status field is touched, so NO credits
    are used. After this runs, no-TOC students show on the No-TOC page and resend works."""
    import mvs_sync
    from fastapi.concurrency import run_in_threadpool
    if not mvs_sync.enabled():
        return {"ok": False, "message": "MVS Portal is not configured."}
    try:
        portal = await run_in_threadpool(mvs_sync.fetch_students_for_tracker, None, True)
    except Exception as e:
        return {"ok": False, "message": f"Could not reach Portal: {e}"}
    if not portal:
        return {"ok": False, "message": "Portal returned 0 students — try again later."}
    conn = get_db(); c = conn.cursor()
    existing = {r["row_key"] for r in
                c.execute("SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0").fetchall()}
    updated = no_count = yes_count = not_in_tracker = 0
    for p in portal:
        toc = (p.get("toc_status") or "").strip().lower()
        rk = p.get("row_key", "")
        if toc not in ("yes", "no"):
            continue
        if rk not in existing:
            not_in_tracker += 1
            continue
        c.execute("UPDATE student_status SET toc_status=? WHERE row_key=?", (toc, rk))
        updated += 1
        if toc == "no": no_count += 1
        else: yes_count += 1
    conn.commit(); conn.close()
    msg = (f"Synced tocStatus for {updated} student(s) from Portal — no CapSolver used. "
           f"{no_count} are no-TOC, {yes_count} with-TOC.")
    if not_in_tracker:
        msg += f" {not_in_tracker} portal student(s) not on the Tracker yet (will arrive in a normal run)."
    return {"ok": True, "message": msg, "updated": updated, "no_count": no_count}

def _block_if_running():
    """Manual runs must NEVER cancel a run that is already going. If one is active, refuse and
    tell the operator to cancel it first. (Scheduled/auto runs are queued instead — they wait.)"""
    from job_runner import run_is_active
    if run_is_active():
        conn = get_db()
        r = conn.execute("SELECT group_type FROM run_logs WHERE status='running' "
                         "ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        what = (r["group_type"] if r else "A run")
        raise HTTPException(status_code=409, detail=(
            f"'{what}' is running right now. A running check is never cancelled automatically. "
            f"To start a different run, cancel the current one first (Stop button), then try again."))


@app.post("/api/run-now-notoc")
def run_now_notoc(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Run ONLY the no-TOC students (tocStatus='no') that are not yet confirmed. Confirmed ones
    already have a final status, so only the pending/active no-TOC students are checked here."""
    conn = get_db()
    rows = conn.execute("SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 "
                        "AND " + _SPECIAL_TOC_CLAUSE + " AND COALESCE(is_confirmed,0)=0 "
                        "AND COALESCE(current_status,'')!='SYC'").fetchall()
    keys = [r["row_key"] for r in rows]
    conn.close()
    if not keys:
        return {"message": "No pending special-TOC students to run.", "count": 0}
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "selected", keys)
    return {"message": f"Running {len(keys)} no-TOC student(s)…", "count": len(keys)}

@app.post("/api/run-now-unknown")
def run_now_unknown(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Re-run ONLY the Unknown-status students (status auto-retries on each)."""
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 "
                     "AND current_status='Unknown'").fetchone()[0]
    conn.close()
    if n == 0:
        return {"message": "No Unknown students to re-check.", "count": 0}
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "unknown")
    return {"message": f"Re-checking {n} Unknown student(s)…", "count": n}

def _auto_fix_toc(rk, nt, subs, src_trace):
    """Apply one NIOS TOC read for a student.
    If it MATCHES the Portal (or nothing to compare / already verified): just record the read
    (nios_toc + evidence), clear any stale mismatch flag/remark. Returns 'match'.
    If it DISAGREES with the Portal's tocStatus: AUTO-CORRECT — the NIOS value is applied on the
    tracker, pushed to the Portal, and the correction is logged in toc_fix_log for the audit
    list. No manual Verify needed. Returns 'fixed'."""
    cc = get_db()
    row = cc.execute("SELECT COALESCE(toc_status,''), COALESCE(toc_verified,0), COALESCE(student_id,''), "
                     "COALESCE(reference_no,''), COALESCE(student_name,''), COALESCE(session,''), "
                     "COALESCE(current_status,'') FROM student_status WHERE row_key=?", (rk,)).fetchone()
    if not row:
        cc.close()
        return "skip"
    ptoc = (row[0] or "").lower()
    sid = row[2]
    subs_json = json.dumps(subs) if subs else ""
    # NIOS is ALWAYS the source of truth. A previous 'verified' mark must NEVER suppress a fresh
    # disagreement — that bug let one bad read lock a student on the wrong TOC forever (and the
    # wrong WhatsApp campaign went out). The only guard kept: never downgrade yes -> no without
    # concrete evidence of what was read, so a blank/partial page can't wipe a real TOC.
    if ptoc == "yes" and nt == "no" and not (src_trace or "").strip():
        cc.close()
        logger.warning(f"TOC: refusing yes->no for {rk} — NIOS read had no evidence; will re-check")
        return "skip"
    if ptoc not in ("yes", "no") or ptoc == nt:
        # In sync (or manually settled) — record the read + evidence, clear stale flags.
        if subs_json:
            cc.execute("UPDATE student_status SET nios_toc=?, toc_subjects=?, toc_src=?, toc_mismatch=0 WHERE row_key=?",
                       (nt, subs_json, (src_trace or "")[:280], rk))
        else:
            cc.execute("UPDATE student_status SET nios_toc=?, toc_src=?, toc_mismatch=0 WHERE row_key=?",
                       (nt, (src_trace or "")[:280], rk))
        cc.execute("UPDATE student_status SET remark='' WHERE row_key=? AND remark LIKE 'mismatch toc status%'", (rk,))
        cc.commit()
        cc.close()
        return "match"
    # MISMATCH -> auto-correct to the NIOS value (NIOS is the source of truth).
    cc.execute("UPDATE student_status SET toc_status=?, nios_toc=?, toc_subjects=?, toc_src=?, "
               "toc_mismatch=0, toc_verified=1, remark=? WHERE row_key=?",
               (nt, nt, subs_json, (src_trace or "")[:280],
                f"TOC auto-corrected: MVS Portal said {ptoc.upper()}, NIOS official site says {nt.upper()} — "
                f"updated and pushed to the Portal", rk))
    pushed = 0
    try:
        import mvs_sync
        if sid and mvs_sync.push_toc(sid, nt, subs or []):
            pushed = 1
    except Exception as _pe:
        logger.warning(f"TOC auto-fix push failed {rk}: {_pe}")
    _wa = cc.execute("SELECT COALESCE(whatsapp_sent,0) FROM student_status WHERE row_key=?", (rk,)).fetchone()
    _wa_before = 1 if (_wa and _wa[0] == 1) else 0
    cc.execute("INSERT INTO toc_fix_log (row_key, reference_no, student_name, session, current_status, "
               "old_toc, new_toc, subjects, fixed_at, pushed, wa_sent_before) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
               (rk, row[3], row[4], row[5], row[6], ptoc, nt, ", ".join(subs or []),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"), pushed, _wa_before))
    cc.commit()
    cc.close()
    logger.info(f"TOC auto-corrected {rk}: Portal {ptoc} -> NIOS {nt} (pushed={pushed})")
    return "fixed"


# One-time "check TOC for ALL confirmed students" job state (confirmed students are never
# re-run, so this button reads their real NIOS TOC once and flags any mismatch). Live progress
# is polled by the dashboard. In-memory is fine — single-process app; resets on restart.
TOC_CHECK_STATE = {"running": False, "total": 0, "done": 0, "no_toc": 0, "yes_toc": 0,
                   "errors": 0, "cancel": False, "message": ""}


@app.post("/api/toc-check-confirmed")
def toc_check_confirmed(background_tasks: BackgroundTasks, body: dict = None, user=Depends(verify_token)):
    """Kick off a TOC read from NIOS.
    • No body: ONE-TIME pass over every confirmed student not yet TOC-checked (nios_toc empty).
    • body {row_keys:[...]}: check exactly the SELECTED students (any status — Active/Confirmed/etc),
      re-checking even if previously checked (an explicit selection means the operator wants it now).
    Each read is one public status-page request (one captcha, no login). Mismatches land in
    'TOC Status Error'. Verified students are never re-flagged. Re-clicking the one-time pass is
    safe — it only picks up whatever's left."""
    if TOC_CHECK_STATE["running"]:
        return {"ok": False, "running": True, "message": "A TOC check is already running."}
    body = body or {}
    sel_keys = [str(k) for k in (body.get("row_keys") or []) if k]
    flagged_only = bool(body.get("flagged"))
    conn = get_db()
    if flagged_only:
        # Re-check exactly the currently-flagged (unverified) mismatches — a fresh NIOS read.
        # Ones that now MATCH auto-unflag (and their stale remark clears); real mismatches stay.
        rows = conn.execute(
            "SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 "
            "AND COALESCE(toc_mismatch,0)=1 AND COALESCE(toc_verified,0)=0 "
            "AND COALESCE(dob,'')!='' "
            "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')").fetchall()
    elif body.get("force"):
        # FORCE: re-read EVERY confirmed student from NIOS, even if already checked/verified.
        # This is how KHUSHI-type cases (locked on a wrong TOC by an old bad read) get caught.
        rows = conn.execute(
            "SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
            "AND COALESCE(dob,'')!='' "
            "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')").fetchall()
    elif sel_keys:
        qm = ",".join("?" * len(sel_keys))
        rows = conn.execute(
            f"SELECT row_key FROM student_status WHERE row_key IN ({qm}) "
            "AND COALESCE(deleted,0)=0 AND COALESCE(dob,'')!='' "
            "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')",
            sel_keys).fetchall()
    else:
        rows = conn.execute(
            "SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
            "AND COALESCE(nios_toc,'')='' AND COALESCE(toc_verified,0)=0 "
            "AND COALESCE(dob,'')!='' "
            "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')").fetchall()
    keys = [r["row_key"] for r in rows]
    conn.close()
    if not keys:
        return {"ok": True, "count": 0,
                "message": ("None of the selected students can be TOC-checked (need a Reference/Enrollment + DOB)."
                            if sel_keys else "All confirmed students are already TOC-checked.")}
    TOC_CHECK_STATE.update({"running": True, "total": len(keys), "done": 0, "no_toc": 0,
                            "yes_toc": 0, "errors": 0, "cancel": False,
                            "message": f"Checking {len(keys)} student(s)…"})
    background_tasks.add_task(_run_toc_check_confirmed, keys)
    return {"ok": True, "count": len(keys),
            "message": f"Checking TOC for {len(keys)} student(s)… progress shows in TOC Status Error."}


@app.get("/api/toc-check-progress")
def toc_check_progress(user=Depends(verify_token)):
    s = dict(TOC_CHECK_STATE)
    s["percent"] = int(s["done"] * 100 / s["total"]) if s["total"] else 0
    return s


@app.post("/api/toc-check-cancel")
def toc_check_cancel(user=Depends(verify_token)):
    TOC_CHECK_STATE["cancel"] = True
    return {"ok": True, "message": "Stopping after the current student…"}


def _run_toc_check_confirmed(keys):
    """Background: read the real NIOS TOC for the given confirmed students via the public status
    page (same request that returns status also returns Previous Subject Details), and flag any
    mismatch with the Portal. Updates TOC_CHECK_STATE live so the dashboard shows a progress bar
    with checked / no-TOC / yes-TOC / errors."""
    try:
        from scraper import scrape_students
        conn = get_db()
        students = []
        for rk in keys:
            r = conn.execute("SELECT row_key, reference_no, enrollment_no, dob, session, toc_status, "
                             "student_name FROM student_status WHERE row_key=?", (rk,)).fetchone()
            if r:
                students.append(dict(r))
        conn.close()

        def _cb(res):
            rk = res.get("row_key")
            nt = (res.get("nios_toc") or "").lower()
            if not nt and res.get("nios_toc_absent"):
                # April/October pages hide the Previous-Subject table for no-TOC students —
                # absence on a successfully-read public page means TOC = No.
                try:
                    import nios_login as _nl
                    if _nl.is_public_session(res.get("session", "")):
                        nt = "no"
                        res["toc_src"] = ("No Previous-Subject-Details table on the NIOS page — "
                                          "for April/October sessions this means TOC = No")
                except Exception:
                    pass
            try:
                cc = get_db()
                cc.close()
                if nt in ("yes", "no"):
                    outcome = _auto_fix_toc(rk, nt, res.get("toc_subjects") or [],
                                            res.get("toc_src") or "")
                    if outcome == "fixed":
                        TOC_CHECK_STATE["errors"] += 1   # shown as "auto-fixed" on the panel
                    if nt == "yes":
                        TOC_CHECK_STATE["yes_toc"] += 1
                    else:
                        TOC_CHECK_STATE["no_toc"] += 1
            except Exception as e:
                logger.warning(f"confirmed-TOC cb error {rk}: {e}")
            TOC_CHECK_STATE["done"] += 1

        scrape_students(students, should_cancel=lambda: TOC_CHECK_STATE["cancel"], on_result=_cb)
        _d, _t = TOC_CHECK_STATE["done"], TOC_CHECK_STATE["total"]
        _read = TOC_CHECK_STATE["yes_toc"] + TOC_CHECK_STATE["no_toc"]
        _noread = max(0, _d - _read)
        _ok = max(0, _read - TOC_CHECK_STATE["errors"])
        TOC_CHECK_STATE["message"] = (
            f"Done — checked {_d} of {_t}: {TOC_CHECK_STATE['yes_toc']} yes-TOC, "
            f"{TOC_CHECK_STATE['no_toc']} no-TOC, {TOC_CHECK_STATE['errors']} auto-corrected & pushed to the Portal "
            f"(see the list below). {_ok} student(s) already matched — no issue. "
            + (f"{_noread} could not be read (no TOC table on their page / check failed) — "
               f"re-run these later or check them via 'Check TOC (selected)'." if _noread else
               "Every checked student's TOC was read successfully."))
    except Exception as e:
        logger.warning(f"confirmed-TOC check error: {e}")
        TOC_CHECK_STATE["message"] = f"Stopped due to an error: {e}"
    finally:
        TOC_CHECK_STATE["running"] = False


@app.get("/api/toc-fixes")
def toc_fixes(search: str = "", page: int = 1, per_page: int = 20, user=Depends(verify_token)):
    """Paginated audit list of automatic TOC corrections (Portal value replaced by NIOS value)."""
    conn = get_db()
    wc, params = [], []
    if search:
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR row_key LIKE ?)")
        params += [f"%{search}%"] * 3
    where = ("WHERE " + " AND ".join(wc)) if wc else ""
    total = conn.execute(f"SELECT COUNT(*) FROM toc_fix_log {where}", params).fetchone()[0]
    per_page = max(1, min(100, per_page))
    page = max(1, page)
    rows = conn.execute(f"SELECT * FROM toc_fix_log {where} ORDER BY id DESC LIMIT ? OFFSET ?",
                        params + [per_page, (page - 1) * per_page]).fetchall()
    conn.close()
    return {"fixes": [dict(r) for r in rows], "total": total, "page": page,
            "pages": max(1, (total + per_page - 1) // per_page)}


@app.get("/api/toc-errors")
def toc_errors(search: str = "", session_filter: str = "", user=Depends(verify_token)):
    """Students where NIOS says TOC=yes but the Portal has tocStatus=no — a mismatch a counsellor
    must verify before the (wrong) no-TOC WhatsApp goes out."""
    conn = get_db()
    wc = ["COALESCE(deleted,0)=0", "COALESCE(toc_mismatch,0)=1"]
    params = []
    if session_filter:
        clause, sp = _session_clause(session_filter); wc.append(clause); params += sp
    if search:
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR enrollment_no LIKE ?)")
        params += [f"%{search}%"] * 3
    where = "WHERE " + " AND ".join(wc)
    rows = conn.execute(f"SELECT row_key, reference_no, enrollment_no, student_name, session, "
                        f"current_status, nios_toc, toc_subjects, toc_verified, last_checked, "
                        f"COALESCE(toc_status,'') AS toc_status, COALESCE(toc_src,'') AS toc_src "
                        f"FROM student_status {where} ORDER BY toc_verified ASC, last_checked DESC "
                        f"LIMIT 500", params).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["toc_subjects"] = json.loads(r["toc_subjects"]) if r["toc_subjects"] else []
        except Exception:
            d["toc_subjects"] = []
        out.append(d)
    return {"students": out, "count": len(out)}

@app.post("/api/toc-verify")
def toc_verify(body: dict, user=Depends(verify_token)):
    """Counsellor resolves a TOC mismatch. body: {row_key, toc_status:'yes'|'no', toc_subjects:[...]}.
    Applies the corrected TOC, pushes it to the Portal, clears the mismatch and marks it verified —
    after which the run/WhatsApp proceeds with the correct campaign."""
    row_key = str(body.get("row_key") or "")
    toc_status = str(body.get("toc_status") or "").lower()
    toc_subjects = body.get("toc_subjects") or []
    if not row_key or toc_status not in ("yes", "no"):
        return {"ok": False, "error": "row_key and toc_status (yes/no) required"}
    conn = get_db()
    r = conn.execute("SELECT student_id FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    if not r:
        conn.close(); return {"ok": False, "error": "student not found"}
    subs_json = json.dumps(toc_subjects) if toc_subjects else ""
    conn.execute("UPDATE student_status SET toc_status=?, toc_verified=1, toc_mismatch=0, "
                 "nios_toc=?, toc_subjects=?, whatsapp_sent=0, whatsapp_info='' WHERE row_key=?",
                 (toc_status, toc_status, subs_json, row_key))
    conn.commit()
    pushed = False
    try:
        import mvs_sync
        if mvs_sync.enabled() and r["student_id"]:
            pushed = mvs_sync.push_toc(r["student_id"], toc_status, toc_subjects)
    except Exception as e:
        logger.warning(f"TOC push error {row_key}: {e}")
    conn.close()
    return {"ok": True, "pushed_to_portal": pushed,
            "message": f"TOC set to '{toc_status}' and verified. WhatsApp will now use the correct campaign."}

@app.post("/api/toc-check")
def toc_check(background_tasks: BackgroundTasks, body: dict = None, user=Depends(verify_token)):
    """Read the REAL TOC from NIOS (Previous Subject Details) for unverified no-TOC students and
    flag any mismatch with the Portal. Runs in the background; results appear in 'TOC Status Error'."""
    body = body or {}
    session_filter = str(body.get("session_filter", "") or "")
    conn = get_db()
    wc = ["COALESCE(deleted,0)=0", "COALESCE(toc_status,'')='no'", "COALESCE(toc_verified,0)=0",
          "COALESCE(dob,'')!=''", "(COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')"]
    params = []
    if session_filter:
        clause, sp = _session_clause(session_filter); wc.append(clause); params += sp
    rows = conn.execute(f"SELECT row_key FROM student_status WHERE {' AND '.join(wc)} LIMIT 1000",
                        params).fetchall()
    keys = [r["row_key"] for r in rows]
    conn.close()
    if not keys:
        return {"message": "No unverified no-TOC students to check.", "count": 0}
    background_tasks.add_task(_run_toc_check, keys)
    return {"message": f"Checking NIOS TOC for {len(keys)} student(s)… results appear in TOC Status Error.",
            "count": len(keys)}

def _run_toc_check(keys):
    """Background: read NIOS 'Previous Subject Details' per no-TOC student and flag a mismatch if
    NIOS shows TOC=yes. Skips while a run is active (no NIOS/captcha clash)."""
    try:
        import nios_login
        conn = get_db()
        if conn.execute("SELECT id FROM run_logs WHERE status='running'").fetchone():
            conn.close()
            logger.info("TOC check: a run is active — skipping to avoid clash; run it again after")
            return
        checked = flagged = 0
        for rk in keys:
            r = conn.execute("SELECT reference_no, enrollment_no, dob FROM student_status WHERE row_key=?",
                             (rk,)).fetchone()
            if not r:
                continue
            nios_toc, subs, ok = nios_login.fetch_nios_toc(r["reference_no"], r["dob"], r["enrollment_no"] or "")
            if not ok or not nios_toc:
                continue
            checked += 1
            outcome = _auto_fix_toc(rk, nios_toc, subs or [],
                                    "Read via NIOS student login (Previous Subject Details)")
            if outcome == "fixed":
                flagged += 1
        conn.close()
        logger.info(f"TOC check done | checked {checked}, auto-corrected {flagged}")
    except Exception as e:
        logger.warning(f"TOC check error: {e}")

@app.get("/api/reconciliation")
def reconciliation(user=Depends(verify_token)):
    """A precise breakdown of the tracker's own numbers so they can be reconciled against the
    MVS Portal dashboard. Explains WHERE the tracker total comes from and which buckets differ."""
    conn = get_db()
    ND = "COALESCE(deleted,0)=0"
    g = lambda q, *p: conn.execute(q, p).fetchone()[0]
    total_live   = g(f"SELECT COUNT(*) FROM student_status WHERE {ND}")
    total_all    = g("SELECT COUNT(*) FROM student_status")
    deleted      = g("SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=1")
    no_ref       = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} "
                     "AND COALESCE(reference_no,'')='' AND COALESCE(enrollment_no,'')=''")
    # source buckets
    enrol        = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} "
                     "AND COALESCE(source,'mvs_tracker')='mvs_portal' "
                     "AND (COALESCE(portal_origin,'')='enrol' "
                     "     OR (COALESCE(portal_origin,'')='' AND COALESCE(cross_dup,0)=0))")
    sheet_trans  = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} "
                     "AND COALESCE(source,'mvs_tracker')='mvs_portal' "
                     "AND NOT (COALESCE(portal_origin,'')='enrol' "
                     "         OR (COALESCE(portal_origin,'')='' AND COALESCE(cross_dup,0)=0))")
    transferred  = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND COALESCE(cross_dup,0)=1")
    # status buckets (two ways of counting 'confirmed' so any gap is visible)
    conf_flag    = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1")
    conf_status  = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Admission Confirmed'")
    conf_flag_nf = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1 "
                     "AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0")
    verified     = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Verified'")
    docverif     = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Documents Verification In Progress'")
    docreq       = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Document Required'")
    unknown      = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status IN ('Unknown','Fetch Error')")
    failed_run   = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} "
                     "AND (COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1)")
    data_err     = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND COALESCE(data_error,0)=1")
    # Confirmed push health — the ACTUAL reasons tracker/Portal confirmed counts can differ:
    #   push_lag = linked students whose latest status hasn't successfully reached the Portal yet
    #              (the 30-min resync sweep clears these; 'Sync Portal now' forces it).
    #   no_link  = confirmed students with NO Portal student-id — the tracker cannot push these
    #              at all until they exist on the Portal (transfer them first).
    push_lag = no_link = 0
    try:
        push_lag = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1 "
                     "AND COALESCE(student_id,'') != '' "
                     "AND COALESCE(portal_pushed,'') != COALESCE(current_status,'')")
        no_link  = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1 "
                     "AND COALESCE(student_id,'') = ''")
    except Exception:
        pass
    # Origin visibility — shows whether the Portal is actually sending the enrol/legacy split.
    org_enrol = org_sheet = org_blank = 0
    try:
        org_enrol = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND COALESCE(source,'')='mvs_portal' AND COALESCE(portal_origin,'')='enrol'")
        org_sheet = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND COALESCE(source,'')='mvs_portal' AND COALESCE(portal_origin,'')='sheet'")
        org_blank = g(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND COALESCE(source,'')='mvs_portal' AND COALESCE(portal_origin,'')=''")
    except Exception:
        pass
    conn.close()
    return {
        "total_live": total_live, "total_including_deleted": total_all, "deleted": deleted,
        "no_reference": no_ref,
        "by_source": {"enrol_portal": enrol, "sheet_transfer": sheet_trans, "transferred": transferred},
        "confirmed": {"by_flag": conf_flag, "by_flag_excluding_failed": conf_flag_nf,
                      "by_nios_status": conf_status},
        "statuses": {"verified": verified, "documents_verification": docverif,
                     "document_required": docreq, "unknown_or_error": unknown},
        "attention": {"failed_to_run": failed_run, "data_error": data_err,
                      "confirmed_push_lag": push_lag, "confirmed_no_link": no_link},
        "origin": {"enrol": org_enrol, "sheet": org_sheet, "blank": org_blank},
    }


PORTAL_SYNC_STATE = {"running": False, "total": 0, "done": 0, "pushed": 0, "message": ""}


@app.post("/api/portal-resync-now")
def portal_resync_now(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Force the tracker->Portal resync immediately, WITH feedback:
    • Nothing pending -> says so right away (plus how many can never push — no Portal link).
    • Something pending -> starts in the background and reports live progress via
      /api/portal-sync-status so the button can show 'Syncing… X/Y'."""
    if PORTAL_SYNC_STATE["running"]:
        return {"ok": False, "running": True, "message": "A Portal sync is already running."}
    conn = get_db()
    # HEAL drift first: students confirmed (flag=1, docs already sent) whose current_status
    # slid below 'Admission Confirmed' before the confirmed-protection existed. The tracker
    # counts them as Confirmed but keeps pushing the lower status — so the Portal shows
    # verifying and the two confirmed totals differ by exactly these. Restore the status;
    # that makes portal_pushed differ, so the push below sends 'Admission Confirmed'.
    drift = conn.execute(
        "SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 "
        "AND is_confirmed=1 AND COALESCE(current_status,'') != 'Admission Confirmed'").fetchone()[0]
    if drift:
        conn.execute("UPDATE student_status SET current_status='Admission Confirmed' "
                     "WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
                     "AND COALESCE(current_status,'') != 'Admission Confirmed'")
        conn.commit()
        logger.info(f"Portal sync: healed {drift} confirmed-status drift row(s)")
    # ── LINK REPAIR ─────────────────────────────────────────────────────────────
    # Old confirmed students never got their Portal student_id saved (confirmed students are
    # skipped by runs, and the id is saved during runs). Without the id their status can NEVER
    # be pushed — the exact cause of tracker-confirmed > Portal-confirmed. Match them to the
    # full Portal list by reference / enrollment / mobile and save the id (+origin).
    linked = 0
    unlinked = []
    try:
        import mvs_sync
        raw = mvs_sync._trackerlist(include_done=True, timeout=60)
        by_ref, by_enr, by_mob = {}, {}, {}
        for s in raw:
            sid = str(s.get("studentId") or "").strip()
            if not sid:
                continue
            org = mvs_sync._detect_origin(s)
            rec = (sid, org)
            rf = str(s.get("referenceNo") or "").strip().upper()
            en = str(s.get("enrollmentNo") or "").strip().upper()
            mb = "".join(ch for ch in str(s.get("mobile") or "") if ch.isdigit())[-10:]
            if rf and mvs_sync._valid_ref(rf):
                by_ref[rf] = rec
            if en and mvs_sync._valid_ref(en):
                by_enr[en] = rec
            if len(mb) == 10:
                by_mob.setdefault(mb, rec)
        rows = conn.execute(
            "SELECT row_key, COALESCE(reference_no,'') r, COALESCE(enrollment_no,'') e, "
            "COALESCE(mobile,'') m, COALESCE(student_name,'') n, is_confirmed "
            "FROM student_status WHERE COALESCE(deleted,0)=0 AND COALESCE(student_id,'')=''").fetchall()
        for t in rows:
            rec = (by_ref.get(t["r"].strip().upper())
                   or by_enr.get(t["e"].strip().upper())
                   or by_mob.get("".join(ch for ch in t["m"] if ch.isdigit())[-10:]))
            if rec:
                conn.execute("UPDATE student_status SET student_id=?, "
                             "portal_origin=CASE WHEN COALESCE(portal_origin,'')='' THEN ? ELSE portal_origin END "
                             "WHERE row_key=?", (rec[0], rec[1], t["row_key"]))
                linked += 1
            elif t["is_confirmed"] == 1:
                unlinked.append({"name": t["n"], "reference": t["r"] or t["e"], "mobile": t["m"]})
        if linked:
            conn.commit()
            logger.info(f"Portal sync: link-repair matched {linked} student(s) to Portal ids")
    except Exception as _le:
        logger.warning(f"Portal link-repair skipped: {_le}")
    pending = conn.execute(
        "SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 "
        "AND COALESCE(student_id,'') != '' "
        "AND current_status IN ('Admission Confirmed','Verified','Documents Verification In Progress','Document Required') "
        "AND COALESCE(portal_pushed,'') != current_status").fetchone()[0]
    no_link = conn.execute(
        "SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
        "AND COALESCE(student_id,'') = ''").fetchone()[0]
    run_active = bool(conn.execute("SELECT id FROM run_logs WHERE status='running'").fetchone())
    conn.close()
    if run_active:
        return {"ok": False, "message": "A status run is active — the sync will happen automatically right after it."}
    if pending == 0:
        msg = "Nothing pending — every linked student's status is already on the Portal."
        if linked:
            msg = f"Re-linked {linked} student(s) to the Portal. " + msg
        if drift:
            msg = f"Healed {drift} drifted confirmed student(s). " + msg
        if no_link:
            msg += (f" Note: {no_link} confirmed student(s) have NO Portal link (tracker-only) and can "
                    f"never be pushed — transfer them to the Portal from the Transfer Data page.")
        return {"ok": True, "nothing": True, "message": msg, "linked": linked, "unlinked": unlinked[:30]}
    PORTAL_SYNC_STATE.update({"running": True, "total": pending, "done": 0, "pushed": 0,
                              "message": f"Re-pushing {pending} student(s) to the Portal…"})

    def _run():
        try:
            from job_runner import portal_resync_sweep
            def _cb(done, total, pushed):
                PORTAL_SYNC_STATE.update({"done": done, "total": total, "pushed": pushed})
            portal_resync_sweep(progress_cb=_cb)
            PORTAL_SYNC_STATE["message"] = (f"Done — {PORTAL_SYNC_STATE['pushed']} of "
                                            f"{PORTAL_SYNC_STATE['total']} pushed to the Portal.")
        except Exception as e:
            PORTAL_SYNC_STATE["message"] = f"Sync stopped: {e}"
        finally:
            PORTAL_SYNC_STATE["running"] = False

    background_tasks.add_task(_run)
    return {"ok": True, "started": True, "count": pending, "linked": linked, "unlinked": unlinked[:30],
            "message": (f"Re-linked {linked} student(s). " if linked else "") +
                       f"Syncing {pending} student(s) to the Portal…"}


@app.get("/api/portal-sync-status")
def portal_sync_status(user=Depends(verify_token)):
    s = dict(PORTAL_SYNC_STATE)
    s["percent"] = int(s["done"] * 100 / s["total"]) if s["total"] else 0
    return s


@app.get("/api/portal-origin-probe")
def portal_origin_probe(user=Depends(verify_token)):
    """Fetch a sample from the Portal's trackerList and report EXACTLY what fields it sends,
    plus what the origin auto-detector makes of them. This is how we see, without guessing,
    whether the Portal is sending the Real-enrolments vs Bulk-imported split — and under which
    field name — so the two dashboard cards can match the Portal's own numbers."""
    import mvs_sync
    try:
        raw = mvs_sync._trackerlist(include_done=True, timeout=60)
    except Exception as e:
        return {"ok": False, "message": f"Could not fetch from the Portal: {e}"}
    if not raw:
        return {"ok": False, "message": "Portal returned no students."}
    sample = raw[:300]
    # union of keys + candidate origin-ish keys with their top values
    from collections import Counter
    all_keys = sorted({k for s in sample for k in s.keys()})
    hints = ("origin", "source", "legacy", "import", "created", "entry", "type", "via", "added")
    candidates = {}
    for k in all_keys:
        kl = k.lower()
        if any(h in kl for h in hints) and kl not in ("tocstatus", "toc_status"):
            vals = Counter(str(s.get(k))[:40] for s in sample if s.get(k) not in (None, ""))
            if vals:
                candidates[k] = dict(vals.most_common(5))
    det = Counter(mvs_sync._detect_origin(s) or "(blank)" for s in sample)
    return {"ok": True, "sampled": len(sample), "total": len(raw),
            "fields": all_keys, "candidate_fields": candidates,
            "detected": dict(det),
            "message": ("Origin detected for the sample — deploy pe agla run cards ko sahi split karega."
                        if det.get("(blank)", 0) == 0 else
                        f"{det.get('(blank)',0)} of {len(sample)} sampled students have NO origin field the "
                        f"tracker can recognise — the Portal's trackerList must include one (see candidate fields).")}


@app.post("/api/portal-origin-apply")
def portal_origin_apply(user=Depends(verify_token)):
    """One-click: fetch the full Portal list once and persist every student's origin
    (real_enrolment vs bulk_imported) into the tracker DB — no NIOS checks, no waiting for the
    next run. After this the Enrol vs MVS Portal cards match the Portal's own split instantly."""
    conn = get_db()
    if conn.execute("SELECT id FROM run_logs WHERE status='running'").fetchone():
        conn.close()
        return {"ok": False, "message": "A status run is active — origin will be saved automatically by the run itself."}
    conn.close()
    import mvs_sync
    try:
        # include_done=True -> the FULL portal list. Without it the Portal omits confirmed/done
        # students, so their origin never got saved (that's why 1990 stayed 'not sent').
        students = mvs_sync.fetch_students_for_tracker(include_done=True)
    except Exception as e:
        return {"ok": False, "message": f"Could not fetch from the Portal: {e}"}
    rows = [(s.get("portal_origin", ""), s["row_key"]) for s in students
            if s.get("row_key") and s.get("portal_origin")]
    enrol = sum(1 for o, _ in rows if o == "enrol")
    sheet = sum(1 for o, _ in rows if o == "sheet")
    conn = get_db()
    if rows:
        conn.executemany("UPDATE student_status SET portal_origin=? WHERE row_key=?", rows)
        conn.commit()
    conn.close()
    none = len(students) - len(rows)
    return {"ok": True, "updated": len(rows), "enrol": enrol, "sheet": sheet, "none": none,
            "message": (f"Origin saved for {len(rows)} students — Real enrolments: {enrol}, "
                        f"Bulk imported: {sheet}" + (f", no origin: {none}" if none else "") +
                        ". The data-source cards now match the Portal's split.")}


@app.get("/api/confirmed-audit")
def confirmed_audit(user=Depends(verify_token)):
    """Find the EXACT students the tracker counts as Confirmed but the Portal does not.
    Pulls the live Portal list, matches each tracker-confirmed (linked) student by studentId,
    and reports those whose Portal-side status isn't 'confirmed' (or who aren't in the Portal
    list at all). Also reports which status field the Portal sends, so we know the comparison
    is real and not a field-name guess."""
    import mvs_sync
    try:
        raw = mvs_sync._trackerlist(include_done=True, timeout=60)
    except Exception as e:
        return {"ok": False, "message": f"Could not fetch from the Portal: {e}", "mismatches": []}
    STATUS_KEYS = ("niosAdmissionStatus", "admissionStatus", "nios_status", "status",
                   "admission_status", "stage", "niosStatus", "admissionStage", "nios_stage")
    by_sid = {}
    portal_has_status = False
    def pstatus(p):
        nonlocal portal_has_status
        for k in STATUS_KEYS:
            v = p.get(k)
            if v not in (None, ""):
                portal_has_status = True
                return str(v)
        return ""
    for s in raw:
        sid = str(s.get("studentId") or "").strip()
        if sid:
            by_sid[sid] = s
    conn = get_db()
    confirmed = conn.execute(
        "SELECT row_key, reference_no, enrollment_no, student_name, mobile, session, "
        "student_id, current_status, COALESCE(portal_pushed,'') AS portal_pushed "
        "FROM student_status WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
        "AND COALESCE(student_id,'') != ''").fetchall()
    conn.close()
    mismatches = []
    for r in confirmed:
        p = by_sid.get(str(r["student_id"]))
        if not p:
            mismatches.append({"name": r["student_name"], "reference_no": r["reference_no"] or r["enrollment_no"],
                               "mobile": r["mobile"], "session": r["session"],
                               "portal_status": "NOT in the Portal list",
                               "tracker_status": r["current_status"], "pushed": r["portal_pushed"]})
        else:
            ps = pstatus(p)
            if "confirm" not in ps.lower():
                mismatches.append({"name": r["student_name"], "reference_no": r["reference_no"] or r["enrollment_no"],
                                   "mobile": r["mobile"], "session": r["session"],
                                   "portal_status": ps or "(no status field sent)",
                                   "tracker_status": r["current_status"], "pushed": r["portal_pushed"]})
    note = ""
    if not portal_has_status:
        note = ("The Portal's trackerList does NOT send a per-student status field, so a per-student "
                "compare isn't possible — the 5-gap is a Portal-side counting rule (its 'Confirmed' "
                "stage needs something beyond the status we push). The students below are all "
                "tracker-confirmed & linked; check any of them on the Portal to see the extra step.")
    return {"ok": True, "total_confirmed_linked": len(confirmed),
            "mismatch_count": len(mismatches), "portal_sends_status": portal_has_status,
            "note": note, "mismatches": mismatches[:200]}


FORCE_CONF_STATE = {"running": False, "total": 0, "done": 0, "ok": 0, "failed": 0,
                    "fail_list": [], "message": ""}


@app.post("/api/portal-force-confirmed")
def portal_force_confirmed(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Re-push EVERY confirmed + linked student to the Portal, ignoring portal_pushed. This is the
    'they show 0 pending but the Portal is still short' fix: it re-sends 'Admission Confirmed' for
    all of them and records EXACTLY which ones the Portal did not accept (with names), so the
    remaining gap becomes a concrete, checkable list instead of a mystery number."""
    if FORCE_CONF_STATE["running"]:
        return {"ok": False, "running": True, "message": "A force re-push is already running."}
    conn = get_db()
    if conn.execute("SELECT id FROM run_logs WHERE status='running'").fetchone():
        conn.close()
        return {"ok": False, "message": "A status run is active — try again once it finishes."}
    rows = conn.execute(
        "SELECT row_key, student_id, reference_no, enrollment_no, session, student_name, remark "
        "FROM student_status WHERE COALESCE(deleted,0)=0 AND is_confirmed=1 "
        "AND COALESCE(student_id,'') != ''").fetchall()
    students = [dict(r) for r in rows]
    conn.close()
    if not students:
        return {"ok": True, "count": 0, "message": "No confirmed linked students to push."}
    FORCE_CONF_STATE.update({"running": True, "total": len(students), "done": 0, "ok": 0,
                             "failed": 0, "fail_list": [], "message": f"Re-pushing {len(students)}…"})

    def _run():
        try:
            import mvs_sync
            for s in students:
                ok = False
                try:
                    ok = mvs_sync.push_student({
                        "student_id": s["student_id"], "row_key": s["row_key"],
                        "reference_no": s["reference_no"], "enrollment_no": s["enrollment_no"],
                        "discovered_ref": s["reference_no"], "session": s["session"] or "",
                        "remark": s["remark"] or ""}, "Admission Confirmed")
                except Exception as e:
                    logger.warning(f"force-confirmed push error {s['row_key']}: {e}")
                if ok:
                    FORCE_CONF_STATE["ok"] += 1
                    try:
                        cx = get_db()
                        cx.execute("UPDATE student_status SET portal_pushed='Admission Confirmed' WHERE row_key=?",
                                   (s["row_key"],))
                        cx.commit(); cx.close()
                    except Exception:
                        pass
                else:
                    FORCE_CONF_STATE["failed"] += 1
                    if len(FORCE_CONF_STATE["fail_list"]) < 200:
                        FORCE_CONF_STATE["fail_list"].append({
                            "name": s["student_name"] or "",
                            "reference": s["reference_no"] or s["enrollment_no"] or "",
                            "student_id": s["student_id"] or ""})
                FORCE_CONF_STATE["done"] += 1
            FORCE_CONF_STATE["message"] = (
                f"Done — re-pushed {FORCE_CONF_STATE['ok']} of {FORCE_CONF_STATE['total']} confirmed. "
                + (f"{FORCE_CONF_STATE['failed']} were NOT accepted by the Portal (listed below) — "
                   f"check these students on the Portal directly." if FORCE_CONF_STATE["failed"] else
                   "All accepted. If the Portal's confirmed count is still lower, those students are "
                   "counted in a different stage on the Portal side."))
        except Exception as e:
            FORCE_CONF_STATE["message"] = f"Stopped: {e}"
        finally:
            FORCE_CONF_STATE["running"] = False

    background_tasks.add_task(_run)
    return {"ok": True, "started": True, "count": len(students),
            "message": f"Re-pushing {len(students)} confirmed students to the Portal…"}


@app.get("/api/portal-force-status")
def portal_force_status(user=Depends(verify_token)):
    s = dict(FORCE_CONF_STATE)
    s["percent"] = int(s["done"] * 100 / s["total"]) if s["total"] else 0
    return s


DETAIL_SYNC_STATE = {"running": False, "total": 0, "done": 0, "ok": 0, "failed": 0,
                     "fail_list": [], "message": ""}


@app.post("/api/portal-sync-details")
def portal_sync_details(background_tasks: BackgroundTasks, body: dict = None,
                        user=Depends(verify_token)):
    """Push every linked student's IDENTITY DETAILS (reference/enrollment no, DOB, mobile,
    alt mobile, email, name, class, session) from the TRACKER to the PORTAL. This is the
    backfill for students already corrected on the tracker (e.g. wrong references fixed
    after a run) whose Portal copy is still stale. Optional body {"row_keys":[...]} limits
    it to specific students; without it, ALL non-deleted linked students are pushed."""
    if DETAIL_SYNC_STATE["running"]:
        return {"ok": False, "running": True, "message": "A detail sync is already running."}
    import mvs_sync
    if not mvs_sync.enabled():
        return {"ok": False, "message": "Portal sync is OFF (MVS_MODE not enabled)."}
    keys = [k for k in ((body or {}).get("row_keys") or []) if str(k).strip()]
    conn = get_db()
    q = ("SELECT row_key, student_name, mobile, alt_mobile, email, dob, reference_no, "
         "enrollment_no, class_level, session, COALESCE(student_id,'') AS student_id "
         "FROM student_status WHERE COALESCE(deleted,0)=0 AND COALESCE(student_id,'') != ''")
    if keys:
        q += " AND row_key IN (%s)" % ",".join("?" * len(keys))
        rows = conn.execute(q, keys).fetchall()
    else:
        rows = conn.execute(q).fetchall()
    students = [dict(r) for r in rows]
    conn.close()
    if not students:
        return {"ok": True, "count": 0, "message": "No linked students to sync."}
    DETAIL_SYNC_STATE.update({"running": True, "total": len(students), "done": 0, "ok": 0,
                              "failed": 0, "fail_list": [],
                              "message": f"Syncing details of {len(students)} student(s)…"})

    def _run():
        from datetime import datetime as _dt
        try:
            for s in students:
                ok, msg = False, ""
                try:
                    ok, msg = mvs_sync.push_details(s["student_id"], s)
                except Exception as e:
                    msg = str(e)
                _ts = _dt.now().strftime("%Y-%m-%d %H:%M")
                _mark = (f"synced {_ts}" if ok else f"failed {_ts}: {msg}"[:300])
                try:
                    cx = get_db()
                    cx.execute("UPDATE student_status SET edit_sync=? WHERE row_key=?",
                               (_mark, s["row_key"]))
                    cx.commit(); cx.close()
                except Exception:
                    pass
                if ok:
                    DETAIL_SYNC_STATE["ok"] += 1
                else:
                    DETAIL_SYNC_STATE["failed"] += 1
                    if len(DETAIL_SYNC_STATE["fail_list"]) < 200:
                        DETAIL_SYNC_STATE["fail_list"].append({
                            "name": s["student_name"] or "",
                            "reference": s["reference_no"] or s["enrollment_no"] or "",
                            "reason": msg[:120]})
                DETAIL_SYNC_STATE["done"] += 1
            DETAIL_SYNC_STATE["message"] = (
                f"Done — details of {DETAIL_SYNC_STATE['ok']} of {DETAIL_SYNC_STATE['total']} "
                f"student(s) synced to the Portal."
                + (f" {DETAIL_SYNC_STATE['failed']} FAILED (listed below)."
                   if DETAIL_SYNC_STATE["failed"] else " All accepted."))
        except Exception as e:
            DETAIL_SYNC_STATE["message"] = f"Stopped: {e}"
        finally:
            DETAIL_SYNC_STATE["running"] = False

    background_tasks.add_task(_run)
    return {"ok": True, "started": True, "count": len(students),
            "message": f"Syncing details of {len(students)} student(s) to the Portal…"}


@app.get("/api/portal-sync-details-status")
def portal_sync_details_status(user=Depends(verify_token)):
    s = dict(DETAIL_SYNC_STATE)
    s["percent"] = int(s["done"] * 100 / s["total"]) if s["total"] else 0
    return s


@app.get("/api/pending-students")
def pending_students(user=Depends(verify_token)):
    """LIVE list of Portal students who have NO usable Reference/Enrollment yet — the Portal has
    them, the tracker can't check them. Read straight from the Portal every time (nothing is
    stored), so the moment a student's reference is filled, the normal run imports them and they
    drop off this list automatically — fetching is never blocked by this view."""
    import mvs_sync
    try:
        raw = mvs_sync._trackerlist(include_done=True, timeout=60)
    except Exception as e:
        return {"ok": False, "message": f"Could not fetch from the Portal: {e}", "students": []}
    import re as _re
    # A REAL NIOS reference is one letter + 10 digits (B0526300834); a real enrollment is 11-12
    # digits (920526301652). Anything else sitting in those fields — "0", "NA", a mobile number,
    # a half-typed value — is NOT checkable. The Portal counts those students as pending too,
    # which is why its pending count was higher than ours.
    _REF = _re.compile(r"^[A-Za-z]\d{10}$")
    _ENR = _re.compile(r"^\d{11,12}$")

    def _classify(ref, enr):
        has_ref, has_enr = bool(ref), bool(enr)
        if _REF.match(ref or "") or _ENR.match(enr or ""):
            return None                                  # properly checkable
        if not has_ref and not has_enr:
            return "No reference/enrollment yet"
        bad = ref or enr
        if "@" in bad:
            return "Email sitting in the reference field"
        return f"Reference looks invalid: \"{bad[:24]}\" (expected 1 letter + 10 digits)"

    out = []
    counts = {"no_ref": 0, "bad_ref": 0}
    for s in raw:
        ref = str(s.get("referenceNo") or "").strip()
        enr = str(s.get("enrollmentNo") or "").strip()
        reason = _classify(ref, enr)
        if not reason:
            continue
        if reason.startswith("No reference"):
            counts["no_ref"] += 1
        else:
            counts["bad_ref"] += 1
        out.append({"student_id": str(s.get("studentId") or ""),
                    "name": str(s.get("name") or "").strip(),
                    "mobile": str(s.get("mobile") or "").strip(),
                    "session": str(s.get("examSession") or "").strip(),
                    "class_level": str(s.get("class") or "").strip(),
                    "reference": ref or enr or "",
                    "reason": reason})
    return {"ok": True, "total": len(out), "students": out,
            "portal_total": len(raw), "counts": counts}


@app.get("/api/confirmed-compare")
def confirmed_compare(user=Depends(verify_token)):
    """Student-by-student comparison of the tracker's confirmed list against the LIVE Portal
    list — finds exactly WHO makes the two confirmed totals differ and why:
      • portal_stage_differs — Portal has the student but its own stage field isn't 'confirmed'
        (push was accepted yet the stage didn't move) — these can be force re-pushed.
      • missing_on_portal   — tracker has them linked, but the Portal list no longer contains
        that studentId (deleted/merged on the Portal side).
      • no_link             — tracker-only students (no Portal id) — transfer to push."""
    import mvs_sync
    try:
        raw = mvs_sync._trackerlist(include_done=True, timeout=90)
    except Exception as e:
        return {"ok": False, "message": f"Could not fetch from the Portal: {e}"}
    stage_field = None
    if raw:
        keys = set()
        for r in raw[:80]:
            keys |= set(r.keys())
        for cand in ("niosStatus", "nios_status", "admissionStatus", "admission_status",
                     "niosStage", "stage", "trackerStatus", "status"):
            if cand in keys:
                stage_field = cand
                break
    pmap = {str(r.get("studentId") or "").strip(): r for r in raw if r.get("studentId")}
    conn = get_db()
    rows = conn.execute("SELECT row_key, student_id, reference_no, student_name, session, "
                        "current_status FROM student_status WHERE COALESCE(deleted,0)=0 "
                        "AND is_confirmed=1").fetchall()
    conn.close()
    missing, differs, no_link = [], [], []
    matched = 0
    for r in rows:
        sid = (r["student_id"] or "").strip()
        item = {"row_key": r["row_key"], "name": r["student_name"] or "",
                "reference": r["reference_no"] or "", "session": r["session"] or ""}
        if not sid:
            no_link.append(item)
            continue
        p = pmap.get(sid)
        if not p:
            missing.append(item)
            continue
        if stage_field:
            pv = str(p.get(stage_field) or "").strip()
            if "confirm" in pv.lower():
                matched += 1
            else:
                item["portal_value"] = pv or "(blank)"
                differs.append(item)
        else:
            matched += 1
    portal_conf = None
    if stage_field:
        portal_conf = sum(1 for r in raw if "confirm" in str(r.get(stage_field) or "").lower())
    return {"ok": True, "stage_field": stage_field, "tracker_confirmed": len(rows),
            "portal_confirmed_by_field": portal_conf, "matched": matched,
            "differs": differs[:60], "differs_count": len(differs),
            "missing": missing[:60], "missing_count": len(missing),
            "no_link": no_link[:60], "no_link_count": len(no_link)}


@app.post("/api/force-push")
def force_push(body: dict, user=Depends(verify_token)):
    """Force re-push the given students' current status to the Portal (ignores portal_pushed)."""
    import mvs_sync
    keys = [str(k) for k in (body.get("row_keys") or []) if k]
    if not keys:
        return {"ok": False, "message": "No students given."}
    conn = get_db()
    qm = ",".join("?" * len(keys))
    rows = conn.execute(f"SELECT row_key, student_id, reference_no, enrollment_no, session, "
                        f"current_status, remark FROM student_status WHERE row_key IN ({qm}) "
                        "AND COALESCE(student_id,'') != ''", keys).fetchall()
    conn.close()
    pushed = 0
    for r in rows:
        try:
            ok = mvs_sync.push_student({
                "student_id": r["student_id"], "row_key": r["row_key"],
                "reference_no": r["reference_no"], "enrollment_no": r["enrollment_no"],
                "discovered_ref": r["reference_no"], "session": r["session"] or "",
                "remark": r["remark"] or ""}, r["current_status"])
            if ok:
                cc = get_db()
                cc.execute("UPDATE student_status SET portal_pushed=? WHERE row_key=?",
                           (r["current_status"], r["row_key"]))
                cc.commit(); cc.close()
                pushed += 1
        except Exception as e:
            logger.warning(f"force-push error {r['row_key']}: {e}")
    return {"ok": True, "pushed": pushed, "total": len(rows),
            "message": f"Force re-pushed {pushed}/{len(rows)} student(s) to the Portal."}


@app.get("/api/run-queue")
def run_queue(user=Depends(verify_token)):
    """Scheduled runs waiting for the current one to finish (they are never cancelled)."""
    from job_runner import queued_runs, run_is_active
    return {"active": run_is_active(), "queued": queued_runs()}


@app.get("/api/check-summary")
def check_summary(session_filter: str = "", source_filter: str = "", view: str = "normal",
                  user=Depends(verify_token)):
    """For the selected session (or All): how many students were checked in the latest run
    ('new') vs not checked ('stale'). Powers the header counts on the Active Students page."""
    conn = get_db()
    boundary = _verify_boundary(conn, view, session_filter, source_filter)
    w_new, p_new = _build_student_where(view, "", "", session_filter, "", "", "", source_filter,
                                        "", "", "new", boundary)
    w_st, p_st = _build_student_where(view, "", "", session_filter, "", "", "", source_filter,
                                      "", "", "stale", boundary)
    new_n = conn.execute(f"SELECT COUNT(*) FROM student_status {w_new}", p_new).fetchone()[0]
    stale_n = conn.execute(f"SELECT COUNT(*) FROM student_status {w_st}", p_st).fetchone()[0]
    conn.close()
    return {"new": new_n, "stale": stale_n, "boundary": boundary}

@app.post("/api/run-now-notchecked")
def run_now_notchecked(background_tasks: BackgroundTasks, body: dict = None,
                       user=Depends(verify_token)):
    """Run ONLY the students the latest run couldn't check (captcha/proxy fell back), for the
    selected session or All. Skips confirmed/SYC and anything without a reference/enrollment."""
    body = body or {}
    session_filter = str(body.get("session_filter", "") or "")
    source_filter = str(body.get("source_filter", "") or "")
    conn = get_db()
    boundary = _verify_boundary(conn, "normal", session_filter, source_filter)
    where, params = _build_student_where("normal", "", "", session_filter, "", "", "",
                                         source_filter, "", "", "stale", boundary)
    rows = conn.execute(
        f"SELECT row_key FROM student_status {where} "
        "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='') "
        "ORDER BY COALESCE(last_verified,'') ASC, COALESCE(last_checked,'') ASC", params).fetchall()
    keys = [r["row_key"] for r in rows]
    conn.close()
    if not keys:
        return {"message": "No not-checked students in this scope — everything is up to date.", "count": 0}
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "selected", keys)
    scope = "all sessions" if not session_filter else session_filter
    return {"message": f"Re-checking {len(keys)} not-checked student(s) in {scope}…", "count": len(keys)}

@app.post("/api/mark-name-verified")
def mark_name_verified(row_key: str = Form(...), user=Depends(verify_token)):
    """Counsellor confirms (after checking the portal) that the Reference No genuinely
    belongs to this student — the wrong-reference warning was a false alarm. Dismiss the
    error, remember it (name_verified=1) so the warning never re-appears, and drop the
    student into its correct bucket."""
    conn = get_db(); c = conn.cursor()
    r = c.execute("SELECT current_status FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    if not r:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")
    is_conf = 1 if (r["current_status"] or "") == "Admission Confirmed" else 0
    c.execute("""UPDATE student_status
                 SET name_verified=1, login_failed=0, check_failed=0, login_remark='',
                     whatsapp_info='', is_confirmed=?
                 WHERE row_key=?""", (is_conf, row_key))
    conn.commit(); conn.close()
    return {"ok": True, "confirmed": bool(is_conf)}

@app.get("/api/nav-count")
def nav_count(user=Depends(verify_token)):
    """Live counts for the sidebar badges (active / confirmed / required / syc / failed)."""
    conn = get_db()
    ND = "COALESCE(deleted,0)=0"
    NF = "COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0"   # only in 'Failed to Run'
    def c(sql, *p):
        return conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND " + sql, p).fetchone()[0]
    active = c(NF + " AND COALESCE(is_confirmed,0)=0 AND COALESCE(current_status,'')!='SYC' "
               "AND (session IS NULL OR session NOT LIKE '%syc%')")
    confirmed = c(NF + " AND is_confirmed=1")
    required = c(NF + " AND current_status='Document Required'")
    syc = c(NF + " AND (session LIKE '%syc%' OR current_status='SYC')")
    failed = c("(COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1) AND COALESCE(current_status,'')!='Unknown'")
    unknown = c("current_status='Unknown'")
    conn.close()
    return {"students": active, "confirmed": confirmed, "required": required,
            "syc": syc, "failed": failed, "unknown": unknown}

@app.get("/api/failed-count")
def failed_count(user=Depends(verify_token)):
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 "
                     "AND (COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1)").fetchone()[0]
    conn.close()
    return {"count": n}

@app.get("/api/export-students")
def export_students(view: str="normal", search: str="", status_filter: str="",
                         session_filter: str="", class_filter: str="",
                         date_from: str="", date_to: str="", source_filter: str="",
                         toc_filter: str="",
                         user=Depends(verify_token)):
    """Export the CURRENTLY FILTERED list (active / confirmed / required) to .xlsx.
    Honours the exact same filters as the on-screen table (search, status, session,
    class 10/12, data type, and date range)."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    conn = get_db()
    where, params = _build_student_where(view, search, status_filter, session_filter,
                                         class_filter, date_from, date_to, source_filter,
                                         toc_filter=toc_filter)
    rows = conn.execute(f"SELECT * FROM student_status {where} ORDER BY student_name", params).fetchall()
    conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Students"
    headers = ["#", "Reference No", "Student Name", "Mobile", "Class", "Email",
               "Session", "Status", "Data Type", "Remark", "Last Checked"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        dt = "MVS Portal" if (r["source"] or "mvs_tracker") == "mvs_portal" else "MVS Tracker"
        ws.append([i, r["reference_no"], r["student_name"], r["mobile"], r["class_level"],
                   r["email"], r["session"], r["current_status"], dt, r["remark"], r["last_checked"]])
    hdr_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    for idx, w in enumerate([5, 18, 26, 14, 8, 28, 22, 30, 44, 20], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    label = {"confirmed": "confirmed", "required": "document_required"}.get(view, "active")
    fname = f"nios_{label}_students.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.get("/api/history")
def get_history(page: int = 1, per_page: int = 10, from_dt: str = "", to_dt: str = "",
                     status: str = "", source: str = "", search: str = "", user=Depends(verify_token)):
    conn = get_db()
    # Prefer the source STORED on the history row (set at write time). Fall back to a
    # reference-no lookup for old rows, then default. This keeps History in sync with
    # the Active/Confirmed pages.
    src_expr = ("COALESCE(NULLIF(status_history.source,''), "
                "(SELECT ss.source FROM student_status ss "
                "WHERE ss.reference_no != '' AND ss.reference_no = status_history.reference_no "
                "AND ss.source IS NOT NULL LIMIT 1), 'mvs_tracker')")
    cols = "id, reference_no, student_name, old_status, new_status, changed_at, run_id"
    wc, params = [], []
    if from_dt:
        wc.append("changed_at >= ?"); params.append(from_dt)
    if to_dt:
        wc.append("changed_at <= ?"); params.append(to_dt)
    if status:
        wc.append("new_status = ?"); params.append(status)
    if source == "both":
        wc.append("EXISTS (SELECT 1 FROM student_status ss2 "
                  "WHERE ss2.reference_no = status_history.reference_no "
                  "AND COALESCE(ss2.cross_dup,0) = 1)")
    elif source:
        wc.append(f"{src_expr} = ?"); params.append(source)
    if search:
        like = f"%{search.strip()}%"
        # Match reference no OR student name (stored on the row) OR the student's email
        # / enrollment (looked up from student_status by reference).
        wc.append("(status_history.reference_no LIKE ? OR status_history.student_name LIKE ? "
                  "OR EXISTS (SELECT 1 FROM student_status ss3 "
                  "WHERE ss3.reference_no = status_history.reference_no "
                  "AND (ss3.email LIKE ? OR ss3.enrollment_no LIKE ?)))")
        params += [like, like, like, like]
    where = ("WHERE " + " AND ".join(wc)) if wc else ""
    total = conn.execute(f"SELECT COUNT(*) FROM status_history {where}", params).fetchone()[0]
    per_page = max(1, min(per_page, 200))
    page = max(1, page)
    offset = (page - 1) * per_page
    rows = conn.execute(
        f"SELECT {cols}, {src_expr} AS source FROM status_history {where} ORDER BY id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]).fetchall()
    conn.close()
    pages = max(1, (total + per_page - 1) // per_page)
    return {"items": [dict(x) for x in rows], "total": total,
            "page": page, "pages": pages, "per_page": per_page}

@app.post("/api/history-delete")
def history_delete(id: int, user=Depends(verify_token)):
    """Delete ONE status-change history entry by id."""
    conn = get_db()
    conn.execute("DELETE FROM status_history WHERE id=?", (id,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.post("/api/history-clear")
def history_clear(user=Depends(verify_token)):
    """Clear ALL status-change history (does not affect students)."""
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM status_history").fetchone()[0]
    conn.execute("DELETE FROM status_history")
    conn.commit(); conn.close()
    return {"ok": True, "deleted": n}

@app.get("/api/export-history")
def export_history(from_dt: str = "", to_dt: str = "", status: str = "",
                         source: str = "", user=Depends(verify_token)):
    """Export the filtered Change History to .xlsx."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    conn = get_db()
    src_expr = ("COALESCE(NULLIF(status_history.source,''), "
                "(SELECT ss.source FROM student_status ss "
                "WHERE ss.reference_no != '' AND ss.reference_no = status_history.reference_no "
                "AND ss.source IS NOT NULL LIMIT 1), 'mvs_tracker')")
    cols = "id, reference_no, student_name, old_status, new_status, changed_at, run_id"
    wc, params = [], []
    if from_dt:
        wc.append("changed_at >= ?"); params.append(from_dt)
    if to_dt:
        wc.append("changed_at <= ?"); params.append(to_dt)
    if status:
        wc.append("new_status = ?"); params.append(status)
    if source == "both":
        wc.append("EXISTS (SELECT 1 FROM student_status ss2 "
                  "WHERE ss2.reference_no = status_history.reference_no "
                  "AND COALESCE(ss2.cross_dup,0) = 1)")
    elif source:
        wc.append(f"{src_expr} = ?"); params.append(source)
    where = ("WHERE " + " AND ".join(wc)) if wc else ""
    rows = conn.execute(f"SELECT {cols}, {src_expr} AS source FROM status_history {where} ORDER BY id DESC", params).fetchall()
    conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Change History"
    headers = ["#", "Reference No", "Student Name", "Old Status", "New Status", "Data Type", "Changed At"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        dt = "MVS Portal" if (r["source"] or "mvs_tracker") == "mvs_portal" else "MVS Tracker"
        ws.append([i, r["reference_no"], r["student_name"], r["old_status"],
                   r["new_status"], dt, r["changed_at"]])
    fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill
    for idx, w in enumerate([5, 18, 26, 28, 28, 16, 20], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nios_change_history.xlsx"})

@app.get("/api/run-logs")
def get_run_logs(limit: int=50, user=Depends(verify_token)):
    conn = get_db()
    l = conn.execute("SELECT * FROM run_logs ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(x) for x in l]

@app.get("/api/transfers")
def get_transfers(page: int = 1, per_page: int = 10, search: str = "",
                        mode: str = "", user=Depends(verify_token)):
    """Tracker -> Portal transfer log, with master search + pagination (10..100/page)."""
    conn = get_db()
    wc, params = [], []
    if search:
        like = f"%{search.strip()}%"
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR enrollment_no LIKE ? OR mobile LIKE ?)")
        params += [like, like, like, like]
    if mode in ("auto", "manual"):
        wc.append("mode=?"); params.append(mode)
    where = ("WHERE " + " AND ".join(wc)) if wc else ""
    total = conn.execute(f"SELECT COUNT(*) FROM transfer_log {where}", params).fetchone()[0]
    per_page = max(10, min(per_page, 100)); page = max(1, page)
    offset = (page - 1) * per_page
    rows = conn.execute(f"SELECT * FROM transfer_log {where} ORDER BY id DESC LIMIT ? OFFSET ?",
                        params + [per_page, offset]).fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows], "total": total, "page": page,
            "per_page": per_page, "pages": max(1, (total + per_page - 1) // per_page)}

@app.post("/api/transfers-clear")
def transfers_clear(mode: str = "", user=Depends(verify_token)):
    """Reset the Tracker -> Portal transfer log. This ONLY clears the history list — students
    already pushed to the Portal stay transferred (their cross_dup flag is untouched). Pass
    mode=auto|manual to clear just that kind, else the whole log is cleared."""
    conn = get_db()
    if mode in ("auto", "manual"):
        n = conn.execute("DELETE FROM transfer_log WHERE mode=?", (mode,)).rowcount
    else:
        n = conn.execute("DELETE FROM transfer_log").rowcount
    conn.commit(); conn.close()
    return {"deleted": int(n or 0)}

@app.post("/api/transfer-sync")
def transfer_sync(user=Depends(verify_token)):
    """Manually push the current NIOS status + document links of every matched (Both /
    cross-source) student to MVS Portal, and log each as a 'manual' transfer."""
    import mvs_sync
    if not mvs_sync.enabled():
        raise HTTPException(status_code=400, detail="MVS Portal bridge is not enabled (set MVS_MODE).")
    try:
        portal = mvs_sync.fetch_students_for_tracker(include_done=True)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not reach MVS Portal: {e}")
    sid_by_key = {p["row_key"]: p.get("student_id", "") for p in portal if p.get("student_id")}
    conn = get_db(); c = conn.cursor()
    rows = c.execute("SELECT * FROM student_status WHERE COALESCE(deleted,0)=0 "
                     "AND COALESCE(cross_dup,0)=1").fetchall()
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pushed = 0
    for r in rows:
        rk = r["row_key"]; sid = sid_by_key.get(rk, "")
        if not sid:
            continue
        student = {"student_id": sid, "row_key": rk,
                   "reference_no": r["reference_no"] or "", "enrollment_no": r["enrollment_no"] or "",
                   "session": r["session"] or "", "remark": r["remark"] or ""}
        try:
            mvs_sync.push_student(student, r["current_status"] or "", conn)
            c.execute("""INSERT INTO transfer_log
                (row_key, reference_no, enrollment_no, student_name, mobile, session,
                 old_status, new_status, transferred_at, mode)
                VALUES (?,?,?,?,?,?,?,?,?,'manual')""",
                (rk, r["reference_no"] or "", r["enrollment_no"] or "", r["student_name"] or "",
                 r["mobile"] or "", r["session"] or "", r["current_status"] or "",
                 r["current_status"] or "", now_s))
            pushed += 1
        except Exception as e:
            logger.warning(f"transfer-sync push failed {rk}: {e}")
    conn.commit(); conn.close()
    return {"message": f"Synced {pushed} matched student(s) to MVS Portal.", "count": pushed}

_MATCH_PROGRESS = {"running": False, "finished": False, "phase": "", "total": 0,
                   "done": 0, "transferred": 0, "result": None, "error": ""}
_MATCH_WORKERS = 4   # parallel pushes to the Portal (each push is a different student row)

def _do_transfer_match():
    """Background worker: match every live MVS Portal student to an ALREADY-CHECKED Tracker student
    by Reference No and push the existing status to the Portal WITHOUT a CapSolver check. Optimised
    for big data: (1) all Tracker students are loaded ONCE into memory (no per-student DB query),
    (2) the slow network pushes run in parallel and hold NO DB lock (so the portal UI stays fast),
    (3) the DB writes (cross_dup + transfer_log) are done in one short batch AFTER the pushes."""
    import mvs_sync
    _MATCH_PROGRESS["phase"] = "Fetching Portal data"
    try:
        portal = mvs_sync.fetch_students_for_tracker(include_done=True)
    except Exception as e:
        return {"ok": False, "message": f"Could not reach MVS Portal: {e}",
                "transferred": 0, "new_fetch": 0, "not_matched": []}
    if not portal:
        return {"ok": False, "message": "MVS Portal returned 0 students — try again later.",
                "transferred": 0, "new_fetch": 0, "not_matched": []}
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    BAD = ("", "unknown", "fetch error")
    not_matched = []

    # 1) CACHE all Tracker students once (reference_no -> best row, row_key -> row). No more
    #    per-student DB queries during the loop, and the read connection is closed immediately.
    _MATCH_PROGRESS["phase"] = "Loading Tracker data"
    conn = get_db()
    trk_by_ref, trk_by_rk = {}, {}
    for r in conn.execute(
            "SELECT row_key, reference_no, enrollment_no, student_name, mobile, session, remark, "
            "current_status FROM student_status WHERE COALESCE(deleted,0)=0").fetchall():
        d = {k: r[k] for k in r.keys()}
        trk_by_rk[d["row_key"]] = d
        ref = (d["reference_no"] or "").strip()
        if ref:
            cur = trk_by_ref.get(ref)
            d_ok = (d["current_status"] or "").strip().lower() not in BAD
            if cur is None or (d_ok and (cur["current_status"] or "").strip().lower() in BAD):
                trk_by_ref[ref] = d
    conn.close()

    # 2) Decide matches fully in memory (no DB, no network) — instant.
    _MATCH_PROGRESS.update({"phase": "Matching", "total": len(portal), "done": 0, "transferred": 0})
    tasks = []
    for p in portal:
        ref  = (p.get("reference_no") or "").strip()
        sid  = (p.get("student_id") or "").strip()
        name = (p.get("student_name") or "").strip() or "—"
        rk_p = p.get("row_key", "")
        trow = (trk_by_ref.get(ref) if ref else None) or (trk_by_rk.get(rk_p) if rk_p else None)
        status = ((trow["current_status"] if trow else "") or "").strip()
        if not sid:
            not_matched.append({"student_name": name, "reference_no": ref,
                                "reason": "Portal record has no studentId — cannot push"}); continue
        if not trow:
            not_matched.append({"student_name": name, "reference_no": ref,
                                "reason": "Not on Tracker / no Reference match — will run in New Fetch"}); continue
        if status.lower() in BAD:
            not_matched.append({"student_name": name, "reference_no": ref,
                                "reason": f"Tracker status still '{status or 'blank'}' — will run in New Fetch"}); continue
        tasks.append({"sid": sid, "status": status, "name": name, "ref": ref, "trow": trow,
                      "student": {"student_id": sid, "row_key": trow["row_key"],
                                  "reference_no": trow["reference_no"] or ref,
                                  "enrollment_no": trow["enrollment_no"] or "",
                                  "session": trow["session"] or "", "remark": trow["remark"] or ""}})

    # 3) Push to the Portal IN PARALLEL — the slow part. No DB lock is held here, so the portal
    #    UI stays responsive. Each push targets a different student row (safe to parallelise).
    import threading as _th
    from concurrent.futures import ThreadPoolExecutor
    _MATCH_PROGRESS.update({"phase": "Pushing", "total": len(tasks), "done": 0, "transferred": 0})
    lock = _th.Lock()
    okrows = []
    cnt = {"done": 0, "ok": 0}

    def _push_one(t):
        try:
            mvs_sync.push_student(t["student"], t["status"])
            ok, err = True, None
        except Exception as e:
            ok, err = False, str(e)
        with lock:
            cnt["done"] += 1
            _MATCH_PROGRESS["done"] = cnt["done"]
            if ok:
                cnt["ok"] += 1
                _MATCH_PROGRESS["transferred"] = cnt["ok"]
                okrows.append(t)
            else:
                not_matched.append({"student_name": t["name"], "reference_no": t["ref"],
                                    "reason": f"Push failed: {err}"})

    if tasks:
        with ThreadPoolExecutor(max_workers=_MATCH_WORKERS) as ex:
            list(ex.map(_push_one, tasks))

    # 4) Persist the successful transfers to the DB in one short batch (fast, no network).
    _MATCH_PROGRESS["phase"] = "Saving"
    conn = get_db(); c = conn.cursor()
    for i, t in enumerate(okrows):
        trow = t["trow"]; rk = trow["row_key"]; status = t["status"]
        c.execute("UPDATE student_status SET source='mvs_portal', cross_dup=1 WHERE row_key=?", (rk,))
        c.execute("DELETE FROM transfer_log WHERE row_key=? AND mode='manual'", (rk,))
        c.execute("""INSERT INTO transfer_log
            (row_key, reference_no, enrollment_no, student_name, mobile, session,
             old_status, new_status, transferred_at, mode)
            VALUES (?,?,?,?,?,?,?,?,?,'manual')""",
            (rk, trow["reference_no"] or t["ref"], trow["enrollment_no"] or "",
             trow["student_name"] or t["name"], trow["mobile"] or "", trow["session"] or "",
             status, status, now_s))
        if (i + 1) % 200 == 0:
            conn.commit()
    conn.commit(); conn.close()

    transferred = cnt["ok"]
    msg = (f"Transferred {transferred} already-checked student(s) to Portal — no CapSolver used. "
           f"{len(not_matched)} left for New Fetch.")
    result = {"ok": True, "message": msg, "transferred": transferred,
              "new_fetch": len(not_matched), "not_matched": not_matched[:1000]}
    try:
        import json as _json
        set_setting("last_transfer_match", _json.dumps({**result, "at": now_s}))
    except Exception:
        pass
    return result


def _run_transfer_match_bg():
    """Background worker so the HTTP request returns instantly (a large match can take minutes —
    holding the request open would hit the gateway timeout = HTTP 502). The UI polls progress."""
    try:
        res = _do_transfer_match()
        _MATCH_PROGRESS["result"] = res
        _MATCH_PROGRESS["error"] = "" if res.get("ok") else (res.get("message") or "")
    except Exception as e:
        logger.warning(f"transfer-match background error: {e}")
        _MATCH_PROGRESS["result"] = None
        _MATCH_PROGRESS["error"] = str(e)
    finally:
        _MATCH_PROGRESS["running"] = False
        _MATCH_PROGRESS["finished"] = True


@app.post("/api/transfer-match")
async def transfer_match(user=Depends(verify_token)):
    """Start a Match & Transfer in the BACKGROUND and return immediately. Matching live MVS Portal
    students to ALREADY-CHECKED Tracker students by Reference No and pushing the existing status to
    the Portal (no CapSolver) can take minutes for large data, so it runs in a background thread
    and the UI polls /api/transfer-match-progress for the % bar and the final result."""
    import mvs_sync, threading
    if not mvs_sync.enabled():
        raise HTTPException(status_code=400, detail="MVS Portal bridge is not enabled (set MVS_MODE).")
    if _MATCH_PROGRESS.get("running"):
        return {"started": False, "message": "A Match & Transfer is already running."}
    _MATCH_PROGRESS.update({"running": True, "finished": False, "phase": "Starting",
                            "total": 0, "done": 0, "transferred": 0, "result": None, "error": ""})
    threading.Thread(target=_run_transfer_match_bg, daemon=True).start()
    return {"started": True}


@app.get("/api/transfer-match-progress")
async def transfer_match_progress(user=Depends(verify_token)):
    """Live progress + final result of the running/last Match & Transfer (polled by the UI)."""
    return dict(_MATCH_PROGRESS)


@app.get("/api/transfers-download")
def transfers_download(search: str = "", mode: str = "", user=Depends(verify_token)):
    """Download the Transfer Data (Tracker -> Portal) log as an .xlsx sheet."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    conn = get_db()
    wc, params = [], []
    if search:
        like = f"%{search.strip()}%"
        wc.append("(student_name LIKE ? OR reference_no LIKE ? OR enrollment_no LIKE ? OR mobile LIKE ?)")
        params += [like, like, like, like]
    if mode in ("auto", "manual"):
        wc.append("mode=?"); params.append(mode)
    where = ("WHERE " + " AND ".join(wc)) if wc else ""
    rows = conn.execute(f"SELECT * FROM transfer_log {where} ORDER BY id DESC", params).fetchall()
    conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Transfer Data"
    ws.append(["#", "Student", "Reference No", "Enroll No", "Mobile", "Session",
               "Old Status", "New Status", "Transferred At", "Mode"])
    for i, r in enumerate(rows, 1):
        ws.append([i, r["student_name"] or "", r["reference_no"] or "", r["enrollment_no"] or "",
                   r["mobile"] or "", r["session"] or "", r["old_status"] or "",
                   r["new_status"] or "", r["transferred_at"] or "", r["mode"] or ""])
    fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill
    for idx, w in enumerate([5, 24, 18, 16, 14, 14, 24, 24, 20, 10], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mvs_transfer_data.xlsx"})

@app.get("/api/transfer-match-last")
async def transfer_match_last(user=Depends(verify_token)):
    """The most recent Match & Transfer result, so the 'not transferred / why' list stays
    visible on the Transfer Data page even after a page refresh."""
    import json as _json
    raw = get_setting("last_transfer_match", "")
    if not raw:
        return {"ok": False}
    try:
        return _json.loads(raw)
    except Exception:
        return {"ok": False}

@app.get("/api/portal-stale")
def portal_stale(user=Depends(verify_token)):
    """Preview OLD / removed MVS Portal students: students we still track as source=mvs_portal
    whose record is NO LONGER present on the live portal (e.g. an old batch you removed).
    Returns the list only — nothing is deleted here. Guarded so a failed/empty fetch never
    flags everyone as stale."""
    import mvs_sync
    if not mvs_sync.enabled():
        raise HTTPException(status_code=400, detail="MVS Portal bridge is not enabled (set MVS_MODE).")
    try:
        portal = mvs_sync.fetch_students_for_tracker(include_done=True)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not reach MVS Portal: {e}")
    live_keys = {p["row_key"] for p in portal if p.get("row_key")}
    # Safety: if the portal returned nothing, refuse to flag anything (avoid mass-delete).
    if not live_keys:
        return {"ok": False, "fetched": 0, "stale": [],
                "message": "Portal returned 0 students — not flagging anything (try again later)."}
    conn = get_db()
    rows = conn.execute(
        "SELECT row_key, student_name, reference_no, enrollment_no, session, current_status, "
        "mobile, last_checked FROM student_status "
        "WHERE COALESCE(deleted,0)=0 AND COALESCE(source,'mvs_tracker')='mvs_portal' "
        "ORDER BY student_name").fetchall()
    conn.close()
    stale = []
    for r in rows:
        if r["row_key"] not in live_keys:
            stale.append({"row_key": r["row_key"], "student_name": r["student_name"] or "—",
                          "reference_no": r["reference_no"] or r["enrollment_no"] or "—",
                          "session": r["session"] or "—", "status": r["current_status"] or "—",
                          "mobile": r["mobile"] or "—", "last_checked": r["last_checked"] or "—"})
    return {"ok": True, "fetched": len(live_keys), "stale": stale, "count": len(stale)}

@app.post("/api/portal-to-tracker")
def portal_to_tracker(body: dict, user=Depends(verify_token)):
    """Move selected students from MVS Portal -> MVS Tracker (keeps the record, just changes
    its data source). Used to preserve old portal records instead of deleting them: they stay
    tracked as Tracker data and are no longer tied to the live portal."""
    keys = [k for k in (body.get("row_keys", []) or []) if k][:1000]
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    conn = get_db()
    ph = ",".join("?" * len(keys))
    cur = conn.execute(
        f"UPDATE student_status SET source='mvs_tracker', cross_dup=0 "
        f"WHERE row_key IN ({ph}) AND COALESCE(deleted,0)=0", keys)
    conn.commit()
    n = cur.rowcount
    conn.close()
    return {"ok": True, "moved": n}

@app.post("/api/run-now")
async def run_now(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    _block_if_running()
    background_tasks.add_task(run_status_check, "all")
    return {"message": "Run triggered for ALL students (Tracker + Portal)!"}

@app.post("/api/run-now-tracker")
async def run_now_tracker(background_tasks: BackgroundTasks, group: str = "all", user=Depends(verify_token)):
    """Manual run: MVS Tracker data, optionally limited to ONE session group so a
    counsellor can run just On Demand / Stream 2 / Public instead of every tracker
    student (saves CapSolver credits)."""
    g = group if group in ("ondemand", "stream2", "public") else "all"
    background_tasks.add_task(run_status_check, g, "mvs_tracker")
    label = {"ondemand": "On Demand", "stream2": "Stream 2", "public": "Public", "all": "all"}[g]
    return {"message": f"Run triggered for MVS Tracker — {label} students!"}

@app.post("/api/run-required")
def run_now_required(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Manual run: ONLY the students whose status is 'Document Required'. Lets a
    counsellor re-check just those after resolving their documents — so a fixed
    student moves out of the Required list without re-running everyone."""
    conn = get_db()
    ND = "COALESCE(deleted,0)=0 AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0"
    rows = conn.execute(
        f"SELECT row_key FROM student_status WHERE {ND} AND current_status='Document Required'"
    ).fetchall()
    conn.close()
    keys = [r["row_key"] for r in rows if r["row_key"]]
    if not keys:
        return {"message": "No Document Required students to run.", "count": 0}
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "required", keys)
    return {"message": f"Re-checking {len(keys)} Document Required student(s)!", "count": len(keys)}

@app.post("/api/run-now-portal")
async def run_now_portal(background_tasks: BackgroundTasks, group: str = "all", user=Depends(verify_token)):
    """Manual run: MVS Portal data, optionally limited to ONE session group so a
    counsellor can run just On Demand / Stream 2 / Public instead of every verification
    student (saves CapSolver credits)."""
    g = group if group in ("ondemand", "stream2", "public") else "all"
    background_tasks.add_task(run_status_check, g, "mvs_portal")
    label = {"ondemand": "On Demand", "stream2": "Stream 2", "public": "Public", "all": "all"}[g]
    return {"message": f"Run triggered for MVS Portal — {label} students!"}

@app.post("/api/run-now-portal-new")
async def run_now_portal_new(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Manual run: MVS Portal — NEW data only. Checks just the students that have arrived
    on the portal since last time (not already in the tracker), skipping all already-active
    students to save CapSolver credits."""
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", "mvs_portal", "new")
    return {"message": "Run triggered for MVS Portal — NEW data only!"}

@app.post("/api/run-now-upload")
async def run_now_upload(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Run ONLY the students in the just-uploaded Excel sheet (MVS Tracker), not the
    whole database or MVS Portal. Used by the 'Run Check Now' button after upload."""
    background_tasks.add_task(run_status_check, "all", None, "upload")
    return {"message": "Checking only the uploaded students!"}

@app.post("/api/run-now-failed")
def run_now_failed(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Re-run EVERY 'Failed to Run' student with auto-fix: the status read auto-retries
    (transient Fetch Errors heal themselves) and a confirmed student's DOB is auto-flipped
    (date<->month) if that swap makes the NIOS login work. One-click backlog cleanup —
    the same fixes also run automatically on every normal run from now on."""
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM student_status WHERE COALESCE(deleted,0)=0 "
                     "AND (COALESCE(login_failed,0)=1 OR COALESCE(check_failed,0)=1)").fetchone()[0]
    conn.close()
    if n == 0:
        return {"message": "No failed students to re-check.", "count": 0}
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "failed")
    return {"message": f"Re-checking {n} failed student(s) with auto-fix…", "count": n}

@app.post("/api/run-selected")
def run_selected(body: dict, background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Run ONLY hand-picked students (1..20) — saves CapSolver credits when a counsellor
    needs just one or a few statuses right now instead of running everyone."""
    keys = body.get("row_keys") or []
    if not isinstance(keys, list):
        raise HTTPException(status_code=400, detail="row_keys must be a list")
    # de-dup + clean
    keys = [str(k) for k in keys if str(k).strip()]
    keys = list(dict.fromkeys(keys))
    if len(keys) < 1:
        raise HTTPException(status_code=400, detail="Select at least 1 student")
    if len(keys) > 20:
        raise HTTPException(status_code=400, detail="You can select a maximum of 20 students at a time")
    # keep only real, non-deleted students
    conn = get_db()
    qmarks = ",".join("?" * len(keys))
    valid = [r["row_key"] for r in conn.execute(
        f"SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 AND row_key IN ({qmarks})",
        keys).fetchall()]
    conn.close()
    if not valid:
        raise HTTPException(status_code=404, detail="No valid students found for the selection")
    _block_if_running()
    background_tasks.add_task(run_status_check, "all", None, "selected", valid)
    return {"message": f"Checking {len(valid)} selected student(s)…", "count": len(valid)}

@app.post("/api/cancel-run")
def cancel_run(run_id: int = Form(...), user=Depends(verify_token)):
    conn = get_db()
    row = conn.execute("SELECT status FROM run_logs WHERE id=?", (run_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Run not found")
    if row["status"] != "running":
        conn.close()
        return {"message": f"Run already {row['status']}", "status": row["status"]}
    conn.execute("UPDATE run_logs SET status='cancelled' WHERE id=?", (run_id,))
    conn.commit()
    conn.close()
    return {"message": "Run cancelled", "status": "cancelled"}

@app.get("/api/progress")
def run_progress(user=Depends(verify_token)):
    """Live progress of the currently running check (for the progress bar)."""
    conn = get_db()
    row = conn.execute("SELECT id, group_type, run_at, progress_current, progress_total, "
                       "progress_changed, progress_same, progress_notchecked, progress_total_mvs, progress_done_mvs, "
                       "progress_total_trk, progress_done_trk, retry_total, retry_done "
                       "FROM run_logs WHERE status='running' ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    if not row:
        return {"running": False}
    cur = row["progress_current"] or 0
    tot = row["progress_total"] or 0
    pct = int(cur * 100 / tot) if tot else 0
    tm = row["progress_total_mvs"] or 0
    dm = row["progress_done_mvs"] or 0
    tt = row["progress_total_trk"] or 0
    dt = row["progress_done_trk"] or 0
    rtot = row["retry_total"] or 0
    rdone = row["retry_done"] or 0
    # Retry phase = main pass finished (all checked) and there are fixable failures being retried.
    in_retry = (rtot > 0 and cur >= tot and tot > 0)
    return {"running": True, "id": row["id"], "group_type": row["group_type"],
            "run_at": row["run_at"], "current": cur, "total": tot, "percent": pct,
            "changed": row["progress_changed"] or 0, "same": row["progress_same"] or 0,
            "not_checked": row["progress_notchecked"] or 0,
            "remaining": max(0, tot - cur),
            "retry_total": rtot, "retry_done": rdone,
            "retry_percent": int(rdone * 100 / rtot) if rtot else 0,
            "phase": "retry" if in_retry else "main",
            "mvs": {"done": dm, "total": tm, "percent": int(dm*100/tm) if tm else 0},
            "trk": {"done": dt, "total": tt, "percent": int(dt*100/tt) if tt else 0}}

GROUP_LABELS = {"ondemand": "On Demand", "stream2": "Stream 2",
                "public": "Public (April / October)", "portalnew": "MVS Portal — New data"}
_GROUP_JID = {"ondemand": "job_ondemand", "stream2": "job_stream2", "public": "job_public", "portalnew": "job_portalnew"}
_GROUP_DH = {"ondemand": 6, "stream2": 6, "public": 12, "portalnew": 3}

def _job_remaining_sec(jid):
    """Seconds until a scheduled job's next run (None if not scheduled)."""
    job = scheduler.get_job(jid)
    if not job or not job.next_run_time:
        return None
    nrt = job.next_run_time
    if nrt.tzinfo is None:
        delta = (nrt - datetime.now()).total_seconds()
    else:
        delta = (nrt - datetime.now(timezone.utc)).total_seconds()
    return max(0, int(delta))

@app.get("/api/next-runs")
async def next_runs(user=Depends(verify_token)):
    """Seconds remaining until the next automatic run of each group, plus that group's
    own paused flag (each group can be paused independently)."""
    out = []
    for grp, jid, _dh in RUN_GROUPS:
        paused = get_setting(f"paused_{grp}", "") == "1"
        if paused:
            try:
                secs = int(get_setting(f"pause_remaining_{grp}_sec", "") or 0)
            except Exception:
                secs = None
        else:
            secs = _job_remaining_sec(jid)
        out.append({"group": grp, "label": GROUP_LABELS.get(grp, grp),
                    "seconds": secs, "paused": paused})
    # MVS Portal — New data timer
    pn_paused = get_setting("paused_portalnew", "") == "1"
    if pn_paused:
        try:
            pn_secs = int(get_setting("pause_remaining_portalnew_sec", "") or 0)
        except Exception:
            pn_secs = None
    else:
        pn_secs = _job_remaining_sec("job_portalnew")
    out.append({"group": "portalnew", "label": GROUP_LABELS["portalnew"],
                "seconds": pn_secs, "paused": pn_paused})
    return {"runs": out}

def _valid_group(g):
    if g not in ("ondemand", "stream2", "public", "portalnew"):
        raise HTTPException(status_code=400, detail="group must be 'ondemand', 'stream2', 'public' or 'portalnew'")
    return g

@app.get("/api/report-settings")
async def get_report_settings(user=Depends(verify_token)):
    return {
        "enabled": get_setting("report_enabled", "") == "1",
        "numbers": get_setting("report_numbers", ""),
        "campaign_set": bool(os.environ.get("AISENSY_CAMPAIGN_REPORT", "").strip()),
        "last_status": get_setting("report_last_status", ""),
    }

@app.post("/api/report-settings")
async def save_report_settings(body: dict, user=Depends(verify_token)):
    enabled = "1" if body.get("enabled") else ""
    raw = str(body.get("numbers", "") or "")
    # Keep only valid-ish numbers (digits, 10–13 long after cleaning), comma-joined.
    import re as _re
    cleaned = []
    for part in raw.replace("\n", ",").split(","):
        d = _re.sub(r"\D", "", part)
        if 10 <= len(d) <= 13:
            cleaned.append(d)
    if len(cleaned) > 10:
        cleaned = cleaned[:10]
    set_setting("report_enabled", enabled)
    set_setting("report_numbers", ",".join(cleaned))
    return {"enabled": enabled == "1", "numbers": ",".join(cleaned), "count": len(cleaned)}

@app.post("/api/report-test")
def report_test(user=Depends(verify_token)):
    """Send a sample report right now to the configured numbers, to verify setup."""
    import whatsapp, links
    raw = (get_setting("report_numbers", "") or "").replace("\n", ",")
    nums = [n.strip() for n in raw.split(",") if n.strip()]
    if not nums:
        raise HTTPException(status_code=400, detail="Add at least one WhatsApp number first")
    today = datetime.now().strftime("%Y-%m-%d")
    when = datetime.now().strftime("%d %b, %I:%M %p")
    url = links.report_url(today)
    # live current counts for a realistic test
    conn = get_db(); ND = "COALESCE(deleted,0)=0"
    conf = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1").fetchone()[0]
    req = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Document Required'").fetchone()[0]
    ver = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Verified'").fetchone()[0]
    dvp = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Documents Verification In Progress'").fetchone()[0]
    err = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status IN ('Unknown','Fetch Error')").fetchone()[0]
    conn.close()
    params = whatsapp.make_report_params(f"TEST report - {when}", conf, req, err, 0,
                                         conf + req + err + ver + dvp, url)
    sent, errs = whatsapp.send_report_to_all(nums, params)
    return {"sent": sent, "total": len(nums), "errors": errs}

@app.get("/api/recent-runs")
def recent_runs(user=Depends(verify_token)):
    """Recent completed runs that have a saved report, for the manual 'send report' picker."""
    import json as _json
    conn = get_db()
    rows = conn.execute("SELECT id, group_type, run_at, report_label, report_json, total_checked "
                        "FROM run_logs WHERE report_json IS NOT NULL AND report_json != '' "
                        "ORDER BY id DESC LIMIT 15").fetchall()
    conn.close()
    out = []
    for r in rows:
        try:
            st = _json.loads(r["report_json"])
        except Exception:
            st = {}
        out.append({
            "id": r["id"],
            "label": r["report_label"] or (r["group_type"] or "Run"),
            "run_at": r["run_at"],
            "confirmed": st.get("confirmed", 0),
            "required": st.get("required", 0),
            "checked": st.get("checked", r["total_checked"] or 0),
        })
    return out

@app.post("/api/send-run-report")
def send_run_report(body: dict, user=Depends(verify_token)):
    """Send a SPECIFIC past run's report (chosen from the picker) to all management numbers."""
    import whatsapp, links, json as _json
    run_id = body.get("run_id")
    raw = (get_setting("report_numbers", "") or "").replace("\n", ",")
    nums = [n.strip() for n in raw.split(",") if n.strip()]
    if not nums:
        raise HTTPException(status_code=400, detail="Add at least one management WhatsApp number in Settings first")
    if not os.environ.get("AISENSY_CAMPAIGN_REPORT", "").strip():
        raise HTTPException(status_code=400, detail="AISENSY_CAMPAIGN_REPORT env var is not set on Railway")
    conn = get_db()
    r = conn.execute("SELECT report_json, report_label FROM run_logs WHERE id=?", (run_id,)).fetchone()
    conn.close()
    if not r or not r["report_json"]:
        raise HTTPException(status_code=404, detail="That run's report was not found")
    st = _json.loads(r["report_json"])
    label = r["report_label"] or "Run report"
    same = max(0, st.get("checked", 0) - st.get("changed", 0))
    url = links.report_url(datetime.now().strftime("%Y-%m-%d"))
    # Confirmed / Required / Error from LIVE totals (match the Excel Summary), like the
    # automatic report. Total-checked + unchanged stay as that run's own numbers.
    import job_runner as _jr
    live = _jr._live_report_counts()
    params = whatsapp.make_report_params(label, live["confirmed"], live["required"],
                                         live["error"], same, st.get("checked", 0), url)
    sent, errs = whatsapp.send_report_to_all(nums, params)
    stamp = datetime.now().strftime("%d %b %I:%M %p")
    set_setting("report_last_status",
                (f"Sent '{label}' to {sent}/{len(nums)} on {stamp}." if sent
                 else f"FAILED on {stamp}: " + ("; ".join(errs)[:200] if errs else "unknown error")))
    return {"sent": sent, "total": len(nums), "errors": errs}

@app.post("/api/send-latest-report")
def send_latest_report(user=Depends(verify_token)):
    """Push the latest run report to ALL configured management numbers right now —
    a manual safety net for when the auto-report did not reach everyone."""
    import whatsapp, links
    raw = (get_setting("report_numbers", "") or "").replace("\n", ",")
    nums = [n.strip() for n in raw.split(",") if n.strip()]
    if not nums:
        raise HTTPException(status_code=400, detail="Add at least one management WhatsApp number in Settings first")
    if not os.environ.get("AISENSY_CAMPAIGN_REPORT", "").strip():
        raise HTTPException(status_code=400, detail="AISENSY_CAMPAIGN_REPORT env var is not set on Railway — add it, then redeploy")
    conn = get_db(); ND = "COALESCE(deleted,0)=0"
    last = conn.execute("SELECT group_type, run_at FROM run_logs "
                        "WHERE status LIKE 'completed%' OR status LIKE 'done%' "
                        "ORDER BY id DESC LIMIT 1").fetchone()
    conf = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND is_confirmed=1").fetchone()[0]
    req = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Document Required'").fetchone()[0]
    ver = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Verified'").fetchone()[0]
    dvp = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status='Documents Verification In Progress'").fetchone()[0]
    err = conn.execute(f"SELECT COUNT(*) FROM student_status WHERE {ND} AND current_status IN ('Unknown','Fetch Error')").fetchone()[0]
    conn.close()
    today = datetime.now().strftime("%Y-%m-%d")
    when = datetime.now().strftime("%d %b, %I:%M %p")
    label = (last["group_type"].title() + " run" if last and last["group_type"] else "Latest run")
    url = links.report_url(today)
    params = whatsapp.make_report_params(f"{label} - {when}", conf, req, err, 0,
                                         conf + req + err + ver + dvp, url)
    sent, errs = whatsapp.send_report_to_all(nums, params)
    stamp = datetime.now().strftime("%d %b %I:%M %p")
    set_setting("report_last_status",
                (f"Sent to {sent}/{len(nums)} on {stamp} (manual)." if sent
                 else f"FAILED on {stamp}: " + ("; ".join(errs)[:200] if errs else "unknown error")))
    return {"sent": sent, "total": len(nums), "errors": errs}

@app.post("/api/intervals-pause")
async def intervals_pause(body: dict, user=Depends(verify_token)):
    """Freeze ONE group's auto-run timer (regular OR public): capture how long is left,
    then stop just that group's schedule. Resuming continues from this remaining time."""
    grp = _valid_group((body or {}).get("group", ""))
    jid = _GROUP_JID[grp]
    dh = _GROUP_DH[grp]
    rem = _job_remaining_sec(jid)
    if rem is None:
        rem = _interval_minutes(grp, dh) * 60
    set_setting(f"pause_remaining_{grp}_sec", rem)
    try:
        scheduler.remove_job(jid)
    except Exception:
        pass
    set_setting(f"paused_{grp}", "1")
    return {"paused": True, "group": grp,
            "message": f"{GROUP_LABELS.get(grp, grp)} paused — timer frozen."}

@app.post("/api/intervals-resume")
async def intervals_resume(body: dict, user=Depends(verify_token)):
    """Resume ONE group: re-schedule it to fire after its FROZEN remaining time, so the
    countdown picks up exactly where it was paused (not reset to a full interval)."""
    grp = _valid_group((body or {}).get("group", ""))
    jid = _GROUP_JID[grp]
    dh = _GROUP_DH[grp]
    mins = _interval_minutes(grp, dh)
    try:
        rem = int(get_setting(f"pause_remaining_{grp}_sec", "") or 0)
    except Exception:
        rem = 0
    if rem <= 0:
        rem = mins * 60
    if grp == "portalnew":
        job_fn = lambda: run_status_check("all", "mvs_portal", "new", is_auto=True)
    else:
        job_fn = lambda g=grp: run_status_check(g, _group_source(g), is_auto=True)
    scheduler.add_job(job_fn,
                      trigger=IntervalTrigger(minutes=mins),
                      id=jid, replace_existing=True,
                      next_run_time=datetime.now() + timedelta(seconds=rem))
    set_setting(f"paused_{grp}", "")
    set_setting(f"pause_remaining_{grp}_sec", "")
    return {"paused": False, "group": grp,
            "message": f"{GROUP_LABELS.get(grp, grp)} resumed — timer continues from where it was paused."}

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...), user=Depends(verify_token)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    async with aiofiles.open(EXCEL_PATH, "wb") as f:
        content = await file.read()
        await f.write(content)

    # Parse the uploaded sheet to report counts + a preview. A student is a
    # "duplicate" if it already exists in the system (matched by reference OR
    # email OR name+mobile) OR appears more than once within this file. Only the
    # genuinely NEW students are counted as new.
    resp = {"message": f"Excel uploaded ({len(content)} bytes)", "filename": file.filename,
            "total": 0, "duplicates": 0, "unique": 0, "preview": []}
    try:
        from excel_handler import read_students_from_excel

        def _norm(v):
            return str(v or "").strip().lower()

        def _real_email(e):
            e = _norm(e)
            return ("@" in e and "." in e.split("@")[-1]
                    and e not in ("temp", "na", "none", "nil", "-", "null", "n/a"))

        def _valid_mobile(m):
            d = "".join(ch for ch in str(m or "") if ch.isdigit())
            return d if len(d) >= 10 else ""

        def _rowkey(ref, enroll, email):
            # Reference number is NIOS's TRUE unique id, so it decides identity FIRST.
            # (Siblings often share an email/phone but have different reference numbers —
            # they must NOT be merged.) Email is only a fallback when there is no reference.
            if ref:
                return "ref:" + ref.strip(), "Reference No"
            if _real_email(email):
                return "email:" + _norm(email), "Email"
            if enroll:
                return "enr:" + enroll.strip(), "Enrollment No"
            return "", ""

        # Existing students: keep enough to SHOW which student a new row matches, so the
        # counsellor can verify a real duplicate vs. a sibling sharing a phone/email.
        conn = get_db()
        existing_by_key, existing_by_ref = {}, {}
        existing_by_mob, existing_by_email = {}, {}
        for r in conn.execute("SELECT row_key, student_name, reference_no, enrollment_no, mobile, email "
                              "FROM student_status WHERE COALESCE(deleted,0)=0").fetchall():
            nm = r["student_name"] or ""
            rf = (r["reference_no"] or r["enrollment_no"] or "").strip()
            if r["row_key"]:
                existing_by_key[r["row_key"]] = (nm, rf)
            if r["reference_no"] and r["reference_no"].strip():
                existing_by_ref[r["reference_no"].strip()] = (nm, rf)
            mb = _valid_mobile(r["mobile"])
            if mb:
                existing_by_mob.setdefault(mb, (nm, rf))
            em = _norm(r["email"])
            if _real_email(em):
                existing_by_email.setdefault(em, (nm, rf))
        conn.close()

        students = read_students_from_excel(EXCEL_PATH)
        seen_keys, seen_ref, seen_mob, seen_email = {}, {}, {}, {}
        preview = []
        dups = 0
        sims = 0
        for s in students:
            ref = str(s.get("reference_no") or "").strip()
            enroll = str(s.get("enrollment_no") or "").strip()
            email = str(s.get("email") or "")
            em = _norm(email)
            mob = _valid_mobile(s.get("mobile"))
            nm = s.get("student_name", "")
            rk, basis = _rowkey(ref, enroll, email)

            # HARD duplicate = SAME reference (true NIOS id) or same unique key.
            # This is genuinely the same student — at import it is merged & re-checked.
            hard = False
            match = None      # (name, ref) of the student this row matches
            if ref and ref in existing_by_ref:
                hard, basis, match = True, "Reference No", existing_by_ref[ref]
            elif ref and ref in seen_ref:
                hard, basis, match = True, "Reference No", seen_ref[ref]
            elif rk and rk in existing_by_key:
                hard, match = True, existing_by_key[rk]
            elif rk and rk in seen_keys:
                hard, match = True, seen_keys[rk]

            # SOFT "similar" = a DIFFERENT student (different reference) who happens to
            # share a mobile OR an email. NOT merged — both run. Shown so the counsellor
            # can confirm (e.g. siblings) or remove a true duplicate afterwards.
            similar = ""
            if not hard:
                if mob and mob in existing_by_mob:
                    similar, match = "Mobile No", existing_by_mob[mob]
                elif mob and mob in seen_mob:
                    similar, match = "Mobile No", seen_mob[mob]
                elif em and _real_email(em) and em in existing_by_email:
                    similar, match = "Email", existing_by_email[em]
                elif em and _real_email(em) and em in seen_email:
                    similar, match = "Email", seen_email[em]

            # Record this row so later rows in the same sheet can match against it.
            if rk:
                seen_keys.setdefault(rk, (nm, ref or enroll))
            if ref:
                seen_ref.setdefault(ref, (nm, ref))
            if mob:
                seen_mob.setdefault(mob, (nm, ref or enroll))
            if em and _real_email(em):
                seen_email.setdefault(em, (nm, ref or enroll))

            if hard:
                dups += 1
            elif similar:
                sims += 1
            preview.append({
                "student_name": nm,
                "reference_no": s.get("reference_no", ""),
                "enrollment_no": s.get("enrollment_no", ""),
                "email": s.get("email", ""),
                "class_level": s.get("class_level", ""),
                "session": s.get("session", ""),
                "dob": s.get("dob", ""),
                "mobile": s.get("mobile", ""),
                "dup": hard,
                "dup_basis": basis if hard else "",
                "similar": similar,
                "match_name": (match[0] if match else ""),
                "match_ref": (match[1] if match else ""),
            })
        total = len(students)
        resp.update({"total": total, "duplicates": dups, "similar": sims,
                     "unique": total - dups - sims, "preview": preview[:2000]})
    except Exception as e:
        resp["parse_error"] = str(e)[:200]
    return resp

@app.get("/api/download-excel")
async def download_excel(user=Depends(verify_token)):
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(status_code=404, detail="Excel file not found")
    return FileResponse(EXCEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="nios_status_updated.xlsx")

@app.get("/api/sample-sheet")
async def sample_sheet(type: str = "regular", user=Depends(verify_token)):
    """Generate a ready-to-fill sample Excel so counsellors always know the format."""
    import io, openpyxl
    from openpyxl.utils import get_column_letter
    from fastapi import Response
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["STUDENT NAME", "MOBILE NO", "ALTERNATE NUMBER (optional)", "CLASS", "REFERENCE NUMBER", "Enrol No", "Email",
               "Date of Birth", "ADMISSION SESSION", "tocStatus", "ADMISSION STATUS", "REMARKS",
               "DOWNLOAD ID CARD", "DOWNLOAD APPLICATION FORM", "HALL TICKET"]
    ws.append(headers)
    if type == "syc":
        rows = [
            ["AYUSH KUMAR", "9876543210", "9988776655", "12TH", "", "220004253089", "", "19-06-2006", "SYC", ""],
            ["PETER RANA", "7428240153", "", "12TH", "", "50258253204", "peter@example.com", "17-05-2001", "SYC", ""],
        ]
    else:
        rows = [
            ["SABBA NOOR", "6205148930", "8123456789", "12TH", "D1026300062", "", "", "05-02-2010", "On Demand", "yes"],
            ["DEVRAJ JAT", "7737485139", "", "10TH", "B0926200020", "", "", "01-07-2004", "Stream 2", "no"],
            ["SANA PARWEEN", "9523534252", "9012345678", "12TH", "A1026300040", "", "sana@example.com", "28-02-2009", "April", ""],
        ]
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 2)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = "MVS_sample_syc.xlsx" if type == "syc" else "MVS_sample_regular.xlsx"
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})

@app.get("/api/sample-alt-sheet")
async def sample_alt_sheet(user=Depends(verify_token)):
    """Two-column sample for the bulk alternate-number upload."""
    import io, openpyxl
    from fastapi import Response
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Alternate Numbers"
    headers = ["Reference No", "Alternate Number"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
    for r in [["D1026300062", "9988776655"], ["B0926200020", "9012345678"], ["A1026300040", "8123456789"]]:
        ws.append(r)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="MVS_alternate_numbers_sample.xlsx"'})

def _send_alt_to_confirmed(row_keys):
    """Send the documents ONLY to the alternate number for already-confirmed students
    (their primary number already received them, so it is not messaged again)."""
    if get_setting("wa_enabled", "0") != "1":
        return
    import whatsapp
    conn = get_db(); c = conn.cursor()
    sent = 0
    for rk in row_keys:
        r = c.execute("SELECT * FROM student_status WHERE row_key=?", (rk,)).fetchone()
        if not r:
            continue
        alt = (r["alt_mobile"] if "alt_mobile" in r.keys() else "") or ""
        if not alt:
            continue
        try:
            ok, info = whatsapp.send_for_student({
                "row_key": rk, "student_name": r["student_name"] or "",
                "mobile": r["mobile"] or "", "alt_mobile": alt,
                "session": r["session"] or "", "reference_no": r["reference_no"] or "",
                "toc_status": (r["toc_status"] if "toc_status" in r.keys() else "") or "",
                "dob": r["dob"] or ""}, only_number=alt)
            if ok:
                sent += 1
                c.execute("UPDATE student_status SET whatsapp_info=? WHERE row_key=?",
                          (f"Sent to 2 numbers (own + alternate {alt})", rk))
                conn.commit()
            logger.info(f"Alt-send {rk} -> {alt}: {'ok' if ok else 'FAIL'} | {info}")
        except Exception as e:
            logger.warning(f"Alt-send error {rk}: {e}")
    conn.close()
    logger.info(f"Alternate-number send complete: {sent}/{len(row_keys)}")

@app.post("/api/upload-alt-numbers")
async def upload_alt_numbers(background_tasks: BackgroundTasks,
                             file: UploadFile = File(...), user=Depends(verify_token)):
    """Bulk add/update ALTERNATE WhatsApp numbers for students already in the system.
    Sheet needs a Reference No column and an Alternate Number column (header names are
    matched flexibly). Matches by reference (then enrollment), updates alt_mobile, and for
    students that are ALREADY confirmed, sends the documents to the new alternate number."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    import io, openpyxl
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the file: {e}")
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return {"message": "The sheet is empty.", "updated": 0, "not_found": 0, "to_send": 0}
    header = [str(h or "").strip().lower() for h in data[0]]

    def _find(cands):
        for i, h in enumerate(header):
            if any(cc in h for cc in cands):
                return i
        return -1
    ref_i = _find(["reference", "ref no", "ref_no", "refno", "ref"])
    enr_i = _find(["enrollment", "enrol"])
    alt_i = _find(["alternate", "alt", "second", "whatsapp", "other"])
    if ref_i < 0 and enr_i < 0:
        return {"message": "No 'Reference No' (or Enrollment) column found in the sheet.",
                "updated": 0, "not_found": 0, "to_send": 0}
    if alt_i < 0:
        return {"message": "No 'Alternate Number' column found in the sheet.",
                "updated": 0, "not_found": 0, "to_send": 0}

    def _digits(v):
        return "".join(ch for ch in str(v or "") if ch.isdigit())

    conn = get_db(); c = conn.cursor()
    updated, not_found, send_keys = 0, 0, []
    for row in data[1:]:
        if not row:
            continue
        ref = str(row[ref_i] or "").strip() if 0 <= ref_i < len(row) else ""
        enr = str(row[enr_i] or "").strip() if 0 <= enr_i < len(row) else ""
        alt = _digits(row[alt_i]) if alt_i < len(row) else ""
        if not alt or len(alt) < 10:
            continue
        rec = None
        if ref:
            rec = c.execute("SELECT row_key, mobile, is_confirmed FROM student_status "
                            "WHERE reference_no=? AND COALESCE(deleted,0)=0 LIMIT 1", (ref,)).fetchone()
        if rec is None and enr:
            rec = c.execute("SELECT row_key, mobile, is_confirmed FROM student_status "
                            "WHERE enrollment_no=? AND COALESCE(deleted,0)=0 LIMIT 1", (enr,)).fetchone()
        if rec is None:
            not_found += 1
            continue
        if alt == _digits(rec["mobile"]):
            continue   # alternate same as primary — skip
        c.execute("UPDATE student_status SET alt_mobile=? WHERE row_key=?", (alt, rec["row_key"]))
        updated += 1
        if rec["is_confirmed"] == 1:
            send_keys.append(rec["row_key"])
    conn.commit(); conn.close()

    wa_on = get_setting("wa_enabled", "0") == "1"
    if send_keys and wa_on:
        background_tasks.add_task(_send_alt_to_confirmed, send_keys)
    msg = f"Updated the alternate number for {updated} student(s)."
    if not_found:
        msg += f" {not_found} reference(s) were not found in the system."
    if send_keys:
        msg += (f" Sending documents to the alternate number for {len(send_keys)} confirmed student(s) in the background…"
                if wa_on else f" {len(send_keys)} are confirmed — turn ON WhatsApp in Settings to send to the alternate number.")
    return {"message": msg, "updated": updated, "not_found": not_found,
            "to_send": (len(send_keys) if wa_on else 0)}

@app.get("/api/intervals")
async def get_intervals(user=Depends(verify_token)):
    return {"ondemand_min": _interval_minutes("ondemand", 6),
            "stream2_min": _interval_minutes("stream2", 6),
            "public_min": _interval_minutes("public", 12),
            "portalnew_min": _interval_minutes("portalnew", 3),
            "ondemand_src": get_setting("runsrc_ondemand", "both") or "both",
            "stream2_src": get_setting("runsrc_stream2", "both") or "both",
            "public_src": get_setting("runsrc_public", "both") or "both"}

@app.post("/api/intervals")
async def set_intervals(body: dict, user=Depends(verify_token)):
    od = int(body.get("ondemand_min", 360))
    st = int(body.get("stream2_min", 360))
    pub = int(body.get("public_min", 720))
    pn = int(body.get("portalnew_min", 180))
    MAX_MIN = 43200   # 30 days — lets counsellors run every few days, not just hours
    for v in (od, st, pub, pn):
        if not (15 <= v <= MAX_MIN):
            raise HTTPException(status_code=400, detail="Interval must be between 15 minutes and 30 days")
    set_setting("interval_ondemand_min", od)
    set_setting("interval_stream2_min", st)
    set_setting("interval_public_min", pub)
    set_setting("interval_portalnew_min", pn)
    set_setting("interval_regular_min", "")   # retire the old combined setting
    # Per-group DATA SOURCE for auto runs: both (default) | mvs_tracker | mvs_portal.
    def _vsrc(x):
        x = str(x or "both")
        return x if x in ("both", "mvs_tracker", "mvs_portal") else "both"
    set_setting("runsrc_ondemand", _vsrc(body.get("ondemand_src")))
    set_setting("runsrc_stream2", _vsrc(body.get("stream2_src")))
    set_setting("runsrc_public", _vsrc(body.get("public_src")))
    # Saving new intervals restarts the timers fresh, so clear any paused state.
    for grp in ("ondemand", "stream2", "public", "portalnew"):
        set_setting(f"paused_{grp}", "")
        set_setting(f"pause_remaining_{grp}_sec", "")
    reschedule_jobs()
    def _fmt(m):
        if m % 1440 == 0:
            d = m // 1440
            return f"{d} day" + ("s" if d != 1 else "")
        if m % 60 == 0:
            return f"{m//60}h"
        if m < 60:
            return f"{m}m"
        return f"{m//60}h {m%60}m"
    return {"message": f"On Demand: every {_fmt(od)}, Stream 2: every {_fmt(st)}, Public: every {_fmt(pub)}",
            "ondemand_min": od, "stream2_min": st, "public_min": pub}

@app.get("/api/source-counts")
def source_counts(user=Depends(verify_token)):
    """How many active students are tagged MVS Portal vs MVS Tracker (for the move tool)."""
    conn = get_db()
    rows = conn.execute("SELECT COALESCE(NULLIF(source,''),'mvs_tracker') AS src, COUNT(*) AS n "
                        "FROM student_status WHERE COALESCE(deleted,0)=0 GROUP BY src").fetchall()
    conn.close()
    out = {"mvs_portal": 0, "mvs_tracker": 0}
    for r in rows:
        key = "mvs_portal" if r["src"] == "mvs_portal" else "mvs_tracker"
        out[key] += r["n"]
    return out

@app.post("/api/change-source-bulk")
def change_source_bulk(body: dict, user=Depends(verify_token)):
    """Move EVERY active student from one data type to another (e.g. a sheet that was
    auto-detected as MVS Portal by mistake but should be MVS Tracker). No student is
    deleted — only the 'source' tag changes, so nothing is lost."""
    frm = body.get("from_source", "")
    to = body.get("to_source", "")
    if frm not in ("mvs_portal", "mvs_tracker") or to not in ("mvs_portal", "mvs_tracker") or frm == to:
        raise HTTPException(status_code=400, detail="pick two different data types")
    conn = get_db()
    cur = conn.execute("UPDATE student_status SET source=? "
                       "WHERE COALESCE(deleted,0)=0 AND COALESCE(NULLIF(source,''),'mvs_tracker')=?",
                       (to, frm))
    conn.commit()
    n = cur.rowcount
    conn.close()
    return {"ok": True, "moved": n, "from": frm, "to": to}

@app.post("/api/change-source-selected")
def change_source_selected(body: dict, user=Depends(verify_token)):
    """Move ONLY the selected students to a data type (precise control from the list)."""
    keys = [k for k in (body.get("row_keys", []) or []) if k][:5000]
    to = body.get("to_source", "")
    if to not in ("mvs_portal", "mvs_tracker"):
        raise HTTPException(status_code=400, detail="invalid data type")
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    conn = get_db()
    ph = ",".join("?" * len(keys))
    cur = conn.execute(f"UPDATE student_status SET source=? WHERE row_key IN ({ph})", [to] + keys)
    conn.commit()
    n = cur.rowcount
    conn.close()
    return {"ok": True, "moved": n, "to": to}

@app.get("/api/source-override")
async def get_source_override(user=Depends(verify_token)):
    return {"value": get_setting("source_override", "")}

@app.post("/api/source-override")
async def set_source_override(value: str = "", user=Depends(verify_token)):
    """Force the data type of the NEXT run's sheet. Empty = auto-detect."""
    if value not in ("", "mvs_portal", "mvs_tracker"):
        raise HTTPException(status_code=400, detail="invalid value")
    set_setting("source_override", value)
    return {"ok": True, "value": value}

@app.get("/api/debug-login")
def debug_login_endpoint(ref: str, dob: str, action: str = "", user=Depends(verify_token)):
    """Phase 2 debug: login with a confirmed student & discover download links."""
    try:
        from nios_login import debug_login
        result = debug_login(ref, dob, action or None)
        return result
    except Exception as e:
        return {"error": str(e)}

def build_report_excel(path):
    """Build the run-report Excel: a Summary sheet plus Confirmed / Document Required /
    Error & Unknown student lists (current actionable state). Returns the counts."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    conn = get_db()
    ND = "COALESCE(deleted,0)=0"
    NF = "COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0"
    cols = ("student_name, reference_no, enrollment_no, dob, session, mobile, email, "
            "current_status, COALESCE(NULLIF(login_remark,''),NULLIF(remark,'')) as remark, last_checked")
    def rows(where):
        return conn.execute(f"SELECT {cols} FROM student_status WHERE {ND} AND {where} "
                            f"ORDER BY student_name").fetchall()
    confirmed = rows(f"is_confirmed=1 AND {NF}")
    # "Confirmed Today" = confirmations since the start of today — BUT if the latest run
    # CROSSED MIDNIGHT (started yesterday ~11 PM, finished after 12 AM), count from that
    # run's start instead, so the whole run's confirmations stay together and aren't split
    # by the date change. Same-day morning/evening runs are unaffected (cutoff = midnight).
    now = datetime.now()
    midnight = now.strftime("%Y-%m-%d") + " 00:00:00"
    cutoff = midnight
    try:
        lr = conn.execute("SELECT run_at FROM run_logs ORDER BY id DESC LIMIT 1").fetchone()
        if lr and lr["run_at"]:
            ra = str(lr["run_at"])
            recent = (now - timedelta(hours=12)).strftime("%Y-%m-%d %H:%M:%S")
            if ra < midnight and ra >= recent:
                cutoff = ra   # run began before midnight & is recent -> include from run start
    except Exception:
        pass
    confirmed_today = rows(f"is_confirmed=1 AND {NF} AND last_changed >= '{cutoff}'")
    required = rows(f"current_status='Document Required' AND {NF}")
    verified = rows(f"current_status='Verified' AND {NF}")
    docsprog = rows(f"current_status='Documents Verification In Progress' AND {NF}")
    error = rows("(current_status IN ('Unknown','Fetch Error') "
                 "OR COALESCE(check_failed,0)=1 OR COALESCE(login_failed,0)=1)")
    # Session-wise breakdown (same definitions as the dashboard) for the Summary sheet.
    sess_raw = conn.execute(
        f"SELECT session, COUNT(*) as cnt, "
        f"SUM(CASE WHEN is_confirmed=1 THEN 1 ELSE 0 END) as confirmed, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND current_status='Verified' THEN 1 ELSE 0 END) as verified, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND current_status='Documents Verification In Progress' THEN 1 ELSE 0 END) as docsv, "
        f"SUM(CASE WHEN current_status='Document Required' THEN 1 ELSE 0 END) as req, "
        f"SUM(CASE WHEN COALESCE(is_confirmed,0)=0 AND COALESCE(current_status,'')!='SYC' "
        f"         AND (session IS NULL OR session NOT LIKE '%syc%') THEN 1 ELSE 0 END) as active "
        f"FROM student_status WHERE {ND} AND {NF} AND session != '' GROUP BY session"
    ).fetchall()
    conn.close()

    sess_agg = {}
    for r in sess_raw:
        norm = normalize_session(r["session"])
        if not norm:
            continue
        d = sess_agg.setdefault(norm, {"total": 0, "confirmed": 0, "verified": 0,
                                       "docsv": 0, "active": 0, "req": 0})
        d["total"] += r["cnt"] or 0
        for k in ("confirmed", "verified", "docsv", "active", "req"):
            d[k] += r[k] or 0
    sess_list = sorted(sess_agg.items(), key=lambda kv: kv[1]["total"], reverse=True)

    wb = Workbook()
    hfill = PatternFill("solid", fgColor="4F46E5"); hfont = Font(bold=True, color="FFFFFF")
    ws = wb.active; ws.title = "Summary"
    ws.append(["NIOS Status Tracker — Run Report"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Generated", datetime.now().strftime("%d %b %Y, %I:%M %p")])
    ws.append([])
    ws.append(["Category", "Count"])
    for c in ("A4", "B4"):
        ws[c].fill = hfill; ws[c].font = hfont
    ws.append(["Confirmed Today", len(confirmed_today)])
    ws.append(["Confirmed (Total)", len(confirmed)])
    ws.append(["Admission Verified", len(verified)])
    ws.append(["Documents Verification In Progress", len(docsprog)])
    ws.append(["Document Required", len(required)])
    ws.append(["Error / Unknown", len(error)])

    # Session-wise breakdown table
    ws.append([])
    ws.append(["Session-wise Breakdown"]); ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
    shdr = ["Session", "Total", "Confirmed", "Verified", "Docs Verification", "Active", "Required"]
    ws.append(shdr)
    hr = ws.max_row
    for ci in range(1, len(shdr) + 1):
        cell = ws.cell(row=hr, column=ci); cell.fill = hfill; cell.font = hfont
    for name, d in sess_list:
        ws.append([name, d["total"], d["confirmed"], d["verified"], d["docsv"], d["active"], d["req"]])
    for col, w in zip("ABCDEFG", [34, 9, 11, 10, 17, 9, 10]):
        ws.column_dimensions[col].width = w

    def sheet(title, data):
        s = wb.create_sheet(title)
        hdr = ["Name", "Reference No", "Enrollment No", "DOB", "Session", "Mobile", "Email",
               "Status", "Remark", "Last Checked"]
        s.append(hdr)
        for i in range(1, len(hdr) + 1):
            cell = s.cell(row=1, column=i); cell.fill = hfill; cell.font = hfont
        for r in data:
            s.append([r["student_name"], r["reference_no"], r["enrollment_no"], r["dob"], r["session"],
                      r["mobile"], r["email"], r["current_status"], r["remark"], r["last_checked"]])
        widths = [22, 15, 16, 13, 15, 13, 24, 20, 30, 18]
        for i, w in enumerate(widths, 1):
            s.column_dimensions[chr(64 + i)].width = w
        s.freeze_panes = "A2"
    sheet("Confirmed Today", confirmed_today)
    sheet("Confirmed (Total)", confirmed)
    sheet("Admission Verified", verified)
    sheet("Document Required", required)
    sheet("Error & Unknown", error)
    wb.save(path)
    return {"confirmed": len(confirmed), "confirmed_today": len(confirmed_today),
            "verified": len(verified), "docs_progress": len(docsprog),
            "required": len(required), "error": len(error)}

@app.get("/report-excel/{token}")
async def report_excel(token: str):
    """Serve the run-report Excel for a signed link (admins open it straight from
    WhatsApp — no login). The signature is the security; data reflects the latest state."""
    from links import verify_report_token
    from fastapi.responses import FileResponse
    day = verify_report_token(token)
    if not day:
        raise HTTPException(status_code=404, detail="Invalid or expired report link")
    import tempfile
    path = os.path.join(tempfile.gettempdir(), f"nios_report_{day}.xlsx")
    build_report_excel(path)
    return FileResponse(path, filename=f"NIOS_Report_{day}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def whatsapp_webhook_token():
    """Unguessable, stable token for the AiSensy delivery webhook URL."""
    import hmac as _h, hashlib as _hh
    return _h.new(SECRET_KEY.encode(), b"aisensy-delivery-webhook", _hh.sha256).hexdigest()[:20]

@app.get("/api/webhook-url")
async def get_webhook_url(user=Depends(verify_token)):
    base = os.environ.get("PUBLIC_BASE_URL", "https://status.mvsfoundation.in").rstrip("/")
    return {"url": f"{base}/webhook/whatsapp/{whatsapp_webhook_token()}"}

@app.post("/webhook/whatsapp/{token}")
async def whatsapp_delivery_webhook(token: str, request):
    """Receives WhatsApp delivery status from AiSensy and records whether a confirmed
    student's documents were actually DELIVERED (not just accepted by the gateway).
    Parsed defensively so it works across AiSensy payload shapes. Matched by mobile."""
    from fastapi import Request as _Req  # noqa
    if token != whatsapp_webhook_token():
        raise HTTPException(status_code=404, detail="not found")
    try:
        raw = await request.body()
        text = raw.decode("utf-8", "ignore").lower()
    except Exception:
        text = ""
    if not text:
        return {"ok": True, "matched": 0}
    # Map the event to a delivery state (ignore plain "sent"/"accepted" — already known)
    if "undelivered" in text or "failed" in text:
        state = "failed"
    elif "read" in text or "delivered" in text:
        state = "delivered"
    else:
        return {"ok": True, "matched": 0}
    import re as _re
    nums = _re.findall(r"\d{10,15}", text)
    phone10 = nums[0][-10:] if nums else ""
    if not phone10:
        return {"ok": True, "matched": 0}
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    matched = 0
    # Remember that the AiSensy webhook is LIVE — the auto-redeliver sweep only trusts
    # "no delivery event = genuinely undelivered" once we know events actually flow.
    try:
        from database import set_setting
        set_setting("wa_webhook_seen", "1")
    except Exception:
        pass
    for r in conn.execute("SELECT row_key, mobile FROM student_status WHERE whatsapp_sent=1").fetchall():
        m = "".join(ch for ch in (r["mobile"] or "") if ch.isdigit())
        if m and m[-10:] == phone10:
            conn.execute("UPDATE student_status SET whatsapp_delivery=?, whatsapp_delivery_at=? WHERE row_key=?",
                         (state, now, r["row_key"]))
            matched += 1
    conn.commit(); conn.close()
    return {"ok": True, "matched": matched, "state": state}

@app.get("/api/download-doc")
def download_doc(ref: str, dob: str, kind: str, user=Depends(verify_token)):
    """Login as the student and return their document (PDF or print-ready HTML)."""
    from fastapi import Response
    from nios_login import fetch_document
    # Public-cycle students (April / October / 'apr-27') normally only have an ID Card, but a
    # Public student who took TOC (tocStatus = 'yes') also gets the Application Form
    # (Registration Summary). Block anything else so a stale/wrong link can never open the wrong
    # file — matches the student-facing links. (Admins still see every document for On Demand / Stream 2.)
    if kind != "id_card" and ref:
        try:
            conn = get_db()
            srow = conn.execute("SELECT session, toc_status FROM student_status WHERE COALESCE(deleted,0)=0 "
                                "AND reference_no=? LIMIT 1", (ref,)).fetchone()
            conn.close()
            sess = ((srow["session"] if srow else "") or "").lower()
            is_s2 = ("stream 2" in sess or "stream2" in sess or "stream-2" in sess)
            is_od = ("on demand" in sess or "ondemand" in sess or "on-demand" in sess or "odes" in sess)
            toc = ((srow["toc_status"] if srow else "") or "").strip().lower()
            if srow is not None and not is_s2 and not is_od:   # public / unknown
                # allow the Application Form only for Public TOC-yes students
                if not (kind == "app_form" and toc == "yes"):
                    raise HTTPException(status_code=404,
                        detail="This document is not available for this Public (April / October) student. Only the ID Card (and the Application Form for TOC-based admissions) can be opened.")
        except HTTPException:
            raise
        except Exception:
            pass
    # Serve OUR saved copy first (no NIOS/CapSolver). Resolve this student's row_key by reference.
    _rk = None
    try:
        conn = get_db()
        _rr = conn.execute("SELECT row_key FROM student_status WHERE COALESCE(deleted,0)=0 "
                           "AND reference_no=? LIMIT 1", (ref,)).fetchone()
        conn.close()
        _rk = _rr["row_key"] if _rr else None
    except Exception:
        _rk = None
    if _rk:
        _cached = load_doc_cache(_rk, kind)
        if _cached is not None:
            content, ctype, filename = _cached
            disp = f'attachment; filename="{filename}"' if "pdf" in (ctype or "") else "inline"
            return Response(content=content, media_type=ctype, headers={"Content-Disposition": disp})
    content, ctype, filename = fetch_document(ref, dob, kind)
    if content is None:
        # Login/bounce failure -> flag this student as Failed to Run so it surfaces in
        # the sidebar (and any pending WhatsApp is blocked) even if it was confirmed
        # earlier. Matched by reference + DOB.
        err = ctype or "NIOS login failed"
        low = err.lower()
        # A captcha/service failure is NOT a data problem — never move the student to
        # 'Failed to Run' for it (that would wrongly pull confirmed students out when the
        # captcha gateway is just busy / out of balance). Only genuine data mismatches flag.
        if err.startswith("CAPTCHA_BUSY"):
            raise HTTPException(status_code=503,
                detail="NIOS captcha/login service is busy right now (not a data problem). Please try again in a minute.")
        if ("login" in low or "rejected" in low or "dob" in low) and ref:
            try:
                conn = get_db()
                conn.execute(
                    "UPDATE student_status SET login_failed=1, login_remark=?, "
                    "whatsapp_info=?, whatsapp_sent=0 "
                    "WHERE COALESCE(deleted,0)=0 AND reference_no=? AND COALESCE(dob,'')=?",
                    (err[:240], ("Not sent — " + err)[:180], ref, dob))
                conn.commit(); conn.close()
            except Exception:
                pass
        raise HTTPException(status_code=404, detail=err)
    # Save our own copy for next time (so future opens skip NIOS/CapSolver).
    if _rk:
        try:
            save_doc_cache(_rk, kind, content, ctype, filename)
        except Exception:
            pass
    # PDF -> attachment download; HTML -> inline (open in tab to print)
    if "pdf" in ctype:
        disp = f'attachment; filename="{filename}"'
    else:
        disp = "inline"
    return Response(content=content, media_type=ctype, headers={"Content-Disposition": disp})

@app.get("/api/diagnose-login")
def diagnose_login_ep(ref: str = "", dob: str = "", enr: str = "", user=Depends(verify_token)):
    """Live NIOS-login diagnostic — run this when logins are failing to see the REAL cause
    (site-key change, captcha token, or NIOS rejection). Use a known-good student's ref + DOB."""
    from nios_login import diagnose_login
    try:
        return diagnose_login((ref or "").strip(), (dob or "").strip(), (enr or "").strip())
    except Exception as e:
        return {"error": str(e)[:200]}

@app.get("/api/nios-reach")
def nios_reach_ep(ref: str = "", dob: str = "", user=Depends(verify_token)):
    """Diagnose NIOS reachability from the SERVER. Checks (1) can the server load the real NIOS
    login page, and (2) if ref+dob given, whether a full login (captcha solve + POST) actually
    succeeds — pinpointing captcha-token vs score/bounce failures."""
    import requests as _rq
    out = {"endpoint_version": "v3", "page_status": "", "page_len": 0,
           "looks_blocked": None, "csrf_found": False, "page_snippet": "",
           "built_in_sitekey": "", "live_sitekey": "", "recaptcha_action": "",
           "form_fields": "", "recaptcha_fields": "", "recaptcha_type": "", "action_from_js": "", "dob_format_test": "",
           "proxy_set": "", "proxy_direct_test": "", "capsolver_create": "", "capsolver_result": "", "verify_login_test": "",
           "captcha_token": "", "login_result": "(no ref/dob — login test skipped)",
           "final_url": "", "snippet": ""}
    try:
        import nios_login as nl
        out["built_in_sitekey"] = getattr(nl, "RECAPTCHA_SITE_KEY", "") or ""
        LOGIN_URL = getattr(nl, "LOGIN_URL", "https://sdmis.nios.ac.in/auth/other-login")
        HEADERS = getattr(nl, "HEADERS", {"User-Agent": "Mozilla/5.0"})
    except Exception as e:
        out["login_result"] = f"module error: {e}"
        return out
    # THE key check: raw fetch of the NIOS login page from the server.
    try:
        rr = _rq.get(LOGIN_URL, headers=HEADERS, timeout=12)
        body = rr.text or ""
        out["page_status"] = rr.status_code
        out["page_len"] = len(body)
        out["final_url"] = str(getattr(rr, "url", ""))
        out["page_snippet"] = body[:400]
        low = body.lower()
        out["csrf_found"] = ('name="_csrf"' in body or "'_csrf'" in body or "_csrf" in low)
        try:
            out["live_sitekey"] = nl._extract_sitekey(body) or ""
        except Exception:
            out["live_sitekey"] = ""
        try:
            out["recaptcha_action"] = nl._extract_action(body) or "(none found on page)"
        except Exception:
            out["recaptcha_action"] = ""
        out["looks_blocked"] = any(x in low for x in [
            "cloudflare", "just a moment", "attention required", "access denied",
            "forbidden", "cf-chl", "captcha-delivery", "are you a human", "ddos",
            "request blocked", "not allowed", "access to this page has been denied"])
        # Extract the CURRENT login-form field names — if NIOS renamed fields, our POST is
        # ignored and the login silently bounces. This shows exactly what the form now wants.
        try:
            from bs4 import BeautifulSoup as _BS
            fsoup = _BS(body, "html.parser")
            names = []
            for inp in fsoup.find_all(["input", "textarea", "select"]):
                nm = inp.get("name") or inp.get("id") or ""
                if nm and nm not in names:
                    names.append(nm)
            out["form_fields"] = ", ".join(names[:40]) or "(no form inputs found)"
            import re as _re
            rc = _re.findall(r'(g-recaptcha-response|google_recapcha_response|recaptcha[\w\-]*|LoginForm\[[\w_]+\])', body, _re.I)
            out["recaptcha_fields"] = ", ".join(sorted(set(rc))[:15]) or "(none)"
        except Exception as e:
            out["form_fields"] = f"parse error: {e}"
        # Detect reCAPTCHA type + find the action from the page's own JS (it isn't in the HTML).
        try:
            import re as _re
            out["recaptcha_type"] = ("ENTERPRISE" if ("recaptcha/enterprise" in low or "enterprise.js" in low)
                                     else ("standard-v3" if ("recaptcha/api.js" in low or "grecaptcha" in low
                                                             or "www.google.com/recaptcha" in low)
                                           else "not in HTML (loaded by JS)"))
            scripts = _re.findall(r'<script[^>]+src=["\']([^"\']+)["\']', body)
            js_blob = ""
            for su in scripts[:10]:
                u = su
                if u.startswith("//"): u = "https:" + u
                elif u.startswith("/"): u = "https://sdmis.nios.ac.in" + u
                elif not u.startswith("http"): u = "https://sdmis.nios.ac.in/" + u
                if "sdmis.nios.ac.in" in u:
                    try:
                        jr = _rq.get(u, headers=HEADERS, timeout=8)
                        js_blob += (jr.text or "")[:30000]
                    except Exception:
                        pass
            jl = js_blob.lower()
            if "enterprise" in jl and out["recaptcha_type"] != "ENTERPRISE":
                out["recaptcha_type"] = "ENTERPRISE (found in JS)"
            em = _re.search(r'execute\(\s*["\'][\w\-]{20,}["\']\s*,\s*\{\s*action\s*:\s*["\']([\w\-/]+)', js_blob)
            am = em or _re.search(r'action\s*[:=]\s*["\']([\w\-/]{2,40})["\']', js_blob)
            out["action_from_js"] = (am.group(1) if am else "(not found in JS)")
        except Exception as e:
            out["recaptcha_type"] = f"js scan error: {e}"
    except Exception as e:
        out["page_status"] = f"FETCH ERROR: {type(e).__name__}: {str(e)[:150]}"
    # Raw CapSolver test — reveals WHY no token (proxy error code, etc.)
    try:
        import time as _time
        proxy = os.environ.get("CAPSOLVER_PROXY", "").strip()
        out["proxy_set"] = ("yes (" + proxy.split(":")[0] + ":" + (proxy.split(":")[1] if ":" in proxy else "?") + ":***)") if proxy else "NO (proxyless)"
        # 1) DIRECT proxy test — does the proxy itself work (creds OK)?
        pf = getattr(nl, "_parse_proxy_fields", lambda x: None)(proxy)
        if pf and pf.get("proxyLogin"):
            purl = f"{pf['proxyType']}://{pf['proxyLogin']}:{pf['proxyPassword']}@{pf['proxyAddress']}:{pf['proxyPort']}"
            try:
                pr = _rq.get("https://ipv4.icanhazip.com", proxies={"http": purl, "https": purl}, timeout=25)
                out["proxy_direct_test"] = f"OK — exit IP = {pr.text.strip()[:40]}"
            except Exception as pe:
                out["proxy_direct_test"] = f"FAILED: {type(pe).__name__}: {str(pe)[:110]}"
        else:
            out["proxy_direct_test"] = "(no proxy creds parsed)"
        # 2) CapSolver test with separate proxy fields
        capkey = os.environ.get("CAPTCHA_API_KEY", "")
        skey = out.get("live_sitekey") or getattr(nl, "RECAPTCHA_SITE_KEY", "")
        ctask = {"type": ("ReCaptchaV3Task" if pf else "ReCaptchaV3TaskProxyLess"),
                 "websiteURL": LOGIN_URL, "websiteKey": skey, "pageAction": "login", "minScore": 0.9}
        if pf:
            ctask.update(pf)
        cr = _rq.post("https://api.capsolver.com/createTask",
                      json={"clientKey": capkey, "task": ctask}, timeout=30).json()
        out["capsolver_create"] = f"errId={cr.get('errorId')} code={cr.get('errorCode','')} {str(cr.get('errorDescription',''))[:120]}"
        tid = cr.get("taskId")
        if tid:
            for _ in range(25):
                _time.sleep(2)
                gr = _rq.post("https://api.capsolver.com/getTaskResult",
                              json={"clientKey": capkey, "taskId": tid}, timeout=30).json()
                if gr.get("errorId") not in (0, None):
                    out["capsolver_result"] = f"code={gr.get('errorCode','')} {str(gr.get('errorDescription',''))[:120]}"
                    break
                if gr.get("status") == "ready":
                    out["capsolver_result"] = "TOKEN OK (proxy works!)"
                    break
            else:
                out["capsolver_result"] = "timeout (proxy too slow?)"
    except Exception as e:
        out["capsolver_create"] = f"err {type(e).__name__}: {str(e)[:120]}"
    # Full login test (only if a real ref+dob is provided) — this reveals the actual failure:
    #  captcha token not obtained  -> CapSolver/key issue
    #  BOUNCED back to login       -> captcha score too low OR wrong data
    #  LOGGED IN OK                -> login works
    if ref and dob and not str(out.get("page_status")).startswith("FETCH"):
        try:
            import re as _re3
            base = nl.format_dob(dob)                      # DD-MM-YYYY
            m = _re3.match(r'(\d{2})-(\d{2})-(\d{4})', base or "")
            cands = [base] if base else []
            if m:
                d, mo, y = m.groups()
                cands += [f"{y}-{mo}-{d}", f"{d}/{mo}/{y}", f"{mo}-{d}-{y}"]   # try 4 formats total
            results = []
            logged_in_fmt = ""
            for fmt in cands[:4]:
                try:
                    s = _rq.Session()
                    csrf = nl.get_login_csrf(s)
                    try:
                        tok = nl.solve_recaptcha_v3(LOGIN_URL, site_key=(out.get("live_sitekey") or None))
                    except TypeError:
                        tok = nl.solve_recaptcha_v3(LOGIN_URL)
                    if not out.get("captcha_token"):
                        out["captcha_token"] = ("obtained (" + str(len(tok or "")) + " chars)") if tok else "NOT obtained"
                    payload = {
                        "_csrf": csrf,
                        "LoginForm[reference_no]": ref,
                        "LoginForm[application_no]": "",
                        "LoginForm[date_of_birth]": fmt,
                        "LoginForm[google_recaptcha_response]": tok,
                        "LoginForm[google_recapcha_response]": tok,
                        "LoginForm[rememberMe]": "0",
                        "login-button": "",
                    }
                    rr2 = s.post(LOGIN_URL, data=payload,
                                 headers={**HEADERS, "Content-Type": "application/x-www-form-urlencoded",
                                          "Origin": "https://sdmis.nios.ac.in"}, timeout=35)
                    ok = nl.is_logged_in(rr2.text)
                    results.append(f"{fmt} = {'LOGGED IN' if ok else 'bounce'}")
                    if ok:
                        logged_in_fmt = fmt
                        break
                except Exception as e:
                    results.append(f"{fmt} = err({str(e)[:30]})")
            out["dob_format_test"] = " | ".join(results)
            out["login_result"] = (f"LOGGED IN OK with DOB format '{logged_in_fmt}'" if logged_in_fmt
                                   else "ALL formats BOUNCED -> likely captcha score (needs proxy)")
            # Also test the REAL path the run uses (login + fetch a protected doc page).
            try:
                okv, msgv = nl.verify_login(ref, nl.format_dob(dob))
                out["verify_login_test"] = "OK (run's login path works — safe to re-run)" if okv else f"FAILED: {str(msgv)[:160]}"
            except Exception as ev:
                out["verify_login_test"] = f"error: {type(ev).__name__}: {str(ev)[:120]}"
        except Exception as e:
            out["login_result"] = f"login error: {type(e).__name__}: {str(e)[:150]}"
    return out

@app.get("/api/syc")
def get_syc(page: int = 1, per_page: int = 20, search: str = "", user=Depends(verify_token)):
    """List SYC students (session contains SYC). No status check is done for these."""
    conn = get_db()
    where = ("WHERE COALESCE(deleted,0)=0 AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0 "
             "AND (session LIKE '%syc%' OR current_status='SYC')")
    params = []
    if search:
        where += " AND (student_name LIKE ? OR enrollment_no LIKE ? OR mobile LIKE ? OR reference_no LIKE ?)"
        like = f"%{search}%"
        params += [like, like, like, like]
    total = conn.execute(f"SELECT COUNT(*) c FROM student_status {where}", params).fetchone()["c"]
    rows = conn.execute(
        f"SELECT * FROM student_status {where} ORDER BY student_name LIMIT ? OFFSET ?",
        params + [per_page, (page - 1) * per_page]).fetchall()
    conn.close()
    return {"students": [dict(r) for r in rows], "total": total, "page": page,
            "pages": max(1, (total + per_page - 1) // per_page)}

@app.get("/api/syc-doc")
async def syc_doc(row_key: str, kind: str = "hall_ticket", user=Depends(verify_token)):
    """Fetch a SYC student's document (enrollment-login aware) for the counsellor."""
    res, err = _fetch_doc_for(row_key, kind)
    if err:
        raise HTTPException(status_code=404, detail=err)
    return _serve_doc(*res)

@app.post("/api/syc-delete")
def syc_delete(row_key: str, user=Depends(verify_token)):
    """Delete a SINGLE SYC student only. Guarded so it can never touch a
    non-SYC (On Demand / Stream 2 / Public) student."""
    conn = get_db()
    row = conn.execute("SELECT session, current_status FROM student_status WHERE row_key=?",
                       (row_key,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")
    is_syc = ("syc" in (row["session"] or "").lower()) or (row["current_status"] == "SYC")
    if not is_syc:
        conn.close()
        raise HTTPException(status_code=400, detail="Not a SYC student — refused")
    # Soft-delete (same as other students) so it can be restored from Settings → Trash.
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE student_status SET deleted=1, deleted_at=? WHERE row_key=?", (now_s, row_key))
    conn.commit()
    conn.close()
    return {"ok": True, "soft": True}

@app.post("/api/student-delete")
def student_delete(row_key: str, user=Depends(verify_token)):
    """SOFT-delete a student: hide it from the portal but keep it in Trash so it can
    be restored from Settings if removed by mistake. Soft-deleted students are also
    skipped by status runs."""
    conn = get_db()
    row = conn.execute("SELECT row_key, student_name FROM student_status WHERE row_key=?",
                       (row_key,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")
    name = row["student_name"] or ""
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE student_status SET deleted=1, deleted_at=? WHERE row_key=?", (now_s, row_key))
    conn.commit()
    conn.close()
    return {"ok": True, "name": name, "soft": True}

@app.post("/api/students-delete-bulk")
def students_delete_bulk(body: dict, user=Depends(verify_token)):
    """SOFT-delete MANY students at once (move to Trash). Restorable from Settings."""
    keys = [k for k in (body.get("row_keys", []) or []) if k][:1000]
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    ph = ",".join("?" * len(keys))
    cur = conn.execute(f"UPDATE student_status SET deleted=1, deleted_at=? "
                       f"WHERE row_key IN ({ph}) AND COALESCE(deleted,0)=0", [now_s] + keys)
    conn.commit()
    n = cur.rowcount
    conn.close()
    return {"ok": True, "deleted": n}

@app.get("/api/deleted-students")
def deleted_students(user=Depends(verify_token)):
    """List soft-deleted students (the Trash) for the Settings restore panel."""
    conn = get_db()
    rows = conn.execute(
        "SELECT row_key, student_name, reference_no, enrollment_no, session, "
        "current_status, deleted_at, COALESCE(source,'mvs_tracker') AS source "
        "FROM student_status WHERE COALESCE(deleted,0)=1 ORDER BY deleted_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/student-restore")
def student_restore(row_key: str, user=Depends(verify_token)):
    """Restore a soft-deleted student back to the portal."""
    conn = get_db()
    conn.execute("UPDATE student_status SET deleted=0, deleted_at=NULL WHERE row_key=?", (row_key,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/student-purge")
def student_purge(row_key: str, user=Depends(verify_token)):
    """Permanently delete a student from Trash (cannot be undone)."""
    conn = get_db()
    conn.execute("DELETE FROM student_status WHERE row_key=?", (row_key,))
    conn.execute("DELETE FROM short_links WHERE row_key=?", (row_key,))
    try:
        conn.execute("DELETE FROM status_history WHERE row_key=?", (row_key,))
    except Exception:
        pass
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/students-restore-bulk")
def students_restore_bulk(body: dict, user=Depends(verify_token)):
    """Restore MANY soft-deleted students from Trash at once."""
    keys = [k for k in (body.get("row_keys", []) or []) if k][:2000]
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    conn = get_db()
    ph = ",".join("?" * len(keys))
    cur = conn.execute(f"UPDATE student_status SET deleted=0, deleted_at=NULL WHERE row_key IN ({ph})", keys)
    conn.commit()
    n = cur.rowcount
    conn.close()
    return {"ok": True, "restored": n}

@app.post("/api/students-purge-bulk")
def students_purge_bulk(body: dict, user=Depends(verify_token)):
    """Permanently delete MANY students from Trash at once (cannot be undone)."""
    keys = [k for k in (body.get("row_keys", []) or []) if k][:2000]
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    conn = get_db()
    ph = ",".join("?" * len(keys))
    cur = conn.execute(f"DELETE FROM student_status WHERE row_key IN ({ph})", keys)
    n = cur.rowcount
    try:
        conn.execute(f"DELETE FROM short_links WHERE row_key IN ({ph})", keys)
    except Exception:
        pass
    try:
        conn.execute(f"DELETE FROM status_history WHERE row_key IN ({ph})", keys)
    except Exception:
        pass
    conn.commit()
    conn.close()
    return {"ok": True, "purged": n}

@app.get("/api/student-get")
def student_get(row_key: str, user=Depends(verify_token)):
    """Fetch ONE student's editable details (for the Edit modal)."""
    conn = get_db()
    r = conn.execute("SELECT row_key, student_name, mobile, alt_mobile, email, dob, reference_no, "
                     "enrollment_no, class_level, session, current_status, "
                     "COALESCE(login_failed,0) AS login_failed, "
                     "COALESCE(check_failed,0) AS check_failed, login_remark, "
                     "COALESCE(edit_sync,'') AS edit_sync, COALESCE(student_id,'') AS student_id "
                     "FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    conn.close()
    if not r:
        raise HTTPException(status_code=404, detail="Student not found")
    return dict(r)

@app.post("/api/student-edit")
def student_edit(body: dict, user=Depends(verify_token)):
    """Edit a student's uploaded details (DOB, Reference/Enrollment No, email, mobile,
    name, session) WITHOUT re-uploading the sheet. Resets the login-failed flag and the
    WhatsApp-sent flag so a re-run can verify the fix and send fresh."""
    row_key = (body.get("row_key") or "").strip()
    if not row_key:
        raise HTTPException(status_code=400, detail="row_key required")
    conn = get_db()
    row = conn.execute("SELECT row_key FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Student not found")
    allowed = ["student_name", "mobile", "alt_mobile", "email", "dob", "reference_no",
               "enrollment_no", "class_level", "session"]
    sets, params = [], []
    for f in allowed:
        if f in body:
            sets.append(f"{f}=?"); params.append((body.get(f) or "").strip())
    if not sets:
        conn.close()
        raise HTTPException(status_code=400, detail="No fields to update")
    # Editing the data invalidates any previous login result / sent flag.
    sets += ["login_failed=0", "login_remark=''", "check_failed=0", "whatsapp_sent=0"]
    params.append(row_key)
    conn.execute(f"UPDATE student_status SET {', '.join(sets)} WHERE row_key=?", params)
    conn.commit()

    # ── MIRROR THE EDIT ON THE MVS PORTAL ──
    # Any detail changed on the tracker (DOB, reference/enrollment no, mobile, name…) is
    # pushed to the Portal immediately, so the two systems NEVER hold different data for
    # the same student. The result is stored in edit_sync so the UI can show whether the
    # Portal actually took it.
    full = conn.execute("SELECT student_name, mobile, alt_mobile, email, dob, reference_no, "
                        "enrollment_no, class_level, session, COALESCE(student_id,'') AS student_id "
                        "FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    portal_ok, portal_msg = False, ""
    try:
        import mvs_sync
        if not mvs_sync.enabled():
            portal_msg = "Portal sync is OFF"
        elif not (full and full["student_id"]):
            portal_msg = "not linked to a Portal student (no studentId)"
        else:
            portal_ok, portal_msg = mvs_sync.push_details(full["student_id"], dict(full))
    except Exception as e:
        portal_msg = str(e)
    from datetime import datetime as _dt
    _ts = _dt.now().strftime("%Y-%m-%d %H:%M")
    _mark = (f"synced {_ts}" if portal_ok else f"failed {_ts}: {portal_msg}"[:300])
    conn.execute("UPDATE student_status SET edit_sync=? WHERE row_key=?", (_mark, row_key))
    conn.commit()
    conn.close()
    return {"ok": True, "portal_synced": portal_ok, "portal_message": portal_msg}

@app.post("/api/student-recheck")
def student_recheck(row_key: str, background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Re-run ONE student (after an edit). Runs in the background; the UI polls the
    student row to see the new status / login result."""
    from job_runner import recheck_one
    conn = get_db()
    row = conn.execute("SELECT row_key FROM student_status WHERE row_key=?", (row_key,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Student not found")
    background_tasks.add_task(recheck_one, row_key)
    return {"ok": True, "message": "Re-checking this student…"}

@app.post("/api/run-log-delete")
def run_log_delete(id: int, user=Depends(verify_token)):
    """Delete ONE run-log entry (e.g. a 'Nothing to check' or failed run). A run that is
    still 'running' cannot be deleted — cancel it first."""
    conn = get_db()
    row = conn.execute("SELECT status FROM run_logs WHERE id=?", (id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Run not found")
    if row["status"] == "running":
        conn.close()
        raise HTTPException(status_code=400, detail="Run is still running — cancel it first")
    conn.execute("DELETE FROM run_logs WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/run-logs-clear")
def run_logs_clear(user=Depends(verify_token)):
    """Delete all run-log entries that are not currently running."""
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM run_logs WHERE status!='running'").fetchone()[0]
    conn.execute("DELETE FROM run_logs WHERE status!='running'")
    conn.commit()
    conn.close()
    return {"ok": True, "deleted": n}

# ─────────────────────────────────────────────────────────────────────────────
# WhatsApp (AiSensy) — settings, test, manual resend
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/api/wa-diagnose")
def wa_diagnose(user=Depends(verify_token)):
    """One-shot health check for the whole WhatsApp pipeline — never raises, returns plain JSON
    so the exact broken piece (module import / API key / campaign env / DB lock) is obvious."""
    import os
    d = {}
    try:
        import whatsapp
        d["whatsapp_module"] = "imported OK"
        try:
            d["is_configured"] = whatsapp.is_configured()
        except Exception as e:
            d["is_configured"] = f"ERROR: {e}"
        for g in ("ondemand", "stream2", "public", "syc"):
            try:
                d[f"campaign_{g}"] = whatsapp.campaign_for(g) or "(empty)"
            except Exception as e:
                d[f"campaign_{g}"] = f"ERROR: {e}"
    except Exception as e:
        d["whatsapp_module"] = f"IMPORT FAILED: {e}"
    d["env_AISENSY_API_KEY_set"] = bool(os.environ.get("AISENSY_API_KEY", "").strip())
    d["env_AISENSY_API_KEY_PUBLIC_set"] = bool(os.environ.get("AISENSY_API_KEY_PUBLIC", "").strip())
    d["env_CAMPAIGN_REQUIRED"] = os.environ.get("AISENSY_CAMPAIGN_REQUIRED", "").strip() or "(empty)"
    d["env_CAMPAIGN_REQUIRED_PUBLIC"] = os.environ.get("AISENSY_CAMPAIGN_REQUIRED_PUBLIC", "").strip() or "(empty)"
    try:
        d["wa_enabled_setting"] = get_setting("wa_enabled", "0")
        d["wa_required_enabled_setting"] = get_setting("wa_required_enabled", "0")
        d["db_read"] = "OK"
    except Exception as e:
        d["db_read"] = f"ERROR: {e}"
    try:
        conn = get_db()
        d["journal_mode"] = conn.execute("PRAGMA journal_mode").fetchone()[0]
        conn.close()
    except Exception as e:
        d["journal_mode"] = f"ERROR: {e}"
    try:
        conn = get_db()
        rows = conn.execute("SELECT student_name, COALESCE(whatsapp_sent,0) AS s, whatsapp_info "
                            "FROM student_status WHERE is_confirmed=1 AND COALESCE(whatsapp_info,'')!='' "
                            "ORDER BY whatsapp_sent_at DESC LIMIT 5").fetchall()
        conn.close()
        d["recent_send_results"] = [{"name": r["student_name"], "sent_flag": r["s"],
                                     "info": (r["whatsapp_info"] or "")[:140]} for r in rows]
    except Exception as e:
        d["recent_send_results"] = f"ERROR: {e}"
    return d


@app.get("/api/wa-settings")
def wa_settings_get(user=Depends(verify_token)):
    """Always returns 200 with valid JSON — campaigns come straight from env vars, so even if
    the DB is busy or the whatsapp module hiccups, the panel still loads and shows what's set."""
    import os
    def env(k):
        try:
            return os.environ.get(k, "").strip()
        except Exception:
            return ""
    out = {
        "enabled": False,
        "required_enabled": False,
        "configured": bool(env("AISENSY_API_KEY") or env("AISENSY_API_KEY_PUBLIC")),
        "campaigns": {"ondemand": "", "stream2": "", "public": "", "syc": "", "ondemand_notoc": "", "stream2_notoc": "", "public_yestoc": ""},
        "campaigns_env": {
            "ondemand": bool(env("AISENSY_CAMPAIGN_ONDEMAND")),
            "stream2": bool(env("AISENSY_CAMPAIGN_STREAM2")),
            "public": bool(env("AISENSY_CAMPAIGN_PUBLIC")),
            "syc": bool(env("AISENSY_CAMPAIGN_SYC")),
            "ondemand_notoc": bool(env("AISENSY_CAMPAIGN_ONDEMAND_NOTOC")),
            "stream2_notoc": bool(env("AISENSY_CAMPAIGN_STREAM2_NOTOC")),
            "public_yestoc": bool(env("AISENSY_CAMPAIGN_PUBLIC_YESTOC")),
        },
        "required_campaigns": {
            "main": env("AISENSY_CAMPAIGN_REQUIRED"),
            "public": env("AISENSY_CAMPAIGN_REQUIRED_PUBLIC"),
        },
    }
    try:
        out["enabled"] = get_setting("wa_enabled", "0") == "1"
        out["required_enabled"] = get_setting("wa_required_enabled", "0") == "1"
    except Exception:
        pass
    try:
        import whatsapp
        out["configured"] = whatsapp.is_configured()
        out["campaigns"] = {
            "ondemand": whatsapp.campaign_for("ondemand"),
            "stream2": whatsapp.campaign_for("stream2"),
            "public": whatsapp.campaign_for("public"),
            "syc": whatsapp.campaign_for("syc"),
            "ondemand_notoc": whatsapp.campaign_for_notoc("ondemand"),
            "stream2_notoc": whatsapp.campaign_for_notoc("stream2"),
            "public_yestoc": whatsapp.campaign_for_yestoc("public"),
        }
    except Exception:
        pass
    return out

@app.post("/api/wa-settings")
def wa_settings_set(body: dict, user=Depends(verify_token)):
    if "enabled" in body:
        set_setting("wa_enabled", "1" if body.get("enabled") else "0")
    if "required_enabled" in body:
        set_setting("wa_required_enabled", "1" if body.get("required_enabled") else "0")
    return {"message": "saved",
            "enabled": get_setting("wa_enabled", "0") == "1",
            "required_enabled": get_setting("wa_required_enabled", "0") == "1"}

@app.post("/api/wa-campaigns")
def wa_campaigns_set(body: dict, user=Depends(verify_token)):
    """Save the confirmed-send campaign names from the Settings page. Stored in the DB and used
    in preference to the Railway env vars, so campaigns can be fixed in-app. Empty value clears
    the override (falls back to the env var)."""
    for g in ("ondemand", "stream2", "public", "syc", "ondemand_notoc", "stream2_notoc", "public_yestoc"):
        if g in body:
            set_setting("wa_campaign_" + g, str(body.get(g) or "").strip())
    import whatsapp
    return {"ok": True, "campaigns": {
        "ondemand": whatsapp.campaign_for("ondemand"),
        "stream2": whatsapp.campaign_for("stream2"),
        "public": whatsapp.campaign_for("public"),
        "syc": whatsapp.campaign_for("syc"),
        "ondemand_notoc": whatsapp.campaign_for_notoc("ondemand"),
        "stream2_notoc": whatsapp.campaign_for_notoc("stream2"),
        "public_yestoc": whatsapp.campaign_for_yestoc("public"),
    }}

@app.post("/api/wa-test")
def wa_test(body: dict, user=Depends(verify_token)):
    """Send a test message of a chosen template group to any number."""
    import whatsapp
    number = body.get("number", "")
    group = body.get("group", "ondemand")
    name = body.get("name", "") or "Test Student"
    ok, info = whatsapp.send_test(number, name, group)
    return {"ok": ok, "info": info}


def _humanize_remark(remark):
    """Turn NIOS's raw RC comment into a simple, warm document request — written the way a
    human counsellor would, so the student actually understands what to send. This is only a
    DRAFT: the counsellor reviews and edits it before anything is sent. Returns a Hinglish
    line for the message ({{2}} param). Empty string -> unrecognised, counsellor writes it."""
    r = (remark or "").lower()
    asks = []
    if "address" in r or "correspondence" in r or "rent agreement" in r:
        if "delhi" in r or "ncr" in r:
            asks.append("Apna Delhi/NCR ka address proof (Aadhaar card, bijli/paani/phone ka bill, "
                        "ration card, voter ID, passport ya rent agreement me se koi ek — jisme aapka "
                        "ya aapke parents ka naam aur poora address ho)")
        else:
            asks.append("Apna sahi address proof (Aadhaar card, bijli/paani/phone ka bill, ration card, "
                        "voter ID, passport ya rent agreement me se koi ek — jisme aapka ya aapke "
                        "parents ka naam aur poora address ho)")
    if "mark sheet" in r or "marksheet" in r or "mark-sheet" in r:
        asks.append("Apni Class 10th ki original marksheet")
    if ("matriculation" in r or "passing certificate" in r or "secondary examination" in r
            or "date of birth" in r or ("10th" in r and "certificate" in r)):
        asks.append("Apna Class 10th passing certificate (jisme aapki date of birth likhi ho)")
    if ("father" in r) and ("not match" in r or "correct" in r or "e-service" in r or "e service" in r):
        asks.append("Apne father ke naam ka sahi record — abhi jo naam hai wo academic record se match "
                    "nahi kar raha; aap apne dashboard ki E-Services se ise theek kar sakte hain")
    if ("mother" in r) and ("not match" in r or "correct" in r or "e-service" in r or "e service" in r):
        asks.append("Apne mother ke naam ka sahi record (dashboard ki E-Services se correction)")
    if "photo" in r or "photograph" in r:
        asks.append("Apni clear passport-size photo")
    if "signature" in r:
        asks.append("Apne signature ki clear photo")
    if not asks:
        return ""
    if len(asks) == 1:
        return asks[0]
    return "; ".join(asks)


# Fixed, approved wrapper around the editable {{2}} document line (kept in sync with the
# AiSensy template the user gets approved). Shown in the portal preview.
DOCREQ_PREFIX = "Namaste {name} \U0001F64F\n\nAapke NIOS admission ko poora karne ke liye humein aapse ye chahiye:\n\n"
DOCREQ_SUFFIX = ("\n\nJab aapko suvidha ho, kripya ye MVS Foundation ko bhej dijiye — koi jaldi ya "
                 "chinta ki baat nahi \U0001F60A Hum aapki poori madad ke liye yahin hain.\n\n"
                 "Kisi bhi sawaal ke liye is number par message kar dijiye.\n\nMVS Foundation Team")


def _docreq_rows(conn, keys=None, pending_only=False):
    q = ("SELECT row_key, student_name, mobile, alt_mobile, session, remark, required_msg, "
         "required_img, required_notified, required_notified_at FROM student_status "
         "WHERE COALESCE(deleted,0)=0 AND current_status='Document Required' "
         "AND COALESCE(is_confirmed,0)=0 AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0")
    params = []
    if pending_only:
        q += " AND COALESCE(required_notified,0)=0"
    if keys:
        q += " AND row_key IN (%s)" % ",".join("?" * len(keys))
        params = keys
    q += " ORDER BY COALESCE(required_notified,0) ASC, student_name"
    return conn.execute(q, params).fetchall()


@app.get("/api/doc-requests")
def doc_requests(user=Depends(verify_token)):
    """List Document-Required students with the (editable) document-request message that
    will be sent. The counsellor reviews/edits these on the portal before sending."""
    conn = get_db()
    rows = _docreq_rows(conn)
    conn.close()
    out = []
    for r in rows:
        saved = (r["required_msg"] or "").strip()
        draft = saved or _humanize_remark(r["remark"] or "")
        out.append({
            "row_key": r["row_key"], "student_name": r["student_name"] or "\u2014",
            "mobile": r["mobile"] or "\u2014", "session": r["session"] or "\u2014",
            "remark": r["remark"] or "", "message": draft,
            "image": (r["required_img"] if ("required_img" in r.keys()) else "") or "",
            "edited": bool(saved), "auto_blank": (not saved and not draft),
            "sent": bool(r["required_notified"] == 1), "sent_at": r["required_notified_at"] or "",
        })
    return {"requests": out, "count": len(out),
            "prefix": DOCREQ_PREFIX, "suffix": DOCREQ_SUFFIX}


@app.post("/api/doc-request-save")
def doc_request_save(body: dict, user=Depends(verify_token)):
    """Save the counsellor's edited document-request message for one student."""
    rk = body.get("row_key", "")
    msg = (body.get("message", "") or "").strip()
    if not rk:
        raise HTTPException(status_code=400, detail="row_key required")
    conn = get_db()
    conn.execute("UPDATE student_status SET required_msg=? WHERE row_key=?", (msg[:1000], rk))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/api/doc-request-send")
def doc_request_send(body: dict, request: Request, user=Depends(verify_token)):
    """Send the reviewed document-request WhatsApp message to selected (or all pending)
    Document-Required students. Routes per session (public -> public API; others -> main API).
    Uses ONE universal image-header template: sends the uploaded screenshot when present,
    otherwise the default MVS banner. Marks each as sent so it isn't messaged twice.
    NOTE: this is a *sync* endpoint on purpose — the AiSensy HTTP calls are blocking, so running
    them in FastAPI's threadpool keeps the event loop free (otherwise the whole app freezes and
    even the banner image can't be served back to AiSensy while a send is in progress)."""
    if get_setting("wa_required_enabled", "0") != "1":
        raise HTTPException(status_code=400, detail="Document-request reminders are turned OFF in Settings — turn them on first.")
    import whatsapp
    if not whatsapp.is_configured():
        raise HTTPException(status_code=400, detail="AISENSY_API_KEY is not set.")
    base = str(request.base_url).rstrip("/")
    if base.startswith("http://") and "localhost" not in base and "127.0.0.1" not in base:
        base = "https://" + base[len("http://"):]
    default_img = f"{base}/media/docreq-banner-v9.png"
    send_all = bool(body.get("all"))
    keys = [k for k in (body.get("row_keys", []) or []) if k][:500]
    conn = get_db()
    rows = _docreq_rows(conn, keys=None if send_all else keys, pending_only=send_all)
    if not send_all and not keys:
        conn.close()
        raise HTTPException(status_code=400, detail="no students selected")
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sent, failed, results = 0, 0, []
    for r in rows:
        msg = (r["required_msg"] or "").strip() or _humanize_remark(r["remark"] or "")
        media = (r["required_img"] if ("required_img" in r.keys()) else "") or None
        ok, info = whatsapp.send_required_reminder({
            "row_key": r["row_key"], "student_name": r["student_name"] or "",
            "mobile": r["mobile"] or "", "session": r["session"] or "",
            "alt_mobile": (r["alt_mobile"] if ("alt_mobile" in r.keys()) else "") or "",
        }, msg, media_url=media, default_img=default_img)
        if ok:
            conn.execute("UPDATE student_status SET required_notified=1, required_notified_at=?, "
                         "required_msg=? WHERE row_key=?", (now_s, msg[:1000], r["row_key"]))
            sent += 1
        else:
            failed += 1
        results.append({"student_name": r["student_name"] or "\u2014", "ok": ok, "info": str(info)[:120]})
    conn.commit()
    conn.close()
    return {"ok": True, "sent": sent, "failed": failed, "results": results}


import os as _os_docreq
DOCREQ_MEDIA_DIR = "/tmp/docreq_media"

@app.get("/media/docreq/{fname}")
async def docreq_media(fname: str):
    """Serve an uploaded screenshot publicly (no auth) so the WhatsApp gateway can fetch it."""
    safe = _os_docreq.path.basename(fname)
    path = _os_docreq.path.join(DOCREQ_MEDIA_DIR, safe)
    if not _os_docreq.path.isfile(path):
        raise HTTPException(status_code=404, detail="not found")
    ext = safe.rsplit(".", 1)[-1].lower() if "." in safe else ""
    mt = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
          "webp": "image/webp"}.get(ext, "application/octet-stream")
    return FileResponse(path, media_type=mt, headers={"Cache-Control": "public, max-age=3600"})

def _make_default_banner(path):
    """A clean, professional MVS-branded banner used as the image header when no screenshot is
    attached: deep-navy backdrop with a soft glow, the real MVS logo, and a large
    'MVS Foundation Team' that fills the frame so it stays readable when WhatsApp scales it."""
    from PIL import Image, ImageDraw, ImageFont, ImageFilter
    W, H = 800, 418
    img = Image.new("RGB", (W, H))
    d = ImageDraw.Draw(img)
    # deep navy gradient — institutional / professional
    top, bot = (11, 18, 44), (28, 44, 88)
    for y in range(H):
        t = y / H
        d.line([(0, y), (W, y)], fill=(int(top[0] + (bot[0]-top[0])*t),
                                       int(top[1] + (bot[1]-top[1])*t),
                                       int(top[2] + (bot[2]-top[2])*t)))
    cx, cy = W // 2, 130
    # soft teal glow behind the logo for depth
    glow = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    ImageDraw.Draw(glow).ellipse([cx - 134, cy - 124, cx + 134, cy + 140], fill=(45, 212, 191, 50))
    glow = glow.filter(ImageFilter.GaussianBlur(46))
    img = Image.alpha_composite(img.convert("RGBA"), glow).convert("RGB")
    d = ImageDraw.Draw(img)
    # refined double border frame
    d.rectangle([18, 18, W - 18, H - 18], outline=(72, 94, 150), width=2)
    d.rectangle([25, 25, W - 25, H - 25], outline=(40, 58, 108), width=1)

    def font(sz):
        import os as _o
        here = _o.path.dirname(_o.path.abspath(__file__))
        for p in (_o.path.join(here, "DejaVuSans-Bold.ttf"),
                  _o.path.join(here, "assets", "DejaVuSans-Bold.ttf"),
                  "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                  "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
                  "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
                  "DejaVuSans-Bold.ttf"):
            try:
                return ImageFont.truetype(p, sz)
            except Exception:
                pass
        try:
            return ImageFont.load_default(size=sz)   # Pillow >= 10.1 scales the default font
        except Exception:
            return ImageFont.load_default()
    # real MVS logo in a white circle with a soft outer ring
    d.ellipse([cx - 82, cy - 82, cx + 82, cy + 82], fill=(36, 52, 98))     # ring
    r = 68
    d.ellipse([cx - r, cy - r, cx + r, cy + r], fill=(255, 255, 255))
    try:
        import base64, io
        from assets import LOGO_B64
        logo = Image.open(io.BytesIO(base64.b64decode(LOGO_B64))).convert("RGBA")
        ls = int(r * 1.6)
        logo = logo.resize((ls, ls))
        mask = Image.new("L", (ls, ls), 0)
        ImageDraw.Draw(mask).ellipse([0, 0, ls, ls], fill=255)
        img.paste(logo, (cx - ls // 2, cy - ls // 2), mask)
    except Exception:
        pass
    # title on a single line, auto-fit to ~90% width with a soft shadow
    txt = "MVS Foundation Team"
    target = int(W * 0.90)
    fsize = 96
    f = font(fsize)
    while fsize > 28:
        f = font(fsize)
        bb = d.textbbox((0, 0), txt, font=f)
        if (bb[2] - bb[0]) <= target:
            break
        fsize -= 2
    bb = d.textbbox((0, 0), txt, font=f)
    tx = (W - (bb[2] - bb[0])) / 2 - bb[0]
    ty = 250
    d.text((tx + 3, ty + 3), txt, font=f, fill=(5, 9, 24))       # shadow
    d.text((tx, ty), txt, font=f, fill=(246, 249, 255))          # main
    # teal accent underline
    ay = ty + (bb[3] - bb[1]) + 26
    d.rounded_rectangle([cx - 84, ay, cx + 84, ay + 6], radius=3, fill=(45, 212, 191))
    img.save(path, "PNG")

@app.get("/media/docreq-banner-v9.png")
async def docreq_default_banner():
    """Default branded banner image (regenerated when missing) for document requests with no
    screenshot attached. Public (no auth) so the WhatsApp gateway can fetch it. The versioned
    URL (v3) forces WhatsApp/AiSensy to fetch the new image instead of a cached old one."""
    path = _os_docreq.path.join(DOCREQ_MEDIA_DIR, "_default_banner_v9.png")
    if not _os_docreq.path.isfile(path):
        _os_docreq.makedirs(DOCREQ_MEDIA_DIR, exist_ok=True)
        try:
            _make_default_banner(path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"banner generation failed: {e}")
    return FileResponse(path, media_type="image/png", headers={"Cache-Control": "public, max-age=3600"})

@app.post("/api/doc-request-image")
async def doc_request_image(request: Request, row_key: str = Form(...),
                            file: UploadFile = File(...), user=Depends(verify_token)):
    """Upload a demo screenshot for a Document-Required student. Stored on disk and served at a
    public URL, which is attached as the WhatsApp message's image header when you send."""
    import uuid as _uuid
    fn = file.filename or ""
    ext = fn.rsplit(".", 1)[-1].lower() if "." in fn else "jpg"
    if ext not in ("jpg", "jpeg", "png", "webp"):
        raise HTTPException(status_code=400, detail="Only JPG / PNG / WEBP images are allowed.")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 5 MB).")
    _os_docreq.makedirs(DOCREQ_MEDIA_DIR, exist_ok=True)
    fname = f"{_uuid.uuid4().hex}.{ext}"
    with open(_os_docreq.path.join(DOCREQ_MEDIA_DIR, fname), "wb") as f:
        f.write(data)
    base = str(request.base_url).rstrip("/")
    if base.startswith("http://") and "localhost" not in base and "127.0.0.1" not in base:
        base = "https://" + base[len("http://"):]
    url = f"{base}/media/docreq/{fname}"
    conn = get_db()
    conn.execute("UPDATE student_status SET required_img=? WHERE row_key=?", (url, row_key))
    conn.commit()
    conn.close()
    return {"ok": True, "url": url}

@app.post("/api/doc-request-image-remove")
def doc_request_image_remove(body: dict, user=Depends(verify_token)):
    rk = body.get("row_key", "")
    if not rk:
        raise HTTPException(status_code=400, detail="row_key required")
    conn = get_db()
    conn.execute("UPDATE student_status SET required_img=NULL WHERE row_key=?", (rk,))
    conn.commit()
    conn.close()
    return {"ok": True}

def _wa_send_one(row_key, verify=False, fresh_toc=False):
    """(Re)send the documents for ONE confirmed student. Opens its own DB connection so it
    is safe to run as a background task. Returns (ok, info, login_failed). Never raises.

    verify=False (default, FAST): the student is already CONFIRMED — their Reference/DOB were
    already proven valid when the status was confirmed, and the WhatsApp message only carries
    signed links (the document is fetched live when the student taps it). So we skip the slow
    NIOS re-login (CapSolver reCAPTCHA, ~15-40s each) and send straight away (~1-2s each).
    verify=True: re-verify the NIOS login first (slow) — used only if explicitly requested."""
    import whatsapp
    try:
        conn = get_db()
        row = conn.execute("SELECT row_key, student_name, mobile, alt_mobile, session, reference_no, dob, enrollment_no, toc_status "
                           "FROM student_status WHERE row_key=?", (row_key,)).fetchone()
        if not row:
            conn.close()
            return False, "student not found", False
        if fresh_toc:
            # Re-read this student's TOC from NIOS right now and apply/push any correction, so a
            # resend can never repeat the wrong campaign. If NIOS can't be read, don't guess.
            conn.close()
            from job_runner import final_toc_verify
            _ft = final_toc_verify(row_key)
            if _ft not in ("yes", "no"):
                cx = get_db()
                cx.execute("UPDATE student_status SET whatsapp_info=? WHERE row_key=?",
                           ("Not sent — TOC could not be read from NIOS just now; try again", row_key))
                cx.commit(); cx.close()
                return False, "TOC could not be read from NIOS — not sending (try again)", False
            conn = get_db()
            row = conn.execute("SELECT row_key, student_name, mobile, alt_mobile, session, reference_no, dob, enrollment_no, toc_status "
                               "FROM student_status WHERE row_key=?", (row_key,)).fetchone()
        if verify:
            from nios_login import verify_login
            ok_login, lmsg = verify_login(row["reference_no"], row["dob"], row["enrollment_no"] or "")
            if not ok_login:
                conn.execute("UPDATE student_status SET login_failed=1, login_remark=?, whatsapp_info=?, "
                             "whatsapp_sent=0, whatsapp_attempts=COALESCE(whatsapp_attempts,0)+1 WHERE row_key=?",
                             (lmsg[:240], ("Not sent — " + lmsg)[:180], row_key))
                conn.commit(); conn.close()
                return False, lmsg, True
            conn.execute("UPDATE student_status SET login_failed=0, login_remark='' WHERE row_key=?", (row_key,))
        # Cache-first: make sure this student's documents are saved on OUR server
        # BEFORE the WhatsApp goes out — the first click then opens instantly from
        # our copy. Already-cached students pass through in milliseconds; if a live
        # fetch fails we still send (links fall back to live fetch with a loader).
        try:
            cache_student_docs(row_key)
        except Exception:
            pass
        # COMPLETENESS GATE: only send the link once ALL of this student's documents are saved,
        # so their very first tap opens instantly (no live NIOS fetch = no error/panic). If the
        # docs still aren't complete, hold the send — the WhatsApp auto-sweep re-tries caching
        # every cycle and will send the moment it's complete. Fallback: after several attempts
        # we send anyway (a rare doc that NIOS never provides shouldn't block the link forever;
        # the link still live-fetches that one doc on demand).
        if not verify:
            # HOLD the WhatsApp if this student has a TOC mismatch that a counsellor hasn't
            # verified yet — sending the wrong (no-TOC) message is exactly what panics students.
            _tm = conn.execute("SELECT COALESCE(toc_mismatch,0), COALESCE(toc_verified,0) "
                               "FROM student_status WHERE row_key=?", (row_key,)).fetchone()
            if _tm and _tm[0] == 1 and _tm[1] == 0:
                conn.execute("UPDATE student_status SET whatsapp_info=? WHERE row_key=?",
                             ("Held — TOC mismatch needs counsellor verification", row_key))
                conn.commit(); conn.close()
                return False, "TOC mismatch — waiting for counsellor to verify before sending", False
            _att = (conn.execute("SELECT COALESCE(whatsapp_attempts,0) FROM student_status WHERE row_key=?",
                                 (row_key,)).fetchone() or [0])[0]
            if (not docs_all_cached(row_key)) and _att < 4:
                conn.execute("UPDATE student_status SET whatsapp_info=?, "
                             "whatsapp_attempts=COALESCE(whatsapp_attempts,0)+1 WHERE row_key=?",
                             ("Waiting — saving documents before sending", row_key))
                conn.commit(); conn.close()
                return False, "documents not fully saved yet — will send once complete", False
        ok, info = whatsapp.send_for_student(dict(row))
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if ok:
            conn.execute("UPDATE student_status SET whatsapp_sent=1, whatsapp_info=?, "
                         "whatsapp_sent_at=?, whatsapp_delivery='' WHERE row_key=?",
                         (str(info)[:180], now, row_key))
        else:
            conn.execute("UPDATE student_status SET whatsapp_info=?, "
                         "whatsapp_attempts=COALESCE(whatsapp_attempts,0)+1 WHERE row_key=?",
                         (str(info)[:180], row_key))
        conn.commit(); conn.close()
        # (Documents are now saved BEFORE the send above — no post-send caching needed.)
        return ok, info, False
    except Exception as e:
        try:
            conn.close()
        except Exception:
            pass
        return False, f"error: {e}", False

WA_PROGRESS = {"running": False, "total": 0, "done": 0, "sent": 0, "failed": 0,
               "started_at": "", "finished_at": "", "label": ""}

def auto_send_pending_whatsapp(batch=None, label="Auto sweep"):
    """Background sweep: for every CONFIRMED student whose documents have not been sent
    on WhatsApp yet, send them automatically — so the counsellor never has to click
    'Resend' for each one. Tracks live progress (WA_PROGRESS) so the UI can show a
    percentage bar. Gives up on a student after several failed attempts."""
    if get_setting("wa_enabled", "0") != "1":
        return {"sent": 0, "checked": 0}
    if WA_PROGRESS["running"]:
        return {"skipped": "already running"}
    # Never compete with an active run — the run itself sends WhatsApp for new
    # confirmations (documents cached first). Leftovers are picked up by the next
    # sweep after the run finishes, so nothing is ever missed.
    try:
        _c = get_db()
        _active = _c.execute("SELECT id FROM run_logs WHERE status='running'").fetchone()
        _c.close()
        if _active:
            logger.info("WhatsApp auto sweep: a run is in progress — skipping this sweep")
            return {"skipped": "run in progress"}
    except Exception:
        pass
    conn = get_db()
    q = ("SELECT row_key FROM student_status WHERE is_confirmed=1 "
         "AND COALESCE(whatsapp_sent,0)=0 AND COALESCE(login_failed,0)=0 "
         "AND COALESCE(deleted,0)=0 AND COALESCE(whatsapp_attempts,0) < 8 "
         "AND mobile IS NOT NULL AND TRIM(mobile) != '' "
         "ORDER BY COALESCE(whatsapp_attempts,0) ASC, last_changed DESC")
    if batch:
        q += f" LIMIT {int(batch)}"
    rows = conn.execute(q).fetchall()
    conn.close()
    keys = [r["row_key"] for r in rows]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    WA_PROGRESS.update({"running": True, "total": len(keys), "done": 0, "sent": 0,
                        "failed": 0, "started_at": now, "finished_at": "", "label": label})
    sent = 0
    try:
        for rk in keys:
            ok, info, _lf = _wa_send_one(rk)
            WA_PROGRESS["done"] += 1
            if ok:
                WA_PROGRESS["sent"] += 1; sent += 1
            else:
                WA_PROGRESS["failed"] += 1
    finally:
        WA_PROGRESS["running"] = False
        WA_PROGRESS["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if keys:
        logger.info(f"WhatsApp {label}: sent {sent}/{len(keys)} pending confirmed student(s)")
    return {"sent": sent, "checked": len(keys)}

CACHE_PROGRESS = {"running": False, "phase": "", "total": 0, "done": 0, "saved": 0, "failed": 0,
                  "confirmed": 0, "last_error": "", "started_at": "", "finished_at": "", "label": ""}
_CACHE_LOCK = _threading.Lock()
try:
    _CACHE_WORKERS = max(1, min(10, int(os.environ.get("CACHE_WORKERS", "3"))))
except Exception:
    _CACHE_WORKERS = 3
_DB_WRITE_LOCK = _threading.Lock()

def _cache_all_confirmed_worker(force=False):
    """Fetch + save every confirmed student's allowed documents into OUR database (in parallel),
    so their WhatsApp links open straight from our copy (no live NIOS/CapSolver, works if NIOS is
    down). Documents of the SAME student are fetched together (one login reused). force=True
    re-fetches everything; else only missing ones."""
    if CACHE_PROGRESS["running"]:
        return
    # Never compete with an active RUN for NIOS/CapSolver — the watchdog retries
    # every 90s, so saving resumes automatically the moment the run finishes.
    try:
        _c = get_db()
        _active = _c.execute("SELECT id FROM run_logs WHERE status='running'").fetchone()
        _c.close()
        if _active:
            logger.info("Doc saving: a run is in progress — waiting (auto-resumes after the run)")
            return
    except Exception:
        pass
    logger.info(f"Doc saving: worker started (force={force})")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    CACHE_PROGRESS.update({"running": True, "phase": "finding confirmed students", "total": 0,
                           "done": 0, "saved": 0, "failed": 0, "confirmed": 0, "last_error": "",
                           "started_at": now, "finished_at": "", "label": "Saving documents"})
    try:
        import whatsapp
        from nios_login import fetch_document
        from concurrent.futures import ThreadPoolExecutor
        conn = get_db()
        rows = conn.execute("SELECT row_key, reference_no, enrollment_no, dob, session, toc_status "
                            "FROM student_status WHERE is_confirmed=1 AND COALESCE(deleted,0)=0 "
                            "AND (COALESCE(reference_no,'')!='' OR COALESCE(enrollment_no,'')!='')").fetchall()
        conn.close()
        CACHE_PROGRESS["confirmed"] = len(rows)
        # Group missing docs BY STUDENT so each student logs in once and reuses the session.
        # Skip docs that have already failed too many times (genuine data problem) unless forcing.
        students = []
        total_docs = 0
        for r in rows:
            allowed = whatsapp.allowed_docs(r["session"], (r["toc_status"] or ""))
            kinds = []
            for k in allowed:
                if load_doc_cache(r["row_key"], k) is not None:
                    continue                      # already saved
                if (not force) and _doc_fail_attempts(r["row_key"], k) >= _MAX_CACHE_ATTEMPTS:
                    continue                      # genuinely failing -> leave for manual fix
                kinds.append(k)
            if kinds:
                students.append((r["row_key"], r["reference_no"] or "", r["enrollment_no"] or "",
                                 r["dob"] or "", kinds))
                total_docs += len(kinds)
        CACHE_PROGRESS.update({"phase": "saving", "total": total_docs})

        def _do_student(item):
            rk, ref, enroll, dob, kinds = item
            for kind in kinds:
                try:
                    content, ctype, filename = fetch_document(ref, dob, kind, enrollment_no=enroll)
                    if content is not None:
                        with _DB_WRITE_LOCK:
                            save_doc_cache(rk, kind, content, ctype, filename)
                        with _CACHE_LOCK:
                            CACHE_PROGRESS["saved"] += 1
                        logger.info(f"Doc saved [{CACHE_PROGRESS['done']+1}/{CACHE_PROGRESS['total']}] "
                                    f"{ref or enroll} {kind}")
                    else:
                        _record_doc_fail(rk, kind, ctype)
                        with _CACHE_LOCK:
                            CACHE_PROGRESS["failed"] += 1
                            CACHE_PROGRESS["last_error"] = str(ctype)[:150]
                        logger.warning(f"Doc save FAILED {ref or enroll} {kind}: {str(ctype)[:100]}")
                except Exception as e:
                    _record_doc_fail(rk, kind, f"{type(e).__name__}: {e}")
                    with _CACHE_LOCK:
                        CACHE_PROGRESS["failed"] += 1
                        CACHE_PROGRESS["last_error"] = f"{type(e).__name__}: {str(e)[:110]}"
                    logger.warning(f"Doc save ERROR {ref or enroll} {kind}: {type(e).__name__}: {str(e)[:100]}")
                with _CACHE_LOCK:
                    CACHE_PROGRESS["done"] += 1

        if students:
            with ThreadPoolExecutor(max_workers=_CACHE_WORKERS) as ex:
                list(ex.map(_do_student, students))
    except Exception as e:
        CACHE_PROGRESS["last_error"] = f"worker: {type(e).__name__}: {str(e)[:110]}"
    finally:
        CACHE_PROGRESS["running"] = False
        CACHE_PROGRESS["phase"] = "done"
        CACHE_PROGRESS["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"Doc saving: pass finished — saved={CACHE_PROGRESS.get('saved',0)} "
                    f"failed={CACHE_PROGRESS.get('failed',0)} of {CACHE_PROGRESS.get('total',0)}")
        try:
            # Keep auto-resuming while there are still docs worth retrying (this run had tasks).
            # When a run finds 0 tasks, everything is either saved or genuinely failed -> stop.
            if CACHE_PROGRESS.get("total", 0) == 0:
                set_setting("cache_docs_active", "0")
        except Exception:
            pass

def _cache_watchdog():
    """Retries the bulk cache job while it's active (set by pressing 'Save all'), and keeps
    retrying failed ones. Respects the CACHE_OFF kill switch. Waits after boot so the app is
    fully up and serving links before any caching starts."""
    import time as _t
    _t.sleep(120)   # let the app boot + serve first
    while True:
        try:
            _t.sleep(90)
            if os.environ.get("CACHE_OFF", "") == "1":
                continue   # kill switch: no caching at all
            if get_setting("cache_docs_active", "0") == "1" and not CACHE_PROGRESS["running"]:
                _cache_all_confirmed_worker(force=False)
        except Exception:
            pass

try:
    _threading.Thread(target=_cache_watchdog, daemon=True).start()
except Exception:
    pass

@app.post("/api/cache-docs")
async def cache_docs(body: dict = None, user=Depends(verify_token)):
    """Start saving all confirmed students' documents into our database (background).
    Pass {"force": true} to RE-FETCH everything (refresh), else only missing ones are fetched.
    Sets an auto-resume flag so the job continues (even across app restarts) until all are saved."""
    force = bool((body or {}).get("force"))
    if os.environ.get("CACHE_OFF", "") == "1":
        return {"ok": False, "message": "Document saving is turned off (CACHE_OFF). Remove that setting to enable it."}
    try:
        set_setting("cache_docs_active", "1")   # keep resuming until everything is saved
    except Exception:
        pass
    if CACHE_PROGRESS["running"]:
        return {"ok": True, "message": "Already saving — it will keep going until done."}
    import threading
    threading.Thread(target=_cache_all_confirmed_worker, kwargs={"force": force}, daemon=True).start()
    return {"ok": True, "message": ("Refreshing (re-fetching) all confirmed documents in the background."
                                    if force else "Saving all confirmed documents to the database in the background. It will keep going (even if the app restarts) until every document is saved.")}

@app.post("/api/cache-docs-stop")
async def cache_docs_stop(user=Depends(verify_token)):
    """Stop the auto-resume of the bulk document save."""
    try:
        set_setting("cache_docs_active", "0")
    except Exception:
        pass
    return {"ok": True, "message": "Auto-save stopped. Saved documents are kept."}

@app.post("/api/refresh-selected-docs")
async def refresh_selected_docs(body: dict, user=Depends(verify_token)):
    """Re-fetch + re-save documents for the SELECTED students (force overwrite)."""
    keys = (body or {}).get("row_keys", []) or []
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    from fastapi.concurrency import run_in_threadpool
    def _do():
        tot_s = tot_f = 0
        for rk in keys[:200]:
            s, f = cache_student_docs(rk, force=True)
            tot_s += s; tot_f += f
        return tot_s, tot_f
    saved, failed = await run_in_threadpool(_do)
    return {"ok": True, "saved": saved, "failed": failed,
            "message": f"Refreshed {saved} document(s)" + (f", {failed} failed" if failed else "")}

@app.post("/api/refresh-student-docs")
async def refresh_student_docs(body: dict, user=Depends(verify_token)):
    """Re-fetch + re-save ONE student's documents (use when their documents changed on NIOS)."""
    row_key = (body or {}).get("row_key", "")
    if not row_key:
        raise HTTPException(status_code=400, detail="row_key required")
    from fastapi.concurrency import run_in_threadpool
    saved, failed = await run_in_threadpool(cache_student_docs, row_key, True)
    return {"ok": True, "saved": saved, "failed": failed,
            "message": (f"Refreshed {saved} document(s)" + (f", {failed} failed" if failed else "") if saved or failed
                        else "No documents to refresh for this student")}

@app.get("/api/cache-docs-progress")
def cache_docs_progress(user=Depends(verify_token)):
    conn = get_db()
    total_cached = conn.execute("SELECT COUNT(*) FROM document_cache").fetchone()[0]
    students_cached = conn.execute("SELECT COUNT(DISTINCT row_key) FROM document_cache").fetchone()[0]
    try:
        genuine_failed = conn.execute("SELECT COUNT(DISTINCT row_key) FROM cache_fail WHERE attempts >= ?",
                                      (_MAX_CACHE_ATTEMPTS,)).fetchone()[0]
        retrying = conn.execute("SELECT COUNT(*) FROM cache_fail WHERE attempts < ?",
                                (_MAX_CACHE_ATTEMPTS,)).fetchone()[0]
    except Exception:
        genuine_failed = retrying = 0
    # How many CONFIRMED students have ALL their documents saved ("fully saved") vs still
    # missing at least one ("pending"). Computed against each student's expected doc set so a
    # student with 2 of 3 docs counts as pending, not done.
    import whatsapp
    confirmed = conn.execute("SELECT row_key, session, toc_status FROM student_status "
                             "WHERE is_confirmed=1 AND COALESCE(deleted,0)=0 "
                             "AND COALESCE(login_failed,0)=0 AND COALESCE(check_failed,0)=0").fetchall()
    have_map = {}
    for x in conn.execute("SELECT row_key, kind FROM document_cache").fetchall():
        have_map.setdefault(x["row_key"], set()).add(x["kind"])
    conn.close()
    total_conf = len(confirmed)
    fully = 0
    for s in confirmed:
        allowed = whatsapp.allowed_docs(s["session"], (s["toc_status"] or ""))
        if not allowed or allowed.issubset(have_map.get(s["row_key"], set())):
            fully += 1
    pending_docs = max(0, total_conf - fully)
    active = (get_setting("cache_docs_active", "0") == "1")
    return {"progress": CACHE_PROGRESS, "total_cached": total_cached, "students_cached": students_cached,
            "genuine_failed": genuine_failed, "retrying": retrying, "auto_resume": active,
            "fully_saved": fully, "confirmed_total": total_conf, "pending_docs": pending_docs}

@app.post("/api/wa-resend")
async def wa_resend(body: dict, user=Depends(verify_token)):
    """Manual (re)send for ONE student. Runs the send in a worker thread (so the event loop
    never blocks) and returns the ACTUAL result, so the counsellor immediately sees whether
    it went through or exactly why it failed (e.g. 'no campaign set for public')."""
    row_key = body.get("row_key", "")
    if not row_key:
        raise HTTPException(status_code=400, detail="row_key required")
    from fastapi.concurrency import run_in_threadpool
    # A manual resend ALWAYS re-reads the TOC from NIOS first (one captcha) and applies/pushes any
    # correction — so a resend can never repeat a wrong campaign, which is the whole point of it.
    ok, info, lf = await run_in_threadpool(_wa_send_one, row_key, False, True)
    return {"ok": ok, "info": info, "login_failed": lf}

@app.post("/api/wa-resend-bulk")
async def wa_resend_bulk(body: dict, background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """Resend WhatsApp to MANY selected students at once (all in the background)."""
    keys = body.get("row_keys", []) or []
    keys = [k for k in keys if k][:200]
    if not keys:
        raise HTTPException(status_code=400, detail="no students selected")
    def _run(klist):
        for rk in klist:
            _wa_send_one(rk)
    background_tasks.add_task(_run, keys)
    return {"ok": True, "queued": len(keys),
            "message": f"Resending to {len(keys)} student(s) in the background"}

@app.post("/api/wa-autosend-now")
async def wa_autosend_now(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """'Send all pending' — sends to ALL confirmed students that are still 'Not sent yet'
    (in the background, with live progress shown on the Confirmed page)."""
    if WA_PROGRESS["running"]:
        return {"ok": True, "already_running": True, "message": "A send is already running"}
    background_tasks.add_task(auto_send_pending_whatsapp, None, "Send all pending")
    return {"ok": True, "message": "Sending to all pending students…"}

def _resend_notoc_worker(keys, label="Resend no-TOC corrected docs"):
    """Background worker: re-send to the given confirmed no-TOC students, with WA_PROGRESS so the
    Confirmed-page progress bar shows live status. Each send now uses the correct no-TOC campaign."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    WA_PROGRESS.update({"running": True, "total": len(keys), "done": 0, "sent": 0,
                        "failed": 0, "started_at": now, "finished_at": "", "label": label})
    try:
        for rk in keys:
            ok, info, _lf = _wa_send_one(rk)
            WA_PROGRESS["done"] += 1
            WA_PROGRESS["sent" if ok else "failed"] += 1
    finally:
        WA_PROGRESS["running"] = False
        WA_PROGRESS["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

@app.post("/api/wa-resend-notoc")
def wa_resend_notoc(background_tasks: BackgroundTasks, user=Depends(verify_token)):
    """One-time correction: re-send the CORRECT documents to every CONFIRMED student whose
    tocStatus is 'no'. They previously received the wrong (TOC-style) documents. The sent flag is
    reset so the message goes again — now via the no-TOC campaign with the right documents — while
    their old document links auto-block at serve time. New confirmations already use the right
    campaign automatically, so this is only needed for students sent before the TOC fix."""
    if WA_PROGRESS["running"]:
        return {"ok": True, "already_running": True, "message": "A send is already running"}
    conn = get_db()
    rows = conn.execute(
        "SELECT row_key FROM student_status WHERE is_confirmed=1 "
        "AND " + _SPECIAL_TOC_CLAUSE + " AND COALESCE(deleted,0)=0 "
        "AND mobile IS NOT NULL AND TRIM(mobile) != ''").fetchall()
    keys = [r["row_key"] for r in rows]
    if keys:
        conn.executemany("UPDATE student_status SET whatsapp_sent=0, whatsapp_attempts=0, "
                         "whatsapp_delivery='' WHERE row_key=?", [(k,) for k in keys])
        conn.commit()
    conn.close()
    if not keys:
        return {"ok": True, "queued": 0, "message": "No confirmed special-TOC students found to resend."}
    background_tasks.add_task(_resend_notoc_worker, keys)
    return {"ok": True, "queued": len(keys),
            "message": f"Resending corrected documents to {len(keys)} no-TOC confirmed student(s)…"}

@app.get("/api/wa-progress")
async def wa_progress(user=Depends(verify_token)):
    """Live progress of the WhatsApp 'send all pending' / auto-sweep run."""
    p = dict(WA_PROGRESS)
    p["remaining"] = max(0, p.get("total", 0) - p.get("done", 0))
    p["pct"] = int(round(100 * p.get("done", 0) / p["total"])) if p.get("total") else (100 if p.get("finished_at") else 0)
    return p

@app.get("/api/debug-doc")
def debug_doc(ref: str, dob: str, kind: str = "app_form", user=Depends(verify_token)):
    """Inspect how a document page embeds & sizes its images (photo/signature/QR),
    so we can match NIOS exactly. Returns image pixel sizes, EXIF orientation & CSS."""
    try:
        from nios_login import inspect_doc_page
        return inspect_doc_page(ref, dob, kind)
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/debug-idcard")
def debug_idcard(ref: str, dob: str, user=Depends(verify_token)):
    """Show the ID card's visible text + best-effort Regional Centre address
    (used to finalise the Stream 2 address parser)."""
    try:
        from nios_login import debug_idcard_text
        return debug_idcard_text(ref, dob)
    except Exception as e:
        return {"error": str(e)}

# ─────────────────────────────────────────────────────────────────────────────
# Danger zone — wipe all student data for a fresh upload (keeps settings)
# ─────────────────────────────────────────────────────────────────────────────
@app.post("/api/reset-data")
def reset_data(body: dict, user=Depends(verify_token)):
    conn = get_db()
    conn.execute("DELETE FROM student_status")
    conn.execute("DELETE FROM status_history")
    conn.execute("DELETE FROM run_logs")
    try:
        conn.execute("DELETE FROM short_links")
    except Exception:
        pass
    conn.commit()
    conn.close()
    removed = False
    try:
        if os.path.exists(EXCEL_PATH):
            os.remove(EXCEL_PATH)
            removed = True
    except Exception:
        pass
    logger.info("All student data cleared via reset endpoint")
    return {"message": "All data has been cleared.", "excel_removed": removed}

# ─────────────────────────────────────────────────────────────────────────────
# Public document links (student opens WITHOUT portal login; token-signed)
# ─────────────────────────────────────────────────────────────────────────────
def _allowed_kinds(session):
    s = (session or "").lower()
    if "stream 2" in s or "stream2" in s or "stream-2" in s:
        return [("id_card", "ID Card"), ("app_form", "Application Form")]
    if "on demand" in s or "ondemand" in s or "on-demand" in s:
        return [("id_card", "ID Card"), ("app_form", "Application Form"), ("hall_ticket", "Hall Ticket")]
    if "april" in s or "october" in s or "public" in s:
        return [("id_card", "ID Card")]
    return [("id_card", "ID Card"), ("app_form", "Application Form"), ("hall_ticket", "Hall Ticket")]

DOC_PAGE_TPL = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Your NIOS Documents — MVS Foundation</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif}
body{background:linear-gradient(135deg,#4F46E5,#7C3AED);min-height:100vh;display:flex;align-items:center;justify-content:center;padding:18px}
.card{background:#fff;border-radius:20px;max-width:440px;width:100%;padding:30px 26px;box-shadow:0 20px 60px rgba(0,0,0,.3)}
.logo{width:58px;height:58px;border-radius:14px;background:linear-gradient(135deg,#4F46E5,#7C3AED);color:#fff;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:20px;margin:0 auto 14px}
h1{font-size:19px;text-align:center;color:#0F172A}
.sub{text-align:center;color:#64748B;font-size:13px;margin:6px 0 22px}
.name{text-align:center;font-weight:700;color:#4F46E5;font-size:16px;margin-bottom:2px}
.btns{display:flex;flex-direction:column;gap:12px;margin-top:8px}
.docbtn{display:flex;align-items:center;gap:12px;padding:15px 18px;border:2px solid #E2E8F0;border-radius:14px;
  color:#0F172A;font-weight:600;font-size:15px;cursor:pointer;background:#F8FAFC;transition:.15s;text-decoration:none}
.docbtn:hover{border-color:#4F46E5;background:#EEF2FF}
.docbtn .ico{width:34px;height:34px;border-radius:9px;background:#4F46E5;color:#fff;display:flex;align-items:center;justify-content:center;flex-shrink:0;font-size:17px}
.note{margin-top:18px;background:#FEF9C3;border:1px solid #FDE68A;border-radius:12px;padding:11px 14px;font-size:12.5px;color:#854D0E}
.foot{text-align:center;color:#94A3B8;font-size:11.5px;margin-top:18px}
.busy{opacity:.55;pointer-events:none}
</style></head><body>
<div class="card">
  <img src="/logo.png" alt="MVS Foundation" style="width:60px;height:60px;border-radius:50%;object-fit:cover;display:block;margin:0 auto 14px">
  <h1>Your NIOS Documents</h1>
  <div class="sub">Admission Confirmed</div>
  <div class="name">__NAME__</div>
  <div class="sub" style="margin-top:0">Ref: __REF__</div>
  <div class="btns">__BUTTONS__</div>
  <div class="note">&#128161; It may take a few seconds to open each document (securely fetched from NIOS). Once it opens, tap <b>"Save as PDF / Print"</b> at the top to save it.</div>
  <div class="foot">MVS Foundation &middot; NIOS Open Schooling</div>
</div>
<script>
function openDoc(el,url){ if(el.dataset.b==='1')return; el.dataset.b='1';
  var t=el.querySelector('.lbl'); var o=t.textContent; t.textContent='Opening...';
  el.classList.add('busy'); window.open(url,'_blank');
  setTimeout(function(){t.textContent=o;el.classList.remove('busy');el.dataset.b='';},16000); }
</script>
</body></html>"""

@app.get("/d/{token}", response_class=HTMLResponse)
def public_doc_page(token: str):
    from links import verify_doc_token
    row_key = verify_doc_token(token)
    if not row_key:
        return HTMLResponse("<h3 style='font-family:sans-serif;text-align:center;margin-top:60px'>Invalid or broken link</h3>", status_code=404)
    conn = get_db()
    row = conn.execute("SELECT student_name, session, reference_no FROM student_status WHERE row_key=?",
                       (row_key,)).fetchone()
    conn.close()
    if not row:
        return HTMLResponse("<h3 style='font-family:sans-serif;text-align:center;margin-top:60px'>Student not found</h3>", status_code=404)
    btns = ""
    for kind, label in _allowed_kinds(row["session"]):
        url = f"/d/{token}/{kind}"
        btns += (f'<a class="docbtn" onclick="openDoc(this,\'{url}\')">'
                 f'<span class="ico">&#128196;</span><span class="lbl">{label}</span></a>')
    html = (DOC_PAGE_TPL
            .replace("__NAME__", (row["student_name"] or "Student"))
            .replace("__REF__", (row["reference_no"] or "—"))
            .replace("__BUTTONS__", btns))
    return HTMLResponse(html)

@app.get("/d/{token}/{kind}")
def public_doc_file(token: str, kind: str):
    from fastapi import Response
    from links import verify_doc_token
    from nios_login import fetch_document
    row_key = verify_doc_token(token)
    if not row_key:
        raise HTTPException(status_code=404, detail="invalid link")
    conn = get_db()
    row = conn.execute("SELECT reference_no, dob, session, toc_status FROM student_status WHERE row_key=?",
                       (row_key,)).fetchone()
    conn.close()
    if not row or not row["reference_no"]:
        raise HTTPException(status_code=404, detail="student not found")
    import whatsapp
    toc_v = (row["toc_status"] if ("toc_status" in row.keys()) else "") or ""
    if not whatsapp.doc_allowed(row["session"], toc_v, kind):
        raise HTTPException(status_code=404, detail="This document is no longer available for your admission type. Please open the latest links we sent you on WhatsApp.")
    # Serve from OUR saved copy first (no NIOS/CapSolver); else fetch live + save.
    cached = load_doc_cache(row_key, kind)
    if cached is not None:
        content, ctype, filename = cached
    else:
        if not _LIVE_FETCH_SEM.acquire(timeout=15):
            raise HTTPException(status_code=503, detail="server is busy right now — please try again in a minute")
        try:
            content, ctype, filename = fetch_document(row["reference_no"], row["dob"], kind)
        finally:
            _LIVE_FETCH_SEM.release()
        if content is None:
            raise HTTPException(status_code=404, detail=ctype)
        save_doc_cache(row_key, kind, content, ctype, filename)
    disp = f'attachment; filename="{filename}"' if "pdf" in ctype else "inline"
    return Response(content=content, media_type=ctype, headers={"Content-Disposition": disp})

DOC_LABELS = {"id_card": "ID Card", "app_form": "Application / Registration Form",
              "hall_ticket": "Hall Ticket"}

LOADING_PAGE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Preparing your document — MVS Foundation</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif}
body{background:linear-gradient(135deg,#4F46E5,#7C3AED);min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}
.box{background:#fff;border-radius:20px;max-width:430px;width:100%;padding:40px 30px;text-align:center;box-shadow:0 20px 60px rgba(0,0,0,.3)}
.spinner{width:52px;height:52px;border:5px solid #EEF2FF;border-top-color:#4F46E5;border-radius:50%;margin:0 auto 22px;animation:spin 1s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
h1{font-size:19px;color:#0F172A;margin-bottom:8px}
.doc{color:#4F46E5;font-weight:700}
p{color:#64748B;font-size:14px;line-height:1.6;margin-top:6px}
.warn{margin-top:22px;background:#FEF3C7;border:1px solid #FDE68A;border-radius:12px;padding:12px 14px;color:#92400E;font-size:13px;font-weight:600}
.err{display:none;margin-top:18px;color:#DC2626;font-size:14px}
.retry{display:none;margin-top:14px;padding:11px 26px;border:none;border-radius:10px;background:#4F46E5;color:#fff;font-weight:600;font-size:15px;cursor:pointer}
</style></head><body>
<div class="box">
  <img src="/logo.png" alt="MVS Foundation" style="width:60px;height:60px;border-radius:50%;object-fit:cover;display:block;margin:0 auto 14px">
  <div class="spinner" id="sp"></div>
  <h1>Preparing your <span class="doc">__LABEL__</span></h1>
  <p>Please wait a few seconds…</p>
  <div class="warn">Please do not refresh this page or press the back button.</div>
  <div class="err" id="err">Sorry, the document could not be loaded. Please try again.</div>
  <button class="retry" id="retry" onclick="loadDoc()">Try Again</button>
</div>
<script>
function loadDoc(){
  document.getElementById('err').style.display='none';
  document.getElementById('retry').style.display='none';
  document.getElementById('sp').style.display='block';
  fetch('__PREP__').then(function(r){ return r.json(); })
  .then(function(d){
    if(d && d.ok){ window.location.replace('__VIEW__'); }
    else { throw new Error((d && d.error) || 'failed'); }
  }).catch(function(e){
    document.getElementById('sp').style.display='none';
    document.getElementById('err').style.display='block';
    document.getElementById('retry').style.display='inline-block';
  });
}
loadDoc();
</script></body></html>"""

# Short-lived in-memory cache so the document is fetched from NIOS once (during the
# loader's "prepare" step) and then served as a REAL page navigation at "view".
# Real navigation (not document.write) is what makes the browser's Save-as-PDF /
# Print work reliably on every device, including phones.
import time as _time
_DOC_CACHE = {}
_DOC_TTL = 600  # 10 minutes

def _cache_put(key, tup):
    _DOC_CACHE[key] = (_time.time(),) + tuple(tup)
    if len(_DOC_CACHE) > 150:
        now = _time.time()
        for k in [k for k, v in list(_DOC_CACHE.items()) if now - v[0] > _DOC_TTL]:
            _DOC_CACHE.pop(k, None)

def _cache_get(key):
    v = _DOC_CACHE.get(key)
    if not v:
        return None
    if _time.time() - v[0] > _DOC_TTL:
        _DOC_CACHE.pop(key, None)
        return None
    return v[1], v[2], v[3]

def _doc_cache_dir():
    try:
        from database import DATA_DIR as _DD
    except Exception:
        _DD = os.environ.get("DATA_DIR", ".")
    d = os.path.join(_DD, "doccache")
    os.makedirs(d, exist_ok=True)
    return d

try:
    _MAX_CACHE_ATTEMPTS = max(3, int(os.environ.get("CACHE_MAX_ATTEMPTS", "6")))
except Exception:
    _MAX_CACHE_ATTEMPTS = 6

def _record_doc_fail(row_key, kind, err):
    """Bump the failure counter for a document (used to tell transient reCAPTCHA misses from a
    genuine data problem)."""
    try:
        with _DB_WRITE_LOCK:
            conn = get_db()
            conn.execute("INSERT INTO cache_fail (row_key, kind, attempts, last_error, updated_at) "
                         "VALUES (?,?,1,?,?) ON CONFLICT(row_key,kind) DO UPDATE SET "
                         "attempts=attempts+1, last_error=excluded.last_error, updated_at=excluded.updated_at",
                         (row_key, kind, str(err)[:200], datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit(); conn.close()
    except Exception:
        pass

def _clear_doc_fail(row_key, kind):
    try:
        with _DB_WRITE_LOCK:
            conn = get_db()
            conn.execute("DELETE FROM cache_fail WHERE row_key=? AND kind=?", (row_key, kind))
            conn.commit(); conn.close()
    except Exception:
        pass

def _doc_fail_attempts(row_key, kind):
    try:
        conn = get_db()
        r = conn.execute("SELECT attempts FROM cache_fail WHERE row_key=? AND kind=?",
                         (row_key, kind)).fetchone()
        conn.close()
        return (r["attempts"] if r else 0)
    except Exception:
        return 0

def save_doc_cache(row_key, kind, content, ctype, filename):
    """Persist a fetched document to disk + record it, so future opens serve from OUR copy."""
    import re as _re
    import hashlib as _hl
    try:
        ext = "pdf" if "pdf" in (ctype or "").lower() else "html"
        safe = _re.sub(r'[^A-Za-z0-9_.-]', '_', str(row_key))[:80]
        h = _hl.md5(str(row_key).encode("utf-8")).hexdigest()[:10]   # unique per student -> never collide
        path = os.path.join(_doc_cache_dir(), f"{safe}_{h}__{kind}.{ext}")
        with open(path, "wb") as f:
            f.write(content)
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO document_cache "
                     "(row_key, kind, file_path, content_type, filename, size, fetched_at) "
                     "VALUES (?,?,?,?,?,?,?)",
                     (row_key, kind, path, ctype, filename, len(content or b""),
                      datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit(); conn.close()
        _clear_doc_fail(row_key, kind)   # success — forget any prior failure
        return True
    except Exception as e:
        logger.warning(f"save_doc_cache failed {row_key}/{kind}: {e}")
        return False

def load_doc_cache(row_key, kind):
    """Return (content_bytes, content_type, filename) from our saved copy, or None."""
    try:
        conn = get_db()
        row = conn.execute("SELECT file_path, content_type, filename FROM document_cache "
                           "WHERE row_key=? AND kind=?", (row_key, kind)).fetchone()
        conn.close()
        if not row or not row["file_path"]:
            return None
        with open(row["file_path"], "rb") as f:
            data = f.read()
        if not data:
            return None
        return data, (row["content_type"] or "application/pdf"), (row["filename"] or f"{kind}.pdf")
    except Exception:
        return None

def docs_all_cached(row_key):
    """True only if EVERY document this student is supposed to have is already saved in our DB.
    Used to gate the WhatsApp/Portal link so the student's first tap opens instantly from our
    copy (no live NIOS fetch, no 'preparing' spinner, no panic)."""
    try:
        import whatsapp
        conn = get_db()
        r = conn.execute("SELECT session, toc_status FROM student_status WHERE row_key=?", (row_key,)).fetchone()
        if not r:
            conn.close(); return False
        allowed = whatsapp.allowed_docs(r["session"], (r["toc_status"] or ""))
        if not allowed:
            conn.close(); return True
        have = {row["kind"] for row in
                conn.execute("SELECT kind FROM document_cache WHERE row_key=?", (row_key,)).fetchall()}
        conn.close()
        return allowed.issubset(have)
    except Exception:
        return False

def cache_student_docs(row_key, force=False):
    """Fetch + save ALL allowed documents for one confirmed student. force=True re-fetches even
    if already saved (use when a student's documents changed on NIOS). Returns (saved, failed)."""
    try:
        import whatsapp
        from nios_login import fetch_document
        conn = get_db()
        r = conn.execute("SELECT reference_no, enrollment_no, dob, session, toc_status "
                         "FROM student_status WHERE row_key=? AND COALESCE(deleted,0)=0", (row_key,)).fetchone()
        conn.close()
        if not r:
            return (0, 0)
        allowed = whatsapp.allowed_docs(r["session"], (r["toc_status"] or ""))
        saved = failed = 0
        for kind in allowed:
            if (not force) and load_doc_cache(row_key, kind) is not None:
                continue
            content, ctype, filename = fetch_document(r["reference_no"] or "", r["dob"] or "", kind,
                                                      enrollment_no=(r["enrollment_no"] or ""))
            if content is not None:
                save_doc_cache(row_key, kind, content, ctype, filename)
                saved += 1
            else:
                failed += 1
        return (saved, failed)
    except Exception as e:
        logger.warning(f"cache_student_docs {row_key}: {e}")
        return (0, 0)

_CACHE_SEM = _threading.Semaphore(2)   # cap background auto-caching so runs stay light

def _bg_cache_student(row_key):
    """Best-effort background cache of one student's docs (used right after a confirmation send).
    Skips silently if two are already caching — the periodic 'Save all' / lazy-on-click covers the rest."""
    if not _CACHE_SEM.acquire(blocking=False):
        return
    try:
        cache_student_docs(row_key)
    finally:
        _CACHE_SEM.release()

import threading as _threading
# Max simultaneous LIVE NIOS fetches from student links. Cache hits are unlimited.
# Prevents a click-stampede from exhausting the threadpool / CapSolver credits.
_LIVE_FETCH_SEM = _threading.BoundedSemaphore(4)

def _fetch_doc_for(row_key, kind):
    """Return ((content, ctype, filename), None) or (None, error_message).
    Never raises — a NIOS/network hiccup becomes a clean error so the student
    sees a 'Try Again' instead of a server crash."""
    try:
        from nios_login import fetch_document
        conn = get_db()
        row = conn.execute("SELECT reference_no, enrollment_no, dob, session, toc_status FROM student_status WHERE row_key=?",
                           (row_key,)).fetchone()
        conn.close()
        ref = (row["reference_no"] if row else "") or ""
        enroll = (row["enrollment_no"] if row else "") or ""
        if not row or (not ref and not enroll):
            return None, "student not found"
        # Serve-time permission: only the documents that match this student's session + tocStatus
        # may be opened. This auto-blocks any OLD/WRONG link sent before (e.g. an Application Form
        # link a no-TOC student received) — only the correct, current documents stay downloadable.
        import whatsapp
        toc_v = (row["toc_status"] if ("toc_status" in row.keys()) else "") or ""
        if not whatsapp.doc_allowed(row["session"], toc_v, kind):
            return None, ("This document is no longer available for your admission type. "
                          "Please open the latest links we sent you on WhatsApp.")
        # 1) Serve from OUR saved copy first — no NIOS/CapSolver, works even if NIOS is down.
        cached = load_doc_cache(row_key, kind)
        if cached is not None:
            return cached, None
        # 2) Not saved yet -> fetch live from NIOS, then save for next time.
        #    Bounded: at most 4 live NIOS fetches at once; others get a clean retry message.
        if not _LIVE_FETCH_SEM.acquire(timeout=15):
            return None, "server is busy right now — please try again in a minute"
        try:
            content, ctype, filename = fetch_document(ref, row["dob"], kind, enrollment_no=enroll)
        finally:
            _LIVE_FETCH_SEM.release()
        if content is None:
            return None, (ctype or "could not load document")
        save_doc_cache(row_key, kind, content, ctype, filename)
        return (content, ctype, filename), None
    except Exception as e:
        logger.warning(f"doc fetch failed for {kind}: {e}")
        return None, "could not load document"

def _serve_doc(content, ctype, filename):
    from fastapi import Response
    # inline -> the browser renders it as a normal page so Print / Save-as-PDF works
    return Response(content=content, media_type=ctype,
                    headers={"Content-Disposition": "inline"})

def _loader_html(kind, prep_url, view_url):
    label = DOC_LABELS.get(kind, "Document")
    return (LOADING_PAGE.replace("__LABEL__", label)
            .replace("__PREP__", prep_url).replace("__VIEW__", view_url))

def _invalid_link_html():
    return HTMLResponse("<h3 style='font-family:sans-serif;text-align:center;margin-top:60px'>"
                        "This link is invalid or has expired.</h3>", status_code=404)

# ── Long signed links: /doc/{token} ──
@app.get("/doc/{token}", response_class=HTMLResponse)
def public_single_doc(token: str):
    from links import verify_doc_link
    row_key, kind = verify_doc_link(token)
    if not row_key:
        return _invalid_link_html()
    # FAST PATH: already saved on our server -> serve instantly. No loader page,
    # no prepare round-trip, no NIOS, no CapSolver. One click = document opens.
    cached = load_doc_cache(row_key, kind)
    if cached is not None:
        return _serve_doc(*cached)
    return HTMLResponse(_loader_html(kind, f"/doc/{token}/prepare", f"/doc/{token}/view"))

@app.get("/doc/{token}/prepare")
def public_doc_prepare(token: str):
    from links import verify_doc_link
    row_key, kind = verify_doc_link(token)
    if not row_key:
        return {"ok": False, "error": "invalid link"}
    res, err = _fetch_doc_for(row_key, kind)
    if err:
        return {"ok": False, "error": err}
    _cache_put("doc:" + token, res)
    return {"ok": True}

@app.get("/doc/{token}/view")
def public_doc_view(token: str):
    cached = _cache_get("doc:" + token)
    if not cached:
        from links import verify_doc_link
        row_key, kind = verify_doc_link(token)
        if not row_key:
            raise HTTPException(status_code=404, detail="invalid link")
        res, err = _fetch_doc_for(row_key, kind)
        if err:
            raise HTTPException(status_code=404, detail=err)
        cached = res
    return _serve_doc(*cached)

# ── Short links: /s/{code} ──
@app.get("/s/{code}", response_class=HTMLResponse)
def short_doc(code: str):
    from shortlinks import resolve_short
    row_key, kind = resolve_short(code)
    if not row_key:
        return _invalid_link_html()
    # FAST PATH: already saved on our server -> serve instantly. No loader page,
    # no prepare round-trip, no NIOS, no CapSolver. One click = document opens.
    cached = load_doc_cache(row_key, kind)
    if cached is not None:
        return _serve_doc(*cached)
    return HTMLResponse(_loader_html(kind, f"/s/{code}/prepare", f"/s/{code}/view"))

@app.get("/s/{code}/prepare")
def short_doc_prepare(code: str):
    from shortlinks import resolve_short
    row_key, kind = resolve_short(code)
    if not row_key:
        return {"ok": False, "error": "invalid link"}
    res, err = _fetch_doc_for(row_key, kind)
    if err:
        return {"ok": False, "error": err}
    _cache_put("s:" + code, res)
    return {"ok": True}

@app.get("/s/{code}/view")
def short_doc_view(code: str):
    cached = _cache_get("s:" + code)
    if not cached:
        from shortlinks import resolve_short
        row_key, kind = resolve_short(code)
        if not row_key:
            raise HTTPException(status_code=404, detail="invalid link")
        res, err = _fetch_doc_for(row_key, kind)
        if err:
            raise HTTPException(status_code=404, detail=err)
        cached = res
    return _serve_doc(*cached)

@app.get("/logo.png")
async def logo_png():
    """MVS Foundation logo (embedded, served for portal + student pages)."""
    import base64
    from fastapi import Response
    try:
        from assets import LOGO_B64
        return Response(content=base64.b64decode(LOGO_B64), media_type="image/png",
                        headers={"Cache-Control": "public, max-age=86400"})
    except Exception:
        raise HTTPException(status_code=404, detail="logo not found")
